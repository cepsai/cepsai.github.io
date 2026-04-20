#!/usr/bin/env python3
"""
build_v15.py — emits digital_lexicon_v15.html on top of v14 (which itself
patches digital_lexicon_v12_draft.html).

v15 changes vs v14 (driven by the user's review-round feedback):

  1. Cluster summary table — replaces the stack of per-subconcept dimension
     tables on each cluster page with a single "term map" table whose rows
     are sub-concepts and whose cells are per-jurisdiction term variants
     (matches the user's reference image). Click a cell → drawer with
     verbatim legal text.

  2. Per-concept dimension table — when the user drills into a subtab, the
     dimensions are now sourced from the ANALYSIS sheet (which has the
     authoritative breakdown like "Term", "Regulatory trigger", "Compute
     threshold") instead of the legal sheet. This eliminates the silent
     fallback-to-verbatim that was corrupting Risk / GPAI / Frontier /
     Foundation tables in v14.

  3. Interpretative notes — read with rich_text=True so bold runs and
     line breaks from Excel survive into the HTML. Rendered once per
     cluster, beneath the summary table.

  4. Term names — column headers in the per-concept dimension table now
     show the jurisdiction-specific variant (e.g. "Intentional and
     substantial modification" for Colorado) above the citation.

  5. Landing + Methodology pages — sourced from the Excel "About the
     Digital AI Lexicon" and "Methodology" sheets instead of hardcoded
     copy. Rich-text bold + paragraph breaks preserved; the methodology
     regulations table is rendered as an HTML table.

  6. Font size — base bumped 15px → 16px.

v14 stays untouched: re-running build_v14.py regenerates v14 unchanged.

Usage:
    python3 build_v15.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

# Reuse v14's helpers + v13's parsers.
sys.path.insert(0, str(Path(__file__).parent))
import importlib.util

_spec14 = importlib.util.spec_from_file_location("build_v14", Path(__file__).parent / "build_v14.py")
_v14 = importlib.util.module_from_spec(_spec14)
_spec14.loader.exec_module(_v14)

_v13 = _v14._v13  # already loaded by v14

SHEET_TO_TAB           = _v13.SHEET_TO_TAB
SHEET_LAYOUT           = _v13.SHEET_LAYOUT
ANALYSIS_SHEET_FOR     = _v13.ANALYSIS_SHEET_FOR
parse_legal_sheet      = _v13.parse_legal_sheet
parse_analysis_sheet   = _v13.parse_analysis_sheet
parse_matrix           = _v13.parse_matrix
parse_glossary         = _v13.parse_glossary
load_laws              = _v13.load_laws
build_ref_map          = _v13.build_ref_map
extract_json_literal_span = _v13.extract_json_literal_span
LAW_URLS               = _v13.LAW_URLS
JID_LABEL              = _v13.JID_LABEL
clean                  = _v13.clean
clean_preserve         = _v13.clean_preserve
cell_runs              = _v13.cell_runs

CLUSTER_FOR_TAB = _v14.CLUSTER_FOR_TAB
JID_LAW_LABEL   = _v14.JID_LAW_LABEL
LAW_METADATA    = _v14.LAW_METADATA
build_laws_metadata = _v14.build_laws_metadata
slugify         = _v14.slugify
replace_const   = _v14.replace_const

HERE = Path(__file__).parent
HTML_IN = HERE / "digital_lexicon_v12_draft.html"
XLSX    = HERE / "AI terminology and taxonomy-5.xlsx"
HTML_OUT = HERE / "digital_lexicon_v15.html"


# -------- bill-code extraction (used in cluster summary cells) --------

# Map a reference string to the short bill / article code shown in the cluster
# summary pills. Lifted from the user's reference image: "Developer (SB24-205)",
# "Frontier developer (SB 53)", etc.
_BILL_PATTERNS = [
    (re.compile(r"\b(SB\s?\d+[-\d]*)", re.I),     lambda m: m.group(1).replace(" ", "")),
    (re.compile(r"\b(AB\s?\d+[-\d]*)", re.I),     lambda m: m.group(1).replace(" ", "")),
    (re.compile(r"\b(HB\s?\d+[-\d]*)", re.I),     lambda m: m.group(1).replace(" ", "")),
    (re.compile(r"\b(S\s?\d{4,5}[A-Z]?)\b"),      lambda m: m.group(1).replace(" ", "")),
    (re.compile(r"\b(A\s?\d{4,5}[A-Z]?)\b"),      lambda m: m.group(1).replace(" ", "")),
    (re.compile(r"\b(AIA|EU AI Act|AI Act)\s+(?:Article|Art\.?|Recital|Annex)\s*([IVX0-9]+)", re.I),
                                                    lambda m: f"AIA Art. {m.group(2)}"),
    (re.compile(r"\bArticle\s+(\d+[a-z]?)", re.I), lambda m: f"Art. {m.group(1)}"),
]


def bill_code_from_ref(ref: str) -> str:
    """Extract a short bill/article code from a reference string. Returns '' if
    nothing recognisable found."""
    if not ref:
        return ""
    for pat, fmt in _BILL_PATTERNS:
        m = pat.search(ref)
        if m:
            return fmt(m)
    return ""


# -------- "New concepts" matrix → cluster summary rows --------

# The 6 term families from the "New concepts" matrix tab. The lexicon tabs
# now follow this structure (was 8 tabs in v14; "Model / system" combines
# the three legal sheets the v14 lexicon had as separate tabs, matching the
# Excel curator's grouping).
MATRIX_FAMILIES = [
    {"family": "Provider / Developer",  "cluster": "Governance",
     "tab_id": "provider-developer",
     "legal_sheets": ["Provider_Developer"]},
    {"family": "Deployer / Supplier",   "cluster": "Governance",
     "tab_id": "deployer-supplier",
     "legal_sheets": ["Deployer_Supplier"]},
    {"family": "Model / system",        "cluster": "Technical system attributes",
     "tab_id": "model-system",
     "legal_sheets": [" High-risk AI system",
                      "GPAI_Frontier_Foundation model",
                      "GPAI system_Generative AI"]},
    {"family": "Risk",                  "cluster": "Measurement",
     "tab_id": "risk",
     "legal_sheets": ["Risk"]},
    {"family": "Modification",          "cluster": "Measurement",
     "tab_id": "modification",
     "legal_sheets": ["Substantial modification"]},
    {"family": "Incident",              "cluster": "Trustworthy",
     "tab_id": "incident",
     "legal_sheets": ["Incident"]},
]

_PILL_RE = re.compile(r"^(.+?)\s*\(([^)]+)\)\s*$")


def _parse_pill_string(s: str) -> dict:
    """Split a New-concepts cell like 'Frontier developer (SB 53)' into
    {name, bill}. Bill stays empty if no parens found."""
    s = (s or "").strip()
    if not s or s in ("-", "–", "—"):
        return {}
    m = _PILL_RE.match(s)
    if m:
        return {"name": m.group(1).strip(), "bill": m.group(2).strip()}
    return {"name": s, "bill": ""}


def cluster_summary_from_matrix(matrix: dict, family: str, canonical_filter: str = None) -> dict:
    """Build the cluster summary table for one concept tab from the parsed
    'New concepts' matrix. Returns a dict shaped like the v15 cluster_summary
    (headers + rows of pill stacks). Returns None if nothing to render."""
    if not matrix or not matrix.get("rows"):
        return None
    if not family:
        return None

    # Matrix headers (after Cluster, Term): EU AI Act, California, Colorado, New York, Texas, Utah.
    headers_full = matrix.get("headers", [])
    juris_headers = headers_full[2:8] if len(headers_full) >= 8 else headers_full[2:]
    juris_jids = ["eu", "ca", "co", "ny", "tx", "ut"][: len(juris_headers)]

    # Group rows: a "group" starts when EU column has a value; continuation rows
    # (EU empty) belong to the same group and add additional variants.
    # parse_matrix only sets `term` on the first row of each family — subsequent
    # rows inherit. Forward-fill here.
    groups: list[dict] = []
    current = None
    current_term_family = ""
    for row in matrix["rows"]:
        row_term = (row.get("term") or "").strip()
        if row_term:
            current_term_family = row_term
        if current_term_family != family:
            current = None
            continue
        cells = row.get("cells", [])
        eu_val = (cells[0] or "").strip() if len(cells) > 0 else ""
        eu_pill = _parse_pill_string(eu_val) if eu_val and eu_val != "-" else {}
        if eu_pill:
            current = {
                "canonical": eu_pill["name"],
                "rows":      [cells],   # list of cell-arrays (one per matrix row)
            }
            groups.append(current)
        elif current is not None:
            current["rows"].append(cells)

    # Optional canonical-term filter (e.g. limit "Model / system" family to
    # one of GPAI model / GPAI system / High-risk AI system).
    if canonical_filter:
        wanted = canonical_filter.strip().lower()
        groups = [g for g in groups if g["canonical"].strip().lower() == wanted
                  or wanted in g["canonical"].strip().lower()]

    if not groups:
        return None

    # Each matrix row becomes its own cluster-summary row, mirroring the
    # Excel layout. Cells get a `rowspan` field so the renderer can merge
    # cells that are empty on continuation rows (e.g. the EU canonical
    # value spans the whole sub-concept block, like in Excel).
    out_rows: list[dict] = []
    for g in groups:
        n = len(g["rows"])
        # Rowspan only the EU AI Act column (first jurisdiction) — its value
        # is the canonical that applies to the whole group, and Excel renders
        # it merged across the group's continuation rows. Other columns
        # render each row's value as-is (so an empty cell stays "—" rather
        # than being merged with the previous row's variant).
        col_rowspans: list[list[int]] = [[1] * n for _ in juris_jids]
        if juris_jids:
            ci = 0
            anchor = None
            for ri in range(n):
                row_cells = g["rows"][ri]
                val = ""
                if ci < len(row_cells):
                    val = (row_cells[ci] or "").strip()
                if val and val != "-":
                    anchor = ri
                    col_rowspans[ci][ri] = 1
                else:
                    if anchor is not None:
                        col_rowspans[ci][anchor] += 1
                        col_rowspans[ci][ri] = 0
                    else:
                        col_rowspans[ci][ri] = 1
        for ri, row_cells in enumerate(g["rows"]):
            cells_by_jid: dict[str, dict] = {}
            for ci, jid in enumerate(juris_jids):
                rs = col_rowspans[ci][ri]
                pill = None
                if ci < len(row_cells):
                    pill = _parse_pill_string((row_cells[ci] or "").strip())
                cells_by_jid[jid] = {
                    "rowspan": rs,
                    "variants": [pill] if pill else [],
                }
            out_rows.append({
                "term_label":   g["canonical"] if ri == 0 else "",
                "term_rowspan": n if ri == 0 else 0,
                "sub_id":       "",      # filled in by caller
                "cells":        cells_by_jid,
            })

    return {
        "headers": [{"jid": j, "label": h} for j, h in zip(juris_jids, juris_headers)],
        "rows":    out_rows,
    }


# -------- per-concept variant extraction --------

# Entry labels from parse_legal_sheet take the form "<dim> — <variant>" where
# the suffix is the actual sub-variant (e.g. "Frontier developer" /
# "Large frontier developer"). Pull the variant slice.
_VARIANT_RE = re.compile(r"\s+[—-]\s+(.+?)\s*$")


def variant_from_label(label: str) -> str:
    if not label:
        return ""
    m = _VARIANT_RE.search(label)
    return m.group(1).strip() if m else ""


_SCOPE_DIMS = {"scope", "definition", "term"}

# Cell text in multi-bill jurisdictions arrives as "[SB942] X [AB2013] Y" —
# the analysis sheet's parse_analysis_sheet collapses both variants into one
# cell with bracketed prefixes. Split them so each variant gets its own
# column ("California (Covered provider — SB942)", etc.).
_BILL_PREFIX_RE = re.compile(r"\[([^\]\n]+?)\]\s*")


def split_bill_segments(text: str) -> list[tuple[str, str]]:
    """Return [(bill_code, segment_text), ...]. If `text` has no bracketed bill
    prefixes, returns [("", text)]. An empty segment is dropped."""
    if not text:
        return []
    parts = _BILL_PREFIX_RE.split(text)
    if len(parts) <= 1:
        s = text.strip()
        return [("", s)] if s else []
    out: list[tuple[str, str]] = []
    leading = (parts[0] or "").strip()
    if leading:
        out.append(("", leading))
    i = 1
    while i + 1 < len(parts):
        bill = parts[i].strip()
        seg  = (parts[i + 1] or "").strip()
        if seg or bill:
            out.append((bill, seg))
        i += 2
    return out


def collect_variants_for_jid(concept: dict, jid: str) -> list[dict]:
    """Walk a concept's SCOPE-like entries (Scope / Definition / Term rows)
    and produce a list of distinct variant pills for one jurisdiction. Each
    pill = {name, bill, ref}.

    A "variant" is a distinct human-readable term that this jurisdiction uses
    for this sub-concept (e.g. CA's "Frontier developer" + "Large frontier
    developer" both fall under the EU's "Provider of GPAI models with
    systemic risk"). Only Scope-like rows define real variants; Obligations
    /Penalties rows would otherwise leak dim sub-labels ("Transparency",
    "AI literacy") into the pill list."""
    fallback_name = concept.get("per_jid_names", {}).get(jid) or concept.get("name") or ""
    by_name: dict[str, dict] = {}

    for entry in concept.get("entries", []):
        rl = (entry.get("rowLabel") or "").strip().lower()
        if rl not in _SCOPE_DIMS:
            continue
        node = (entry.get("jdata") or {}).get(jid)
        if not node:
            continue
        ref = node.get("reference") or ""
        bill = bill_code_from_ref(ref)
        v_name = variant_from_label(node.get("label") or "")
        if not v_name or v_name.strip().lower() == rl:
            v_name = fallback_name
        key = v_name.strip().lower()
        slot = by_name.setdefault(key, {"name": v_name, "bills": [], "refs": []})
        if bill and bill not in slot["bills"]:
            slot["bills"].append(bill)
        if ref:
            slot["refs"].append(ref)

    # Edge case: jurisdiction is in per_jid_names but had no Scope-like row.
    if not by_name and fallback_name:
        # Try to harvest bills from any entry to give the pill a citation.
        bills_any: list[str] = []
        first_ref = ""
        for entry in concept.get("entries", []):
            node = (entry.get("jdata") or {}).get(jid)
            if not node:
                continue
            ref = node.get("reference") or ""
            b = bill_code_from_ref(ref)
            if b and b not in bills_any:
                bills_any.append(b)
            if ref and not first_ref:
                first_ref = ref
        if bills_any or first_ref:
            return [{"name": fallback_name, "bill": " / ".join(bills_any), "ref": first_ref}]
        return []

    return [
        {"name": s["name"],
         "bill": " / ".join(s["bills"]),
         "ref":  s["refs"][0] if s["refs"] else ""}
        for s in by_name.values()
    ]


# -------- analysis-sheet → dimensions (the v14 verbatim-mixup fix) --------

def build_sub_concepts_v15(concept: dict, tab_analyses: list, jid_order: list[str],
                            concept_index: int = -1) -> dict:
    """Return a single sub-concept dict with dim_list sourced from the
    analysis sheet (not the legal sheet)."""
    # Card matching: prefer positional (Excel keeps both lists in the same
    # top-to-bottom order). Fall back to substring match on section title.
    ana_card = None
    if 0 <= concept_index < len(tab_analyses):
        ana_card = tab_analyses[concept_index]
    if ana_card is None:
        for a in tab_analyses:
            sec = (a.get("section") or "").strip().lower()
            cn = (concept.get("name") or "").strip().lower()
            if sec == cn or (sec and cn and (sec in cn or cn in sec)):
                ana_card = a
                break
        if ana_card is None and len(tab_analyses) == 1:
            ana_card = tab_analyses[0]

    # Per-jurisdiction header data: term name + bill code(s).
    juris: dict[str, dict] = {}
    per_jid_names = concept.get("per_jid_names") or {}
    bills_seen: dict[str, set] = {}
    for entry in concept.get("entries", []):
        for jid, node in (entry.get("jdata") or {}).items():
            if not node:
                continue
            bill = bill_code_from_ref(node.get("reference") or "")
            if bill:
                bills_seen.setdefault(jid, set()).add(bill)
            if jid not in juris:
                juris[jid] = {
                    "term": per_jid_names.get(jid) or concept.get("name") or "",
                    "law":  JID_LAW_LABEL.get(jid, jid.upper()),
                }
    # Attach a comma-joined bill list to each jurisdiction header.
    for jid in juris:
        bills = sorted(bills_seen.get(jid, []))
        juris[jid]["bills"] = ", ".join(bills) if bills else ""

    # Build dim_list. Source: analysis card rows when present.
    dim_list: list[dict] = []

    # Index legal entries by (jid, normalized_dim) for verbatim lookup.
    legal_by_dim: dict[tuple[str, str], list[dict]] = {}
    for entry in concept.get("entries", []):
        d = (entry.get("rowLabel") or "").strip().lower()
        for jid, node in (entry.get("jdata") or {}).items():
            if node:
                legal_by_dim.setdefault((jid, d), []).append(node)

    # Aliases bridge the analysis-sheet labels to legal-sheet labels when
    # the concepts are equivalent. Keep this small and surgical.
    DIM_ALIASES = {
        "exemptions": "exclusions",
        "exemption":  "exclusions",
        "term":       "definition",
    }

    # Tab-wide verbatim fallback per jid: concat of ALL legal entries' text for
    # the jid. Used when a specific analysis dim has no matching legal row
    # (most concepts have a Scope row but no Term/Definition/Compute-threshold
    # row in the legal sheet, since those are analytical breakdowns).
    full_verbatim_by_jid: dict[str, str] = {}
    full_ref_by_jid: dict[str, str] = {}
    for entry in concept.get("entries", []):
        for jid, node in (entry.get("jdata") or {}).items():
            if not node or not node.get("text"):
                continue
            full_verbatim_by_jid.setdefault(jid, "")
            sep = "\n\n" if full_verbatim_by_jid[jid] else ""
            full_verbatim_by_jid[jid] += sep + node["text"]
            r = (node.get("reference") or "").strip()
            if r and r not in (full_ref_by_jid.get(jid) or ""):
                full_ref_by_jid[jid] = ((full_ref_by_jid.get(jid) or "") + ("; " if full_ref_by_jid.get(jid) else "") + r)

    if ana_card and ana_card.get("rows"):
        for ridx, row in enumerate(ana_card["rows"]):
            dim_label = (row.get("dim") or "").strip()
            if not dim_label:
                continue
            cells: dict[str, dict] = {}
            for jid in juris:
                analysis_text = row.get(jid) or ""
                # Find verbatim by trying exact dim match, then alias, then the
                # concept-wide fallback so the drawer always shows real legal text.
                key_a = dim_label.lower()
                key_b = DIM_ALIASES.get(key_a, key_a)
                nodes = legal_by_dim.get((jid, key_a)) or legal_by_dim.get((jid, key_b)) or []
                verbatim = "\n\n".join((n.get("text") or "") for n in nodes if n.get("text")).strip()
                reference = "; ".join((n.get("reference") or "") for n in nodes if n.get("reference")).strip("; ")
                if not verbatim:
                    verbatim = full_verbatim_by_jid.get(jid, "")
                    reference = reference or full_ref_by_jid.get(jid, "")
                if not analysis_text and not verbatim:
                    continue
                cells[jid] = {
                    "analysis":  analysis_text,
                    "verbatim":  verbatim,
                    "reference": reference,
                }
            if not cells:
                continue
            dim_list.append({
                "id":    slugify(dim_label)[:32] + f"-{ridx}",
                "label": dim_label,
                "cells": cells,
            })
    else:
        # No analysis card found — fall back to the legal-sheet rows. Merge
        # entries that share a rowLabel (e.g. multiple "Scope" rows) so we
        # don't render the same dim twice with different sub-aspects.
        merged: dict[str, dict] = {}
        order: list[str] = []
        for entry in concept.get("entries", []):
            dim_label = (entry.get("rowLabel") or "").strip()
            if not dim_label:
                continue
            key = dim_label.lower()
            slot = merged.setdefault(key, {"label": dim_label, "cells": {}})
            if key not in order:
                order.append(key)
            for jid, node in (entry.get("jdata") or {}).items():
                if not node:
                    continue
                prev = slot["cells"].get(jid)
                txt = node.get("text") or ""
                ref = node.get("reference") or ""
                if prev:
                    if txt and txt not in prev["verbatim"]:
                        prev["verbatim"] = (prev["verbatim"] + "\n\n" + txt).strip()
                    if ref and ref not in prev["reference"]:
                        prev["reference"] = (prev["reference"] + "; " + ref).strip("; ")
                else:
                    slot["cells"][jid] = {"analysis": "", "verbatim": txt, "reference": ref}
        for key in order:
            slot = merged[key]
            if slot["cells"]:
                dim_list.append({
                    "id":    slugify(slot["label"])[:32] + f"-{len(dim_list)}",
                    "label": slot["label"],
                    "cells": slot["cells"],
                })

    # Post-process: split multi-bill jurisdictions into per-variant lanes so
    # each variant ("Developer (AB2013)" vs "Covered provider (SB942)") gets
    # its own dimension-table column.
    juris, dim_list = _split_multi_bill_lanes(juris, dim_list)

    return {"jurisdictions": juris, "dimensions": dim_list}


def _split_multi_bill_lanes(juris: dict, dim_list: list) -> tuple[dict, list]:
    """Split jurisdictions whose analysis cells contain multiple variants
    (either `[BILLA] X [BILLB] Y` or `[SB 53] Frontier developer\\n\\n[SB 53]
    Large frontier developer`) into per-variant "lanes" — each lane is a
    separate jurisdiction key with its own term name + bill code, so
    `renderAnalysisTable` shows them as distinct columns.

    Strategy: the **Term** dim row defines the variant list per jid (one
    segment per variant). Subsequent rows' segments map positionally to
    those variants."""

    # Locate the "Term" row to use as the variant template.
    term_row = next(
        (d for d in dim_list if (d.get("label") or "").strip().lower() == "term"),
        None,
    )

    # Per jid, build the variant list: ordered [(bill, term_name), ...]
    variants_per_jid: dict[str, list[tuple[str, str]]] = {}
    if term_row:
        for jid, cell in (term_row.get("cells") or {}).items():
            if jid not in juris:
                continue
            segs = split_bill_segments(cell.get("analysis") or "")
            # Only treat as multi-variant if 2+ segments exist.
            if len(segs) >= 2:
                variants_per_jid[jid] = [(bill, seg) for bill, seg in segs if seg or bill]

    # Build lane map.
    new_juris: dict = {}
    jid_lane_map: dict[str, list[tuple[str, str, str]]] = {}
    for jid, jdata in juris.items():
        variants = variants_per_jid.get(jid, [])
        if len(variants) >= 2:
            lanes: list[tuple[str, str, str]] = []  # (lane_id, bill, term)
            for idx, (bill, term) in enumerate(variants):
                lane = f"{jid}-{idx}-{slugify(term or bill or str(idx))}"
                lanes.append((lane, bill, term))
                new_juris[lane] = {
                    "term":        term or jdata.get("term") or "",
                    "law":         jdata.get("law", jid.upper()),
                    "bills":       bill or jdata.get("bills", ""),
                    "_parent_jid": jid,
                }
            jid_lane_map[jid] = lanes
        else:
            new_juris[jid] = jdata
            jid_lane_map[jid] = [(jid, "", "")]

    # Rebuild each dim's cells against lane keys. Segments map POSITIONALLY
    # to the variants defined by the Term row.
    new_dim_list: list = []
    for dim in dim_list:
        new_cells: dict = {}
        for jid, cell in (dim.get("cells") or {}).items():
            lanes = jid_lane_map.get(jid, [(jid, "", "")])
            text      = cell.get("analysis") or ""
            verbatim  = cell.get("verbatim") or ""
            reference = cell.get("reference") or ""
            if len(lanes) == 1:
                lane_id, _, _ = lanes[0]
                # Single lane: strip any [BILL] prefixes for cleaner display.
                segments = split_bill_segments(text)
                if len(segments) <= 1:
                    new_cells[lane_id] = cell
                else:
                    new_cells[lane_id] = {
                        "analysis":  "\n\n".join(seg for _, seg in segments if seg),
                        "verbatim":  verbatim,
                        "reference": reference,
                    }
            else:
                segments = split_bill_segments(text)
                # Map positionally — segment index → lane index.
                for idx, (lane_id, bill, _) in enumerate(lanes):
                    seg = ""
                    if idx < len(segments):
                        s_bill, s_text = segments[idx]
                        # Sanity: if the segment's bill prefix matches expected,
                        # use its text. Otherwise still take it positionally.
                        seg = s_text
                    if seg or verbatim:
                        new_cells[lane_id] = {
                            "analysis":  seg,
                            "verbatim":  verbatim,
                            "reference": reference,
                        }
        new_dim_list.append({**dim, "cells": new_cells})

    return new_juris, new_dim_list


# -------- rich-text interpretative notes --------

def collect_rich_notes(wb_rt, sheet_name: str) -> list[dict]:
    """Walk the analysis sheet with rich-text loading and emit notes as
    arrays of {t, b} runs. Output: [{title, body_runs}], deduped by body
    plain-text."""
    if sheet_name not in wb_rt.sheetnames:
        return []
    ws = wb_rt[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    # Locate header rows + their notes column (mirrors _is_analysis_header_row).
    notes_locs: list[tuple[int, int, int]] = []  # (header_row, notes_col, end_row)
    header_rows: list[tuple[int, int]] = []
    for r in range(1, max_row + 1):
        notes_col = None
        jr_count = 0
        for c in range(1, max_col + 1):
            v = clean(ws.cell(r, c).value)
            if not v:
                continue
            if re.search(r"interpretative notes?", v, re.I):
                notes_col = c
            elif _v13.parse_jid_header(v):
                jr_count += 1
        if jr_count >= 2 and notes_col is not None:
            header_rows.append((r, notes_col))
    for i, (hr, nc) in enumerate(header_rows):
        end_r = (header_rows[i + 1][0] - 2) if i + 1 < len(header_rows) else max_row
        notes_locs.append((hr, nc, end_r))

    notes_out: list[dict] = []
    seen_keys: set[str] = set()
    for (hr, nc, end_r) in notes_locs:
        for r in range(hr + 1, end_r + 1):
            cell = ws.cell(r, nc)
            runs = cell_runs(cell)
            if not runs:
                continue
            plain = "".join(r["t"] for r in runs).strip()
            if not plain or plain in seen_keys:
                continue
            seen_keys.add(plain)
            # Title = first non-empty bold run, or the cell's row label, or a generic header.
            title = ""
            for r0 in runs:
                if r0["b"] and r0["t"].strip():
                    title = r0["t"].strip().lstrip("- \t").rstrip(":")
                    break
            if not title:
                title = clean(ws.cell(r, 1).value) or "Note"
            notes_out.append({"title": title, "body_runs": runs})
    return notes_out


# -------- About + Methodology prose extraction --------

def extract_prose_blocks(ws_rt, ws_plain) -> dict:
    """Read a prose sheet (e.g. About / Methodology). Returns:
        {"intro": [run_arr_per_paragraph],
         "table": {"headers": [...], "rows": [[cell, cell, ...]]} | None,
         "text":  "<full A1 text, plaintext>"}

    The intro comes from cell A1 (which holds the full prose as a multi-
    paragraph rich-text blob in this workbook). The optional table is
    detected by looking for a row with multiple bold header cells in the
    rest of the sheet, and reading subsequent rows. `text` is a plain-text
    shadow of A1 so the coverage test can substring-match without tripping
    on JSON keys interleaved between runs.
    """
    a1_runs = cell_runs(ws_rt.cell(1, 1))
    # Split into paragraphs at \n\n.
    paragraphs: list[list[dict]] = []
    if a1_runs:
        cur: list[dict] = []
        for run in a1_runs:
            parts = run["t"].split("\n\n")
            for i, p in enumerate(parts):
                if i > 0:
                    if cur:
                        paragraphs.append(cur)
                    cur = []
                if p:
                    cur.append({"t": p, "b": run["b"]})
        if cur:
            paragraphs.append(cur)

    # Find an inline table by scanning for a row with ≥3 bold header cells.
    table = None
    max_row = ws_plain.max_row
    max_col = ws_plain.max_column
    header_row = None
    for r in range(2, max_row + 1):
        bold_count = 0
        for c in range(1, max_col + 1):
            v = ws_plain.cell(r, c).value
            font = ws_plain.cell(r, c).font
            if v and font and font.b and clean(v):
                bold_count += 1
        if bold_count >= 3:
            header_row = r
            break
    if header_row:
        headers = []
        for c in range(1, max_col + 1):
            v = clean(ws_plain.cell(header_row, c).value)
            if v:
                headers.append(v)
            elif headers:
                break
        rows: list[list] = []
        carry: list[str] = ["" for _ in headers]
        for r in range(header_row + 1, max_row + 1):
            row_cells: list = []
            non_empty = False
            for ci in range(len(headers)):
                cell = ws_rt.cell(r, ci + 1)
                runs = cell_runs(cell)
                if not runs:
                    row_cells.append([])
                else:
                    non_empty = True
                    row_cells.append(runs)
            if non_empty:
                rows.append(row_cells)
        if rows:
            table = {"headers": headers, "rows": rows}

    plain_text = clean_preserve(ws_plain.cell(1, 1).value)
    return {"intro": paragraphs, "table": table, "text": plain_text}


# -------- main build --------

def main():
    html = HTML_IN.read_text(encoding="utf-8")
    # Use rich_text=True so notes / About / Methodology survive with bold runs.
    wb = openpyxl.load_workbook(XLSX, data_only=True, rich_text=True)
    wb_plain = openpyxl.load_workbook(XLSX, data_only=True)  # for parsers that don't need rich text

    tab_order = [
        "Provider_Developer", "Deployer_Supplier",
        "GPAI_Frontier_Foundation model", "GPAI system_Generative AI",
        " High-risk AI system", "Risk",
        "Substantial modification", "Incident",
    ]
    tabs = [parse_legal_sheet(wb_plain[sn], sn) for sn in tab_order if sn in wb_plain.sheetnames]

    # Group analysis cards by tabId.
    analyses: list = []
    for sn in tab_order:
        ana_sn = ANALYSIS_SHEET_FOR.get(sn)
        if ana_sn and ana_sn in wb_plain.sheetnames:
            analyses.extend(parse_analysis_sheet(wb_plain[ana_sn], ana_sn))
    ana_by_tab: dict[str, list] = {}
    for a in analyses:
        ana_by_tab.setdefault(a.get("tab", ""), []).append(a)

    # Concept_ids back-link (for v11/v12 renderer compatibility).
    tabs_by_id = {t["tabId"]: t for t in tabs}
    for card in analyses:
        parent = tabs_by_id.get(card.get("tab"))
        if parent:
            card["concept_ids"] = [c["id"] for c in parent.get("concepts", [])]

    ref_map = build_ref_map(tabs)

    # Parse "New concepts" matrix early — it drives the cluster summary table.
    matrix = parse_matrix(wb_plain["New concepts"]) if "New concepts" in wb_plain.sheetnames else {"headers": [], "rows": []}

    # Build v15 concept structure — one entry per matrix term family.
    # The "Model / system" family combines three legal sheets that v14 used
    # to surface as separate top-level tabs.
    concepts_v15: list[dict] = []
    JURIS_ORDER = ["eu", "ca", "co", "ny", "tx", "ut"]

    tabs_by_sheetname = {sn: t for sn, t in zip(tab_order, tabs) if sn in wb_plain.sheetnames}

    for fam in MATRIX_FAMILIES:
        family_label = fam["family"]
        tab_id       = fam["tab_id"]
        cluster_name = fam["cluster"]

        sub_concepts: list[dict] = []
        rich_notes: list[dict] = []
        framing_pool: list[str] = []

        for legal_sn in fam["legal_sheets"]:
            tab = tabs_by_sheetname.get(legal_sn)
            if not tab:
                continue
            tab_id_legal  = tab["tabId"]
            tab_analyses  = ana_by_tab.get(tab_id_legal, [])

            for ci, concept in enumerate(tab.get("concepts", [])):
                built = build_sub_concepts_v15(concept, tab_analyses, JURIS_ORDER, concept_index=ci)
                juris = built["jurisdictions"]
                dims  = built["dimensions"]
                if not juris or not dims:
                    continue
                sub_id = slugify(concept.get("name") or concept.get("id") or f"sub-{ci}")
                base, n = sub_id, 1
                while any(sc["id"] == sub_id for sc in sub_concepts):
                    n += 1
                    sub_id = f"{base}-{n}"
                sub_concepts.append({
                    "id":    sub_id,
                    "title": concept.get("name") or "",
                    "jurisdictions": juris,
                    "dimensions":    dims,
                })

            ana_sn = ANALYSIS_SHEET_FOR.get(legal_sn)
            if ana_sn:
                rich_notes.extend(collect_rich_notes(wb, ana_sn))

            for a in tab_analyses:
                for row in (a.get("rows") or []):
                    if row.get("notes"):
                        framing_pool.append(row["notes"])
                        break

        if not sub_concepts:
            continue

        # Cluster summary table from the curated "New concepts" matrix.
        matrix_summary = cluster_summary_from_matrix(matrix, family_label)
        if matrix_summary:
            sub_titles = [(sc["id"], (sc.get("title") or "").strip().lower()) for sc in sub_concepts]
            for row in matrix_summary["rows"]:
                wanted = row["term_label"].strip().lower()
                best = ""
                for sid, title in sub_titles:
                    if title == wanted:
                        best = sid
                        break
                if not best:
                    for sid, title in sub_titles:
                        if title and (title in wanted or wanted in title):
                            best = sid
                            break
                if not best and sub_titles:
                    best = sub_titles[0][0]
                row["sub_id"] = best
                for jid, cell in row["cells"].items():
                    for p in (cell.get("variants") or []):
                        p["sub_id"] = best
                        p["jid"]    = jid
            cluster_summary = matrix_summary
        else:
            cluster_summary = {"headers": [], "rows": []}

        ceps_framing = framing_pool[0] if framing_pool else "Comparative analysis across jurisdictions."

        concepts_v15.append({
            "id":              tab_id,
            "cluster":         cluster_name,
            "title":           family_label,
            "ceps_framing":    ceps_framing,
            "cluster_summary": cluster_summary,
            "ceps_notes_rich": rich_notes,
            "sub_concepts":    sub_concepts,
        })

    laws_meta = build_laws_metadata()
    laws_full = load_laws()

    # Wire up inline JSON law blobs.
    blob_tags: list[str] = []
    for lid, law in laws_full.items():
        blob = json.dumps(law, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
        blob_tags.append(f'<script type="application/json" id="law-blob-{lid}">{blob}</script>')

    # Replace CONCEPTS + LAWS in the draft.
    html = replace_const(html, "CONCEPTS", json.dumps(concepts_v15, ensure_ascii=False, separators=(",", ":")))
    html = replace_const(html, "LAWS",     json.dumps(laws_meta,    ensure_ascii=False, separators=(",", ":")))

    # Stubs + REF_MAP for the law drawer (lifted from v14).
    stubs = {}
    for lid, law in laws_full.items():
        stub = {"title": law.get("title", ""), "url": law.get("url", LAW_URLS.get(lid, ""))}
        stub["articles"] = [{"id": a.get("id", ""), "title": a.get("title", "")}
                            for a in (law.get("articles") or [])]
        stub["sections"] = [{"id": s.get("id", ""), "title": s.get("title", "") or ("Section " + s.get("id", ""))}
                            for s in (law.get("sections") or [])]
        stubs[lid] = stub

    glossary = parse_glossary(wb_plain["Second edition terminology"]) if "Second edition terminology" in wb_plain.sheetnames else []
    about    = extract_prose_blocks(wb["About the Digital AI Lexicon"], wb_plain["About the Digital AI Lexicon"]) \
               if "About the Digital AI Lexicon" in wb.sheetnames else {"intro": [], "table": None}
    methodology = extract_prose_blocks(wb["Methodology"], wb_plain["Methodology"]) \
                  if "Methodology" in wb.sheetnames else {"intro": [], "table": None}

    extra = (
        "\n<script>\n"
        "const LAW_STUBS = " + json.dumps(stubs,       ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const REF_MAP   = " + json.dumps(ref_map,     ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const DATA          = " + json.dumps(tabs,     ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const ANALYSIS_DATA = " + json.dumps(analyses, ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const MATRIX        = " + json.dumps(matrix,   ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const GLOSSARY      = " + json.dumps(glossary, ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const ABOUT_PROSE       = " + json.dumps(about,       ensure_ascii=False, separators=(",", ":")) + ";\n"
        "const METHODOLOGY_PROSE = " + json.dumps(methodology, ensure_ascii=False, separators=(",", ":")) + ";\n"
        "</script>\n"
    )
    tail = "\n".join(blob_tags) + extra
    html = html.replace("</body>", tail + "\n</body>", 1)

    # Theme toggle (lifted verbatim from v14).
    html = html.replace("</head>", _v14_theme_head() + "\n</head>", 1)

    # Law drawer + cross-page navigation (lifted from v14).
    html = html.replace("</body>", _v14_ext_js() + "\n</body>", 1)

    # v15-specific overrides: cluster summary renderer, term-name headers,
    # rich-text notes renderer, About + Methodology prose renderers, font.
    html = html.replace("</body>", _v15_overrides() + "\n</body>", 1)

    HTML_OUT.write_text(html, encoding="utf-8")
    print(f"Wrote {HTML_OUT}  ({len(html):,} bytes)")
    print(f"  concepts:           {len(concepts_v15)} top-level")
    print(f"  sub-concepts:       {sum(len(c['sub_concepts']) for c in concepts_v15)}")
    print(f"  cluster-summary rows: {sum(len(c['cluster_summary']['rows']) for c in concepts_v15)}")
    print(f"  rich notes:         {sum(len(c['ceps_notes_rich']) for c in concepts_v15)}")
    print(f"  analyses (raw):     {len(analyses)}")
    print(f"  laws embedded:      {len(laws_full)}")
    print(f"  ref map:            {sum(1 for v in ref_map.values() if v.get('law'))}/{len(ref_map)} resolved")


# Lift the v14 head/body injections via source extraction so we don't duplicate them.
def _v14_theme_head() -> str:
    """Return the theme <style>+<script> block that v14 injects into <head>."""
    src = (HERE / "build_v14.py").read_text(encoding="utf-8")
    m = re.search(r'theme_head\s*=\s*"""(.*?)"""', src, re.S)
    return m.group(1) if m else ""


def _v14_ext_js() -> str:
    """Return the law-drawer + nav extension that v14 injects before </body>."""
    src = (HERE / "build_v14.py").read_text(encoding="utf-8")
    m = re.search(r'ext_js\s*=\s*r"""(.*?)"""', src, re.S)
    return m.group(1) if m else ""


def _v15_overrides() -> str:
    """v15-specific CSS + JS overrides injected before </body>:

      - Font 16px on html,body
      - .cluster-table CSS (the new image-style summary)
      - .pill-stack / .variant-pill CSS
      - .j-term CSS (new term-name span in dimension table headers)
      - .ceps-note-rich CSS for bold-preserved interpretative notes
      - .prose-section CSS for About + Methodology paragraphs
      - JS overrides:
          renderConceptPage()    → swap detail body to renderClusterSummary
          renderClusterSummary() → builds the term-map table
          renderAnalysisTable()  → enriched headers (term + bills)
          renderHome()           → reads ABOUT_PROSE
          renderMethodology()    → reads METHODOLOGY_PROSE
          go()                   → call renderHome / renderMethodology on nav
    """
    return r"""
<style>
/* v15 — bump base font for readability */
html, body { font-size: 16px !important; }

/* --- Cluster summary table (replaces stack of dim tables on cluster page) --- */
.cluster-summary-wrap { margin: 24px 0 16px; }
.cluster-summary-table {
  width: 100%; border-collapse: collapse; background: var(--card);
  border: 1px solid var(--bd); border-radius: var(--r-xl); overflow: hidden;
  font-family: var(--serif); font-size: 14px;
}
.cluster-summary-table th, .cluster-summary-table td {
  border-top: 1px solid var(--bd-s); padding: 12px 14px;
  vertical-align: top; text-align: left;
}
.cluster-summary-table thead th {
  background: var(--surf); border-top: none; border-bottom: 2px solid var(--bd);
  font-size: 12px; text-transform: uppercase; letter-spacing: .06em;
  color: var(--ink-s); font-weight: 600;
}
.cluster-summary-table .term-cell {
  font-weight: 600; color: var(--ink); width: 28%;
  background: var(--bg2);
}
.cluster-summary-table .pill-stack {
  display: flex; flex-direction: column; gap: 8px;
}
.cluster-summary-table .variant-pill {
  display: inline-block; padding: 6px 10px; border-radius: var(--r-md);
  background: var(--surf); border: 1px solid var(--bd-s);
  cursor: pointer; transition: background .12s, border-color .12s, transform .08s;
  color: var(--ink); font-size: 13px; line-height: 1.3;
  text-align: left; font-family: var(--serif);
}
.cluster-summary-table .variant-pill:hover {
  background: var(--accent-l); border-color: var(--accent); color: var(--accent-d);
}
.cluster-summary-table .variant-pill .pill-bill {
  display: block; font-family: var(--mono); font-size: 11px;
  color: var(--ink-h); margin-top: 2px;
}
.cluster-summary-table .variant-pill:hover .pill-bill { color: var(--accent); }
.cluster-summary-table .empty { color: var(--bd); }

/* --- Browse Matrix: family-rowspan separator + family cell styling --- */
.matrix-table tbody tr.family-first td { border-top: 2px solid var(--bd); }
.matrix-table tbody tr:first-child td { border-top: none; }
.matrix-table .matrix-family-cell {
  background: var(--bg2); vertical-align: top; min-width: 180px;
}

/* --- Term-name span in per-concept dimension table headers (sub-tabs view) --- */
.analysis-table th .j-term {
  display: block; font-style: italic; font-weight: 500; font-size: 12px;
  color: var(--ink-s); margin-top: 2px; line-height: 1.3;
}

/* --- Rich-text interpretative notes (bold runs + line breaks preserved) --- */
.ceps-notes-rich-wrap {
  margin-top: 28px; background: var(--ceps-bg); border-left: 4px solid var(--ceps-border);
  border-radius: var(--r-md); padding: 16px 20px;
}
.ceps-notes-rich-wrap .ceps-notes-header {
  font-weight: 600; color: var(--ceps-text); font-size: 13px;
  text-transform: uppercase; letter-spacing: .06em; margin-bottom: 10px;
  display: flex; align-items: center; gap: 8px;
}
.ceps-note-rich { margin-top: 14px; }
.ceps-note-rich:first-of-type { margin-top: 0; }
.ceps-note-rich .nrt-title {
  font-weight: 600; color: var(--ceps-text); font-size: 14px; margin-bottom: 6px;
}
.ceps-note-rich .nrt-body {
  font-size: 14px; color: var(--ink); line-height: 1.65; white-space: pre-wrap;
}
.ceps-note-rich .nrt-body strong { color: var(--ceps-text); }

/* --- Landing page + Methodology prose (rendered from xlsx) --- */
.prose-section { font-size: 16px; line-height: 1.7; color: var(--ink); }
.prose-section .prose-paragraph { margin-bottom: 14px; white-space: pre-wrap; }
.prose-section .prose-paragraph strong { color: var(--ink); font-weight: 600; }
.prose-section h1.prose-title {
  font-size: 32px; font-weight: 700; letter-spacing: -.02em;
  margin: 0 0 16px; color: var(--ink);
}
.prose-section h2.prose-section-head {
  font-size: 18px; font-weight: 600; margin: 28px 0 8px; color: var(--ink);
}
.prose-table-wrap { margin: 18px 0 6px; overflow-x: auto; }
.prose-table {
  width: 100%; border-collapse: collapse; background: var(--card);
  border: 1px solid var(--bd); border-radius: var(--r-md); overflow: hidden;
  font-size: 13px;
}
.prose-table th, .prose-table td {
  border-top: 1px solid var(--bd-s); padding: 10px 12px;
  vertical-align: top; text-align: left;
}
.prose-table thead th {
  background: var(--surf); border-top: none; font-weight: 600;
  text-transform: uppercase; font-size: 11px; letter-spacing: .05em;
  color: var(--ink-s);
}
.prose-table td.region-cell { font-weight: 600; color: var(--ink); }
.prose-table td strong { font-weight: 600; }
</style>
<script>
(function(){
  /* ---- Helpers ---- */
  function escH(s){return String(s||'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
  function runsToHtml(runs){
    if (!runs || !runs.length) return '';
    return runs.map(r => {
      const txt = escH(r.t || '').replace(/\n/g, '<br>');
      return r.b ? '<strong>'+txt+'</strong>' : txt;
    }).join('');
  }
  function runsToPlain(runs){
    if (!runs || !runs.length) return '';
    return runs.map(r => r.t || '').join('');
  }

  /* ---- Browse-page Matrix view (Concepts page) ----
     Override v12_draft's renderMatrix so the browse table mirrors the
     same New-concepts layout: each sub-concept is its own row, the family
     name rowspans across the family's rows, and each jurisdiction column
     shows the variant pill for that row. Click a pill → drill into the
     concept page Overview. */
  const _origRenderMatrix = window.renderMatrix;
  window.renderMatrix = function(filtered, activeJuris){
    const allJ = (activeJuris && activeJuris.length)
      ? JURIS_ORDER.filter(j => activeJuris.includes(j))
      : JURIS_ORDER;

    let html = '<div class="matrix-wrap"><table class="matrix-table"><thead><tr><th>Concept</th>';
    allJ.forEach(j => { html += '<th class="th-'+j+'">'+escH(JURIS_LABELS[j] || j)+'</th>'; });
    html += '</tr></thead><tbody>';

    filtered.forEach(c => {
      const cs = c.cluster_summary;
      const rows = (cs && cs.rows && cs.rows.length) ? cs.rows : [{
        term_label: '', term_rowspan: 1, sub_id: (c.sub_concepts[0] || {}).id || '',
        cells: {}
      }];
      const familyRowspan = rows.length;

      rows.forEach((row, rowIdx) => {
        html += '<tr'+(rowIdx === 0 ? ' class="family-first"' : '')+'>';
        // Family / Concept cell — only on the first row of the family.
        if (rowIdx === 0){
          const rs = familyRowspan > 1 ? ' rowspan="'+familyRowspan+'"' : '';
          html += '<td'+rs+' class="matrix-family-cell" style="cursor:pointer" onclick="go(\'concept\',\''+escH(c.id)+'\')">'
                +    '<span style="font-size:14px;font-weight:600">'+escH(c.title)+'</span>'
                +    '<br><span class="cluster-chip">'+escH(c.cluster)+'</span>'
                +  '</td>';
        }

        allJ.forEach(j => {
          const cell = row.cells && row.cells[j];
          if (!cell || cell.rowspan === 0){
            // Either no cell at all, or covered by a previous row's rowspan.
            if (cell && cell.rowspan === 0) return;
            html += '<td><button class="matrix-empty">—</button></td>';
            return;
          }
          const rsAttr = cell.rowspan > 1 ? ' rowspan="'+cell.rowspan+'"' : '';
          const variants = cell.variants || [];
          if (!variants.length){
            html += '<td'+rsAttr+'><button class="matrix-empty">—</button></td>';
            return;
          }
          const v = variants[0];
          const label = v.name + (v.bill ? ' ('+v.bill+')' : '');
          html += '<td'+rsAttr+'><button class="j-pill '+j+'" '
                +    'onclick="go(\'concept\',\''+escH(c.id)+'\')">'
                +    escH(label)
                +  '</button></td>';
        });
        html += '</tr>';
      });
    });
    html += '</tbody></table></div>';

    const view = document.getElementById('matrix-view');
    if (view) view.innerHTML = html;
  };

  /* ---- Cluster summary (matches the New-concepts Excel sheet layout) ----
     Each matrix row is its own table row. The Term column rowspans across
     a group's continuation rows so the EU canonical term appears once per
     sub-concept. Each jurisdiction cell holds one variant pill (or '—'). */
  function renderClusterSummary(c){
    const cs = c && c.cluster_summary;
    if (!cs || !cs.rows || !cs.rows.length){
      return '<p style="color:var(--ink-s);font-size:14px;">No comparison data available.</p>';
    }
    let html = '<div class="cluster-summary-wrap"><table class="cluster-summary-table"><thead><tr>';
    html += '<th>Term</th>';
    cs.headers.forEach(h => { html += '<th class="th-'+h.jid+'">'+escH(h.label)+'</th>'; });
    html += '</tr></thead><tbody>';
    cs.rows.forEach(row => {
      html += '<tr>';
      // Term cell: skip if covered by a previous row's rowspan.
      if (row.term_rowspan > 0){
        const rs = row.term_rowspan > 1 ? ' rowspan="'+row.term_rowspan+'"' : '';
        html += '<td class="term-cell"'+rs+'>'+escH(row.term_label || '')+'</td>';
      }
      cs.headers.forEach(h => {
        const cell = row.cells[h.jid] || {rowspan: 1, variants: []};
        // rowspan=0 means this cell is covered by a rowspan from a previous row.
        if (cell.rowspan === 0) return;
        const rsAttr = cell.rowspan > 1 ? ' rowspan="'+cell.rowspan+'"' : '';
        const variants = cell.variants || [];
        if (!variants.length){
          html += '<td'+rsAttr+'><span class="empty">—</span></td>';
          return;
        }
        const v = variants[0];
        const ref = v.ref || '';
        html += '<td'+rsAttr+'><button class="variant-pill" '
              +    'data-sub="'+escH(row.sub_id)+'" '
              +    'data-jid="'+escH(h.jid)+'" '
              +    'data-bill="'+escH(v.bill || '')+'" '
              +    'data-ref="'+escH(ref)+'" '
              +    'onclick="__openVariantDrawer(\''+escH(row.sub_id)+'\',\''+escH(h.jid)+'\',\''+escH(v.bill || '')+'\')">'
              +    escH(v.name || '')
              +    (v.bill ? '<span class="pill-bill">'+escH(v.bill)+'</span>' : '')
              +  '</button></td>';
      });
      html += '</tr>';
    });
    html += '</tbody></table></div>';
    return html;
  }

  /* ---- Rich-text interpretative notes ---- */
  function renderRichNotes(c){
    const notes = (c && c.ceps_notes_rich) || [];
    if (!notes.length) return '';
    let html = '<div class="ceps-notes-rich-wrap">';
    html += '<div class="ceps-notes-header">CEPS Interpretative Notes <span class="ceps-badge">CEPS</span></div>';
    notes.forEach(n => {
      html += '<div class="ceps-note-rich">';
      if (n.title) html += '<div class="nrt-title">'+escH(n.title)+'</div>';
      html += '<div class="nrt-body">'+runsToHtml(n.body_runs)+'</div>';
      html += '</div>';
    });
    html += '</div>';
    return html;
  }

  /* ---- Variant pill click: open verbatim drawer ----
     Multi-bill jurisdictions are split into lane keys like 'ca-sb942'. Match
     the pill (which carries jid + bill from the matrix) to the right lane
     by comparing bill codes case-insensitively, ignoring whitespace. */
  function _normBill(s){ return String(s || '').toLowerCase().replace(/\s+/g, ''); }

  window.__openVariantDrawer = function(subId, jid, bill){
    const c = (typeof getConcept === 'function') ? getConcept(state.conceptId) : null;
    if (!c) return;
    const sc = c.sub_concepts.find(s => s.id === subId) || c.sub_concepts[0];
    if (!sc) return;
    let useKey = jid;
    if (sc.jurisdictions && !sc.jurisdictions[jid]){
      const lanes = Object.keys(sc.jurisdictions).filter(k => k === jid || k.startsWith(jid + '-'));
      let match = null;
      if (bill){
        const want = _normBill(bill);
        match = lanes.find(k => _normBill((sc.jurisdictions[k] || {}).bills) === want);
      }
      useKey = match || lanes[0] || jid;
    }
    let dim = sc.dimensions.find(d => d.cells && d.cells[useKey] && d.cells[useKey].verbatim)
           || sc.dimensions.find(d => d.cells && d.cells[useKey])
           || sc.dimensions[0];
    if (!dim) return;
    state.subConceptIdx = c.sub_concepts.indexOf(sc);
    saveState();
    if (typeof openDrawer === 'function') openDrawer(dim.id, useKey);
  };

  /* ---- Override renderConceptPage to show cluster summary by default ----
     Sub-tabs still reveal the per-concept analysis table (now sourced from
     the analysis sheet, with term names in headers — see renderAnalysisTable
     override below). When state.subConceptIdx === -1 we show the cluster
     summary; otherwise the dimension table for that sub-concept. */
  const _origRenderConceptPage = window.renderConceptPage;
  window.renderConceptPage = function(){
    const c = getConcept(state.conceptId);
    if (!c) return;

    const allJuris = (typeof getAllJurisForConcept === 'function') ? getAllJurisForConcept(c) : [];
    const dots = (typeof JURIS_ORDER !== 'undefined' ? JURIS_ORDER : ['eu','ca','co','ny','tx','ut'])
      .map(j => '<span class="j-dot '+(allJuris.includes(j) ? j : '')+'" style="width:10px;height:10px" title="'+JURIS_LABELS[j]+'"></span>').join('');
    document.getElementById('concept-header').innerHTML =
      '<button class="concept-cluster-chip" onclick="go(\'concepts\')">'+escH(c.cluster)+'</button>' +
      '<h1 class="concept-title">'+escH(c.title)+'</h1>' +
      '<div class="concept-coverage">'+dots+'</div>' +
      (c.ceps_framing ? '<p class="concept-framing">'+escH(c.ceps_framing)+'</p>' : '');

    // Sub-tabs: prepend an "Overview" tab that shows the cluster summary.
    const tabsEl = document.getElementById('sub-tabs');
    const subs = c.sub_concepts || [];
    const idx = (typeof state.subConceptIdx === 'number') ? state.subConceptIdx : -1;
    let tabsHtml = '<button class="sub-tab '+(idx === -1 ? 'active' : '')+'" onclick="__selectSub(-1)">Overview</button>';
    subs.forEach((sc, i) => {
      tabsHtml += '<button class="sub-tab '+(i === idx ? 'active' : '')+'" onclick="__selectSub('+i+')">'+escH(sc.title)+'</button>';
    });
    tabsEl.innerHTML = tabsHtml;
    tabsEl.style.display = '';

    // Body: cluster summary OR dim table for selected sub.
    const wrap = document.querySelector('.analysis-wrap');
    const notesEl = document.getElementById('ceps-notes');
    if (idx === -1){
      wrap.innerHTML = renderClusterSummary(c);
      notesEl.innerHTML = renderRichNotes(c);
    } else {
      // Restore the original analysis-table DOM (the v12_draft template uses
      // #analysis-thead and #analysis-tbody) and render the per-concept dim table.
      wrap.innerHTML = '<table class="analysis-table" id="analysis-table"><thead id="analysis-thead"></thead><tbody id="analysis-tbody"></tbody></table>';
      renderAnalysisTable();
      notesEl.innerHTML = renderRichNotes(c);
    }
  };

  // Default landing on a concept page = Overview (cluster summary).
  window.__selectSub = function(idx){
    state.subConceptIdx = idx;
    saveState();
    renderConceptPage();
    window.scrollTo(0, 0);
  };

  /* ---- Override renderAnalysisTable to enrich headers with term names ----
     Lane keys (e.g. 'ca-sb942') from multi-bill splitting are not in
     JURIS_LABELS — fall back to jd.law for the visible jurisdiction label
     and use jd._parent_jid for CSS color theming. */
  window.renderAnalysisTable = function(){
    const c = getConcept(state.conceptId);
    if (!c) return;
    const idx = (typeof state.subConceptIdx === 'number' && state.subConceptIdx >= 0) ? state.subConceptIdx : 0;
    const sc = c.sub_concepts[idx] || c.sub_concepts[0];
    if (!sc) return;
    const juris = Object.keys(sc.jurisdictions);

    let thead = '<tr><th>Dimension</th>';
    juris.forEach(j => {
      const jd = sc.jurisdictions[j] || {};
      const themeJid = jd._parent_jid || j;
      const visibleLabel = JURIS_LABELS[j] || jd.law || j;
      thead += '<th class="th-'+themeJid+'">' + escH(visibleLabel)
            +  (jd.term ? '<span class="j-term">'+escH(jd.term)+'</span>' : '')
            +  (jd.bills ? '<span class="j-law">'+escH(jd.bills)+'</span>'
                          : (jd.law && jd.law !== visibleLabel
                             ? '<span class="j-law">'+escH(jd.law)+'</span>' : ''))
            +  '</th>';
    });
    thead += '</tr>';
    const theadEl = document.getElementById('analysis-thead');
    if (theadEl) theadEl.innerHTML = thead;

    let tbody = '';
    sc.dimensions.forEach(dim => {
      tbody += '<tr><td>'+escH(dim.label)+'</td>';
      juris.forEach(j => {
        const cell = dim.cells[j];
        if (cell && (cell.analysis || cell.verbatim)){
          const display = cell.analysis || (cell.verbatim ? cell.verbatim.split('\n')[0].slice(0, 280) : '');
          tbody += '<td><span class="analysis-cell" onclick="openDrawer(\''+dim.id+'\',\''+j+'\')" title="Click for verbatim text">'+escH(display)+'</span></td>';
        } else {
          tbody += '<td><span class="cell-null">—</span></td>';
        }
      });
      tbody += '</tr>';
    });
    const tbodyEl = document.getElementById('analysis-tbody');
    if (tbodyEl) tbodyEl.innerHTML = tbody;
  };

  /* ---- About + Methodology renderers (data from xlsx) ---- */
  function renderProseIntro(prose){
    if (!prose || !prose.intro || !prose.intro.length) return '';
    let html = '<div class="prose-section">';
    prose.intro.forEach((para, i) => {
      // First paragraph with a wholly-bold first run = treat as title.
      if (i === 0 && para.length && para[0].b){
        const titleTxt = para[0].t.trim();
        const restRuns = para.slice(1);
        html += '<h1 class="prose-title">'+escH(titleTxt)+'</h1>';
        if (restRuns.length){
          html += '<div class="prose-paragraph">'+runsToHtml(restRuns).replace(/^(\s*<br>\s*)+/, '')+'</div>';
        }
      } else {
        // If a paragraph starts with a bold "Step N." or single-line bold, treat as section head.
        if (para.length && para[0].b && /^(step\s+\d|overview\b)/i.test(para[0].t.trim())){
          const headTxt = para[0].t.trim();
          const restRuns = para.slice(1);
          html += '<h2 class="prose-section-head">'+escH(headTxt)+'</h2>';
          if (restRuns.length) html += '<div class="prose-paragraph">'+runsToHtml(restRuns).replace(/^(\s*<br>\s*)+/, '')+'</div>';
        } else {
          html += '<div class="prose-paragraph">'+runsToHtml(para)+'</div>';
        }
      }
    });
    html += '</div>';
    return html;
  }

  function renderProseTable(table){
    if (!table || !table.headers || !table.headers.length) return '';
    let html = '<div class="prose-table-wrap"><table class="prose-table"><thead><tr>';
    table.headers.forEach(h => { html += '<th>'+escH(h)+'</th>'; });
    html += '</tr></thead><tbody>';
    table.rows.forEach(row => {
      html += '<tr>';
      row.forEach((runs, ci) => {
        const cls = (ci === 0 && runsToPlain(runs).trim()) ? ' class="region-cell"' : '';
        html += '<td'+cls+'>'+runsToHtml(runs)+'</td>';
      });
      html += '</tr>';
    });
    html += '</tbody></table></div>';
    return html;
  }

  function renderHomeFromProse(){
    const home = document.getElementById('p-home');
    if (!home) return;
    const landing = home.querySelector('.landing');
    // ABOUT_PROSE is a script-scoped const (not on window). Access by name.
    const aboutProse = (typeof ABOUT_PROSE !== 'undefined') ? ABOUT_PROSE : null;
    if (!landing || !aboutProse) return;
    const proseHtml = renderProseIntro(aboutProse);
    if (!proseHtml) return;
    // Find the existing inst-pill / h1 / landing-sub trio and replace with prose.
    const pill = landing.querySelector('.inst-pill');
    const h1   = landing.querySelector('h1');
    const sub  = landing.querySelector('.landing-sub');
    if (pill && pill.dataset.v15 === '1') return; // already rendered
    if (pill) pill.dataset.v15 = '1';
    const replacement = document.createElement('div');
    replacement.innerHTML = proseHtml;
    if (h1 && sub){
      h1.parentNode.insertBefore(replacement.firstChild, h1);
      h1.remove();
      sub.remove();
    } else {
      landing.insertBefore(replacement.firstChild, landing.firstChild);
    }
  }

  function renderMethodologyFromProse(){
    const page = document.getElementById('p-methodology');
    if (!page) return;
    const methProse = (typeof METHODOLOGY_PROSE !== 'undefined') ? METHODOLOGY_PROSE : null;
    if (!methProse) return;
    const wrap = page.querySelector('.method-page');
    if (!wrap) return;
    if (wrap.dataset.v15 === '1') return;
    wrap.dataset.v15 = '1';
    const intro = renderProseIntro(methProse);
    const table = renderProseTable(methProse.table);
    wrap.innerHTML = intro + table;
  }

  /* ---- Wire navigation: render prose on first visit to home/methodology ---- */
  const _origGo2 = window.go;
  window.go = function(page, conceptId, subIdx){
    if (page === 'concept' && (subIdx === undefined || subIdx === null)){
      // First time landing on a concept → show Overview.
      state.subConceptIdx = -1;
    }
    _origGo2.apply(this, arguments);
    if (page === 'home') renderHomeFromProse();
    if (page === 'methodology') renderMethodologyFromProse();
  };

  // On initial load, the v12_draft init already called go(). Re-render
  // home / methodology from prose if we landed on them.
  function _initialRender(){
    if (document.getElementById('p-home') && document.getElementById('p-home').classList.contains('active')){
      renderHomeFromProse();
    }
    if (document.getElementById('p-methodology') && document.getElementById('p-methodology').classList.contains('active')){
      renderMethodologyFromProse();
    }
  }
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', _initialRender);
  else _initialRender();

  // Force home page on initial load (already done by v14, but our go() override
  // above replaces v14's; rerun renderConceptPage if user is mid-concept).
  if (typeof state !== 'undefined' && state.page === 'concept'){
    renderConceptPage();
  }
})();
</script>
"""


if __name__ == "__main__":
    main()
