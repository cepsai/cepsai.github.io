"""load_lexicon_sources.py — Canonical loader for both Excel inputs.

US-001: a single loader module that reads both the cross-checked analysis
Excel and the verbatim Excel into normalised pandas DataFrames so that every
downstream task works from one canonical representation.

Public API
----------
- load_analyses(path: Path | None = None) -> pandas.DataFrame
- load_verbatim(path: Path | None = None) -> pandas.DataFrame

Each row is keyed by ``(term, law_id, article_id)`` so the two frames can be
joined directly. Empty cells are returned as Python ``None``, never as the
string "nan" or numpy ``NaN``.

The loader inspects the actual sheet/column layout for each file rather than
assuming a fixed schema:
* The analysis Excel has both single-section sheets (one Term row at the top)
  and multi-section sheets (Provider/Deployer with several Term blocks).
* The verbatim Excel has 4-column blocks (Risk, Substantial modification) and
  5-column blocks (Provider, Deployer, etc.) with the jurisdiction header on
  either row 1 or row 2 depending on the sheet.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ITER_DIR = Path(__file__).resolve().parent
if str(ITER_DIR) not in sys.path:
    sys.path.insert(0, str(ITER_DIR))

from build_reference_lookup import (  # noqa: E402
    SECTION_RE,
    parse_atomic,
    parse_para_list,
    split_atomic,
)

DEFAULT_ANALYSIS = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis-2.xlsx"
)
DEFAULT_VERBATIM = Path(
    "/Users/robertpraas/Downloads/"
    "AI terminology and taxonomy_verbatim excel.xlsx"
)

JURIS_KEYS = (
    "eu", "u.s.", "colorado", "utah", "texas", "california", "new york",
)

# Fallback law mapping when an analysis cell has no inline reference. The keys
# are lowercase substrings of the jurisdiction header text in the Excel.
JURIS_TO_LAW = {
    "eu": "eu-ai-act",
    "eu (aia)": "eu-ai-act",
    "colorado": "co-sb24205",
    "colorado (sb 24-205)": "co-sb24205",
    "u.s. - colorado": "co-sb24205",
    "utah": "ut-sb226",
    "utah (sb 226)": "ut-sb226",
    "u.s. - utah": "ut-sb226",
    "texas": "tx-hb149",
    "texas (hb 149)": "tx-hb149",
    "u.s. - texas": "tx-hb149",
    "california": "ca-sb53",
    "california (sb 53)": "ca-sb53",
    "california (sb 942)": "ca-sb942",
    "california (ab 2013)": "ca-ab2013",
    "u.s. - california": "ca-sb53",
    "new york": "ny-s8828",
    "new york (s8828)": "ny-s8828",
    "new york (a6453b)": "ny-a6453",
    "u.s. - new york": "ny-s8828",
}

ANALYSIS_COLUMNS = [
    "term", "law_id", "article_id", "dim_label", "analysis_text",
    "reference", "jurisdiction_header", "sheet", "addr",
]

VERBATIM_COLUMNS = [
    "term", "law_id", "article_id", "dim_label", "sub_dim_label",
    "verbatim_text", "reference", "tags", "jurisdiction_header",
    "sheet", "addr",
]


# --------------------------------------------------------------------------- #
# Cell helpers                                                                #
# --------------------------------------------------------------------------- #

def _col_letter(col_idx: int) -> str:
    """1-based column index → A1 letters."""
    s = ""
    n = col_idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(ord("A") + rem) + s
    return s


# Sentinel set: cells whose stripped text equals one of these are treated as
# empty (the lexicon source uses "-" / "—" as "nothing applies"). Note that
# the ISO 3166-1 alpha-2 country code "NA" is *not* in this set: per project
# convention, "NA" is Namibia and must be preserved as a real value.
_EMPTY_TOKENS = {"", "-", "—", "–"}


def _norm_str(v) -> str | None:
    """Strip whitespace from a cell value and return None for empty cells.

    Preserves "NA" (Namibia) and any other non-blank token literally so the
    caller can decide whether to treat sentinels like "-" as empty.
    """
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    if s == "":
        return None
    return s


def _juris_to_law(label: str | None) -> str | None:
    """Map a jurisdiction header label (e.g. ``EU (AIA)``) to a law_id."""
    if not label:
        return None
    return JURIS_TO_LAW.get(label.strip().lower())


# --------------------------------------------------------------------------- #
# Reference parsing                                                           #
# --------------------------------------------------------------------------- #

_PAREN_RE = re.compile(r"\(([^()]+)\)")


def _parse_atomic_with_inheritance(
    atom: str, last_law: str | None,
) -> dict | None:
    """Parse one atomic citation; inherit ``last_law`` for orphan sections."""
    parsed = parse_atomic(atom)
    if parsed.get("law") is None and last_law and not parsed.get("anchor"):
        m = SECTION_RE.search(atom)
        if m:
            parsed["law"] = last_law
            parsed["kind"] = "section"
            parsed["article_id"] = m.group("id")
            parsed["anchor"] = m.group("id")
            if m.group("para"):
                parsed["paragraphs"] = parse_para_list(m.group("para"))
    if parsed.get("law") and parsed.get("anchor"):
        return parsed
    return None


def _explode_reference_cell(
    text: str | None, fallback_law: str | None,
) -> list[dict]:
    """Parse a reference column value into one row per atomic citation."""
    if not text:
        return [{
            "law_id": fallback_law,
            "article_id": None,
            "kind": None,
            "raw_ref": None,
        }]
    out: list[dict] = []
    last_law = fallback_law
    for atom in split_atomic(text):
        parsed = _parse_atomic_with_inheritance(atom, last_law)
        if parsed is None:
            continue
        last_law = parsed["law"]
        out.append({
            "law_id": parsed["law"],
            "article_id": str(parsed["anchor"]),
            "kind": parsed["kind"],
            "raw_ref": parsed["raw"],
        })
    if not out:
        out.append({
            "law_id": fallback_law,
            "article_id": None,
            "kind": None,
            "raw_ref": None,
        })
    return out


def _scan_inline_references(
    text: str, fallback_law: str | None,
) -> list[dict]:
    """Find every ``(...)`` citation embedded in an analysis text and parse it.

    Analysis cells store free-form text with parenthesised citations inline
    (e.g. ``"AI system... (Article 6, Annex III) (§6-1-1701)"``). We pull each
    parenthesised span and feed it through ``_parse_atomic_with_inheritance``.
    """
    out: list[dict] = []
    seen: set[tuple[str, str]] = set()
    last_law = fallback_law
    for m in _PAREN_RE.finditer(text or ""):
        inner = m.group(1)
        if not any(c.isdigit() for c in inner):
            continue
        for atom in split_atomic(inner):
            parsed = _parse_atomic_with_inheritance(atom, last_law)
            if parsed is None:
                continue
            last_law = parsed["law"]
            key = (parsed["law"], str(parsed["anchor"]))
            if key in seen:
                continue
            seen.add(key)
            out.append({
                "law_id": parsed["law"],
                "article_id": str(parsed["anchor"]),
                "kind": parsed["kind"],
                "raw_ref": parsed["raw"],
            })
    if not out:
        out.append({
            "law_id": fallback_law,
            "article_id": None,
            "kind": None,
            "raw_ref": None,
        })
    return out


# --------------------------------------------------------------------------- #
# Analysis sheet parsing                                                      #
# --------------------------------------------------------------------------- #

def _juris_columns(ws, header_row: int) -> list[tuple[int, str]]:
    """Return [(col, juris_header), ...] for a given header row."""
    out = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        s = _norm_str(v)
        if not s:
            continue
        low = s.lower()
        if low.startswith("interpretative"):
            continue
        if any(kw in low for kw in JURIS_KEYS):
            out.append((c, s))
    return out


def _find_term_rows(ws) -> list[int]:
    """Rows where column A holds the literal text ``Term`` (any casing)."""
    out = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v).strip().lower() == "term":
            out.append(r)
    return out


def _walk_analysis_section(
    ws, term_row: int, end_row: int, juris_cols: list[tuple[int, str]],
) -> list[dict]:
    """Yield one record per (dim_label, jurisdiction_column) cell.

    Continuation rows (col A blank, juris columns populated) are concatenated
    into the previous dimension's text — this matches the source convention
    where multi-paragraph definitions span several Excel rows.
    """
    term_per_col = {
        c: _norm_str(ws.cell(row=term_row, column=c).value)
        for c, _ in juris_cols
    }
    last_dim: str | None = None
    accum: dict[tuple[str, int], list[str]] = {}
    accum_addr: dict[tuple[str, int], str] = {}

    for r in range(term_row + 1, end_row + 1):
        a = _norm_str(ws.cell(row=r, column=1).value)
        # Footnote lines like "[1] Note that..." reset the dimension context.
        if a and a.startswith("["):
            last_dim = None
            continue
        if a:
            last_dim = a
        if not last_dim:
            continue
        for c, _ in juris_cols:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            sv = str(v).replace("\xa0", " ").strip()
            if sv == "":
                continue
            key = (last_dim, c)
            accum.setdefault(key, []).append(sv)
            accum_addr.setdefault(key, f"{_col_letter(c)}{r}")

    records = []
    juris_lookup = dict(juris_cols)
    for (dim_label, c), parts in accum.items():
        term = term_per_col.get(c)
        if not term:
            continue
        text = "\n".join(parts)
        records.append({
            "term": term,
            "dim_label": dim_label,
            "analysis_text": text,
            "jurisdiction_header": juris_lookup.get(c),
            "addr": accum_addr[(dim_label, c)],
        })
    return records


def _load_analysis_sheet(ws) -> list[dict]:
    """Parse one analysis sheet into raw per-cell records."""
    term_rows = _find_term_rows(ws)
    if not term_rows:
        return []
    out = []
    for i, tr in enumerate(term_rows):
        # Header row is the row immediately above the Term row that carries
        # jurisdiction labels in cols B+. Walk up to 3 rows back to find it.
        header_row = None
        for r in range(tr - 1, max(tr - 4, 0), -1):
            if _juris_columns(ws, r):
                header_row = r
                break
        if header_row is None:
            continue
        juris_cols = _juris_columns(ws, header_row)
        if not juris_cols:
            continue
        end_row = (
            term_rows[i + 1] - 1 if i + 1 < len(term_rows) else ws.max_row
        )
        out.extend(_walk_analysis_section(ws, tr, end_row, juris_cols))
    return out


def load_analyses(path: Path | str | None = None) -> pd.DataFrame:
    """Read the cross-checked analysis Excel into a normalised DataFrame.

    Each row corresponds to one (term, law_id, article_id) — analysis cells
    that cite multiple atomic references (e.g. ``Article 6`` plus ``Annex III``)
    are exploded into one row per citation so that the frame can be joined
    against ``load_verbatim()`` directly.
    """
    src = Path(path) if path else DEFAULT_ANALYSIS
    wb = load_workbook(src, data_only=True, read_only=True)
    rows: list[dict] = []
    for sn in wb.sheetnames:
        # Match "analysis" plus its ~31-char Excel-truncated forms
        # ("…_Analys", "…_ANALY") so the GPAI sheets are not silently dropped.
        if "_analy" not in sn.lower() and "analysis" not in sn.lower():
            continue
        ws = wb[sn]
        for cell_rec in _load_analysis_sheet(ws):
            fallback_law = _juris_to_law(cell_rec["jurisdiction_header"])
            refs = _scan_inline_references(
                cell_rec["analysis_text"], fallback_law,
            )
            for ref in refs:
                rows.append({
                    "term": cell_rec["term"],
                    "law_id": ref["law_id"],
                    "article_id": ref["article_id"],
                    "dim_label": cell_rec["dim_label"],
                    "analysis_text": cell_rec["analysis_text"],
                    "reference": ref["raw_ref"],
                    "jurisdiction_header": cell_rec["jurisdiction_header"],
                    "sheet": sn,
                    "addr": cell_rec["addr"],
                })
    wb.close()
    df = pd.DataFrame(rows, columns=ANALYSIS_COLUMNS)
    # Replace pandas NaN with Python None for downstream code that distinguishes
    # "no value" from "the string 'nan'".
    return df.astype(object).where(df.notna(), None)


# --------------------------------------------------------------------------- #
# Verbatim sheet parsing                                                      #
# --------------------------------------------------------------------------- #

def _detect_verbatim_layout(ws) -> dict | None:
    """Locate the jurisdiction header row and per-block column layout.

    Returns ``{header_row, term_row, data_start, blocks: [...]}`` where each
    block is ``{start, end, juris, dim_col, sub_dim_col, verbatim_col, ref_col,
    tags_col}``. Returns None when the sheet doesn't look like a verbatim
    sheet (e.g. an empty section).
    """
    header_row = None
    for r in range(1, min(ws.max_row + 1, 5)):
        hits = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and any(
                kw in str(v).strip().lower() for kw in JURIS_KEYS
            ):
                hits += 1
        if hits >= 2:
            header_row = r
            break
    if header_row is None:
        return None

    block_starts: list[tuple[int, str]] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if not v:
            continue
        s = str(v).replace("\xa0", " ").strip()
        if any(kw in s.lower() for kw in JURIS_KEYS):
            block_starts.append((c, s))
    if not block_starts:
        return None

    blocks = []
    for i, (c, juris) in enumerate(block_starts):
        end = (
            block_starts[i + 1][0] - 1
            if i + 1 < len(block_starts) else ws.max_column
        )
        # Find the Reference column inside [c, end]: it may be on header_row,
        # header_row+1, or header_row+2 (sheet conventions vary).
        ref_col = None
        for r in (header_row, header_row + 1, header_row + 2):
            if r > ws.max_row:
                break
            for cc in range(c, end + 1):
                v = ws.cell(row=r, column=cc).value
                if v and "reference" in str(v).strip().lower():
                    ref_col = cc
                    break
            if ref_col:
                break

        if ref_col is None:
            # Default to 4-col layout: dim, verbatim, ref, tags.
            dim_col, sub_dim_col = c, None
            verbatim_col, ref_col_resolved, tags_col = c + 1, c + 2, c + 3
        else:
            offset = ref_col - c
            if offset == 2:
                dim_col, sub_dim_col = c, None
                verbatim_col, ref_col_resolved, tags_col = (
                    c + 1, ref_col, ref_col + 1,
                )
            elif offset >= 3:
                dim_col, sub_dim_col = c, c + 1
                verbatim_col, ref_col_resolved, tags_col = (
                    ref_col - 1, ref_col, ref_col + 1,
                )
            else:
                # offset 1 or weird — fall back to 4-col defaults.
                dim_col, sub_dim_col = c, None
                verbatim_col, ref_col_resolved, tags_col = (
                    c + 1, c + 2, c + 3,
                )

        blocks.append({
            "start": c,
            "end": end,
            "juris": juris,
            "dim_col": dim_col,
            "sub_dim_col": sub_dim_col,
            "verbatim_col": verbatim_col,
            "ref_col": ref_col_resolved,
            "tags_col": tags_col,
        })

    term_row = header_row + 1
    data_start = term_row + 1
    return {
        "header_row": header_row,
        "term_row": term_row,
        "data_start": data_start,
        "blocks": blocks,
    }


def _load_verbatim_sheet(ws) -> list[dict]:
    layout = _detect_verbatim_layout(ws)
    if not layout:
        return []
    out = []
    for blk in layout["blocks"]:
        term = _norm_str(
            ws.cell(row=layout["term_row"], column=blk["dim_col"]).value
        )
        if not term:
            continue
        for r in range(layout["data_start"], ws.max_row + 1):
            dim = _norm_str(ws.cell(row=r, column=blk["dim_col"]).value)
            sub_dim = (
                _norm_str(ws.cell(row=r, column=blk["sub_dim_col"]).value)
                if blk["sub_dim_col"] is not None else None
            )
            verbatim = _norm_str(
                ws.cell(row=r, column=blk["verbatim_col"]).value
            )
            reference = _norm_str(ws.cell(row=r, column=blk["ref_col"]).value)
            tags = _norm_str(ws.cell(row=r, column=blk["tags_col"]).value)

            # Skip rows that are entirely empty for this block.
            if not (dim or sub_dim or verbatim or reference or tags):
                continue
            # A row only counts as a verbatim record when it has actual quoted
            # text. Header/spacer rows with just "Reference" labels are skipped.
            if not verbatim:
                continue

            fallback_law = _juris_to_law(blk["juris"])
            refs = _explode_reference_cell(reference, fallback_law)
            addr = f"{_col_letter(blk['verbatim_col'])}{r}"
            for ref in refs:
                out.append({
                    "term": term,
                    "law_id": ref["law_id"],
                    "article_id": ref["article_id"],
                    "dim_label": dim,
                    "sub_dim_label": sub_dim,
                    "verbatim_text": verbatim,
                    "reference": reference,
                    "tags": tags,
                    "jurisdiction_header": blk["juris"],
                    "sheet": ws.title,
                    "addr": addr,
                })
    return out


def load_verbatim(path: Path | str | None = None) -> pd.DataFrame:
    """Read the verbatim Excel into a normalised DataFrame.

    Each row corresponds to one (term, law_id, article_id) verbatim citation.
    Cells that are empty in the source remain Python ``None`` rather than the
    string ``"nan"`` or numpy ``NaN``.
    """
    src = Path(path) if path else DEFAULT_VERBATIM
    wb = load_workbook(src, data_only=True, read_only=True)
    rows: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows.extend(_load_verbatim_sheet(ws))
    wb.close()
    df = pd.DataFrame(rows, columns=VERBATIM_COLUMNS)
    return df.astype(object).where(df.notna(), None)


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #

def _print_summary(df: pd.DataFrame, name: str) -> None:
    print(f"\n=== {name} ({len(df)} rows) ===")
    if len(df):
        with pd.option_context(
            "display.max_columns", None,
            "display.max_colwidth", 60,
            "display.width", 200,
        ):
            print(df.head(5).to_string(index=False))
        unique_terms = df["term"].dropna().nunique()
        unique_laws = df["law_id"].dropna().nunique()
        print(
            f"  unique terms: {unique_terms} | "
            f"unique law_ids: {unique_laws} | "
            f"rows with article_id: "
            f"{int(df['article_id'].notna().sum())}"
        )


def main(argv: list[str] | None = None) -> int:
    import argparse
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--analysis", default=str(DEFAULT_ANALYSIS))
    p.add_argument("--verbatim", default=str(DEFAULT_VERBATIM))
    args = p.parse_args(argv)

    print(f"Loading analyses: {args.analysis}")
    a = load_analyses(args.analysis)
    _print_summary(a, "ANALYSES")

    print(f"\nLoading verbatim: {args.verbatim}")
    v = load_verbatim(args.verbatim)
    _print_summary(v, "VERBATIM")
    return 0


if __name__ == "__main__":
    sys.exit(main())
