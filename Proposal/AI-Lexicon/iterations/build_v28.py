"""build_v28.py — Digital AI Lexicon v28.

Scaffold-only release: identical behaviour to v27 (byte-equivalent output)
so that subsequent user stories (article-link audit, superscript FLOPs,
terminology fixes) can land as targeted modifications on a fresh build
file rather than a heavily-overloaded v27 module.

Lands on top of v26 with the same xlsx-driven copy swaps as v27:

  1. Home — replace the intro/landing prose with the three body
     paragraphs from About!A1 (verbatim, drop the heading line).
  2. Regulations — replace each `law-card-v2-desc` with the matching
     Description cell from Methodology rows 18-29; refresh the
     `Effective` meta date from column D.
  3. Methodology — replace each `method-step` body with the matching
     paragraph + numbered list from Methodology!A1 (steps 1-6).
     Step 3 keeps the v26-added GL bullet as the 5th list-item.

v28 (this scaffold) operates as a pure HTML post-process on v26:
    1. Read v26 HTML.
    2. Read xlsx sheets via openpyxl.
    3. Perform precise `html.replace()` swaps with idempotency guards.
    4. Write digital_lexicon_v28.html and mirror to ../final_tool.html
       and ../final_lexicon_tool.html.

Build chain:
    v13 → v15 → v16 → v17 → v18 → v20 → v21 → v22 → v23 → v24 → v25 →
    v26 → v27 → **v28**.

Run:
    python3 build_v28.py
"""
from __future__ import annotations

import html as _html
import re
import shutil
import sys
from pathlib import Path

import openpyxl

HERE          = Path(__file__).parent
HTML_V26      = HERE / "digital_lexicon_v26.html"
HTML_V28      = HERE / "digital_lexicon_v28.html"
FINAL_TOOL    = HERE.parent / "final_tool.html"
FINAL_LEXICON = HERE.parent / "final_lexicon_tool.html"
XLSX          = Path("/Users/robertpraas/Downloads/"
                     "Cross-checked_AI terminology and taxonomy_analysis.xlsx")


# --------------------------------------------------------------------------- #
# xlsx helpers                                                                #
# --------------------------------------------------------------------------- #

def _esc(text: str) -> str:
    """HTML-escape a body of plain xlsx text. Preserve curly quotes,
    em-dashes and Unicode superscripts as-is; only convert reserved
    HTML characters (`&`, `<`, `>`, `"`)."""
    return _html.escape(text, quote=True)


def _date_str(dt) -> str:
    """Render an Effective-from date as 'D MMM YYYY'."""
    if dt is None:
        return ""
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{dt.day} {months[dt.month - 1]} {dt.year}"


# --------------------------------------------------------------------------- #
# Exponent / superscript rendering                                            #
# --------------------------------------------------------------------------- #

_SUP_DIGITS  = str.maketrans("0123456789-+", "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺")
_INV_SUP     = str.maketrans("⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺", "0123456789-+")

# `10(^25)` parenthesised caret — used in some law-blob verbatim cells.
_RE_PAREN_CARET = re.compile(r'(\d+)\(\^(-?\d+)\)')
# `10^25` and `10**25`.
_RE_PLAIN_EXP   = re.compile(r'(\d+)(?:\^|\*\*)(-?\d+)')
# Run of unicode superscript digits / signs.
_RE_UNI_SUP     = re.compile(r'([⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺]+)')

# `<script ...>...</script>` blocks (non-greedy, multiline).
_RE_SCRIPT_BLOCK = re.compile(
    r'(<script\b[^>]*>)(.*?)(</script\s*>)',
    re.IGNORECASE | re.DOTALL,
)


def _exponents_to_sup_html(text: str) -> str:
    """Convert every exponent form to `<sup>…</sup>` markup.

    Handles:
      * `10²⁵` (Unicode-superscript run)
      * `10^25`, `10**25`
      * `10(^25)`
    Existing `<sup>` markup is left untouched (the patterns only match
    digit-bounded forms, never literal `<sup>`)."""
    text = _RE_UNI_SUP.sub(
        lambda m: f'<sup>{m.group(1).translate(_INV_SUP)}</sup>',
        text,
    )
    text = _RE_PAREN_CARET.sub(
        lambda m: f'{m.group(1)}<sup>{m.group(2)}</sup>',
        text,
    )
    text = _RE_PLAIN_EXP.sub(
        lambda m: f'{m.group(1)}<sup>{m.group(2)}</sup>',
        text,
    )
    return text


def _ascii_exp_to_unicode(text: str) -> str:
    """Convert ASCII exponent forms to Unicode-superscript characters.

    Used inside `<script>` blocks: the law-blob verbatim drawer renders
    via `textContent` (see `drawer-verbatim` assignment), so embedded
    `<sup>` markup would show literally. Unicode superscripts render
    correctly in both `innerHTML` and `textContent` paths."""
    text = _RE_PAREN_CARET.sub(
        lambda m: m.group(1) + m.group(2).translate(_SUP_DIGITS),
        text,
    )
    text = _RE_PLAIN_EXP.sub(
        lambda m: m.group(1) + m.group(2).translate(_SUP_DIGITS),
        text,
    )
    return text


def apply_superscripts(html: str) -> tuple[str, dict]:
    """Two-pass exponent rendering pipeline.

    Outside `<script>` blocks: convert ASCII (`10^25`, `10**25`,
    `10(^25)`) and Unicode (`10²⁵`) forms to `<sup>` markup.

    Inside `<script>` blocks (CONCEPTS literal + law-blob JSON):
    convert ASCII forms to Unicode superscripts. Unicode renders as
    proper superscript characters via both innerHTML and textContent.
    Existing Unicode forms are left untouched.

    Returns (html, stats) so the build log can report what changed.
    """
    stats = {
        "static_unicode_to_sup":  0,
        "static_ascii_to_sup":    0,
        "script_ascii_to_unicode": 0,
    }

    parts: list[str] = []
    pos = 0
    for m in _RE_SCRIPT_BLOCK.finditer(html):
        # ---- Outside-script segment ------------------------------------ #
        outside = html[pos:m.start()]
        stats["static_unicode_to_sup"] += len(_RE_UNI_SUP.findall(outside))
        stats["static_ascii_to_sup"]   += (
            len(_RE_PAREN_CARET.findall(outside))
            + len(_RE_PLAIN_EXP.findall(outside))
        )
        parts.append(_exponents_to_sup_html(outside))

        # ---- Script segment -------------------------------------------- #
        open_tag, body, close_tag = m.groups()
        stats["script_ascii_to_unicode"] += (
            len(_RE_PAREN_CARET.findall(body))
            + len(_RE_PLAIN_EXP.findall(body))
        )
        parts.append(open_tag)
        parts.append(_ascii_exp_to_unicode(body))
        parts.append(close_tag)
        pos = m.end()

    # Tail (after the last script block, or full text if no scripts).
    tail = html[pos:]
    stats["static_unicode_to_sup"] += len(_RE_UNI_SUP.findall(tail))
    stats["static_ascii_to_sup"]   += (
        len(_RE_PAREN_CARET.findall(tail))
        + len(_RE_PLAIN_EXP.findall(tail))
    )
    parts.append(_exponents_to_sup_html(tail))

    return ''.join(parts), stats


def load_xlsx_data():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ---- About sheet ------------------------------------------------------- #
    about_a1 = wb["About the Digital AI Lexicon"]["A1"].value
    # Split into paragraphs separated by blank lines. The first is the H1.
    blocks = [b.strip() for b in about_a1.split("\n\n") if b.strip()]
    # blocks[0] = "About the Digital AI Lexicon tool"
    # blocks[1..3] = the three body paragraphs
    # The xlsx contains "NDICI FPN FPI /2022/432 -762" with an extra
    # space; normalise to the canonical "NDICI FPN FPI /2022/432-762".
    body_paras = [
        p.replace("/2022/432 -762", "/2022/432-762").rstrip()
        for p in blocks[1:4]
    ]

    # ---- Methodology sheet ------------------------------------------------- #
    ws = wb["Methodology"]

    # The A1 cell holds all six step blocks.
    method_a1 = ws["A1"].value

    # Slice A1 by step header. Headers begin with "Step N. ".
    steps_text: dict[int, str] = {}
    headers = [f"Step {i}." for i in range(1, 7)]
    for i, hdr in enumerate(headers):
        start = method_a1.index(hdr)
        end = (method_a1.index(headers[i + 1])
               if i + 1 < len(headers) else len(method_a1))
        steps_text[i + 1] = method_a1[start:end].strip()

    # Regulations table rows 18-29.
    regs: dict[int, dict] = {}
    for r in range(18, 30):
        regs[r] = {
            "title":  ws.cell(row=r, column=2).value,
            "desc":   ws.cell(row=r, column=3).value,
            "date":   ws.cell(row=r, column=4).value,
        }

    return {
        "about_paras":  body_paras,
        "steps":        steps_text,
        "regs":         regs,
    }


# --------------------------------------------------------------------------- #
# Step text -> HTML rendering                                                 #
# --------------------------------------------------------------------------- #

def _render_step_body(step_text: str) -> tuple[str, str | None, list[str], str | None]:
    """Parse a step block (e.g. 'Step 1. Title\\n<lead>\\n1. Foo\\n...').

    Return (heading_title, lead_paragraph, items, trailing_paragraph).
    `items` is a list of plain-text numbered-list entries.
    """
    lines = step_text.splitlines()
    # First line: "Step N. <heading title>"
    head_line = lines[0].strip()
    # heading title: text after the first ". "
    _, _, head_title = head_line.partition(". ")
    head_title = head_title.strip()

    body_lines = lines[1:]

    lead_parts: list[str] = []
    items: list[str] = []
    trail_parts: list[str] = []
    seen_first_item = False
    seen_last_item = False

    def _is_numbered(s: str) -> bool:
        s = s.lstrip()
        if len(s) < 2 or not s[0].isdigit():
            return False
        # support "1." "1\t" "1.\t" "1)"
        for sep in (".", "\t", ")"):
            if sep in s[:3]:
                return True
        return False

    # In the xlsx, each numbered item lives on its own line. As soon as we
    # see a non-numbered line after we've started the list, close the list
    # and treat the remaining lines as trailing paragraph content.
    in_list = False
    for raw in body_lines:
        line = raw.rstrip()
        if not line.strip():
            # blank line — separator only
            continue

        if _is_numbered(line):
            in_list = True
            seen_first_item = True
            # strip the leading "N." or "N.\t" or "N)" then whitespace
            stripped = line.lstrip()
            i = 0
            while i < len(stripped) and stripped[i].isdigit():
                i += 1
            if i < len(stripped) and stripped[i] in ".)\t":
                i += 1
            stripped = stripped[i:].lstrip()
            items.append(stripped)
            continue

        if not in_list:
            lead_parts.append(line.strip())
        else:
            # Non-numbered line after numbered list — trailing paragraph.
            trail_parts.append(line.strip())
            seen_last_item = True

    lead = " ".join(lead_parts).strip() or None
    trail = " ".join(trail_parts).strip() or None
    return head_title, lead, items, trail


def _step_body_html(num: int, lead: str | None,
                    items: list[str], trail: str | None,
                    extra_li_html: str | None = None) -> str:
    """Render the inner HTML of a `.method-step-body` (without the
    enclosing `<h2>` heading — the script preserves the existing h2)."""
    out: list[str] = []
    if lead:
        out.append(f"        <p>{_esc(lead)}</p>")
    if items:
        out.append("        <ol>")
        for it in items:
            out.append(f"          <li>{_esc(it)}</li>")
        if extra_li_html:
            out.append(extra_li_html.rstrip())
        out.append("        </ol>")
    if trail:
        out.append(f"        <p>{_esc(trail)}</p>")
    return "\n".join(out)


# --------------------------------------------------------------------------- #
# v26 anchor strings -> v28 replacement strings                               #
# --------------------------------------------------------------------------- #

# ---- HOME ----------------------------------------------------------------- #
HOME_TAGLINE_OLD = (
    '    <p class="landing-tagline">The Digital AI Lexicon is a searchable '
    'digital database on key AI governance terms present in regulatory '
    'frameworks on AI in the European Union and the United States.</p>\n'
    '    <div class="landing-sub about-body"><p>The Digital AI Lexicon takes '
    'the EU\'s AI Act as a benchmark, juxtaposing and organising the other '
    'jurisdictions\' terms based on the EU regulatory framework for AI. It '
    'covers a total of 43 terms across 12 regulatory frameworks and relevant '
    'guidance documents. For each term, it catalogues the definition, scope, '
    'enforcement context, and authoritative source, with side-by-side EU–US '
    'comparisons and interpretive notes.</p>\n'
    '<p>The Digital AI Lexicon allows the user to consult key AI governance '
    'terms in two modes:</p>\n'
    '<ul class="about-bullets"><li>Comparative analysis: provides at-a-glance '
    'comparative tables alongside interpretative notes.</li>\n'
    '<li>Legal text: provides side-by-side verbatim regulatory text, keeping '
    'the original detail and language of selected regulations.</li></ul>\n'
    '<p>The Digital AI Lexicon was produced by the Centre for European Policy '
    'Studies (CEPS) in the context of the EU-US Trade and Technology Dialogue '
    '(TTD), funded by the European Commission. Study reference: NDICI FPN FPI '
    '/2022/432 -762.</p></div>\n'
)


def _build_home_block(about_paras: list[str]) -> str:
    p1, p2, p3 = about_paras  # body paragraphs from About!A1
    tagline_html = (
        f'    <p class="landing-tagline">{_esc(p1)}</p>\n'
    )
    sub_html = (
        '    <div class="landing-sub about-body">'
        f'<p>{_esc(p2)}</p>\n'
        f'<p>{_esc(p3)}</p></div>\n'
    )
    return tagline_html + sub_html


# ---- REGULATION CARDS ----------------------------------------------------- #
# We swap each card by its unique title anchor inside <article class=...>.
# Each card maps to an xlsx Methodology row.
CARD_ANCHORS: list[tuple[str, int]] = [
    # (title-substring used to locate the card, xlsx Methodology row)
    ('Artificial Intelligence Act (2024)', 18),
    ('Guidelines on the scope of the obligations for general-purpose '
     'AI models established by AIA (2025)', 20),
    ('Code of Practice for General-purpose AI Models (2025)', 19),
    ('SB 53 (2025) — Transparency in Frontier Artificial Intelligence (TFAI)', 21),
    ('SB 942 (2024) — California AI Transparency Act', 22),
    ('SB 24-205 (2024) — Colorado AI Act (CAIA)', 23),
    ('SB 25B-004 (2026)', 24),
    ('A6453B (2025) — Responsible AI Safety and Education Act (RAISE Act)', 25),
    ('S8828 (2025)', 26),
    ('HB 149 (2025) — Texas Responsible Artificial Intelligence '
     'Governance Act (TRAIGA)', 27),
    ('SB 149 (2024) — Artificial Intelligence Policy Act (AIPA)', 28),
    ('SB 226 (2025)', 29),
]


def _format_desc_html(text: str) -> str:
    """Render an xlsx Description cell as one or more <p
    class="law-card-v2-desc"> blocks. Multi-paragraph cells are split
    on blank lines."""
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    if not paragraphs:
        paragraphs = [text.strip()]
    out_paras = []
    for p in paragraphs:
        # Replace any internal single newlines with a space.
        clean = " ".join(line.strip() for line in p.splitlines()).strip()
        out_paras.append(f'<p class="law-card-v2-desc">{_esc(clean)}</p>')
    return "\n        ".join(out_paras)


def _replace_card(html: str, title_substr: str, row: dict) -> str:
    """Locate the law-card-v2 article whose title contains `title_substr`,
    then replace its description and Effective-date meta entry."""
    title_pos = html.find(title_substr)
    if title_pos == -1:
        raise RuntimeError(
            f"build_v28: card title not found: {title_substr!r}"
        )
    # Walk back to the <article class="law-card-v2"...> open tag.
    art_open = html.rfind("<article class=\"law-card-v2", 0, title_pos)
    if art_open == -1:
        raise RuntimeError(
            f"build_v28: <article> open not found for {title_substr!r}"
        )
    art_close = html.find("</article>", title_pos)
    if art_close == -1:
        raise RuntimeError(
            f"build_v28: </article> close not found for {title_substr!r}"
        )
    card = html[art_open:art_close]

    # ---- Replace description block --------------------------------------- #
    new_desc_html = _format_desc_html(row["desc"])
    # The description block is a single <p class="law-card-v2-desc">...</p>
    # in v26. Find and replace it.
    desc_open = card.find('<p class="law-card-v2-desc">')
    if desc_open == -1:
        raise RuntimeError(
            f"build_v28: <p class=\"law-card-v2-desc\"> not found in card "
            f"for {title_substr!r}"
        )
    desc_close = card.find("</p>", desc_open) + len("</p>")
    new_card = card[:desc_open] + new_desc_html + card[desc_close:]

    # ---- Replace Effective date entry ------------------------------------ #
    if row["date"] is not None:
        new_eff = _date_str(row["date"])
        eff_idx = new_card.find("<strong>Effective:</strong>")
        if eff_idx == -1:
            raise RuntimeError(
                f"build_v28: <strong>Effective:</strong> not found in card "
                f"for {title_substr!r}"
            )
        # The line is e.g. '<strong>Effective:</strong> 1 Aug 2024</span>'.
        # Replace from after the </strong>+space up to the next </span>.
        after_label = new_card.index("</strong>", eff_idx) + len("</strong>")
        # skip exactly one space if present
        i = after_label
        if i < len(new_card) and new_card[i] == " ":
            i += 1
        end_span = new_card.index("</span>", i)
        new_card = (
            new_card[:after_label] + " " + new_eff + new_card[end_span:]
        )

    return html[:art_open] + new_card + html[art_close:]


# ---- METHODOLOGY STEPS ---------------------------------------------------- #
# Each method-step body is replaced by the parsed step text. Step 3 keeps
# the v26 GL bullet as the 5th `<li>`.

GL_LI_HTML = (
    '          <li>Cross-reference the <strong>Commission Guidelines '
    'on the scope of obligations for providers of general-purpose '
    'AI models</strong> (C(2025) 7719, 19 Nov 2025; cited as '
    '&quot;(GL)&quot; in the comparative analysis).</li>'
)


def _replace_step_body(html: str, num: int, step_text: str,
                       extra_li: str | None = None) -> str:
    """Replace the body content (everything between </h2> and the closing
    </div></section>) of method-step #step-N with rendered xlsx text."""
    anchor = f'id="step-{num}"'
    a_pos = html.find(anchor)
    if a_pos == -1:
        raise RuntimeError(f"build_v28: step-{num} anchor not found")
    sec_open = html.rfind("<section", 0, a_pos)
    sec_close = html.find("</section>", a_pos)
    if sec_open == -1 or sec_close == -1:
        raise RuntimeError(f"build_v28: step-{num} section bounds not found")

    section = html[sec_open:sec_close]
    # Find </h2> to preserve the heading.
    h2_close = section.find("</h2>")
    if h2_close == -1:
        raise RuntimeError(f"build_v28: step-{num} </h2> not found")
    after_h2 = h2_close + len("</h2>")

    # Find the closing </div> of method-step-body. The structure is:
    #   <section ...>
    #     <div class="method-step-num"...>NN</div>
    #     <div class="method-step-body">
    #       <h2>...</h2>
    #       <p>...</p>
    #       ...
    #     </div>
    #   </section>
    # So we need the LAST </div> before </section>.
    body_div_close = section.rfind("</div>")
    if body_div_close == -1 or body_div_close < after_h2:
        raise RuntimeError(f"build_v28: step-{num} body </div> not found")

    head_title, lead, items, trail = _render_step_body(step_text)
    body_html = _step_body_html(num, lead, items, trail, extra_li_html=extra_li)

    new_section = (
        section[:after_h2]
        + "\n"
        + body_html
        + "\n      "
        + section[body_div_close:]
    )
    return html[:sec_open] + new_section + html[sec_close:]


# --------------------------------------------------------------------------- #
# Terminology fixes (US-004)                                                  #
# --------------------------------------------------------------------------- #

# Opening category-table EU cell for the limited-risk Deployer row carries the
# bare label "Deployer" in v26. The Excel reference (Provider_Developer /
# Deployer_Supplier_Analysis sheets) names this anchor "Deployer of
# limited-risk AI systems", matching the Provider row's EU cell which already
# uses the full label. The Colorado/Texas Deployer cells (bill="SB24-205" /
# bill="HB149") stay as bare "Deployer" — those are the U.S.-state equivalents
# per Excel.
LIMITED_RISK_DEPLOYER_EU_OLD = (
    '"name":"Deployer","bill":"","sub_id":"deployer","jid":"eu"'
)
LIMITED_RISK_DEPLOYER_EU_NEW = (
    '"name":"Deployer of limited-risk AI systems",'
    '"bill":"","sub_id":"deployer","jid":"eu"'
)


def apply_terminology_fixes(html: str) -> tuple[str, dict]:
    """Apply opening-category-table terminology corrections.

    Currently only fixes the EU cell for the limited-risk Deployer row.
    The Provider row's EU cell already reads "Provider of limited-risk AI
    systems" in v26, and detail (per-jurisdiction) tables are out of scope
    for this story.
    """
    stats = {"limited_risk_deployer_eu": 0}
    if LIMITED_RISK_DEPLOYER_EU_OLD in html:
        html = html.replace(
            LIMITED_RISK_DEPLOYER_EU_OLD,
            LIMITED_RISK_DEPLOYER_EU_NEW,
            1,
        )
        stats["limited_risk_deployer_eu"] = 1
    elif LIMITED_RISK_DEPLOYER_EU_NEW in html:
        # Idempotent — already applied.
        pass
    else:
        raise RuntimeError(
            "build_v28: limited-risk Deployer EU cell anchor not found"
        )
    return html, stats


# --------------------------------------------------------------------------- #
# Limited-risk Provider table cell fixes (US-005)                              #
# --------------------------------------------------------------------------- #
#
# Excel `Provider_Developer_Analysis` rows 1-19 (sub-table §4.1, "Limited-risk
# Provider / Developer") is the source of truth. The v26 baseline matches
# Excel for analysis text in every cell, but several cells have empty or
# wrong verbatim/reference fields, breaking the popup ("verbatim drawer"):
#
#   * Scope EU — verbatim is Article 3 (3) (general "provider" definition);
#     the analysis cell cites Article 50, so the popup links to the wrong
#     article. Replace with Article 50 (1) text.
#   * Reg trigger TX — analysis cites 552.103, but verbatim/ref are empty.
#   * Transparency EU — analysis cites Article 50, verbatim/ref empty.
#   * Transparency CO — analysis cites §6-1-1704, verbatim/ref empty
#     (CO law-blob has no embedded sections, so we add the reference label
#      only and rely on the drawer fallback for the verbatim).
#   * General info disclosure TX — analysis cites 552.103, verbatim/ref empty.
#   * Risk management TX — analysis cites 552.103, verbatim/ref empty.
#   * AI literacy EU — analysis cites Article 4, verbatim/ref empty.
#
# Provider/dev info TX cell cites "(§6-1-1702.)" — which is a Colorado
# section appearing in the Texas column. This is an Excel-side oddity that
# we preserve verbatim per the Excel-wins rule (see v28_excel_inventory.md
# §4.1); leaving verbatim/ref empty avoids fabricating a non-existent
# Texas section.

# EU AI Act Article 50 (1) — limited-risk transparency obligation.
_AIA_ART_50_1 = (
    '1. Providers shall ensure that AI systems intended to interact directly '
    'with natural persons are designed and developed in such a way that the '
    'natural persons concerned are informed that they are interacting with '
    'an AI system, unless this is obvious from the point of view of a '
    'natural person who is reasonably well-informed, observant and '
    'circumspect, taking into account the circumstances and the context of '
    'use. This obligation shall not apply to AI systems authorised by law to '
    'detect, prevent, investigate or prosecute criminal offences, subject to '
    'appropriate safeguards for the rights and freedoms of third parties, '
    'unless those systems are available for the public to report a criminal '
    'offence.'
)

# EU AI Act Article 50 (1) + (2) — interaction disclosure + machine-readable
# marking of synthetic content (the two transparency obligations applicable
# to limited-risk providers).
_AIA_ART_50_1_2 = (
    _AIA_ART_50_1
    + '\n\n2. Providers of AI systems, including general-purpose AI '
    'systems, generating synthetic audio, image, video or text content, '
    'shall ensure that the outputs of the AI system are marked in a '
    'machine-readable format and detectable as artificially generated or '
    'manipulated. Providers shall ensure their technical solutions are '
    'effective, interoperable, robust and reliable as far as this is '
    'technically feasible, taking into account the specificities and '
    'limitations of various types of content, the costs of implementation '
    'and the generally acknowledged state of the art, as may be reflected '
    'in relevant technical standards.'
)

# EU AI Act Article 4 — AI literacy obligation (full text).
_AIA_ART_4 = (
    'Providers and deployers of AI systems shall take measures to ensure, '
    'to their best extent, a sufficient level of AI literacy of their '
    'staff and other persons dealing with the operation and use of AI '
    'systems on their behalf, taking into account their technical '
    'knowledge, experience, education and training and the context the AI '
    'systems are to be used in, and considering the persons or groups of '
    'persons on whom the AI systems are to be used.'
)

# Texas HB 149 §552.103 (a)+(b) — investigative authority + the disclosure
# items the AG can request. The same section underpins three different
# cells (Reg trigger, General info, Risk management); the cells differ
# only in which subsection of (b) their analysis text emphasises.
_TX_552_103_A = (
    '(a) If the attorney general receives a complaint through the online '
    'mechanism under Section 552.102 alleging a violation of this chapter, '
    'the attorney general may issue a civil investigative demand to '
    'determine if a violation has occurred. The attorney general shall '
    'issue demands in accordance with and under the procedures established '
    'under Section 15.10.'
)
_TX_552_103_B_GENERAL = (
    '(b) The attorney general may request from the person reported through '
    'the online mechanism, pursuant to a civil investigative demand issued '
    'under Subsection (a):\n'
    '(1) a high-level description of the purpose, intended use, deployment '
    'context, and associated benefits of the artificial intelligence '
    'system with which the person is affiliated;\n'
    '(2) a description of the type of data used to program or train the '
    'artificial intelligence system;\n'
    '(3) a high-level description of the categories of data processed as '
    'inputs for the artificial intelligence system;\n'
    '(4) a high-level description of the outputs produced by the '
    'artificial intelligence system; (5) any metrics the person uses to '
    'evaluate the performance of the artificial intelligence system.'
)
_TX_552_103_B_RISK = (
    '(b) The attorney general may request from the person reported through '
    'the online mechanism, pursuant to a civil investigative demand issued '
    'under Subsection (a):\n'
    '(6) a high-level description of any post-deployment monitoring of the '
    'artificial intelligence system, where applicable;\n'
    '(7) a high-level description of user safeguards in place for the '
    'artificial intelligence system.'
)


def _json_str(value: str) -> str:
    """Encode a Python string as a JSON-string LITERAL (without surrounding
    quotes), matching the encoding used by the v26 CONCEPTS literal —
    backslash-escape `"` and `\\`, encode line-feed as `\\n`."""
    return (
        value
        .replace('\\', '\\\\')
        .replace('"', '\\"')
        .replace('\n', '\\n')
    )


# Each fix is (description, old_substring, new_substring). The old_substring
# is the unique JSON cell triple `{"analysis":...,"verbatim":...,"reference":...}`
# wrapped in its jurisdiction key so that the cell is unambiguously located.
# Constructed programmatically to keep escape-sequence handling explicit.

def _build_lr_fixes() -> list[tuple[str, str, str]]:
    fixes: list[tuple[str, str, str]] = []

    # 1. Scope EU — replace Article 3 (3) verbatim with Article 50 (1).
    scope_eu_analysis = (
        'A person that develops or has developed or makes a substantial '
        'modification to and places on the market of an AI system that '
        'interacts with natural persons (Article 50)'
    )
    scope_eu_old_verbatim = (
        '(3) "provider" means a natural or legal person, public authority, '
        'agency or other body that develops an AI system or a '
        'general-purpose AI model or that has an AI system or a '
        'general-purpose AI model developed and places it on the market or '
        'puts the AI system into service under its own name or trademark, '
        'whether for payment or free of charge.'
    )
    fixes.append((
        "Scope EU: verbatim Article 3 (3) -> Article 50 (1)",
        f'"eu":{{"analysis":"{_json_str(scope_eu_analysis)}",'
        f'"verbatim":"{_json_str(scope_eu_old_verbatim)}",'
        f'"reference":"EU AI Act, Article 3 (3)"}}',
        f'"eu":{{"analysis":"{_json_str(scope_eu_analysis)}",'
        f'"verbatim":"{_json_str(_AIA_ART_50_1)}",'
        f'"reference":"EU AI Act, Article 50 (1)"}}',
    ))

    # 2. Regulatory trigger TX — add 552.103 (a) verbatim and reference.
    rt_tx_analysis = (
        'Upon request by attorney general, following investigative demand '
        '(552.103.)'
    )
    fixes.append((
        "Reg trigger TX: add 552.103 (a) verbatim/ref",
        f'"tx":{{"analysis":"{_json_str(rt_tx_analysis)}",'
        f'"verbatim":"","reference":""}}',
        f'"tx":{{"analysis":"{_json_str(rt_tx_analysis)}",'
        f'"verbatim":"{_json_str(_TX_552_103_A)}",'
        f'"reference":"Texas HB149, 552.103. (a)"}}',
    ))

    # 3. Transparency EU — add Article 50 (1)+(2) verbatim and reference.
    transp_eu_analysis = (
        'Disclose AI interaction;\n'
        'If generating synthetic content, put in place machine-readable '
        'disclosure that content is AI-generated (Article 50)'
    )
    fixes.append((
        "Transparency EU: add Article 50 (1)+(2) verbatim/ref",
        f'"eu":{{"analysis":"{_json_str(transp_eu_analysis)}",'
        f'"verbatim":"","reference":""}}',
        f'"eu":{{"analysis":"{_json_str(transp_eu_analysis)}",'
        f'"verbatim":"{_json_str(_AIA_ART_50_1_2)}",'
        f'"reference":"EU AI Act, Article 50 (1, 2)"}}',
    ))

    # 4. Transparency CO — add reference only. Colorado law-blob has no
    #    embedded sections, so we don't fabricate a verbatim quote; the
    #    drawer falls back to showing the analysis text.
    transp_co_analysis = (
        'Must disclose to consumers interaction with AI system (§6-1-1704)'
    )
    fixes.append((
        "Transparency CO: add §6-1-1704 reference label",
        f'"co":{{"analysis":"{_json_str(transp_co_analysis)}",'
        f'"verbatim":"","reference":""}}',
        f'"co":{{"analysis":"{_json_str(transp_co_analysis)}",'
        f'"verbatim":"",'
        f'"reference":"Colorado SB24-205, 6-1-1704"}}',
    ))

    # 5. General information disclosure TX — add 552.103 (b)(1)-(5) verbatim.
    gid_tx_analysis = (
        'High-level description of the system, training data, outputs and '
        'metrics for system evaluation (552.103.)'
    )
    fixes.append((
        "Gen info disclosure TX: add 552.103 (b)(1-5) verbatim/ref",
        f'"tx":{{"analysis":"{_json_str(gid_tx_analysis)}",'
        f'"verbatim":"","reference":""}}',
        f'"tx":{{"analysis":"{_json_str(gid_tx_analysis)}",'
        f'"verbatim":"{_json_str(_TX_552_103_B_GENERAL)}",'
        f'"reference":"Texas HB149, 552.103. (b)"}}',
    ))

    # 6. Risk management TX — add 552.103 (b)(6)-(7) verbatim.
    rm_tx_analysis = (
        'High-level description of the post-deployment monitoring and user '
        'safeguards (552.103.)'
    )
    fixes.append((
        "Risk management TX: add 552.103 (b)(6-7) verbatim/ref",
        f'"tx":{{"analysis":"{_json_str(rm_tx_analysis)}",'
        f'"verbatim":"","reference":""}}',
        f'"tx":{{"analysis":"{_json_str(rm_tx_analysis)}",'
        f'"verbatim":"{_json_str(_TX_552_103_B_RISK)}",'
        f'"reference":"Texas HB149, 552.103. (b)"}}',
    ))

    # 7. AI literacy EU — add Article 4 verbatim/ref.
    ail_eu_analysis = (
        'Providers must ensure people operating / using AI systems on their '
        'behalf have a sufficient level of AI literacy (Article 4)'
    )
    fixes.append((
        "AI literacy EU: add Article 4 verbatim/ref",
        f'"eu":{{"analysis":"{_json_str(ail_eu_analysis)}",'
        f'"verbatim":"","reference":""}}',
        f'"eu":{{"analysis":"{_json_str(ail_eu_analysis)}",'
        f'"verbatim":"{_json_str(_AIA_ART_4)}",'
        f'"reference":"EU AI Act, Article 4"}}',
    ))

    return fixes


_LR_PROVIDER_START = (
    '"id":"provider","title":"Provider of limited-risk AI systems"'
)
_LR_PROVIDER_END = '"id":"provider-of-high-risk-ai-systems"'


def apply_limited_risk_provider_fixes(html: str) -> tuple[str, dict]:
    """Patch the seven cells in the limited-risk Provider/Developer
    analysis table whose verbatim/reference fields disagree with the
    article references cited in their analysis text.

    Replacements are scoped to the limited-risk Provider sub-concept JSON
    segment — bounded above by the sub-concept header (`"id":"provider",
    "title":"Provider of limited-risk AI systems"`) and below by the next
    sub-concept's header (`"id":"provider-of-high-risk-ai-systems"`). This
    is necessary because several cell triples (analysis text + empty
    verbatim/reference) recur in the limited-risk Deployer table; scoping
    by sub-concept disambiguates them without coupling the anchors to
    dimension ids that encode row position.

    Each fix is idempotent: if the new triple is already present, the
    replacement is silently skipped.
    """
    stats: dict[str, int] = {}
    s = html.find(_LR_PROVIDER_START)
    if s == -1:
        raise RuntimeError(
            "build_v28: limited-risk Provider sub-concept header not found"
        )
    e = html.find(_LR_PROVIDER_END, s)
    if e == -1:
        raise RuntimeError(
            "build_v28: high-risk Provider sub-concept header (used as "
            "limited-risk end-marker) not found"
        )
    section = html[s:e]

    fixes = _build_lr_fixes()
    for desc, old, new in fixes:
        key = desc.split(":")[0].strip().lower().replace(" ", "_")
        if old in section:
            count = section.count(old)
            if count != 1:
                raise RuntimeError(
                    f"build_v28: limited-risk fix '{desc}' anchor matched "
                    f"{count} times within the Provider sub-concept "
                    "(expected exactly 1)"
                )
            section = section.replace(old, new, 1)
            stats[key] = 1
        elif new in section:
            stats[key] = 0  # already applied (idempotent)
        else:
            raise RuntimeError(
                f"build_v28: limited-risk fix '{desc}' anchor not found "
                "within the Provider sub-concept"
            )
    return html[:s] + section + html[e:], stats


# --------------------------------------------------------------------------- #
# EU article-link audit fixes (US-006)                                         #
# --------------------------------------------------------------------------- #
#
# Audited every EU cell in the v28 CONCEPTS literal against the Excel
# inventory (`v28_excel_inventory.md` §4–§6); see `outputs/us006_eu_audit.md`
# for the per-cell findings. Two real mismatches surfaced:
#
#   1. provider-of-high-risk-ai-systems / scope-1-0 — Excel says
#      "Article 3, Article 6, Annex III" but the v26 reference reads
#      "AI Act, Article 6; AI Act, Annex III; EU AI Act, Article 25 (1)".
#      Article 3 (3) (the definition of "provider") is missing, leaving an
#      asymmetry with the deployer analog (which already includes Art 3 (4)).
#      Fix: prepend Article 3 (3) verbatim text + reference label.
#
#   2. provider-of-general-purpose-ai-models / scope-system-1-2 (3rd scope
#      sub-row) — analysis cites "(GL (59))" and "(GL (60))" (the modification
#      criteria for GPAI providers per Excel §4.3) but the reference field
#      reads only "(GL, (17))" — the same string used by the unrelated
#      compute-threshold cell. Fix: replace the reference with
#      "(GL, (59)); (GL, (60))" so the popup label points at the right GL
#      paragraphs. Verbatim stays empty (the GL law-blob has no embedded
#      sections; popup falls back to the analysis text).
#
# Both fixes are scoped via long, unique `analysis` text anchors so the
# replacement is naturally restricted to the intended cell.

# Article 3 (3) — definition of "provider" (used in fix #1 only).
_AIA_ART_3_3_PROVIDER = (
    '(3) "provider" means a natural or legal person, public authority, '
    'agency or other body that develops an AI system or a general-purpose '
    'AI model or that has an AI system or a general-purpose AI model '
    'developed and places it on the market or puts the AI system into '
    'service under its own name or trademark, whether for payment or '
    'free of charge.'
)


def _build_eu_article_link_fixes() -> list[tuple[str, str, str]]:
    """Build the (description, old, new) tuples for the two EU article-link
    audit mismatches identified in US-006."""
    fixes: list[tuple[str, str, str]] = []

    # Fix #1: provider-of-high-risk-ai-systems / scope.
    # Anchor on the existing reference string, which appears exactly once
    # in the whole HTML (verified: `grep -o … | wc -l` == 1).
    old_ref_1 = (
        '"reference":"AI Act, Article 6; AI Act, Annex III; '
        'EU AI Act, Article 25 (1)"'
    )
    new_ref_1 = (
        '"reference":"EU AI Act, Article 3 (3); AI Act, Article 6; '
        'AI Act, Annex III; EU AI Act, Article 25 (1)"'
    )
    fixes.append((
        "Provider HR / Scope: add Article 3 (3) to reference",
        old_ref_1,
        new_ref_1,
    ))

    # Fix #2: provider-of-general-purpose-ai-models / scope-system-1-2.
    # The analysis text for this cell is unique and ends with "...the
    # original model (GL (60))". We anchor on the full cell triple
    # (analysis + empty verbatim + reference) so the swap can't collide
    # with the model-system / compute-threshold cell, which uses the same
    # `"reference":"(GL, (17))"` string with a different analysis.
    scope_sys_analysis = (
        'Person that develops a GPAI model that displays significant '
        'generality and is capable of competently performing a wide range '
        'of distinct tasks (Article 3)\n'
        'Indicative criterion: > 1023 FLOPs in compute (GL, (17))\n'
        'Or person that modifies an existent GPAI model and that '
        'modification results in significant change in the model’s '
        'generality, capabilities, or systemic risk (GL (59))\n'
        'Indicative criterion: training compute used for the modification '
        'is greater than 1/3 of the training compute of the original model '
        '(GL (60))'
    )
    old_cell_2 = (
        f'{{"analysis":"{_json_str(scope_sys_analysis)}",'
        f'"verbatim":"","reference":"(GL, (17))"}}'
    )
    new_cell_2 = (
        f'{{"analysis":"{_json_str(scope_sys_analysis)}",'
        f'"verbatim":"","reference":"(GL, (59)); (GL, (60))"}}'
    )
    fixes.append((
        "Provider GPAI / Scope (modification sub-row): "
        "GL (17) -> GL (59), (60) reference",
        old_cell_2,
        new_cell_2,
    ))

    return fixes


def apply_eu_article_link_fixes(html: str) -> tuple[str, dict]:
    """Apply the two cell-level reference fixes flagged by the US-006
    EU-article-link audit. Idempotent: each fix detects an already-applied
    state via the new-substring check and silently skips."""
    fixes = _build_eu_article_link_fixes()
    stats: dict[str, int] = {}
    for desc, old, new in fixes:
        key = desc.split(":")[0].strip().lower().replace(" ", "_").replace("/", "_")
        # Drop double underscores from `__`-runs caused by " / " in desc.
        while "__" in key:
            key = key.replace("__", "_")
        if old in html:
            count = html.count(old)
            if count != 1:
                raise RuntimeError(
                    f"build_v28: EU-link fix '{desc}' anchor matched "
                    f"{count} times (expected exactly 1)"
                )
            html = html.replace(old, new, 1)
            stats[key] = 1
        elif new in html:
            stats[key] = 0  # already applied (idempotent)
        else:
            raise RuntimeError(
                f"build_v28: EU-link fix '{desc}' anchor not found"
            )
    return html, stats


# --------------------------------------------------------------------------- #
# Main build                                                                  #
# --------------------------------------------------------------------------- #

def main() -> None:
    print("== v28 build ==")
    if not HTML_V26.exists():
        print("  digital_lexicon_v26.html missing — running build_v26 …")
        sys.path.insert(0, str(HERE))
        import build_v26 as _v26
        _v26.main()

    html = HTML_V26.read_text(encoding="utf-8")
    print(f"  read v26:                  {len(html):,} bytes")

    data = load_xlsx_data()

    # ---- HOME ------------------------------------------------------------- #
    new_home = _build_home_block(data["about_paras"])
    if HOME_TAGLINE_OLD in html:
        html = html.replace(HOME_TAGLINE_OLD, new_home, 1)
        print("  home:                      swapped tagline + sub block")
    elif new_home in html:
        print("  home:                      already updated (idempotent)")
    else:
        raise RuntimeError("build_v28: home tagline anchor not found")

    # ---- REGULATIONS CARDS ------------------------------------------------ #
    updated: list[str] = []
    for title_substr, row_idx in CARD_ANCHORS:
        row = data["regs"][row_idx]
        if row["desc"] is None:
            continue
        # Idempotency: if the new desc is already present in the card,
        # skip. Use the first 60 chars of the new desc as a fingerprint.
        fingerprint = _esc(row["desc"].split("\n\n")[0].strip())[:60]
        # Find the card and check if its current description starts with
        # the fingerprint.
        title_pos = html.find(title_substr)
        if title_pos == -1:
            raise RuntimeError(f"build_v28: card not found: {title_substr!r}")
        art_close = html.find("</article>", title_pos)
        card = html[html.rfind("<article", 0, title_pos):art_close]
        if fingerprint in card and _date_str(row["date"]) in card:
            updated.append(f"{title_substr.split(' ')[0]}(skip)")
            continue
        html = _replace_card(html, title_substr, row)
        # Short label
        short = title_substr.split(" ")[0] + " " + title_substr.split(" ")[1] \
            if " " in title_substr else title_substr
        updated.append(short)
    print(f"  regs:                      updated {len(updated)} cards")

    # ---- METHODOLOGY STEPS ----------------------------------------------- #
    for n in range(1, 7):
        extra = GL_LI_HTML if n == 3 else None
        html = _replace_step_body(html, n, data["steps"][n], extra_li=extra)
    print("  methodology:               swapped steps 1-6")

    # ---- TERMINOLOGY FIXES ----------------------------------------------- #
    html, term_stats = apply_terminology_fixes(html)
    print(
        "  terminology:               "
        f"limited-risk Deployer EU label fix: {term_stats['limited_risk_deployer_eu']}"
    )

    # ---- LIMITED-RISK PROVIDER TABLE FIXES (US-005) ---------------------- #
    html, lr_stats = apply_limited_risk_provider_fixes(html)
    applied = sum(lr_stats.values())
    print(
        "  limited-risk provider:     "
        f"applied {applied} cell fix(es) "
        f"({len(lr_stats)} target cells)"
    )

    # ---- EU ARTICLE-LINK AUDIT FIXES (US-006) ---------------------------- #
    html, eu_stats = apply_eu_article_link_fixes(html)
    eu_applied = sum(eu_stats.values())
    print(
        "  EU article-link audit:     "
        f"applied {eu_applied} reference fix(es) "
        f"({len(eu_stats)} target cells)"
    )

    # ---- SUPERSCRIPT RENDERING ------------------------------------------- #
    html, sup_stats = apply_superscripts(html)
    print(
        "  superscripts:              "
        f"static unicode→<sup>: {sup_stats['static_unicode_to_sup']}, "
        f"static ASCII→<sup>: {sup_stats['static_ascii_to_sup']}, "
        f"script ASCII→unicode: {sup_stats['script_ascii_to_unicode']}"
    )

    HTML_V28.write_text(html, encoding="utf-8")
    shutil.copy2(HTML_V28, FINAL_TOOL)
    shutil.copy2(HTML_V28, FINAL_LEXICON)
    print(f"\n  wrote {HTML_V28.name}        ({len(html):,} bytes)")
    print(f"  wrote {FINAL_TOOL.name}            (mirror)")
    print(f"  wrote {FINAL_LEXICON.name}    (mirror)")


if __name__ == "__main__":
    main()
