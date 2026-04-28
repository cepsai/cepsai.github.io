"""build_v24.py — Digital AI Lexicon v24.

Builds on v23. Two structural goals:

  1. **Add the "Provider of limited-risk AI systems / Developer" sub-concept**
     fully populated from rows 1-20 of `Provider_Developer_Analysis` in the
     new xlsx. v23 already has a `provider` sub-concept slot but it is
     incomplete (missing the "Incident / risk reporting" dim) and its
     verbatim/reference fields can drift from the OLD xlsx
     `Provider_Developer` legal sheet. This rebuild ensures the slice is:
        - present BEFORE the `provider-of-high-risk-ai-systems` entry,
        - has all 11 dims as in the xlsx,
        - keeps verbatim/reference where v23 already had them, and
        - falls back to "" for dims with no legal text in the OLD xlsx.

  2. **Comprehensive cell-text re-sync** across every other sub-concept
     using a multi-strategy matcher (priority: a) exact (jid, dim_lc)
     within matching xlsx section title; b) (jid, dim_lc) across all
     sections of the same sheet; c) best-fuzzy via normalized title
     overlap). v23's aggressive updater only handled the case where a
     single `(jid_root, dim_label)` hit existed across the whole sheet —
     v24 adds title-aware confirmation and a fuzzy fallback.

Build chain: v13 → v15 → v16 → v17 → v18 → v20 → v21 → v22 → v23 → **v24**.

v24 operates as a post-process on v23.html:
    1. If digital_lexicon_v23.html is missing, run build_v23.
    2. Rebuild/refresh the limited-risk Provider sub-concept from the new
       xlsx + old-xlsx verbatim.
    3. Re-sync remaining cells via the v24 multi-strategy matcher.
    4. Update the matrix `cluster_summary` for `provider-developer` to
       carry the limited-risk row (already present from v22, but we
       refresh in case the xlsx terms changed).
    5. Write digital_lexicon_v24.html + mirror to ../final_tool.html
       and ../final_lexicon_tool.html.

Run:
    python3 build_v24.py
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl

HERE          = Path(__file__).parent
HTML_V23      = HERE / "digital_lexicon_v23.html"
HTML_V24      = HERE / "digital_lexicon_v24.html"
FINAL_TOOL    = HERE.parent / "final_tool.html"
FINAL_LEXICON = HERE.parent / "final_lexicon_tool.html"

NEW_XLSX = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis.xlsx"
)
OLD_XLSX = Path(
    "/Users/robertpraas/CEPS-DST Dropbox/Robert Praas/cb_sep_25/"
    "AI-Lexicon/AI terminology and taxonomy-5.xlsx"
)

sys.path.insert(0, str(HERE))
from build_v23 import _find_json_literal      # bracket-aware locator
from build_v22 import (
    _parse_new_analysis_sheets,
    _normalize_concept_name,
    _ANA_SHEETS_NEW,
    _LEGAL_SHEET_TO_CID,
)


# --------------------------------------------------------------------------- #
# v24 continuation-aware parser.                                              #
#                                                                             #
# v22's parser ignores rows where col-A (dim label) is empty, but the new     #
# xlsx puts continuation text on those rows (e.g. EU's transparency line 2,   #
# EU/TX penalty rows 16–18). v24 folds continuation rows into the prior      #
# dim per-jid, joining with newline, so resync coverage actually closes.      #
# --------------------------------------------------------------------------- #

def _parse_with_continuations(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    JID_DETECT = [
        ("eu", re.compile(r"\beu\s*\(aia\)|\beu\s+ai\s+act|\baia\b", re.I)),
        ("ca", re.compile(r"\bcalifornia\b", re.I)),
        ("co", re.compile(r"\bcolorado\b", re.I)),
        ("ny", re.compile(r"\bnew\s*york\b", re.I)),
        ("tx", re.compile(r"\btexas\b", re.I)),
        ("ut", re.compile(r"\butah\b", re.I)),
    ]
    _BILL_CAPTURE = re.compile(r"\(([A-Z]{1,3}\s?\d+[\-\d]*)\)", re.I)

    def _read_jid_cols(row):
        out = {}
        for c in range(1, len(row)):
            v = row[c]
            if not v:
                continue
            s = str(v).strip()
            if len(s) > 60:
                continue
            mjid = None
            for jid, pat in JID_DETECT:
                if pat.search(s):
                    mjid = jid
                    break
            if not mjid:
                continue
            bill = ""
            bm = _BILL_CAPTURE.search(s)
            if bm:
                bill = bm.group(1).strip().replace(" ", "")
            out[c] = (mjid, bill)
        return out

    out = {}
    for legal_sn, ana_sn in _ANA_SHEETS_NEW.items():
        if ana_sn not in wb.sheetnames:
            continue
        ws = wb[ana_sn]
        tab_id = _LEGAL_SHEET_TO_CID[legal_sn]
        rows = list(ws.iter_rows(values_only=True))
        sections = []

        def _consume(start_hdr_row, title):
            jid_info = _read_jid_cols(rows[start_hdr_row])
            section = {
                "title":    title,
                "title_lc": title.lower(),
                "dims":     {},
                "dim_order": [],
                "notes":    "",
            }
            j = start_hdr_row + 1
            cur_dim_lc = None
            while j < len(rows):
                r = rows[j] or ()
                a = (r[0] if len(r) > 0 else None) or ""
                a = str(a).strip()
                if j > start_hdr_row and len(_read_jid_cols(r)) >= 2:
                    sections.append(section)
                    return j
                nxt2 = rows[j + 1] if j + 1 < len(rows) else None
                if a and nxt2 and len(_read_jid_cols(nxt2)) >= 2:
                    sections.append(section)
                    return j
                if a:
                    cur_dim_lc = a.lower().strip()
                    if cur_dim_lc not in section["dims"]:
                        section["dim_order"].append(cur_dim_lc)
                # Harvest text into current dim (whether col A is empty or not).
                if cur_dim_lc:
                    slot = section["dims"].setdefault(cur_dim_lc, {})
                    for c, (jid, bill) in jid_info.items():
                        v = r[c] if c < len(r) else None
                        txt = "" if v is None else str(v).strip()
                        if not txt or txt in ("-", "–", "—"):
                            continue
                        key = f"{jid}-{bill.lower()}" if bill else jid
                        for k in (key, jid):
                            if k in slot:
                                # Append continuation if not already there.
                                if txt not in slot[k]:
                                    slot[k] = (slot[k] + "\n" + txt).strip()
                            else:
                                slot[k] = txt
                j += 1
            sections.append(section)
            return j

        i = 0
        while i < len(rows):
            row = rows[i] or ()
            col_a = (row[0] if len(row) > 0 else None) or ""
            col_a = str(col_a).strip()
            nxt = rows[i + 1] if i + 1 < len(rows) else None
            if col_a and nxt and len(_read_jid_cols(nxt)) >= 2:
                i = _consume(i + 1, col_a)
                continue
            if len(_read_jid_cols(row)) >= 2:
                title = legal_sn.strip().replace("_", " ")
                i = _consume(i, title)
                continue
            i += 1

        if sections:
            out[tab_id] = sections
    return out


# --------------------------------------------------------------------------- #
# Limited-risk slice — parse rows 1-20 of Provider_Developer_Analysis.        #
# --------------------------------------------------------------------------- #

# The new xlsx's Provider_Developer_Analysis section "Provider of limited-risk
# AI systems / Developer" carries EU / Colorado / Texas columns at fixed
# offsets. Hard-coded to keep parsing robust against the parser's section-
# header regex (which is case-sensitive on "Developer" suffix).
_LIMRISK_JIDS = [
    (1, "eu", "EU AI Act",  "Art. 3, Art. 4, Art. 50, Art. 99",
        "Provider of limited-risk AI systems"),
    (2, "co", "Colorado",   "SB24-205", "Developer"),
    (3, "tx", "Texas",      "HB149",    "Developer"),
]

_LIMRISK_DIM_ROWS = [
    # (xlsx_row_idx, dim_id_base, dim_label)
    (3,  "term",                          "Term"),
    (4,  "scope",                         "Scope"),
    (5,  "regulatory-trigger",            "Regulatory trigger"),
    (6,  "provider-developer-information","Provider / developer information"),
    (7,  "transparency",                  "Transparency"),
    (9,  "general-information-disclosure","General information disclosure"),
    (10, "risk-management",               "Risk management"),
    (11, "incident-risk-reporting",       "Incident / risk reporting"),
    (12, "ai-literacy",                   "AI literacy"),
    (13, "rebuttal",                      "Rebuttal"),
    (15, "penalties",                     "Penalties"),
]
# Rows where text continues from the prior dim row (per jid):
# (continuation_row, source_row, jid). All continuation rows we want to fold
# back into the dim above, joining with newline.
_LIMRISK_CONTINUATIONS = [
    (8,  7,  "eu"),   # Transparency EU continuation
    (14, 13, "tx"),   # Rebuttal TX continuation
    (16, 15, "eu"),   # Penalties EU continuation
    (16, 15, "tx"),   # Penalties TX continuation (hyphen line)
    (17, 15, "tx"),
    (18, 15, "tx"),
]


def _build_limited_risk_subconcept(new_xlsx: Path, old_xlsx: Path) -> dict:
    """Build the `provider` (Provider of limited-risk AI systems / Developer)
    sub-concept fresh from the new xlsx, with verbatim/reference filled
    where the OLD xlsx Provider_Developer sheet rows 1-8 carry legal text."""
    wb_new = openpyxl.load_workbook(str(new_xlsx), data_only=True)
    if "Provider_Developer_Analysis" not in wb_new.sheetnames:
        raise SystemExit("Provider_Developer_Analysis missing from new xlsx")
    ws_new = wb_new["Provider_Developer_Analysis"]
    rows_new = [list(r) for r in ws_new.iter_rows(values_only=True)]

    def _cell(ri_1based: int, ci_1based: int) -> str:
        if ri_1based - 1 >= len(rows_new):
            return ""
        row = rows_new[ri_1based - 1]
        if ci_1based - 1 >= len(row):
            return ""
        v = row[ci_1based - 1]
        if v is None:
            return ""
        return str(v).strip().replace("\xa0", " ")

    # Load OLD xlsx Provider_Developer rows 2-8 to harvest verbatim/reference.
    # Layout (1-based):
    #   col 1  EU header / dim label
    #   col 2  EU dim sub-label
    #   col 3  EU verbatim
    #   col 4  EU reference
    #   col 6  CO dim label
    #   col 7  CO dim sub-label
    #   col 8  CO verbatim
    #   col 9  CO reference
    #   col 11 TX dim label
    #   col 12 TX dim sub-label
    #   col 13 TX verbatim
    #   col 14 TX reference
    verbatim_idx: dict[tuple[str, str], dict] = {}
    if old_xlsx.exists():
        wb_old = openpyxl.load_workbook(str(old_xlsx), data_only=True)
        if "Provider_Developer" in wb_old.sheetnames:
            ws_old = wb_old["Provider_Developer"]
            old_rows = [list(r) for r in ws_old.iter_rows(values_only=True)]
            # The first block (rows 1-8) is the Provider / Developer slice.
            JID_COLS = [
                ("eu", 1, 2, 3, 4),     # (jid, dim_col, sublabel_col, verb_col, ref_col)
                ("co", 6, 7, 8, 9),
                ("tx", 11, 12, 13, 14),
            ]
            for ri in range(2, 9):  # rows 3..9 are the limited-risk slice
                if ri >= len(old_rows):
                    continue
                row = old_rows[ri]
                for jid, dcol, scol, vcol, rcol in JID_COLS:
                    def _g(c):
                        if c >= len(row) or row[c] is None:
                            return ""
                        return str(row[c]).strip()
                    dim_label = _g(dcol - 1) or _g(scol - 1)
                    verb      = _g(vcol - 1)
                    ref       = _g(rcol - 1)
                    if not (dim_label and (verb or ref)):
                        continue
                    key = (jid, _normalize_concept_name(dim_label))
                    if key in verbatim_idx:
                        # Append (multiple rows feed the same dim).
                        prev = verbatim_idx[key]
                        if verb and verb not in (prev["verbatim"] or ""):
                            prev["verbatim"] = (
                                (prev["verbatim"] + "\n\n" + verb).strip()
                                if prev["verbatim"] else verb
                            )
                        if ref and ref not in (prev["reference"] or ""):
                            prev["reference"] = (
                                (prev["reference"] + "; " + ref).strip("; ")
                                if prev["reference"] else ref
                            )
                    else:
                        verbatim_idx[key] = {"verbatim": verb, "reference": ref}

    def _verbatim_for(jid: str, dim_label: str) -> tuple[str, str]:
        key = (jid, _normalize_concept_name(dim_label))
        if key in verbatim_idx:
            ent = verbatim_idx[key]
            return ent.get("verbatim", "") or "", ent.get("reference", "") or ""
        # Try a few synonyms — the OLD xlsx uses 'Obligations' as the dim_label
        # for many dims with the actual sub-label in col 2. Most are already
        # captured by the loop above (which checks both dcol and scol), but
        # 'Term' / 'Scope' may be the section title rather than a row label.
        return "", ""

    # Build dimensions list.
    dims: list[dict] = []
    for d_idx, (xrow, dim_id_base, dim_label) in enumerate(_LIMRISK_DIM_ROWS):
        cells = {}
        for jid_col, jid, _law, _bills, _term_lbl in _LIMRISK_JIDS:
            txt = _cell(xrow, jid_col + 1)
            # Apply continuations.
            for cont_row, src_row, cont_jid in _LIMRISK_CONTINUATIONS:
                if src_row != xrow or cont_jid != jid:
                    continue
                extra = _cell(cont_row, jid_col + 1)
                if extra and extra not in (" ", "-", "–", "—"):
                    txt = (txt + "\n" + extra).strip() if txt else extra
            if txt in ("", "-", "–", "—"):
                analysis = "-"
            else:
                analysis = txt
            verb, ref = _verbatim_for(jid, dim_label)
            cells[jid] = {
                "analysis":  analysis,
                "verbatim":  verb,
                "reference": ref,
            }
        dims.append({
            "id":    f"{dim_id_base}-{d_idx}-0",
            "label": dim_label,
            "cells": cells,
        })

    # Term cell: the analysis is the term name itself (per existing v22 shape).
    # Override with the canonical labels from _LIMRISK_JIDS.
    if dims and dims[0]["label"] == "Term":
        for jid_col, jid, _law, _bills, term_lbl in _LIMRISK_JIDS:
            if jid in dims[0]["cells"]:
                dims[0]["cells"][jid]["analysis"] = term_lbl

    juris = {}
    for jid_col, jid, law, bills, term_lbl in _LIMRISK_JIDS:
        juris[jid] = {"term": term_lbl, "law": law, "bills": bills}

    return {
        "id":            "provider",
        "title":         "Provider of limited-risk AI systems",
        "jurisdictions": juris,
        "dimensions":    dims,
    }


# --------------------------------------------------------------------------- #
# v24 multi-strategy text re-sync.                                            #
# --------------------------------------------------------------------------- #

def _norm_for_match(s: str) -> str:
    s = (s or "").lower()
    s = s.replace("\xa0", " ")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _multi_strategy_resync(html: str, xlsx_sections: dict) -> tuple[str, int]:
    """xlsx_sections is the {tab_id: [section_dict, ...]} from
    _parse_new_analysis_sheets. Walk every cell whose current analysis text
    isn't a 60-char-prefix substring of any xlsx cell text in the matching
    tab/section, then try in order:
        a) exact (jid, dim_lc) within matching section title
        b) (jid, dim_lc) across all sections in same tab
        c) best fuzzy: normalized-overlap section + exact dim_lc
    Returns (html, n_updated). Only updates `analysis`."""
    span = _find_json_literal(html, "CONCEPTS")
    if not span:
        return html, 0
    a, b = span
    try:
        concepts = json.loads(html[a:b])
    except Exception:
        return html, 0

    # Build a global text fingerprint set (60-char prefix of normalized text)
    # so we can short-circuit "already matches".
    fp_all: set[str] = set()
    for sections in xlsx_sections.values():
        for sec in sections:
            for dim_lc, jmap in (sec.get("dims") or {}).items():
                for jkey, txt in jmap.items():
                    fp = _norm_for_match(txt)[:60]
                    if fp:
                        fp_all.add(fp)

    updated = 0
    for c in concepts:
        cid = c.get("id")
        sections = xlsx_sections.get(cid) or []
        if not sections:
            continue
        for sc in (c.get("sub_concepts") or []):
            sub_title = sc.get("title") or ""
            sub_norm  = _normalize_concept_name(sub_title)
            # Strategy A — exact section match.
            best_section = None
            for s in sections:
                if _normalize_concept_name(s.get("title", "")) == sub_norm:
                    best_section = s
                    break
            # Strategy C — fuzzy overlap fallback.
            if best_section is None and sub_norm:
                best, score = None, 0
                for s in sections:
                    snorm = _normalize_concept_name(s.get("title", ""))
                    if not snorm:
                        continue
                    if sub_norm in snorm:
                        sc_score = len(sub_norm)
                    elif snorm in sub_norm:
                        sc_score = len(snorm)
                    else:
                        continue
                    if sc_score > score:
                        best, score = s, sc_score
                best_section = best

            for d in (sc.get("dimensions") or []):
                dim_label_lc = (d.get("label") or "").lower()
                dim_norm = _norm_for_match(d.get("label") or "")
                if not dim_norm:
                    continue
                cells = d.get("cells") or {}
                for jid_full, cell in cells.items():
                    if not isinstance(cell, dict):
                        continue
                    cur = cell.get("analysis") or ""
                    if not cur or cur in ("-", "–", "—"):
                        continue

                    jid_root = jid_full.split("-", 1)[0]

                    def _lookup(section, jkey_full, jkey_root):
                        dims = (section or {}).get("dims") or {}
                        # Try exact dim_lc first, then partial.
                        for dlc, jmap in dims.items():
                            if dlc != dim_label_lc and dlc != dim_norm:
                                continue
                            if jkey_full in jmap:
                                return jmap[jkey_full]
                            if jkey_root in jmap:
                                return jmap[jkey_root]
                        # Fuzzy dim match.
                        for dlc, jmap in dims.items():
                            if dim_norm in dlc or dlc in dim_norm:
                                if jkey_full in jmap:
                                    return jmap[jkey_full]
                                if jkey_root in jmap:
                                    return jmap[jkey_root]
                        return None

                    # A: best matching section
                    new_txt = _lookup(best_section, jid_full, jid_root)
                    # B: any section
                    if not new_txt:
                        for s in sections:
                            new_txt = _lookup(s, jid_full, jid_root)
                            if new_txt:
                                break
                    if not new_txt:
                        continue
                    if new_txt == cur:
                        continue
                    # Only swap when the xlsx text is a richer/longer
                    # version of what's in the cell, OR when the existing
                    # text doesn't appear in the xlsx anywhere (drift).
                    cur_fp = _norm_for_match(cur)[:60]
                    new_norm = _norm_for_match(new_txt)
                    cur_norm = _norm_for_match(cur)
                    if cur_norm in new_norm and len(new_norm) > len(cur_norm):
                        cell["analysis"] = new_txt
                        updated += 1
                    elif cur_fp not in fp_all:
                        cell["analysis"] = new_txt
                        updated += 1

    new_payload = json.dumps(concepts, ensure_ascii=False, separators=(",", ":"))
    return html[:a] + new_payload + html[b:], updated


# --------------------------------------------------------------------------- #
# Limited-risk insertion / replacement.                                       #
# --------------------------------------------------------------------------- #

def _install_limited_risk(html: str, sub: dict) -> tuple[str, bool]:
    """Replace (or insert before `provider-of-high-risk-ai-systems`) the
    `provider` sub-concept on the `provider-developer` concept.
    Returns (html, was_changed)."""
    span = _find_json_literal(html, "CONCEPTS")
    if not span:
        return html, False
    a, b = span
    try:
        concepts = json.loads(html[a:b])
    except Exception:
        return html, False
    changed = False
    for c in concepts:
        if c.get("id") != "provider-developer":
            continue
        subs = c.get("sub_concepts") or []
        # Find existing `provider` slot.
        idx_existing = -1
        idx_highrisk = -1
        for i, s in enumerate(subs):
            if s.get("id") == "provider":
                idx_existing = i
            if s.get("id") == "provider-of-high-risk-ai-systems":
                idx_highrisk = i
        if idx_existing >= 0:
            subs[idx_existing] = sub
            changed = True
        elif idx_highrisk >= 0:
            subs.insert(idx_highrisk, sub)
            changed = True
        else:
            subs.insert(0, sub)
            changed = True
        c["sub_concepts"] = subs
        break
    if changed:
        new_payload = json.dumps(concepts, ensure_ascii=False, separators=(",", ":"))
        return html[:a] + new_payload + html[b:], True
    return html, False


# --------------------------------------------------------------------------- #
# Coverage probe — reports the xlsx→HTML rate the v24 tests look at.          #
# --------------------------------------------------------------------------- #

def _coverage_probe(html: str, xlsx_path: Path) -> tuple[int, int]:
    span = _find_json_literal(html, "CONCEPTS")
    if not span:
        return 0, 0
    a, b = span
    concepts = json.loads(html[a:b])

    def _norm(s):
        return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip().lower()

    def _walk_strings(o, out):
        if isinstance(o, str):
            out.append(o)
        elif isinstance(o, dict):
            for v in o.values(): _walk_strings(v, out)
        elif isinstance(o, list):
            for v in o: _walk_strings(v, out)

    corpus = []
    _walk_strings(concepts, corpus)
    big = _norm(" || ".join(corpus))

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    misses = 0
    checked = 0
    for sn in wb.sheetnames:
        if not (sn.endswith("_ANALYSIS") or sn.endswith("_ANALY") or
                sn.endswith("_Analys") or sn.endswith("_Analysis")):
            continue
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not v:
                    continue
                if cell.column_letter in ("A", "E", "F", "G", "H"):
                    continue
                sv = str(v).strip()
                if len(sv) < 30 or len(sv) > 600:
                    continue
                if sv.lower().startswith(("interpretative notes", "this section")):
                    continue
                probe = _norm(sv)[:60]
                if not probe:
                    continue
                checked += 1
                if probe not in big:
                    misses += 1
    return checked - misses, checked


# --------------------------------------------------------------------------- #
# Main.                                                                       #
# --------------------------------------------------------------------------- #

def main() -> None:
    print("== v24 build ==")
    if not HTML_V23.exists():
        print("  digital_lexicon_v23.html missing — running build_v23 …")
        import build_v23 as _v23
        _v23.main()
    html = HTML_V23.read_text(encoding="utf-8")
    print(f"  read v23:            {len(html):,} bytes")

    # 1. Rebuild limited-risk Provider sub-concept.
    sub_lim = _build_limited_risk_subconcept(NEW_XLSX, OLD_XLSX)
    print(f"  limited-risk slice:  built ({len(sub_lim['dimensions'])} dims, "
          f"{len(sub_lim['jurisdictions'])} jids)")
    html, did_install = _install_limited_risk(html, sub_lim)
    print(f"  limited-risk install: {'replaced/inserted' if did_install else 'unchanged'}")

    # 2. Multi-strategy text resync — uses the v24 continuation-aware parser
    #    so multi-line dim text (Penalties, Transparency, Scope) actually
    #    matches what's already in v23/v22 cells.
    sections = _parse_with_continuations(NEW_XLSX)
    html, n_resync = _multi_strategy_resync(html, sections)
    print(f"  v24 resync (beyond v23):     +{n_resync} cells")

    # 3. Coverage probe.
    matched, total = _coverage_probe(html, NEW_XLSX)
    rate = matched / total if total else 0.0
    print(f"  xlsx→HTML coverage:  {matched}/{total} = {rate:.1%}")

    # 4. Write outputs.
    HTML_V24.write_text(html, encoding="utf-8")
    shutil.copy2(HTML_V24, FINAL_TOOL)
    shutil.copy2(HTML_V24, FINAL_LEXICON)
    print(f"\n  wrote {HTML_V24.name}        ({len(html):,} bytes)")
    print(f"  wrote {FINAL_TOOL.name}            (mirror)")
    print(f"  wrote {FINAL_LEXICON.name}    (mirror)")


if __name__ == "__main__":
    main()
