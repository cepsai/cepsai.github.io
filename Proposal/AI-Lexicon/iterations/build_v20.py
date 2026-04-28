"""build_v20.py — Digital AI Lexicon v20

Reshapes v19's Verbatim feature from a top-level nav tab into an in-page
mode-tab pair on each concept page (Analysis / Verbatim). Rationale: a
separate top-level Verbatim nav entry confused users because its landing
page looked identical to Analysis. The mode is now a per-concept toggle,
same as a sub-concept tab — visually colocated, obvious choice.

Build chain: v13 -> v15 -> v16 -> v17 -> v18 -> v20. (v19 is skipped —
v20 reads v18's output directly, reuses v19's xlsx dim-parent parse but
not its HTML mutations.)

v20 changes:
    * Top nav: back to 'Concepts' (no Verbatim button, no Analysis
      rename).
    * Home cards: back to v17's three cards (no Browse verbatim card).
    * Inside each concept detail page:
        - a new row of mode tabs above the sub-concept tabs:
          [Analysis] [Verbatim]
        - Analysis tab keeps all v18 behavior (CEPS notes, inline
          verbatim drawer on cell click, v18's Incident two-col).
        - Verbatim tab renders the same 2-col Dimension / Sub-dimension
          layout as v19, with verbatim text in cells. Cell click opens
          the inline drawer then auto-expands "Explore in full law".
    * URL: mode is encoded as ?view=verbatim on #/concept/<id>. Default
      (absent param) is Analysis.

Predecessors (v13..v18) untouched.

Note: downstream HTML files cache aggressively. If xlsx or law JSON
changed, wipe before rebuild:
    rm -f digital_lexicon_v1{6,7,8}.html digital_lexicon_v20.html
    python3 build_v20.py
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE     = Path(__file__).parent
HTML_V18 = HERE / "digital_lexicon_v18.html"
HTML_V20 = HERE / "digital_lexicon_v20.html"
XLSX     = HERE / "AI terminology and taxonomy-final.xlsx"


# --------------------------------------------------------------------------- #
# xlsx re-parse — same as v19. Copied in full here rather than imported       #
# from build_v19 so v20 remains independent.                                  #
# --------------------------------------------------------------------------- #

_SHEET_TO_CID: dict[str, str] = {
    "Provider_Developer":              "provider-developer",
    "Deployer_Supplier":               "deployer-supplier",
    " High-risk AI system":            "model-system",
    "GPAI_Frontier_Foundation model":  "model-system",
    "GPAI system_Generative AI":       "model-system",
    "Risk":                            "risk",
    "Substantial modification":        "modification",
    "Incident":                        "incident",
}

_HARDCODED_PARENTS: dict[str, str] = {
    # Scope family.
    "regulatory trigger":              "Scope",
    "temporal trigger":                "Scope",
    "compute threshold":               "Scope",
    "harm thresholds":                 "Scope",
    "continuous learning":             "Scope",
    "provider / developer information":"Scope",
    # Definition family.
    "definition approach":             "Definition",
    "approach":                        "Definition",
    # Obligations family.
    "general information disclosure":  "Obligations",
    "specific information disclosure": "Obligations",
    "risk management - review":        "Obligations",
    "risk management - reporting":     "Obligations",
    "impact assessment - review":      "Obligations",
    "cooperation with authorities":    "Obligations",
    "communication to deployers":      "Obligations",
    "incident / risk reporting":       "Obligations",
    "documentation keeping":           "Obligations",
    "compliance check":                "Obligations",
    "whistleblower protections":       "Obligations",
    "obligations triggered":           "Obligations",
    # Reporting (Incident).
    "reporting timeline":              "Reporting",
    "reporting timelines":             "Reporting",
    "reporting mechanism":             "Reporting",
    "reporting obligations":           "Reporting",
    # Exemption singular/plural normalization.
    "exemption":                       "Exemptions",
}


def _norm(s) -> str:
    if not s:
        return ""
    return str(s).replace("\xa0", " ").strip()


def _is_verbatim_like(s: str) -> bool:
    if not s:
        return False
    if "\n" in s:
        return True
    if len(s) > 100:
        return True
    if s.count(".") >= 2 or s.count(";") >= 1:
        return True
    return False


def _build_parent_lookup() -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    if not XLSX.exists():
        print(f"  [v20] xlsx missing at {XLSX}; parent lookup will be empty")
        return {}, {}
    wb = load_workbook(XLSX, data_only=True)
    result: dict[str, dict[str, str]] = {}
    global_counts: dict[str, dict[str, int]] = {}
    for sheet_name, cid in _SHEET_TO_CID.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_a_has_text = any(
            _norm(ws.cell(r, 1).value)
            for r in range(1, min(ws.max_row, 20) + 1)
        )
        if not col_a_has_text:
            continue
        pmap = result.setdefault(cid, {})
        for r in range(1, ws.max_row + 1):
            a = _norm(ws.cell(r, 1).value)
            b = _norm(ws.cell(r, 2).value)
            if not a or not b:
                continue
            if _is_verbatim_like(b):
                continue
            key = b.lower()
            pmap.setdefault(key, a)
            bucket = global_counts.setdefault(key, {})
            bucket[a] = bucket.get(a, 0) + 1
    global_fallback: dict[str, str] = {}
    for key, votes in global_counts.items():
        best = sorted(votes.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
        global_fallback[key] = best
    for k, v in _HARDCODED_PARENTS.items():
        global_fallback[k] = v
    return result, global_fallback


def _inject_dim_parent_blob(html: str, parent_lookup: dict, global_fallback: dict) -> str:
    per_concept = json.dumps(parent_lookup, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    glob        = json.dumps(global_fallback, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    tag = (
        '<script type="application/json" id="__v20_dim_parents__">' + per_concept + "</script>\n"
        '<script type="application/json" id="__v20_dim_parents_global__">' + glob + "</script>\n"
        "<script>(function(){"
        "try{window.V20_DIM_PARENTS=JSON.parse(document.getElementById('__v20_dim_parents__').textContent);}"
        "catch(e){window.V20_DIM_PARENTS={};}"
        "try{window.V20_DIM_PARENTS_GLOBAL=JSON.parse(document.getElementById('__v20_dim_parents_global__').textContent);}"
        "catch(e){window.V20_DIM_PARENTS_GLOBAL={};}"
        "})();</script>"
    )
    return html.replace("</body>", tag + "\n</body>", 1)


# --------------------------------------------------------------------------- #
# v20 overrides — in-page mode tabs + verbatim renderer.                      #
# --------------------------------------------------------------------------- #

def _v20_overrides_css_js() -> str:
    return r"""
<style>
/* ── v20 mode tabs (Analysis | Legal text) inside each concept page ────
   Styled as underlined tabs (same grammar as sub-tabs, slightly smaller)
   placed below the sub-concept sub-tabs. */
.v20-mode-bar{
  display:flex;
  gap:0;
  border-bottom:1px solid var(--bd-s);
  margin:18px 0 22px 0;
  padding:0;
  align-items:flex-end;
}
.v20-mode-tab{
  font-family:var(--sans);
  font-size:13px;
  font-weight:500;
  color:var(--ink-s);
  background:none;
  border:none;
  border-top:none;
  border-left:none;
  border-right:none;
  border-bottom:2px solid transparent;
  padding:8px 14px;
  cursor:pointer;
  white-space:nowrap;
  transition:color .15s, border-color .15s;
  line-height:1.4;
  margin-bottom:-1px; /* overlap the parent border-bottom */
}
.v20-mode-tab:hover{ color:var(--ink); }
.v20-mode-tab.active{
  color:var(--accent);
  font-weight:600;
  border-bottom-color:var(--accent);
}

/* Verbatim cell — line-clamped preview, full text reads in the drawer */
.v20-verbatim-cell{
  display:-webkit-box;
  -webkit-line-clamp:6;
  -webkit-box-orient:vertical;
  overflow:hidden;
  text-overflow:ellipsis;
  cursor:pointer;
  font-family:var(--serif);
  font-size:13px;
  line-height:1.55;
  color:var(--ink);
  padding:6px 8px;
  border-radius:var(--r-sm,4px);
  transition:background-color .12s;
}
.v20-verbatim-cell:hover,
.v20-verbatim-cell:focus{ background:var(--accent-l); outline:none; }
.v20-analysis-only{ color:var(--ink-s); font-style:italic; }

/* Dim / Sub-dim hierarchy (Verbatim view only, always on) */
.analysis-table th.analysis-dim-cell.v20-dim-col{ min-width:140px }
.analysis-table td.v20-dim-cell{
  vertical-align:top;
  font-family:var(--mono);
  font-size:11px;
  text-transform:uppercase;
  letter-spacing:.05em;
  background:var(--bg2);
  color:var(--ink);
  font-weight:600;
  padding-top:12px;
}
.analysis-table th.v20-subdim-cell{
  min-width:140px;
  background:var(--card);
  border-left:1px dashed var(--bd-s);
}
.analysis-table td.v20-subdim-cell{
  font-family:var(--serif);
  font-weight:500;
  font-size:12px;
  color:var(--ink-s);
  font-style:italic;
  border-left:1px dashed var(--bd-s);
  background:var(--card);
  vertical-align:top;
}

/* The Analysis view's "open verbatim" icon on each cell (per user spec:
   an icon that opens the inline drawer). The shell already makes the
   whole analysis cell clickable; the icon just telegraphs the action. */
.analysis-cell .v20-open-ic{
  display:inline-block;
  width:12px; height:12px;
  margin-left:6px;
  vertical-align:middle;
  opacity:.5;
  color:currentColor;
  transition:opacity .12s;
}
.analysis-cell:hover .v20-open-ic{ opacity:1; }
</style>
<script>
(function(){
  if (typeof state === 'undefined') return;
  // state.mode in {'analysis', 'verbatim'}. Default: analysis.
  state.mode = (state.mode === 'verbatim') ? 'verbatim' : 'analysis';

  // ── Helpers ──────────────────────────────────────────────────────────
  function _escH(s){
    return String(s==null?'':s).replace(/[&<>"']/g, function(c){
      return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];
    });
  }

  function _getViewFromHash(){
    var h = location.hash || '';
    var q = h.indexOf('?');
    if (q < 0) return 'analysis';
    var params = new URLSearchParams(h.slice(q + 1));
    var v = params.get('view');
    return (v === 'verbatim') ? 'verbatim' : 'analysis';
  }

  function _writeViewToHash(mode){
    var h = location.hash || '';
    var q = h.indexOf('?');
    var base = q >= 0 ? h.slice(0, q) : h;
    var query = q >= 0 ? h.slice(q + 1) : '';
    var params = new URLSearchParams(query);
    if (mode === 'verbatim') params.set('view', 'verbatim');
    else params.delete('view');
    var qs = params.toString();
    var newH = base + (qs ? ('?' + qs) : '');
    if (newH !== h) history.replaceState(null, '', newH);
  }

  // ── Mode tab strip installation ──────────────────────────────────────
  function _ensureModeBar(){
    var pc = document.getElementById('p-concept');
    if (!pc) return null;
    var bar = document.getElementById('v20-mode-bar');
    if (bar) return bar;
    bar = document.createElement('div');
    bar.id = 'v20-mode-bar';
    bar.className = 'v20-mode-bar';
    bar.setAttribute('role', 'tablist');
    bar.setAttribute('aria-label', 'Content view mode');
    bar.innerHTML =
      '<button type="button" class="v20-mode-tab" data-mode="analysis" role="tab" aria-selected="false" title="CEPS comparative analysis summary. Click a cell to open the verbatim quote in the side drawer.">' +
      'Analysis' +
      '</button>' +
      '<button type="button" class="v20-mode-tab" data-mode="verbatim" role="tab" aria-selected="false" title="Full legal text per jurisdiction. Click a cell to open the article in the full-law view.">' +
      'Legal text' +
      '</button>';
    bar.addEventListener('click', function(e){
      var btn = e.target && e.target.closest && e.target.closest('.v20-mode-tab');
      if (!btn) return;
      var mode = btn.getAttribute('data-mode');
      if (!mode || mode === state.mode) return;
      _switchMode(mode);
    });

    // Insert the bar AFTER the sub-concept sub-tabs so the visual order
    // reads: concept title → pick sub-concept → pick view mode → table.
    var subTabs = document.getElementById('sub-tabs');
    if (subTabs && subTabs.parentNode) {
      if (subTabs.nextSibling) {
        subTabs.parentNode.insertBefore(bar, subTabs.nextSibling);
      } else {
        subTabs.parentNode.appendChild(bar);
      }
    } else {
      pc.appendChild(bar);
    }
    return bar;
  }

  function _updateModeBar(){
    var bar = document.getElementById('v20-mode-bar');
    if (!bar) return;
    bar.querySelectorAll('.v20-mode-tab').forEach(function(b){
      var on = (b.getAttribute('data-mode') === state.mode);
      b.classList.toggle('active', on);
      b.setAttribute('aria-selected', on ? 'true' : 'false');
    });
  }

  function _switchMode(newMode){
    state.mode = newMode;
    try { if (typeof saveState === 'function') saveState(); } catch(e) {}
    _writeViewToHash(newMode);
    _updateModeBar();
    if (typeof renderAnalysisTable === 'function') renderAnalysisTable();
  }

  // ── Verbatim renderer (same shape as v19, renamed v20_*) ─────────────
  function _dimParent(c, dim){
    if (dim && dim.parent_label) return dim.parent_label;
    var key = (dim && (dim.label || '')).toLowerCase().trim();
    var m = window.V20_DIM_PARENTS || {};
    var cm = m[c && c.id] || {};
    if (cm[key]) return cm[key];
    var g = window.V20_DIM_PARENTS_GLOBAL || {};
    if (g[key]) return g[key];
    return (dim && (dim.label || '')) || '';
  }
  function _dimSub(dim){ return (dim && (dim.sub_label || dim.label || '')) || ''; }

  function _renderVerbatimTable(){
    var c  = getConcept(state.conceptId);
    var sc = c && c.sub_concepts && (c.sub_concepts[state.subConceptIdx] || c.sub_concepts[0]);
    if (!sc) return;
    var thead = document.getElementById('analysis-thead');
    var tbody = document.getElementById('analysis-tbody');
    if (!thead || !tbody) return;
    var juris = Object.keys(sc.jurisdictions);
    var JL = (typeof JURIS_LABELS !== 'undefined') ? JURIS_LABELS : {};

    var head = '<tr>';
    head += '<th scope="col" class="analysis-dim-cell v20-dim-col">Dimension</th>';
    head += '<th scope="col" class="v20-subdim-cell">Sub-dimension</th>';
    juris.forEach(function(j){
      var jid = String(j).split('-')[0];
      var jd = sc.jurisdictions[j];
      head += '<th scope="col" class="th-' + jid + '">' + _escH(JL[j] || JL[jid] || j) +
              '<span class="j-law">' + _escH((jd && jd.law) || '') + '</span></th>';
    });
    head += '</tr>';
    thead.innerHTML = head;

    var rowsPlan = (sc.dimensions || []).map(function(dim){
      return {dim: dim, parent: _dimParent(c, dim), sub: _dimSub(dim)};
    });
    var groups = [];
    rowsPlan.forEach(function(r, i){
      var last = groups[groups.length - 1];
      if (last && last.parent === r.parent) { last.end = i; }
      else { groups.push({parent: r.parent, start: i, end: i}); }
    });
    var rowspanMap = {};
    groups.forEach(function(g){ rowspanMap[g.start] = g.end - g.start + 1; });

    var bodyParts = [];
    rowsPlan.forEach(function(r, i){
      var tr = '<tr>';
      if (i in rowspanMap) {
        tr += '<td class="v20-dim-cell" rowspan="' + rowspanMap[i] + '">' + _escH(r.parent) + '</td>';
      }
      tr += '<td class="v20-subdim-cell">' + _escH(r.sub) + '</td>';
      juris.forEach(function(j){
        var cell = r.dim.cells[j];
        if (cell && cell.verbatim) {
          tr += '<td><span class="v20-verbatim-cell addressed" tabindex="0" role="button" ' +
                'data-dimid="' + _escH(r.dim.id) + '" data-juris="' + _escH(j) + '" ' +
                'aria-label="Open full law: ' + _escH(JL[j] || j) + ' on ' + _escH(r.dim.label) + '">' +
                _escH(cell.verbatim) + '</span></td>';
        } else if (cell && cell.analysis) {
          tr += '<td><span class="cell-null v20-analysis-only" ' +
                'title="Analysis only \u2014 no verbatim quote in source law">\u2014</span></td>';
        } else {
          tr += '<td><span class="cell-null" aria-label="Not addressed">\u2014</span></td>';
        }
      });
      tr += '</tr>';
      bodyParts.push(tr);
    });
    tbody.innerHTML = bodyParts.join('');

    tbody.addEventListener('click', _verbatimCellClick);
    tbody.addEventListener('keydown', function(e){
      if (e.key !== 'Enter' && e.key !== ' ') return;
      var t = e.target && e.target.closest && e.target.closest('.v20-verbatim-cell');
      if (!t) return;
      e.preventDefault();
      _verbatimCellClick({target: t});
    });
  }

  function _verbatimCellClick(ev){
    var el = ev && ev.target && (ev.target.closest ? ev.target.closest('.v20-verbatim-cell') : null);
    if (!el) return;
    var dimId = el.getAttribute('data-dimid');
    var j     = el.getAttribute('data-juris');
    var c  = getConcept(state.conceptId);
    var sc = c && c.sub_concepts && (c.sub_concepts[state.subConceptIdx] || c.sub_concepts[0]);
    var dim = sc && (sc.dimensions || []).find(function(d){ return d.id === dimId; });
    if (!dim) return;
    _openFullLawForCell(dim, j);
  }

  function _openFullLawForCell(dim, juris){
    if (typeof openDrawer !== 'function') return;
    openDrawer(dim.id, juris);
    // After the drawer renders, click the v17 Explore-in-full-law button
    // programmatically so .v17-full-article appears inline — same DOM
    // state the user reaches manually from Analysis mode.
    requestAnimationFrame(function(){
      var btn = document.querySelector('.drawer-actions .v17-explore-btn');
      if (btn) btn.click();
    });
  }

  // ── Wrap renderAnalysisTable + renderConceptPage + handleHash ────────
  function _installHooks(){
    if (window.__v20_installed) return;
    window.__v20_installed = true;

    if (typeof window.renderAnalysisTable === 'function') {
      var origRAT = window.renderAnalysisTable;
      window.renderAnalysisTable = function(){
        if (state.mode === 'verbatim') {
          // Still call orig first so CEPS notes + any shell-side setup runs.
          origRAT.apply(this, arguments);
          // Now replace the thead/tbody with the verbatim layout.
          _renderVerbatimTable();
          _updateModeBar();
          return;
        }
        return origRAT.apply(this, arguments);
      };
    }

    if (typeof window.renderConceptPage === 'function') {
      var origRCP = window.renderConceptPage;
      window.renderConceptPage = function(){
        origRCP.apply(this, arguments);
        _ensureModeBar();
        _updateModeBar();
      };
    }

    // Wrap handleHash to grab view=verbatim BEFORE the shell's go()
    // rewrites the hash. The shell's hashMap for the concept page only
    // emits ?sub=N — our ?view=verbatim gets stripped during go(). So we
    // read the view param here, update state.mode, then let the shell
    // process the hash. After the shell's replaceState strips our
    // param, _writeViewToHash puts it back.
    if (typeof window.handleHash === 'function') {
      var origHH = window.handleHash;
      try { window.removeEventListener('hashchange', origHH); } catch(e){}
      window.handleHash = function(){
        var mode = _getViewFromHash();
        if (mode === 'verbatim') state.mode = 'verbatim';
        else state.mode = 'analysis';
        origHH.apply(this, arguments);
        // After shell's go() has written its canonical hash, re-apply
        // ?view=verbatim if needed. Do this on the next frame so the
        // shell's replaceState is already done.
        requestAnimationFrame(function(){
          _writeViewToHash(state.mode);
          _updateModeBar();
        });
      };
      window.addEventListener('hashchange', window.handleHash);
    }
  }

  // Run as early as possible.
  if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', _installHooks);
  } else {
    _installHooks();
  }

  // Also sync mode from URL on first load (no hashchange fires for
  // the initial hash).
  (function(){
    state.mode = _getViewFromHash();
  })();
})();
</script>
"""


# --------------------------------------------------------------------------- #
# Main                                                                        #
# --------------------------------------------------------------------------- #

def main() -> None:
    import build_v18

    print("== v20 build ==")
    if not HTML_V18.exists():
        build_v18.main()
    html = HTML_V18.read_text(encoding="utf-8")

    parent_lookup, global_fallback = _build_parent_lookup()
    total = sum(len(v) for v in parent_lookup.values())
    print(f"  dim parent lookup:   {len(parent_lookup)} concepts, {total} sub->parent (per-concept)")
    print(f"  dim parent fallback: {len(global_fallback)} sub->parent (global + hardcoded)")
    html = _inject_dim_parent_blob(html, parent_lookup, global_fallback)

    html = html.replace("</body>", _v20_overrides_css_js() + "\n</body>", 1)
    print("  overrides:           CSS + JS appended (mode tabs + verbatim renderer)")

    html = html.replace(
        "<!-- DAL v18 (final-edits pass on v17) -->",
        "<!-- DAL v20 (in-page Analysis/Verbatim mode tabs) -->",
        1,
    )
    html = html.replace(
        "<title>Digital AI Lexicon v18 \u2014 CEPS</title>",
        "<title>Digital AI Lexicon v20 \u2014 CEPS</title>",
        1,
    )

    HTML_V20.write_text(html, encoding="utf-8")
    print(f"\nWrote {HTML_V20}  ({len(html):,} bytes)")

    test_path = HERE / "test_lexicon_v20.py"
    if test_path.exists():
        print("\n[v20 smoke test]")
        rc = subprocess.run([sys.executable, str(test_path)], cwd=HERE).returncode
        if rc != 0:
            raise SystemExit(f"v20 smoke test failed (rc={rc})")


if __name__ == "__main__":
    main()
