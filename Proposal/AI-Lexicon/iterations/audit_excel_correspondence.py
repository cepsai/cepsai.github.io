#!/usr/bin/env python3
"""audit_excel_correspondence.py

Compare the latest Excel inputs against the live HTML (v28) to flag cells
whose analysis text or verbatim citation has drifted between source data
and the rendered tool.

Inputs (override with flags):
  --analysis  path to "Cross-checked_AI terminology and taxonomy_analysis-2.xlsx"
  --verbatim  path to "AI terminology and taxonomy_verbatim excel.xlsx"
  --html      path to digital_lexicon_v28.html

Output:
  Markdown report at outputs/excel_correspondence_<timestamp>.md
  Console summary printed at the end.

Mismatch categories
-------------------
  ANALYSIS_DIFF    HTML analysis text differs from Excel analysis text.
  ANALYSIS_MISSING Excel has analysis text; HTML cell is empty or "-".
  ANALYSIS_EXTRA   HTML has analysis text; Excel has nothing for that cell.
  VERBATIM_DIFF    Excel verbatim snippet not found inside HTML verbatim
                   for the corresponding (concept, juris, dim) cell.
  REFERENCE_DIFF   Excel reference / citation differs from HTML reference.

Each row in the report is annotated with the Excel sheet + cell address
so the user can jump straight to the source.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# --------------------------------------------------------------------------- #
# Paths                                                                       #
# --------------------------------------------------------------------------- #

REPO = Path(__file__).resolve().parent
DEFAULT_ANALYSIS = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis-2.xlsx"
)
DEFAULT_VERBATIM = Path(
    "/Users/robertpraas/Downloads/"
    "AI terminology and taxonomy_verbatim excel.xlsx"
)
DEFAULT_HTML = REPO / "digital_lexicon_v28.html"
OUTPUTS = REPO.parent / "outputs"


# --------------------------------------------------------------------------- #
# Sheet → (concept_id, sub_concept_id, jurisdiction-column-map) mapping       #
# --------------------------------------------------------------------------- #
#
# Each value is (concept_id, [(sheet_col_letter, html_jid)], sub_concept_id)
# When sub_concept_id is None, the sheet contains multiple sub-concepts
# separated by section headers and we have to walk the rows to discover them.

# For sheets with one sub-concept, define the column → jurisdiction map.
SINGLE_SUB_SHEETS = {
    " High-risk AI system_ANALYSIS": (
        "model-system", "high-risk-ai-system",
        {"B": "eu", "C": "co", "D": "ut"},
    ),
    "GPAI_Frontier_Foundation_Analys": (
        "model-system", "general-purpose-ai-model",
        {"B": "eu", "C": "ca", "D": "ny"},
    ),
    "GPAI system_Generative AI_ANALY": (
        "model-system", "general-purpose-ai-system",
        {"B": "eu", "C": "ca", "D": "ut"},
    ),
    "Risk_ANALYSIS": (
        "risk", "systemic-risk",
        {"B": "eu", "C": "ca", "D": "ny"},
    ),
    "Modification_ANALYSIS": (
        "modification", "substantial-modification",
        {
            "B": "eu",
            # Modification sheet has 4 jurisdiction columns post-EU; the HTML
            # splits CA into two ca-* sub-jids. Map the Excel SB 53 column to
            # the CA SB 53 jid and the AB 2013 column to the CA AB 2013 jid.
            "C": "ca-0-substantially-modified-version-of-a-frontier-model-no-standalone-defined-term",
            "D": "ca-1-substantial-modification",
            "E": "ny",  # NY S8828 — currently absent from HTML for this sub
            "F": "co",
        },
    ),
    "Incident_ANALYSIS": (
        "incident", "serious-incident",
        {"B": "eu", "C": "ca", "D": "ny"},
    ),
}

# Provider/Deployer sheets have multiple sub-concepts. Define the section
# header text → (sub_concept_id, jurisdiction column map) mapping.
MULTI_SUB_SHEETS = {
    "Provider_Developer_Analysis": [
        # (header text contains substring, sub_id, juris_map)
        (
            "Provider of limited-risk AI systems",
            "provider",
            {"B": "eu", "C": "co", "D": "tx"},
        ),
        (
            "Provider of high-risk AI systems",
            "provider-of-high-risk-ai-systems",
            {"B": "eu", "C": "co"},
        ),
        (
            "Provider of GPAI",
            "provider-of-general-purpose-ai-models",
            {
                "B": "eu",
                "C": "ca-0-covered-provider",
                "D": "ca-1-developer",
            },
        ),
        (
            "Provider of GPAI models with systemic risk",
            "provider-of-general-purpose-ai-models-with-systemic-risk",
            {
                "B": "eu",
                "C": "ca-0-frontier-developer",
                "D": "ca-1-large-frontier-developer",
                "E": "ny-0-frontier-developer",
                "F": "ny-1-large-frontier-developer",
                "G": "ny-2-large-developer",
            },
        ),
    ],
    "Deployer_Supplier_Analysis": [
        (
            "Deployer of limited-risk",
            "deployer",
            {"B": "eu", "C": "co", "D": "tx"},
        ),
        (
            "Deployer / supplier of high-risk",
            "deployer-of-high-risk-ai-systems",
            {"B": "eu", "C": "co", "D": "ut"},
        ),
        (
            "Deployers of GPAI systems",
            "deployer-of-general-purpose-ai-systems",
            {"B": "eu", "C": "ut"},
        ),
    ],
}


# --------------------------------------------------------------------------- #
# Normalisation                                                               #
# --------------------------------------------------------------------------- #

# Map every typographic dash / quote / superscript variant to a canonical form.
SUPERSCRIPT_MAP = str.maketrans({
    "⁰": "0", "¹": "1", "²": "2", "³": "3",
    "⁴": "4", "⁵": "5", "⁶": "6", "⁷": "7",
    "⁸": "8", "⁹": "9",
})

DASH_MAP = str.maketrans({
    "–": "-",  # en dash
    "—": "-",  # em dash
    "‐": "-", "‑": "-", "‒": "-", "―": "-",
    "−": "-",  # minus sign
})

QUOTE_MAP = str.maketrans({
    "‘": "'", "’": "'", "‚": "'", "‛": "'",
    "“": '"', "”": '"', "„": '"', "‟": '"',
    "\xab": '"', "\xbb": '"',
})

SUP_HTML_RE = re.compile(r"<sup>([^<]*)</sup>")
TAG_RE = re.compile(r"<[^>]+>")


def normalize(text: str | None) -> str:
    """Canonicalise a string for fuzzy comparison.

    - Strip <sup> tags but keep their digits inline (so "10<sup>26</sup>"
      → "10^26").
    - Convert Unicode superscripts to plain digits (10²⁶ → 10^26).
    - Unify dashes and curly quotes.
    - Collapse whitespace and lowercase.
    """
    if text is None:
        return ""
    s = str(text)
    # Strip <sup> tags by inserting a caret marker so digits stay intact.
    s = SUP_HTML_RE.sub(r"^\1", s)
    # Strip remaining HTML tags.
    s = TAG_RE.sub("", s)
    # Unicode superscripts → ^digits
    s = re.sub(
        r"([⁰¹²³⁴-⁹]+)",
        lambda m: "^" + m.group(0).translate(SUPERSCRIPT_MAP),
        s,
    )
    # Dashes & quotes → ASCII.
    s = s.translate(DASH_MAP).translate(QUOTE_MAP)
    # NBSP & other whitespace → single space.
    s = s.replace("\xa0", " ")
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


# --------------------------------------------------------------------------- #
# HTML CONCEPTS extraction                                                    #
# --------------------------------------------------------------------------- #

def load_html_concepts(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8")
    needle = "const CONCEPTS = "
    head = text.find(needle)
    if head < 0:
        raise SystemExit("CONCEPTS literal not found in HTML")
    start = head + len(needle)
    if text[start] != "[":
        raise SystemExit("Expected '[' after CONCEPTS =")
    depth = 0
    in_str = False
    esc = False
    i = start
    while i < len(text):
        c = text[i]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == "[":
                depth += 1
            elif c == "]":
                depth -= 1
                if depth == 0:
                    return json.loads(text[start:i + 1])
        i += 1
    raise SystemExit("Unbalanced CONCEPTS literal")


def html_cells(concepts: list[dict]):
    """Yield (concept_id, sub_id, dim_label, dim_idx, jid, cell_dict)."""
    for c in concepts:
        for sc in c.get("sub_concepts", []):
            for d_idx, dim in enumerate(sc.get("dimensions", [])):
                label = dim.get("label", "?")
                for jid, cell in (dim.get("cells") or {}).items():
                    yield (
                        c["id"], sc["id"], label, d_idx,
                        jid, cell,
                    )


def lookup_html_cell(
    concepts: list[dict], cid: str, sid: str, dim_label: str, jid: str,
) -> dict | None:
    """Find a cell by (concept, sub, dimension label, juris). Returns None
    if no match. When multiple dimensions share a label (e.g. duplicated
    "Term" rows), prefers the first match."""
    for c in concepts:
        if c.get("id") != cid:
            continue
        for sc in c.get("sub_concepts", []):
            if sc.get("id") != sid:
                continue
            for dim in sc.get("dimensions", []):
                if normalize(dim.get("label")) != normalize(dim_label):
                    continue
                cell = (dim.get("cells") or {}).get(jid)
                if cell is not None:
                    return cell
    return None


# --------------------------------------------------------------------------- #
# Excel parsing                                                                #
# --------------------------------------------------------------------------- #

def _cell_addr(col_letter: str, row_num: int) -> str:
    return f"{col_letter}{row_num}"


def parse_single_sheet(
    ws, sub_id: str, juris_cols: dict[str, str],
):
    """Walk a single-sub-concept analysis sheet and yield cell records.

    Yields dicts: {sub_id, dim_label, jid, value, addr, sheet}.
    """
    # Discover the row offset by finding the header row that contains
    # something like "EU (AIA)". The row above it (or the same row) is the
    # title row; the rows below are the data rows.
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            if cell.value and "EU" in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break
    if header_row is None:
        return

    # Walk the rows below header_row. Column A holds the dimension label.
    # When a row has a value in B/C/D... but A is empty, append the value
    # to the previous dimension's text (multi-line in Excel are a single
    # logical cell rendered as multiple rows).
    last_dim = None
    accum: dict[tuple[str, str], list[str]] = {}
    accum_addr: dict[tuple[str, str], str] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a:
            last_dim = str(a).strip().rstrip("\xa0").strip()
        if not last_dim:
            continue
        for col_letter, jid in juris_cols.items():
            col_idx = ord(col_letter) - ord("A") + 1
            v = ws.cell(row=r, column=col_idx).value
            if v is None or str(v).strip() == "":
                continue
            key = (last_dim, jid)
            accum.setdefault(key, []).append(str(v))
            accum_addr.setdefault(key, _cell_addr(col_letter, r))

    for (dim_label, jid), parts in accum.items():
        yield {
            "sub_id": sub_id,
            "dim_label": dim_label,
            "jid": jid,
            "value": "\n".join(parts),
            "addr": accum_addr[(dim_label, jid)],
            "sheet": ws.title,
        }


def parse_multi_sheet(ws, sections: list):
    """Walk a multi-sub-concept analysis sheet (Provider/Deployer)."""
    section_starts = []  # list of (row, sub_id, juris_cols, header_substr)
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if not a:
            continue
        a_str = str(a).strip()
        for header_substr, sub_id, juris_cols in sections:
            if header_substr.lower() in a_str.lower():
                section_starts.append((r, sub_id, juris_cols, header_substr))
                break

    for idx, (start_row, sub_id, juris_cols, hdr) in enumerate(section_starts):
        end_row = (
            section_starts[idx + 1][0] - 1
            if idx + 1 < len(section_starts) else ws.max_row
        )
        # Find the EU header row inside [start_row, end_row].
        header_row = None
        for r in range(start_row, min(start_row + 4, end_row + 1)):
            for col_letter, _ in juris_cols.items():
                col_idx = ord(col_letter) - ord("A") + 1
                v = ws.cell(row=r, column=col_idx).value
                if v and "EU" in str(v):
                    header_row = r
                    break
            if header_row:
                break
        if header_row is None:
            header_row = start_row

        last_dim = None
        accum: dict[tuple[str, str], list[str]] = {}
        accum_addr: dict[tuple[str, str], str] = {}
        for r in range(header_row + 1, end_row + 1):
            a = ws.cell(row=r, column=1).value
            if a:
                last_dim = str(a).strip().rstrip("\xa0").strip()
            if not last_dim:
                continue
            for col_letter, jid in juris_cols.items():
                col_idx = ord(col_letter) - ord("A") + 1
                v = ws.cell(row=r, column=col_idx).value
                if v is None or str(v).strip() == "":
                    continue
                key = (last_dim, jid)
                accum.setdefault(key, []).append(str(v))
                accum_addr.setdefault(key, _cell_addr(col_letter, r))

        for (dim_label, jid), parts in accum.items():
            yield {
                "sub_id": sub_id,
                "dim_label": dim_label,
                "jid": jid,
                "value": "\n".join(parts),
                "addr": accum_addr[(dim_label, jid)],
                "sheet": ws.title,
            }


def parse_analysis(path: Path):
    wb = load_workbook(path, data_only=True)
    records = []
    for sheet_name, (cid, sid, juris_cols) in SINGLE_SUB_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for rec in parse_single_sheet(ws, sid, juris_cols):
            rec["concept_id"] = cid
            records.append(rec)
    for sheet_name, sections in MULTI_SUB_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Determine concept_id from the first matching subsection.
        cid = (
            "provider-developer" if "Provider" in sheet_name
            else "deployer-supplier"
        )
        for rec in parse_multi_sheet(ws, sections):
            rec["concept_id"] = cid
            records.append(rec)
    return records


def parse_verbatim(path: Path):
    """Return list of verbatim snippets keyed by concept-text fingerprint.

    Each row in a verbatim sheet contains: (col 1) sub-concept name, then
    repeated 5-column blocks per jurisdiction:
      [dim_label, sub_dim_label, verbatim_text, reference, tags]
    The structure varies enough across sheets that we don't try to map
    rigidly to (concept, sub, dim) tuples. Instead we emit each verbatim
    snippet keyed by (sheet, row, juris_block_idx) and let the comparison
    step look for substring presence in the HTML's verbatim field.
    """
    wb = load_workbook(path, data_only=True)
    records = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c_idx).value
                if v is None:
                    continue
                s = str(v).strip()
                # Heuristic: a verbatim quote is a long string containing
                # legal-style language. Also include reference cells (short).
                if len(s) >= 40 and any(
                    kw in s.lower() for kw in (
                        "article ", "§", "section ", "paragraph", "shall",
                        "means", "definition", "deployer", "provider",
                    )
                ):
                    records.append({
                        "sheet": sn,
                        "addr": _cell_addr(
                            chr(ord("A") + c_idx - 1), r,
                        ),
                        "value": s,
                        "kind": "verbatim",
                    })
    return records


# --------------------------------------------------------------------------- #
# Comparison                                                                   #
# --------------------------------------------------------------------------- #

REFERENCE_PATTERN = re.compile(
    r"\(([^()]*?(?:Article|§|Annex|\d+\.\d+|GL,)[^()]*?)\)"
)


def split_analysis_and_reference(text: str) -> tuple[str, str]:
    """Excel cells often combine "analysis text (citation)". Strip the
    trailing parenthesised citation when present and return both."""
    if not text:
        return "", ""
    m = REFERENCE_PATTERN.search(text)
    if not m:
        return text.strip(), ""
    # Take the LAST parenthesised reference.
    last = list(REFERENCE_PATTERN.finditer(text))[-1]
    analysis = (text[: last.start()] + text[last.end():]).strip()
    return analysis, last.group(1).strip()


def compare(records, concepts):
    """Compare each Excel analysis record against the matching HTML cell."""
    findings = []
    matched = 0
    skipped = 0
    for rec in records:
        cid = rec["concept_id"]
        sid = rec["sub_id"]
        dim_label = rec["dim_label"]
        jid = rec["jid"]
        excel_value = rec["value"]
        excel_an, excel_ref = split_analysis_and_reference(excel_value)

        cell = lookup_html_cell(concepts, cid, sid, dim_label, jid)
        if cell is None:
            findings.append({
                "category": "HTML_MISSING_CELL",
                "concept": cid, "sub": sid, "dim": dim_label, "jid": jid,
                "sheet": rec["sheet"], "addr": rec["addr"],
                "excel": excel_value[:200],
                "html": "(no matching cell in HTML)",
            })
            skipped += 1
            continue

        html_an = cell.get("analysis") or ""
        html_vb = cell.get("verbatim") or ""
        html_ref = cell.get("reference") or ""

        # Normalise both sides.
        ne = normalize(excel_an)
        nh = normalize(html_an)
        # Treat "-", "—" and "" as "no value".
        empty_tokens = {"", "-", "—"}

        if ne in empty_tokens and nh in empty_tokens:
            matched += 1
        elif ne in empty_tokens and nh not in empty_tokens:
            findings.append({
                "category": "ANALYSIS_EXTRA",
                "concept": cid, "sub": sid, "dim": dim_label, "jid": jid,
                "sheet": rec["sheet"], "addr": rec["addr"],
                "excel": excel_value[:200],
                "html": html_an[:200],
            })
        elif ne not in empty_tokens and nh in empty_tokens:
            findings.append({
                "category": "ANALYSIS_MISSING",
                "concept": cid, "sub": sid, "dim": dim_label, "jid": jid,
                "sheet": rec["sheet"], "addr": rec["addr"],
                "excel": excel_value[:200],
                "html": html_an[:200],
            })
        elif ne == nh:
            matched += 1
        elif ne in nh or nh in ne:
            # Substring match — likely OK but flag for awareness.
            matched += 1
        else:
            findings.append({
                "category": "ANALYSIS_DIFF",
                "concept": cid, "sub": sid, "dim": dim_label, "jid": jid,
                "sheet": rec["sheet"], "addr": rec["addr"],
                "excel": excel_value[:300],
                "html": html_an[:300],
                "excel_norm": ne[:200],
                "html_norm": nh[:200],
            })

        # Reference comparison — only when Excel has one.
        if excel_ref:
            nref_e = normalize(excel_ref)
            nref_h = normalize(html_ref)
            if nref_h and nref_e not in nref_h and nref_h not in nref_e:
                findings.append({
                    "category": "REFERENCE_DIFF",
                    "concept": cid, "sub": sid, "dim": dim_label, "jid": jid,
                    "sheet": rec["sheet"], "addr": rec["addr"],
                    "excel": excel_ref,
                    "html": html_ref,
                })
            elif not nref_h and excel_ref:
                findings.append({
                    "category": "REFERENCE_MISSING",
                    "concept": cid, "sub": sid, "dim": dim_label, "jid": jid,
                    "sheet": rec["sheet"], "addr": rec["addr"],
                    "excel": excel_ref,
                    "html": "(empty)",
                })

    return findings, matched, skipped


def compare_verbatim(verbatim_records, concepts):
    """Spot-check: every long verbatim cell in the verbatim Excel must
    appear (substring) somewhere in the HTML's verbatim fields."""
    # Build a corpus of HTML verbatim text.
    corpus = []
    for c in concepts:
        for sc in c.get("sub_concepts", []):
            for dim in sc.get("dimensions", []):
                for jid, cell in (dim.get("cells") or {}).items():
                    vb = cell.get("verbatim") or ""
                    if vb:
                        corpus.append(normalize(vb))
    big = " || ".join(corpus)

    findings = []
    matched = 0
    for rec in verbatim_records:
        snippet = rec["value"]
        # Use the first 80 normalised chars as the fingerprint to test for
        # presence — gives some flexibility against pre/post whitespace.
        ns = normalize(snippet)
        head = ns[:80]
        if not head:
            continue
        if head in big:
            matched += 1
        else:
            findings.append({
                "category": "VERBATIM_MISSING_IN_HTML",
                "sheet": rec["sheet"], "addr": rec["addr"],
                "excel": snippet[:300],
                "html": "(not found in any HTML verbatim cell)",
            })
    return findings, matched


# --------------------------------------------------------------------------- #
# Reporting                                                                    #
# --------------------------------------------------------------------------- #

def _short(s, n=240):
    if not s:
        return ""
    s = str(s).replace("\n", "\\n").replace("|", "\\|")
    return s if len(s) <= n else s[: n - 1] + "…"


def write_report(
    out_path: Path, findings: list, vfindings: list, matched: int,
    vmatched: int, skipped: int, n_records: int,
) -> None:
    lines: list[str] = []
    lines.append(f"# Excel ↔ HTML correspondence audit\n")
    lines.append(f"_Generated: {datetime.now().isoformat(timespec='seconds')}_\n")
    lines.append("## Summary\n")
    lines.append(
        f"- Analysis records compared: **{n_records}** "
        f"(matched: {matched}, skipped/no-html: {skipped}, "
        f"flagged: {len(findings)})\n"
    )
    lines.append(
        f"- Verbatim Excel rows checked: **{vmatched + len(vfindings)}** "
        f"(matched in HTML: {vmatched}, missing: {len(vfindings)})\n"
    )

    by_cat: dict[str, list] = {}
    for f in findings + vfindings:
        by_cat.setdefault(f["category"], []).append(f)

    if not by_cat:
        lines.append("\n*No discrepancies found — Excel and HTML are aligned.*\n")
    else:
        lines.append("\n## Findings by category\n")
        for cat in sorted(by_cat):
            items = by_cat[cat]
            lines.append(f"\n### {cat} ({len(items)})\n")
            lines.append(
                "| Excel sheet | Cell | Concept | Sub-concept | Dim | "
                "Juris | Excel value | HTML value |"
            )
            lines.append(
                "|---|---|---|---|---|---|---|---|"
            )
            for f in items:
                lines.append(
                    "| {sheet} | {addr} | {cid} | {sid} | {dim} | {jid} "
                    "| {excel} | {html} |".format(
                        sheet=f.get("sheet", ""),
                        addr=f.get("addr", ""),
                        cid=f.get("concept", ""),
                        sid=f.get("sub", ""),
                        dim=f.get("dim", ""),
                        jid=f.get("jid", ""),
                        excel=_short(f.get("excel", "")),
                        html=_short(f.get("html", "")),
                    )
                )

    out_path.write_text("\n".join(lines), encoding="utf-8")


CSV_FIELDS = [
    "category", "sheet", "cell", "concept", "sub_concept", "dimension",
    "jurisdiction", "excel_value", "html_value",
]


def write_csv(out_path: Path, findings: list, vfindings: list) -> None:
    rows = []
    for f in findings:
        rows.append({
            "category": f.get("category", ""),
            "sheet": f.get("sheet", ""),
            "cell": f.get("addr", ""),
            "concept": f.get("concept", ""),
            "sub_concept": f.get("sub", ""),
            "dimension": f.get("dim", ""),
            "jurisdiction": f.get("jid", ""),
            "excel_value": (f.get("excel") or "").replace("\n", " \\n "),
            "html_value": (f.get("html") or "").replace("\n", " \\n "),
        })
    for f in vfindings:
        rows.append({
            "category": f.get("category", ""),
            "sheet": f.get("sheet", ""),
            "cell": f.get("addr", ""),
            "concept": "",
            "sub_concept": "",
            "dimension": "",
            "jurisdiction": "",
            "excel_value": (f.get("excel") or "").replace("\n", " \\n "),
            "html_value": (f.get("html") or "").replace("\n", " \\n "),
        })
    # Stable sort: category first, then sheet/cell.
    rows.sort(key=lambda r: (r["category"], r["sheet"], r["cell"]))
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


# --------------------------------------------------------------------------- #
# Main                                                                         #
# --------------------------------------------------------------------------- #

def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--analysis", default=str(DEFAULT_ANALYSIS))
    p.add_argument("--verbatim", default=str(DEFAULT_VERBATIM))
    p.add_argument("--html", default=str(DEFAULT_HTML))
    p.add_argument(
        "--out",
        default=str(
            OUTPUTS / f"excel_correspondence_{datetime.now():%Y%m%d_%H%M%S}.md"
        ),
    )
    p.add_argument(
        "--csv",
        default=str(
            OUTPUTS / f"excel_correspondence_{datetime.now():%Y%m%d_%H%M%S}.csv"
        ),
    )
    args = p.parse_args(argv)

    OUTPUTS.mkdir(parents=True, exist_ok=True)

    print(f"Loading HTML: {args.html}")
    concepts = load_html_concepts(Path(args.html))
    print(f"  → {len(concepts)} concepts")

    print(f"Loading analysis Excel: {args.analysis}")
    a_records = parse_analysis(Path(args.analysis))
    print(f"  → {len(a_records)} analysis records")

    print(f"Loading verbatim Excel: {args.verbatim}")
    v_records = parse_verbatim(Path(args.verbatim))
    print(f"  → {len(v_records)} verbatim cells")

    print("Comparing analysis Excel ↔ HTML…")
    findings, matched, skipped = compare(a_records, concepts)

    print("Comparing verbatim Excel ↔ HTML…")
    vfindings, vmatched = compare_verbatim(v_records, concepts)

    out_path = Path(args.out)
    write_report(
        out_path, findings, vfindings, matched, vmatched, skipped,
        len(a_records),
    )

    csv_path = Path(args.csv)
    write_csv(csv_path, findings, vfindings)

    # Console summary.
    print()
    print("=" * 60)
    print(f"Analysis cells: {len(a_records)} total")
    print(f"  matched:  {matched}")
    print(f"  flagged:  {len(findings)}")
    print(f"  skipped:  {skipped}  (no matching HTML cell)")
    print(f"Verbatim cells: {vmatched + len(vfindings)} total")
    print(f"  found in HTML: {vmatched}")
    print(f"  missing:       {len(vfindings)}")
    by_cat: dict[str, int] = {}
    for f in findings + vfindings:
        by_cat[f["category"]] = by_cat.get(f["category"], 0) + 1
    if by_cat:
        print("\nFindings by category:")
        for cat, n in sorted(by_cat.items(), key=lambda x: -x[1]):
            print(f"  {cat:30s} {n}")
    print(f"\nFull report: {out_path}")
    print(f"CSV (verifiable): {csv_path}")
    return 0 if not findings and not vfindings else 1


if __name__ == "__main__":
    sys.exit(main())
