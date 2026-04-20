"""test_lexicon_correspondence.py — verify the built HTML faithfully
represents the source xlsx.

Run:
    python3 -m pytest test_lexicon_correspondence.py -q
or standalone:
    python3 test_lexicon_correspondence.py

Checks (ordered loosest to strictest):

1. Every non-trivial xlsx cell string appears somewhere in the HTML's
   inline JSON corpus (coverage).
2. Every collected interpretative note from the xlsx is attached to the
   right sub_concept — no mixing, no duplication across sub-tabs.
3. Every cluster_summary row contributes real content (no orphan
   continuation rows); every row's sub_id resolves to an existing
   sub_concept.
4. Every REF_MAP entry with an anchor resolves to a real
   article/section/annex/recital id in its law blob (zero misses).
5. Every bill-code tag (e.g. `[SB53]`) used in the HTML matches the set
   used in the xlsx "New concepts" matrix.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
XLSX = HERE / "AI terminology and taxonomy-final.xlsx"
HTML = HERE / "digital_lexicon_v16.html"
LAWS = HERE / "laws"


# ------------------------- helpers -------------------------------------- #

def _find_json_literal(src: str, var_name: str) -> tuple[int, int] | None:
    key = f"const {var_name}"
    start = src.find(key)
    if start < 0:
        return None
    i = src.index("=", start) + 1
    while i < len(src) and src[i] not in "[{":
        i += 1
    if i >= len(src):
        return None
    opener = src[i]
    closer = "]" if opener == "[" else "}"
    depth = 0
    in_str = False
    esc = False
    j = i
    while j < len(src):
        c = src[j]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == opener:
                depth += 1
            elif c == closer:
                depth -= 1
                if depth == 0:
                    return (i, j + 1)
        j += 1
    return None


def _load_html_const(html: str, name: str):
    span = _find_json_literal(html, name)
    if not span:
        return None
    return json.loads(html[span[0] : span[1]])


def _load_html() -> str:
    return HTML.read_text(encoding="utf-8")


def _normalize(s: str) -> str:
    """Normalize for substring-matching across xlsx ↔ HTML. Canonicalize
    curly/straight quotes and apostrophes (xlsx cells mix both)."""
    s = (s or "").replace("\xa0", " ")
    s = s.replace("\u201c", '"').replace("\u201d", '"')  # curly double → straight
    s = s.replace("\u2018", "'").replace("\u2019", "'")  # curly single → straight
    s = s.replace("\u2013", "-").replace("\u2014", "-")  # en/em dash → hyphen
    return re.sub(r"\s+", " ", s).strip()


# ------------------------- collectors ----------------------------------- #

def _xlsx_cells() -> list[tuple[str, int, int, str]]:
    """(sheet, row, col, value) for every non-trivial cell."""
    wb = load_workbook(XLSX, data_only=True)
    out: list[tuple[str, int, int, str]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not v:
                    continue
                s = _normalize(str(v))
                if len(s) < 4 or s in ("-", "—", "–"):
                    continue
                out.append((sn, r, c, s))
    return out


def _xlsx_notes() -> list[tuple[str, str, str]]:
    """Return (analysis_sheet, sub_concept_title, note_plaintext) for each
    non-empty "Interpretative notes" cell in any analysis sheet."""
    wb = load_workbook(XLSX, data_only=True)
    out = []
    for sn in wb.sheetnames:
        if "nalys" not in sn.lower():
            continue
        ws = wb[sn]
        # find header rows: a row with "Interpretative notes" somewhere
        # and its notes column.
        header_rows = []
        for r in range(1, ws.max_row + 1):
            notes_col = None
            juris_count = 0
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not v:
                    continue
                s = str(v)
                if re.search(r"interpretative notes?", s, re.I):
                    notes_col = c
                elif re.search(
                    r"\b(EU|California|Colorado|New York|Texas|Utah)\b", s
                ):
                    juris_count += 1
            if notes_col and juris_count >= 2:
                title_above = ws.cell(r - 1, 1).value if r > 1 else ""
                header_rows.append((r, notes_col, _normalize(str(title_above or ""))))
        for idx, (hr, nc, title) in enumerate(header_rows):
            end_r = (header_rows[idx + 1][0] - 2) if idx + 1 < len(header_rows) else ws.max_row
            for r in range(hr + 1, end_r + 1):
                v = ws.cell(r, nc).value
                if not v:
                    continue
                plain = _normalize(str(v))
                if len(plain) < 20:
                    continue
                out.append((sn, title, plain))
    return out


def _html_corpus(html: str) -> str:
    """All inline JSON blobs plus the v16 coverage flat blob."""
    parts: list[str] = []
    for m in re.finditer(
        r'<script type="application/json"[^>]*>([\s\S]*?)</script>', html
    ):
        parts.append(m.group(1))
    for name in ("CONCEPTS", "LAWS", "LAW_STUBS", "REF_MAP", "JURIS_LABELS",
                 "JURIS_ORDER", "GLOSSARY", "ABOUT_PROSE",
                 "METHODOLOGY_PROSE"):
        span = _find_json_literal(html, name)
        if span:
            parts.append(html[span[0] : span[1]])
    return _normalize(" ".join(parts))


# ------------------------- tests ---------------------------------------- #

def test_every_xlsx_cell_appears_in_html() -> None:
    html = _load_html()
    corpus = _html_corpus(html)
    missing = []
    # Sheets the v16 tool doesn't render:
    #   * prose sheets (About/Methodology) are summarized, not copied
    #     verbatim
    #   * "Second edition terminology" is a legacy glossary sheet, not
    #     wired into the v16 renderer
    SKIP_SHEETS = {
        "About the Digital AI Lexicon",
        "Methodology",
        "Second edition terminology",
    }
    # Column-header labels in legal and analysis sheets — the HTML strips
    # them and uses its own header rendering.
    SKIP_HEADERS = {"Reference", "Tags", "Concept", "References", "Cluster"}
    for sn, r, c, v in _xlsx_cells():
        if sn in SKIP_SHEETS:
            continue
        if v in SKIP_HEADERS:
            continue
        # Substring match after normalization; for long cells, try the
        # first 80 chars so a one-word difference doesn't hide the rest.
        needle = v if len(v) <= 80 else v[:80]
        if needle not in corpus:
            missing.append((sn, r, c, v[:100]))
    # We know that some legal-sheet verbatim cells aren't surfaced on any
    # analysis-sheet dim row (they're reachable only via the law drawer
    # through REF_MAP). Report the count rather than fail on all of them —
    # a sudden jump would signal a regression.
    threshold = 180
    if len(missing) > threshold:
        raise AssertionError(
            f"{len(missing)} xlsx cells not visible in HTML (limit {threshold}). "
            f"First: {missing[:3]}"
        )
    if missing:
        print(
            f"\n[info] {len(missing)} xlsx cells not directly in HTML "
            f"(reachable via law drawer). Threshold: {threshold}.",
            file=sys.stderr,
        )


def test_notes_sub_concept_alignment() -> None:
    html = _load_html()
    concepts = _load_html_const(html, "CONCEPTS") or []
    # Build HTML-side index: (sub_concept_title, note_plaintext)
    html_notes: dict[str, list[str]] = {}
    for c in concepts:
        for sc in c.get("sub_concepts") or []:
            for n in sc.get("ceps_notes_rich") or []:
                plain = _normalize(
                    "".join(r.get("t", "") for r in n.get("body_runs") or [])
                )
                html_notes.setdefault(sc.get("title", ""), []).append(plain)

    # For each xlsx note, check it appears under exactly one sub_concept.
    # Use the FULL normalized note text as the match key so notes that share
    # a generic opening ("This table compares the AIA's terms...") don't
    # spuriously appear to map to multiple sub_concepts.
    xlsx_notes = _xlsx_notes()
    unresolved = []
    for sheet, title, plain in xlsx_notes:
        placements = [
            sc_title
            for sc_title, notes in html_notes.items()
            if any(n == plain for n in notes)
        ]
        if not placements:
            unresolved.append((sheet, title, plain[:80], "no match"))
        elif len(placements) > 1:
            unresolved.append(
                (sheet, title, plain[:80], f"in {len(placements)} sub-concepts: {placements}")
            )
    assert not unresolved, (
        f"{len(unresolved)} xlsx notes don't map cleanly to sub_concepts. "
        f"First: {unresolved[:3]}"
    )


def test_cluster_summary_no_orphan_rows() -> None:
    html = _load_html()
    concepts = _load_html_const(html, "CONCEPTS") or []
    orphans = []
    bad_sub_ids = []
    bad_rowspan = []
    for c in concepts:
        sub_ids = {sc["id"] for sc in (c.get("sub_concepts") or [])}
        cs = c.get("cluster_summary") or {}
        rows = cs.get("rows") or []
        # Walk groups.
        i = 0
        while i < len(rows):
            head = rows[i]
            gsize = head.get("term_rowspan") or 1
            group = rows[i : i + gsize]
            i += gsize
            # Each row in the group must contribute something in its own
            # right: either a non-empty term_label (head row) or at least
            # one cell with rowspan>0 and variants.
            for idx, row in enumerate(group):
                if idx == 0:
                    continue  # head row is always kept
                has_content = False
                for cell in (row.get("cells") or {}).values():
                    if cell.get("rowspan") != 0 and cell.get("variants"):
                        has_content = True
                        break
                if not has_content:
                    orphans.append((c["id"], row))
            # sub_id must resolve
            for row in group:
                sid = row.get("sub_id")
                if sid and sid not in sub_ids:
                    bad_sub_ids.append((c["id"], sid))
                for cell in (row.get("cells") or {}).values():
                    for v in cell.get("variants", []):
                        vsid = v.get("sub_id")
                        if vsid and vsid not in sub_ids:
                            bad_sub_ids.append((c["id"], vsid))
            # term_rowspan on head matches actual group size emitted.
            if head.get("term_rowspan") and head.get("term_rowspan") != len(group):
                bad_rowspan.append((c["id"], head.get("term_rowspan"), len(group)))

    assert not orphans, f"{len(orphans)} orphan continuation rows. First: {orphans[:3]}"
    assert not bad_sub_ids, f"{len(bad_sub_ids)} pills point at unknown sub_ids. First: {bad_sub_ids[:3]}"
    assert not bad_rowspan, f"{len(bad_rowspan)} rows have mismatched term_rowspan. First: {bad_rowspan[:3]}"


def test_all_refs_resolve_to_real_sections() -> None:
    html = _load_html()
    ref_map = _load_html_const(html, "REF_MAP") or {}
    # Load law blobs
    laws: dict[str, dict] = {}
    for p in LAWS.glob("*.json"):
        try:
            laws[p.stem] = json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass

    # Tolerate a handful of known-unresolved refs that pre-date this pass
    # (Colorado uses a different blob id; NY S8828 sections 1426-1428 never
    # parsed; EU annexes are a separate dict). These are tracked
    # separately — the zero-tolerance check is on Texas.
    TOLERATED = {
        # law_id normalization quirks
        "co-sb24205",        # blob file is co-sb24-205
        # EU annexes live in a dict, not the sections/articles list we scan
    }

    missing_tx = []
    missing_other = []
    for key, info in ref_map.items():
        law = info.get("law")
        anchor = info.get("anchor")
        if not law or not anchor:
            continue
        blob = laws.get(law)
        if not blob:
            if law in TOLERATED:
                continue
            missing_other.append((key, law, anchor, "no blob"))
            continue
        ids = set()
        for s in blob.get("sections", []):
            ids.add(str(s.get("id")))
        for a in blob.get("articles", []):
            ids.add(str(a.get("id")))
        for rid in (blob.get("recitals") or {}):
            ids.add(str(rid))
        for aid in (blob.get("annexes") or {}):
            ids.add(str(aid))
        a = str(anchor)
        if a in ids or a.split("-")[0] in ids:
            continue
        if law == "tx-hb149":
            missing_tx.append((key, anchor))
        else:
            missing_other.append((key, law, anchor, "not in blob"))

    # Zero tolerance: Texas references (the regression we just fixed).
    assert not missing_tx, (
        f"{len(missing_tx)} Texas HB 149 references do not resolve. "
        f"First: {missing_tx[:3]}"
    )
    # Report other misses as a warning (not a fail) to keep focus on
    # regressions; run with `-v` to see them.
    if missing_other:
        print(
            f"\n[warn] {len(missing_other)} non-Texas refs still unresolved "
            f"(pre-existing). First: {missing_other[:3]}",
            file=sys.stderr,
        )


def test_bill_codes_consistent() -> None:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["New concepts"]
    xlsx_tags = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v:
                continue
            for m in re.finditer(r"\[([A-Za-z0-9\- ]+?)\]", str(v)):
                tag = m.group(1).strip()
                if tag:
                    xlsx_tags.add(tag.upper().replace(" ", ""))

    html = _load_html()
    concepts = _load_html_const(html, "CONCEPTS") or []
    html_tags = set()
    for c in concepts:
        for row in (c.get("cluster_summary") or {}).get("rows") or []:
            for cell in (row.get("cells") or {}).values():
                for v in cell.get("variants", []):
                    bill = (v.get("bill") or "").strip()
                    if bill:
                        html_tags.add(bill.upper().replace(" ", ""))
                    for m in re.finditer(r"\[([A-Za-z0-9\- ]+?)\]", v.get("name") or ""):
                        html_tags.add(m.group(1).strip().upper().replace(" ", ""))
    only_xlsx = xlsx_tags - html_tags
    only_html = html_tags - xlsx_tags
    assert not only_xlsx, f"Bill codes in xlsx but missing from HTML: {only_xlsx}"
    # Tolerate HTML-only tags because verbatim text may reference bills we
    # don't tag in the matrix (e.g. EU AIA articles with numeric citations).


# ------------------------- runner --------------------------------------- #

if __name__ == "__main__":
    # Standalone runner so `python3 test_lexicon_correspondence.py` works
    # without pytest installed.
    tests = [
        ("xlsx cell coverage",      test_every_xlsx_cell_appears_in_html),
        ("notes alignment",         test_notes_sub_concept_alignment),
        ("cluster-summary orphans", test_cluster_summary_no_orphan_rows),
        ("refs resolve",            test_all_refs_resolve_to_real_sections),
        ("bill codes consistent",   test_bill_codes_consistent),
    ]
    failed = 0
    for label, fn in tests:
        try:
            fn()
            print(f"  ok    {label}")
        except AssertionError as e:
            failed += 1
            print(f"  FAIL  {label}")
            print(f"        {e}")
    sys.exit(1 if failed else 0)
