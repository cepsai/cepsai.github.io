"""build_v23.py — Digital AI Lexicon v23.

User feedback round 2026-04-26 (post-v22 review):

  1. **Regulation #12** — `eu-guidelines-gpai-scope` (Commission Guidelines
     on the Scope of the Obligations for Providers of General-Purpose AI
     Models, C(2025) 7719 final, 19.11.2025; PDF GUIDEL~2.PDF).
     The JSON has been ingested into `laws/eu-guidelines-gpai-scope.json`
     (6 sections + 99,870 chars raw_text) but is **not embedded** in v22.
     v23 wires it into the law-blob list, the runtime LAW_STUBS map, and
     the European Union region of the Regulations / Law Sources page.

  2. **"See verbatim in full law" / Explore-in-full-law buttons** — verify
     they all resolve to a real article/section/anchor in the embedded
     blob. Anchor mismatches in v22 carry over from v21's parser (e.g.
     California references where the parser greedily picked a 4-digit bill
     number instead of the section). v23 patches `openLawDrawerById` at
     the JS layer with progressive-suffix-strip lookup so anchors like
     `22757.11-c` correctly resolve to the existing `22757.11` section.

  3. **All article references correct** — extends correspondence-test
     coverage with a new `test_lexicon_v23.py` suite that walks every
     dimension cell's `reference` field, parses it, and asserts the
     resolved article id matches a stated number in the reference text.

Build chain: v13 → v15 → v16 → v17 → v18 → v20 → v21 → v22 → **v23**.

v23 operates as a post-process on v22:
    1. If digital_lexicon_v22.html is missing, rebuild via build_v22.
    2. Read v22 HTML.
    3. Inject the eu-guidelines-gpai-scope law blob.
    4. Extend the runtime `window.LAW_STUBS` map.
    5. Insert the law into the LAWS nav for the European Union region
       (after the AI-system definition Guidelines; before the CoP entry).
    6. Append the v23 anchor-resolution patch script to <body>.
    7. Write digital_lexicon_v23.html + copy to ../final_tool.html and
       ../final_lexicon_tool.html.

Run:
    python3 build_v23.py
"""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

HERE          = Path(__file__).parent
HTML_V22      = HERE / "digital_lexicon_v22.html"
HTML_V23      = HERE / "digital_lexicon_v23.html"
FINAL_TOOL    = HERE.parent / "final_tool.html"
FINAL_LEXICON = HERE.parent / "final_lexicon_tool.html"
LAWS_DIR      = HERE / "laws"

NEW_XLSX = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis.xlsx"
)

REG12_LAW_ID = "eu-guidelines-gpai-scope"
REG12_META = {
    "code":      "Comm. Guidelines — GPAI provider obligations",
    "title":     "Commission Guidelines on the Scope of the Obligations for "
                 "Providers of General-Purpose AI Models",
    "effective": "2025-11-19",
    "desc":      "Interpretive guidance on which actors are 'providers placing "
                 "on the market' GPAI models and their Article 53/55 "
                 "obligations (C(2025) 7719 final).",
    "url":       "https://digital-strategy.ec.europa.eu/en/library/"
                 "guidelines-scope-obligations-providers-general-purpose-"
                 "ai-models-under-ai-act",
    "law_id":    REG12_LAW_ID,
}


# --------------------------------------------------------------------------- #
# JSON-literal locator (bracket-aware) — same logic as v22.                    #
# --------------------------------------------------------------------------- #

def _find_json_literal(src: str, var_name: str) -> tuple[int, int] | None:
    """Locate the JSON literal that follows `const <var_name> =` (or
    `window.<var_name> =`). Returns (start, end) byte offsets such that
    src[start:end] is the literal text including its outer brackets."""
    for prefix in (f"const {var_name}", f"window.{var_name}",
                   f"var {var_name}", f"let {var_name}"):
        start = src.find(prefix)
        if start < 0:
            continue
        i = src.index("=", start) + 1
        while i < len(src) and src[i] not in "[{":
            i += 1
        if i >= len(src):
            continue
        opener = src[i]
        closer = "]" if opener == "[" else "}"
        depth, in_str, esc, j = 0, False, False, i
        quote = ""
        while j < len(src):
            c = src[j]
            if in_str:
                if esc:
                    esc = False
                elif c == "\\":
                    esc = True
                elif c == quote:
                    in_str = False
            elif c in '"\'':
                in_str = True
                quote = c
            elif c == opener:
                depth += 1
            elif c == closer:
                depth -= 1
                if depth == 0:
                    return (i, j + 1)
            j += 1
    return None


def _replace_json_literal(html: str, var_name: str, new_value) -> str:
    span = _find_json_literal(html, var_name)
    if not span:
        return html
    a, b = span
    payload = json.dumps(new_value, ensure_ascii=False, separators=(",", ":"))
    return html[:a] + payload + html[b:]


# --------------------------------------------------------------------------- #
# v23 best-effort cell-text updater                                           #
# --------------------------------------------------------------------------- #

def _norm_for_match(s: str) -> str:
    s = (s or "").lower()
    s = s.replace(" ", " ")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _build_xlsx_text_index(xlsx_path: Path):
    """Index every analysis-text cell in the new xlsx by
    (jurisdiction_label_lower, dimension_label_lower) → list of texts.

    Returns a dict {(jid_label, dim_label): [text, ...]}. We don't use
    sub-concept_title here — that's where build_v22's fuzzy match misses;
    instead we rely on (jid, dim) overlap to confirm a candidate match
    when the cell's current text doesn't appear in the xlsx anywhere.
    """
    if not xlsx_path.exists():
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    JID_LABEL = {
        "european union": "eu", "eu": "eu", "ai act": "eu", "aia": "eu",
        "california": "ca", "ca": "ca",
        "colorado": "co", "co": "co",
        "new york": "ny", "ny": "ny",
        "texas": "tx", "tx": "tx",
        "utah": "ut", "ut": "ut",
    }
    out: dict = {}
    for sn in wb.sheetnames:
        if not (sn.endswith("_ANALYSIS") or sn.endswith("_ANALY") or
                sn.endswith("_Analys") or sn.endswith("_Analysis")):
            continue
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # Walk: locate rows that look like jid headers, then map each
        # subsequent cell as (jid, dim) → text.
        for ri, row in enumerate(rows):
            if not row:
                continue
            # Detect a jid-header row: a row where ≥2 cells map to JID labels.
            jid_cols: dict[int, str] = {}
            for ci, v in enumerate(row):
                if v is None:
                    continue
                lab = _norm_for_match(str(v))
                for key, jid in JID_LABEL.items():
                    if lab == key or lab.startswith(key + " ") or lab.endswith(" " + key):
                        jid_cols[ci] = jid
                        break
            if len(jid_cols) < 2:
                continue
            # Walk subsequent rows until next header / blank break.
            for rj in range(ri + 1, len(rows)):
                drow = rows[rj]
                if not drow:
                    continue
                col_a = (drow[0] if len(drow) > 0 else None) or \
                        (drow[1] if len(drow) > 1 else None)
                if col_a is None:
                    continue
                dim_label = _norm_for_match(str(col_a))
                if not dim_label:
                    continue
                # Stop on next header row.
                next_jids = sum(1 for ci, v in enumerate(drow)
                                if v is not None and
                                _norm_for_match(str(v)) in JID_LABEL)
                if next_jids >= 2:
                    break
                for ci, jid in jid_cols.items():
                    if ci >= len(drow):
                        continue
                    val = drow[ci]
                    if val is None:
                        continue
                    txt = str(val).strip()
                    if not txt or txt in ("-", "–", "—") or len(txt) < 8:
                        continue
                    out.setdefault((jid, dim_label), []).append(txt)
    return out


def _v23_aggressive_update(html: str, index: dict) -> tuple[str, int]:
    """For each dim cell whose current analysis text isn't a substring of
    *any* xlsx cell, see if there's a uniquely-matching (jid, dim) text in
    the xlsx and substitute it. Returns (html, count_updated)."""
    span = _find_json_literal(html, "CONCEPTS")
    if not span:
        return html, 0
    a, b = span
    try:
        concepts = json.loads(html[a:b])
    except Exception:
        return html, 0

    # Build a set of all xlsx text fingerprints.
    all_xlsx_norm = set()
    for tlist in index.values():
        for t in tlist:
            all_xlsx_norm.add(_norm_for_match(t)[:80])

    updated = 0
    for c in concepts:
        for sc in (c.get("sub_concepts") or []):
            for d in (sc.get("dimensions") or []):
                dim_label_norm = _norm_for_match(d.get("label") or "")
                if not dim_label_norm:
                    continue
                cells = d.get("cells") or {}
                for jid_full, cell in cells.items():
                    if not isinstance(cell, dict):
                        continue
                    cur = cell.get("analysis") or ""
                    if not cur:
                        continue
                    cur_n = _norm_for_match(cur)[:80]
                    if cur_n in all_xlsx_norm:
                        continue   # already matches somewhere
                    # Lookup candidates by (jid_root, dim_label)
                    jid_root = jid_full.split("-", 1)[0]
                    cands = index.get((jid_root, dim_label_norm), [])
                    if len(cands) == 1:
                        cell["analysis"] = cands[0]
                        updated += 1
                        continue
                    # Try partial dim match
                    for (j2, d2), tlist in index.items():
                        if j2 != jid_root:
                            continue
                        if d2 in dim_label_norm or dim_label_norm in d2:
                            if len(tlist) == 1:
                                cell["analysis"] = tlist[0]
                                updated += 1
                                break
    new_payload = json.dumps(concepts, ensure_ascii=False, separators=(",", ":"))
    return html[:a] + new_payload + html[b:], updated


# --------------------------------------------------------------------------- #
# Reg #12 wiring                                                              #
# --------------------------------------------------------------------------- #

def _inject_law_blob(html: str, law_id: str, blob: dict) -> tuple[str, bool]:
    marker = f'id="law-blob-{law_id}"'
    if marker in html:
        return html, False
    payload = json.dumps(blob, ensure_ascii=False,
                         separators=(",", ":")).replace("</", "<\\/")
    tag = f'<script type="application/json" id="law-blob-{law_id}">{payload}</script>'
    return html.replace("</body>", tag + "\n</body>", 1), True


def _add_to_law_stubs(html: str, law_id: str, blob: dict) -> tuple[str, bool]:
    """Append a stub entry to the runtime LAW_STUBS map. The map is built
    by the page itself from the inline JSON blobs, so usually no edit is
    needed — but if v22 hard-codes a const/window assignment we handle it.
    Returns (html, was_changed)."""
    if "LAW_STUBS" not in html:
        return html, False
    span = _find_json_literal(html, "LAW_STUBS")
    if not span:
        return html, False
    a, b = span
    try:
        stubs = json.loads(html[a:b])
    except Exception:
        return html, False
    if law_id in stubs:
        return html, False
    stubs[law_id] = {
        "title":   blob.get("title", ""),
        "url":     blob.get("url", ""),
        "articles": [{"id": a.get("id", ""), "title": a.get("title", "")}
                     for a in (blob.get("articles") or [])],
        "sections": [{"id": s.get("id", ""),
                      "title": s.get("title", "") or "Section " + s.get("id", "")}
                     for s in (blob.get("sections") or [])],
    }
    return html[:a] + json.dumps(stubs, ensure_ascii=False, separators=(",", ":")) + html[b:], True


def _add_to_laws_nav(html: str, entry: dict) -> tuple[str, bool]:
    """Insert the new law into the European Union region of LAWS, just
    before the GPAI CoP entry (so EU ordering: AIA → AI-def Guidelines →
    GPAI provider obligations → CoP). Returns (html, was_changed)."""
    span = _find_json_literal(html, "LAWS")
    if not span:
        return html, False
    a, b = span
    try:
        laws_nav = json.loads(html[a:b])
    except Exception:
        return html, False
    for region in laws_nav:
        if region.get("region") != "European Union":
            continue
        existing = list(region.get("laws") or [])
        if any(l.get("law_id") == entry["law_id"] for l in existing):
            return html, False
        # find first CoP entry to insert before
        insert_at = len(existing)
        for i, l in enumerate(existing):
            lid = (l.get("law_id") or "").lower()
            code = (l.get("code") or "").lower()
            if "cop" in lid or "cop" in code or "code of practice" in code:
                insert_at = i
                break
        existing.insert(insert_at, entry)
        region["laws"] = existing
        new_payload = json.dumps(laws_nav, ensure_ascii=False, separators=(",", ":"))
        return html[:a] + new_payload + html[b:], True
    return html, False


# --------------------------------------------------------------------------- #
# v23 JS overrides — anchor-resolution progressive strip.                     #
# --------------------------------------------------------------------------- #

def _v23_overrides() -> str:
    return r"""
<script>
/* v23 — anchor resolution patch.
   v22's openLawDrawerById exact-matches the article/section id, so anchors
   like "22757.11-c" never resolved against blobs keyed by "22757.11" and
   silently fell back to the first item. This wrap retries with the
   trailing "-token" stripped (up to 2x) before giving up. Safe for both
   EU AI Act articles and US-state section ids. */
(function(){
  function _strip(s){
    if (!s) return s;
    var i = s.lastIndexOf('-');
    return i > 0 ? s.slice(0, i) : '';
  }
  function _findAnchor(items, want){
    if (!items || !items.length || !want) return null;
    var hit = items.find(function(it){ return it && it.id === want; });
    if (hit) return hit;
    var s1 = _strip(want);
    if (s1){
      hit = items.find(function(it){ return it && it.id === s1; });
      if (hit) return hit;
      var s2 = _strip(s1);
      if (s2){
        hit = items.find(function(it){ return it && it.id === s2; });
        if (hit) return hit;
      }
    }
    return null;
  }
  function patch(){
    if (typeof window.openLawDrawerById !== 'function' || window.__v23_anchor_patched) return;
    window.__v23_anchor_patched = true;
    var orig = window.openLawDrawerById;
    window.openLawDrawerById = function(lawId, articleId){
      try {
        if (articleId && typeof window.LAW_STUBS !== 'undefined'){
          var stub = window.LAW_STUBS && window.LAW_STUBS[lawId];
          if (stub){
            var pool = (stub.articles && stub.articles.length) ? stub.articles
                     : (stub.sections || []);
            var hit = _findAnchor(pool, String(articleId));
            if (hit) return orig.call(this, lawId, hit.id);
          }
        }
      } catch(e){ console.warn('v23 anchor wrap:', e); }
      return orig.call(this, lawId, articleId);
    };
  }
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', patch);
  else patch();
})();
</script>
"""


# --------------------------------------------------------------------------- #
# Main                                                                        #
# --------------------------------------------------------------------------- #

def main() -> None:
    print("== v23 build ==")
    if not HTML_V22.exists():
        print("  digital_lexicon_v22.html missing — running build_v22 …")
        import build_v22 as _v22
        _v22.main()

    html = HTML_V22.read_text(encoding="utf-8")
    print(f"  read v22:            {len(html):,} bytes")

    # 1. Load reg #12 JSON.
    reg12_path = LAWS_DIR / f"{REG12_LAW_ID}.json"
    if not reg12_path.exists():
        raise SystemExit(f"Missing {reg12_path} — run ingest first.")
    blob = json.loads(reg12_path.read_text(encoding="utf-8"))
    print(f"  reg #12 JSON:        {blob.get('title','?')[:60]}…")
    print(f"                       sections={len(blob.get('sections') or [])}, "
          f"raw_text={len(blob.get('raw_text') or ''):,} chars")

    # 2. Inject law blob.
    html, did_blob = _inject_law_blob(html, REG12_LAW_ID, blob)
    print(f"  law blob injected:   {'yes' if did_blob else 'already present'}")

    # 3. Extend LAW_STUBS if hard-coded; otherwise the page builds it from blobs.
    html, did_stubs = _add_to_law_stubs(html, REG12_LAW_ID, blob)
    print(f"  LAW_STUBS extended:  {'yes' if did_stubs else 'n/a (built at runtime)'}")

    # 4. Add nav entry to the EU region of LAWS.
    html, did_nav = _add_to_laws_nav(html, REG12_META)
    print(f"  LAWS nav extended:   {'+1 entry' if did_nav else 'already present'}")

    # 5. v23 best-effort cell-text update — close the gap left by v22's
    #    fuzzy concept-name matcher when the xlsx sub-section title differs
    #    from the v21 sub-concept name.
    xlsx_index = _build_xlsx_text_index(NEW_XLSX)
    html, n_updated = _v23_aggressive_update(html, xlsx_index)
    print(f"  cells re-synced from xlsx: {n_updated}")

    # 6. v23 JS overrides.
    overrides = _v23_overrides()
    if "v23_anchor_patched" not in html:
        html = html.replace("</body>", overrides + "\n</body>", 1)
        print(f"  v23 overrides:       injected ({len(overrides)} chars)")

    # 6. Write outputs.
    HTML_V23.write_text(html, encoding="utf-8")
    shutil.copy2(HTML_V23, FINAL_TOOL)
    shutil.copy2(HTML_V23, FINAL_LEXICON)
    print(f"\n  wrote {HTML_V23.name}        ({len(html):,} bytes)")
    print(f"  wrote {FINAL_TOOL.name}            (mirror)")
    print(f"  wrote {FINAL_LEXICON.name}    (mirror)")


if __name__ == "__main__":
    main()
