"""build_v19.py — Digital AI Lexicon v19

Adds a top-level Verbatim tab alongside a renamed "Analysis" tab (was
"Concepts"). Analysis keeps today's behavior (analysis text in cells;
click opens the inline verbatim drawer). Verbatim renders the same
dimension x jurisdiction table with the raw legal text in each cell
and a two-column Dimension / Sub-dimension layout (sourced from the
xlsx legal sheets). Clicking a verbatim cell opens the full-law drawer
directly, skipping the intermediate inline verbatim drawer.

Build chain stays additive: v13 -> v15 -> v16 -> v17 -> v18 -> v19.
v18's output HTML is read, mutated, and written back as
digital_lexicon_v19.html. v13 .. v18 are untouched.

Note: downstream HTML files cache aggressively. If xlsx or law JSON
changed, wipe before rebuild:
    rm -f digital_lexicon_v1{6,7,8,9}.html
    python3 build_v19.py
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE      = Path(__file__).parent
HTML_V18  = HERE / "digital_lexicon_v18.html"
HTML_V19  = HERE / "digital_lexicon_v19.html"
XLSX      = HERE / "AI terminology and taxonomy-final.xlsx"


# --------------------------------------------------------------------------- #
# xlsx re-parse: Dimension parent lookup                                      #
# --------------------------------------------------------------------------- #

# Map legal sheet name -> concept id (mirrors build_v15.MATRIX_FAMILIES).
_SHEET_TO_CID: dict[str, str] = {
    "Provider_Developer":              "provider-developer",
    "Deployer_Supplier":               "deployer-supplier",
    " High-risk AI system":            "model-system",
    "GPAI_Frontier_Foundation model":  "model-system",
    "GPAI system_Generative AI":       "model-system",
    "Risk":                            "risk",
    "Substantial modification":        "modification",
    "Incident":                        "incident",
}


def _norm(s: str | None) -> str:
    if not s:
        return ""
    return str(s).replace("\xa0", " ").strip()


def _is_verbatim_like(s: str) -> bool:
    """Heuristic: col B carrying legal verbatim (not a sub-dim label).
    Sub-dim labels are short phrases; verbatim is long / multi-line."""
    if not s:
        return False
    if "\n" in s:
        return True
    if len(s) > 100:
        return True
    # A sub-dim label rarely contains sentence-ending punctuation.
    if s.count(".") >= 2 or s.count(";") >= 1:
        return True
    return False


# Curator-renamed analysis dims that don't appear in the legal sheets
# verbatim but always inherit a specific parent. Lowercased keys.
_HARDCODED_PARENTS: dict[str, str] = {
    # Scope family — analysis-only refinements of the Scope parent.
    "regulatory trigger":              "Scope",
    "temporal trigger":                "Scope",
    "compute threshold":               "Scope",
    "harm thresholds":                 "Scope",
    "continuous learning":             "Scope",
    "provider / developer information":"Scope",
    # Definition family — meta-dims about the Definition itself.
    "definition approach":             "Definition",
    "approach":                        "Definition",
    # Obligations family — all the compliance actions.
    "general information disclosure":  "Obligations",
    "specific information disclosure": "Obligations",
    "risk management - review":        "Obligations",
    "risk management - reporting":     "Obligations",
    "impact assessment - review":      "Obligations",
    "cooperation with authorities":    "Obligations",
    "communication to deployers":      "Obligations",
    "incident / risk reporting":       "Obligations",
    "documentation keeping":           "Obligations",
    "compliance check":                "Obligations",
    "whistleblower protections":       "Obligations",
    "obligations triggered":           "Obligations",
    # Reporting — Incident concept.
    "reporting timeline":              "Reporting",
    "reporting timelines":             "Reporting",
    "reporting mechanism":             "Reporting",
    "reporting obligations":           "Reporting",
    # Exemption / Exemptions normalization.
    "exemption":                       "Exemptions",
}


def _build_parent_lookup() -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    """Walk each legal sheet's col A / col B to build:
      - per-concept map: concept_id -> {sub_dim_lower: parent_dim_label}.
      - global fallback: {sub_dim_lower: parent_dim_label} (most-common
        parent across all sheets, ties broken by first seen).

    Returns (per_concept, global_fallback).

    Handles the two column layouts observed in the xlsx:
      * col A = parent, col B = sub-dim (most sheets).
      * col A blank, col B = dim, col C = verbatim ("Substantial
        modification"). No hierarchy emitted for those sheets.
    """
    if not XLSX.exists():
        print(f"  [v19] xlsx missing at {XLSX}; parent lookup will be empty")
        return {}, {}
    wb = load_workbook(XLSX, data_only=True)
    result: dict[str, dict[str, str]] = {}
    global_counts: dict[str, dict[str, int]] = {}
    for sheet_name, cid in _SHEET_TO_CID.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_a_has_text = any(
            _norm(ws.cell(r, 1).value)
            for r in range(1, min(ws.max_row, 20) + 1)
        )
        if not col_a_has_text:
            continue
        pmap = result.setdefault(cid, {})
        for r in range(1, ws.max_row + 1):
            a = _norm(ws.cell(r, 1).value)
            b = _norm(ws.cell(r, 2).value)
            if not a or not b:
                continue
            if _is_verbatim_like(b):
                continue
            key = b.lower()
            pmap.setdefault(key, a)
            # Track global vote.
            bucket = global_counts.setdefault(key, {})
            bucket[a] = bucket.get(a, 0) + 1
    # Compute global fallback by majority vote.
    global_fallback: dict[str, str] = {}
    for key, votes in global_counts.items():
        # Pick the parent with the highest count; ties broken by
        # alphabetical order for determinism.
        best = sorted(votes.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
        global_fallback[key] = best
    # Merge hardcoded overrides into the global fallback. Hardcoded
    # values win since they encode curator intent not in the xlsx.
    for k, v in _HARDCODED_PARENTS.items():
        global_fallback[k] = v
    return result, global_fallback


# --------------------------------------------------------------------------- #
# HTML mutations                                                              #
# --------------------------------------------------------------------------- #

def _rename_concepts_nav_to_analysis(html: str) -> str:
    """Rename the 'Concepts' nav button visible text and onclick target,
    keeping the id stable (so existing CSS selectors still work)."""
    old = '<button class="nav-link" id="nav-concepts" onclick="go(\'concepts\')">Concepts</button>'
    new = '<button class="nav-link" id="nav-concepts" onclick="go(\'analysis\')">Analysis</button>'
    if old not in html:
        raise SystemExit("v19: could not locate Concepts nav button")
    return html.replace(old, new, 1)


def _insert_verbatim_nav_button(html: str) -> str:
    """Insert a Verbatim nav button immediately after the Analysis one."""
    anchor = '<button class="nav-link" id="nav-concepts" onclick="go(\'analysis\')">Analysis</button>'
    new_btn = '<button class="nav-link" id="nav-verbatim" onclick="go(\'verbatim\')">Verbatim</button>'
    if anchor not in html:
        raise SystemExit("v19: could not find nav anchor for Verbatim button")
    return html.replace(anchor, anchor + "\n  " + new_btn, 1)


def _rewrite_home_cards(html: str) -> str:
    """Rename the first v17 home card 'Browse concepts' to 'Browse analysis'
    (and route it through go('analysis')) + insert a new 'Browse verbatim'
    card right after it. Keeps Methodology and Regulations cards intact."""

    # Rename first card.
    browse_concepts_old = (
        "        <a class=\"v17-home-card v17-home-card-primary\"\n"
        "           onclick=\"go('concepts');return false\"\n"
        "           href=\"#/concepts\"\n"
        "           aria-label=\"Browse concepts\">\n"
        "          <span class=\"v17-home-card-title\">Browse concepts <span aria-hidden=\"true\">\u2192</span></span>\n"
        "          <span class=\"v17-home-card-desc\">Explore 6 concept families across EU and US AI legislation with CEPS comparative analysis</span>\n"
        "        </a>"
    )
    browse_analysis_new = (
        "        <a class=\"v17-home-card v17-home-card-primary\"\n"
        "           onclick=\"go('analysis');return false\"\n"
        "           href=\"#/analysis\"\n"
        "           aria-label=\"Browse analysis\">\n"
        "          <span class=\"v17-home-card-title\">Browse analysis <span aria-hidden=\"true\">\u2192</span></span>\n"
        "          <span class=\"v17-home-card-desc\">CEPS comparative analysis across 6 concept families, with a summary per dimension and jurisdiction</span>\n"
        "        </a>\n"
        "        <a class=\"v17-home-card\"\n"
        "           onclick=\"go('verbatim');return false\"\n"
        "           href=\"#/verbatim\"\n"
        "           aria-label=\"Browse verbatim legal text\">\n"
        "          <span class=\"v17-home-card-title\">Browse verbatim <span aria-hidden=\"true\">\u2192</span></span>\n"
        "          <span class=\"v17-home-card-desc\">Source legal text side-by-side, jurisdiction by jurisdiction</span>\n"
        "          <span class=\"v17-home-card-cta\" aria-hidden=\"true\">\u2192</span>\n"
        "        </a>"
    )
    if browse_concepts_old not in html:
        raise SystemExit("v19: could not locate v17 Browse-concepts home card")
    html = html.replace(browse_concepts_old, browse_analysis_new, 1)

    # Update the grid to 4 columns via a targeted CSS override (appended
    # in _v19_overrides_css_js).
    return html


def _inject_dim_parent_blob(html: str, parent_lookup: dict, global_fallback: dict) -> str:
    """Inline the per-concept parent lookup (V19_DIM_PARENTS) + global
    fallback (V19_DIM_PARENTS_GLOBAL) as inline JSON blobs."""
    per_concept = json.dumps(parent_lookup, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    glob        = json.dumps(global_fallback, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    tag = (
        '<script type="application/json" id="__v19_dim_parents__">' + per_concept + "</script>\n"
        '<script type="application/json" id="__v19_dim_parents_global__">' + glob + "</script>\n"
        "<script>(function(){"
        "try{window.V19_DIM_PARENTS=JSON.parse(document.getElementById('__v19_dim_parents__').textContent);}"
        "catch(e){window.V19_DIM_PARENTS={};}"
        "try{window.V19_DIM_PARENTS_GLOBAL=JSON.parse(document.getElementById('__v19_dim_parents_global__').textContent);}"
        "catch(e){window.V19_DIM_PARENTS_GLOBAL={};}"
        "})();</script>"
    )
    return html.replace("</body>", tag + "\n</body>", 1)


def _v19_overrides_css_js() -> str:
    """Appended before </body>: CSS + JS that implements Analysis / Verbatim
    mode routing and the Verbatim-tab table renderer."""
    return r"""
<style>
/* v19: 4-column home-card grid */
.v17-home-card-grid{grid-template-columns:1.2fr 1fr 1fr 1fr!important;max-width:1080px!important}
@media (max-width:960px){.v17-home-card-grid{grid-template-columns:1fr 1fr!important}}
@media (max-width:560px){.v17-home-card-grid{grid-template-columns:1fr!important}}

/* Verbatim cell — line-clamped preview, full text reads in the law drawer */
.v19-verbatim-cell{
  display:-webkit-box;
  -webkit-line-clamp:6;
  -webkit-box-orient:vertical;
  overflow:hidden;
  text-overflow:ellipsis;
  cursor:pointer;
  font-family:var(--serif);
  font-size:13px;
  line-height:1.55;
  color:var(--ink);
  padding:6px 8px;
  border-radius:var(--r-sm,4px);
  transition:background-color .12s;
}
.v19-verbatim-cell:hover,
.v19-verbatim-cell:focus{ background:var(--accent-l); outline:none; }
.v19-analysis-only{ color:var(--ink-s); font-style:italic; }

/* Dim / Sub-dim hierarchy (Verbatim tab, always on) */
.analysis-table th.analysis-dim-cell.v19-dim-col{min-width:140px}
.analysis-table td.v19-dim-cell{
  vertical-align:top;
  font-family:var(--mono);
  font-size:11px;
  text-transform:uppercase;
  letter-spacing:.05em;
  background:var(--bg2);
  color:var(--ink);
  font-weight:600;
  padding-top:12px;
}
.analysis-table th.v19-subdim-cell{
  min-width:140px;
  background:var(--card);
  border-left:1px dashed var(--bd-s);
}
.analysis-table td.v19-subdim-cell{
  font-family:var(--serif);
  font-weight:500;
  font-size:12px;
  color:var(--ink-s);
  font-style:italic;
  border-left:1px dashed var(--bd-s);
  background:var(--card);
  vertical-align:top;
}
</style>
<script>
(function(){
  if (typeof state === 'undefined') return;
  state.mode = state.mode || 'analysis';

  // ── Router ───────────────────────────────────────────────────────────
  // Wrap go() for programmatic navigation (nav buttons, pill clicks, etc.)
  var _origGo = window.go;
  window.go = function(page, conceptId, subIdx, opts){
    if (page === 'analysis') { state.mode = 'analysis'; page = 'concepts'; }
    else if (page === 'verbatim') { state.mode = 'verbatim'; page = 'concepts'; }
    var ret = _origGo.call(this, page, conceptId, subIdx, opts);
    // Rewrite hash to reflect mode (only the concepts/concept routes).
    try {
      var h = location.hash;
      if (h === '#/concepts' || h.indexOf('#/concepts?') === 0) {
        history.replaceState(null, '', h.replace('#/concepts', '#/' + state.mode));
      } else if (h.indexOf('#/concept/') === 0) {
        history.replaceState(null, '', h.replace('#/concept/', '#/' + state.mode + '/concept/'));
      }
    } catch(e){}
    _updateNavMode();
    return ret;
  };

  // Replace the shell's hashchange listener. We can't rely on capture-phase
  // ordering for window events (it's registration-order in practice), so
  // we remove the shell's listener and install our own wrapper that runs
  // our mode-detection logic before delegating to the shell's handleHash.
  function _v19HashNormalize(h){
    // Detects a mode-prefixed hash, updates state.mode, and returns the
    // canonical hash the shell expects. Hash is mutated via replaceState
    // (silent, no hashchange re-fire).
    if (h.indexOf('#/verbatim/concept/') === 0) {
      state.mode = 'verbatim';
      var newH = '#/concept/' + h.slice('#/verbatim/concept/'.length);
      history.replaceState(null, '', newH);
      return newH;
    }
    if (h.indexOf('#/analysis/concept/') === 0) {
      state.mode = 'analysis';
      var newH = '#/concept/' + h.slice('#/analysis/concept/'.length);
      history.replaceState(null, '', newH);
      return newH;
    }
    if (h === '#/verbatim' || h.indexOf('#/verbatim?') === 0) {
      state.mode = 'verbatim';
      var newH = '#/concepts' + h.slice('#/verbatim'.length);
      history.replaceState(null, '', newH);
      return newH;
    }
    if (h === '#/analysis' || h.indexOf('#/analysis?') === 0) {
      state.mode = 'analysis';
      var newH = '#/concepts' + h.slice('#/analysis'.length);
      history.replaceState(null, '', newH);
      return newH;
    }
    if (h === '#/concepts' || h.indexOf('#/concepts?') === 0 ||
        h.indexOf('#/concept/') === 0) {
      // Legacy permalinks — force Analysis mode.
      state.mode = 'analysis';
    }
    return h;
  }

  function _v19PostRewrite(){
    // After the shell's handleHash runs and state.page is set, put the
    // mode prefix back onto the URL so users see a canonical sharable
    // link in the address bar.
    var h = location.hash;
    if (state.page === 'concepts') {
      if (h === '#/concepts' || h.indexOf('#/concepts?') === 0) {
        history.replaceState(null, '', h.replace('#/concepts', '#/' + state.mode));
      }
    } else if (state.page === 'concept') {
      if (h.indexOf('#/concept/') === 0) {
        history.replaceState(null, '', h.replace('#/concept/', '#/' + state.mode + '/concept/'));
      }
    }
  }

  // Remove shell's listener; install ours.
  if (typeof window.handleHash === 'function') {
    try { window.removeEventListener('hashchange', window.handleHash); } catch(e){}
    var _origHandleHash = window.handleHash;
    window.handleHash = function(){
      _v19HashNormalize(location.hash);
      _origHandleHash.call(this);
      _v19PostRewrite();
      _updateNavMode();
    };
    window.addEventListener('hashchange', window.handleHash);
  }

  // Initial-load hash (no hashchange fires for the initial URL).
  (function(){
    _v19HashNormalize(location.hash);
    // After the shell has done its init (it sets state.page based on
    // hash on first load via its own bootstrap), re-apply mode prefix.
    requestAnimationFrame(_v19PostRewrite);
  })();

  function _updateNavMode(){
    var cm = document.getElementById('nav-concepts');
    var vm = document.getElementById('nav-verbatim');
    if (!cm || !vm) return;
    if (state.page === 'concepts' || state.page === 'concept'){
      if (state.mode === 'verbatim'){
        cm.classList.remove('active'); cm.removeAttribute('aria-current');
        vm.classList.add('active'); vm.setAttribute('aria-current','page');
      } else {
        vm.classList.remove('active'); vm.removeAttribute('aria-current');
        cm.classList.add('active'); cm.setAttribute('aria-current','page');
      }
    } else {
      cm.classList.remove('active'); cm.removeAttribute('aria-current');
      vm.classList.remove('active'); vm.removeAttribute('aria-current');
    }
  }

  // ── Concepts landing h2 ──────────────────────────────────────────────
  var _origRB = window.renderBrowse;
  if (typeof _origRB === 'function') {
    window.renderBrowse = function(){
      _origRB.apply(this, arguments);
      var title = document.querySelector('#p-concepts .browse-title');
      if (title) title.textContent = (state.mode === 'verbatim') ? 'Verbatim' : 'Analysis';
      _updateNavMode();
    };
  }

  // ── Verbatim-tab table renderer ──────────────────────────────────────
  function _escH(s){
    return String(s==null?'':s).replace(/[&<>"']/g, function(c){
      return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];
    });
  }

  function _resolveRef(refStr){
    if (!refStr || typeof REF_MAP === 'undefined') return null;
    var s = String(refStr).trim();
    var hit = REF_MAP[s];
    if (hit && hit.law) return hit;
    var parts = s.split(/\s*;\s*/);
    for (var i = 0; i < parts.length; i++){
      hit = REF_MAP[parts[i].trim()];
      if (hit && hit.law) return hit;
    }
    return null;
  }

  function _openFullLawForCell(cell, dim, juris, sc){
    // v19 skip-the-middle-step: open the inline drawer then auto-click
    // the "Explore in full law" button, which expands `.v17-full-article`
    // inside the drawer — same DOM state as if the user manually
    // clicked through from the Analysis tab.
    if (typeof openDrawer !== 'function') return;
    openDrawer(dim.id, juris);
    // After the drawer renders (microtask is enough; the shell's
    // render is synchronous), click the Explore button if present.
    requestAnimationFrame(function(){
      var btn = document.querySelector('.drawer-actions .v17-explore-btn');
      if (btn) {
        btn.click();
      }
      // Else the ref didn't resolve to a known law — the inline drawer
      // is still shown, which is the best we can do.
    });
  }

  function _dimParent(c, dim){
    // Prefer an explicit parent_label emitted by build_v15.
    if (dim && dim.parent_label) return dim.parent_label;
    var key = (dim && (dim.label || '')).toLowerCase().trim();
    // Per-concept lookup first.
    var m = window.V19_DIM_PARENTS || {};
    var cm = m[c && c.id] || {};
    if (cm[key]) return cm[key];
    // Global fallback (union across concepts + curator-hardcoded).
    var g = window.V19_DIM_PARENTS_GLOBAL || {};
    if (g[key]) return g[key];
    // Last resort — no known parent, single-column effect.
    return (dim && (dim.label || '')) || '';
  }

  function _dimSub(dim){
    return (dim && (dim.sub_label || dim.label || '')) || '';
  }

  function _installVerbatimTable(){
    if (typeof window.renderAnalysisTable !== 'function' || window.__v19_rat_patched) return;
    window.__v19_rat_patched = true;
    var orig = window.renderAnalysisTable;
    window.renderAnalysisTable = function(){
      if (state.mode !== 'verbatim') {
        return orig.apply(this, arguments);
      }
      // Call original first so it renders CEPS notes + sets up DOM ids.
      orig.apply(this, arguments);
      try {
        var c  = getConcept(state.conceptId);
        var sc = c && c.sub_concepts && (c.sub_concepts[state.subConceptIdx] || c.sub_concepts[0]);
        if (!sc) return;
        var thead = document.getElementById('analysis-thead');
        var tbody = document.getElementById('analysis-tbody');
        if (!thead || !tbody) return;
        var juris = Object.keys(sc.jurisdictions);
        var JL = (typeof JURIS_LABELS !== 'undefined') ? JURIS_LABELS : {};

        // Header: Dimension / Sub-dimension / ...jurisdictions
        var head = '<tr>';
        head += '<th scope="col" class="analysis-dim-cell v19-dim-col">Dimension</th>';
        head += '<th scope="col" class="v19-subdim-cell">Sub-dimension</th>';
        juris.forEach(function(j){
          var jid = String(j).split('-')[0];
          var jd = sc.jurisdictions[j];
          head += '<th scope="col" class="th-' + jid + '">' + _escH(JL[j] || JL[jid] || j) +
                  '<span class="j-law">' + _escH((jd && jd.law) || '') + '</span></th>';
        });
        head += '</tr>';
        thead.innerHTML = head;

        // Build per-row (parent, sub) tuples in sc.dimensions order.
        var rowsPlan = (sc.dimensions || []).map(function(dim){
          return {dim: dim, parent: _dimParent(c, dim), sub: _dimSub(dim)};
        });

        // Compute rowspan groups for the Dimension column.
        var groups = [];
        rowsPlan.forEach(function(r, i){
          var last = groups[groups.length - 1];
          if (last && last.parent === r.parent) { last.end = i; }
          else { groups.push({parent: r.parent, start: i, end: i}); }
        });
        var rowspanMap = {};
        groups.forEach(function(g){
          rowspanMap[g.start] = g.end - g.start + 1;
        });

        var bodyParts = [];
        rowsPlan.forEach(function(r, i){
          var tr = '<tr>';
          if (i in rowspanMap) {
            tr += '<td class="v19-dim-cell" rowspan="' + rowspanMap[i] + '">' +
                  _escH(r.parent) + '</td>';
          }
          tr += '<td class="v19-subdim-cell">' + _escH(r.sub) + '</td>';
          juris.forEach(function(j){
            var cell = r.dim.cells[j];
            if (cell && cell.verbatim) {
              tr += '<td><span class="v19-verbatim-cell addressed" tabindex="0" role="button" ' +
                    'data-dimid="' + _escH(r.dim.id) + '" data-juris="' + _escH(j) + '" ' +
                    'aria-label="Open full law: ' + _escH(JL[j] || j) +
                    ' on ' + _escH(r.dim.label) + '">' +
                    _escH(cell.verbatim) + '</span></td>';
            } else if (cell && cell.analysis) {
              tr += '<td><span class="cell-null v19-analysis-only" ' +
                    'title="Analysis only \u2014 no verbatim quote in source law">\u2014</span></td>';
            } else {
              tr += '<td><span class="cell-null" aria-label="Not addressed">\u2014</span></td>';
            }
          });
          tr += '</tr>';
          bodyParts.push(tr);
        });
        tbody.innerHTML = bodyParts.join('');

        // Event delegation for verbatim cell clicks / keypress.
        tbody.addEventListener('click', _verbatimCellClick);
        tbody.addEventListener('keydown', function(e){
          if (e.key !== 'Enter' && e.key !== ' ') return;
          var t = e.target && e.target.closest && e.target.closest('.v19-verbatim-cell');
          if (!t) return;
          e.preventDefault();
          _verbatimCellClick({target: t});
        });
      } catch(e){ console.error('v19 verbatim table:', e); }
    };
  }

  function _verbatimCellClick(ev){
    var el = ev && ev.target && (ev.target.closest ? ev.target.closest('.v19-verbatim-cell') : null);
    if (!el) return;
    var dimId = el.getAttribute('data-dimid');
    var j     = el.getAttribute('data-juris');
    var c  = getConcept(state.conceptId);
    var sc = c && c.sub_concepts && (c.sub_concepts[state.subConceptIdx] || c.sub_concepts[0]);
    var dim = sc && (sc.dimensions || []).find(function(d){ return d.id === dimId; });
    if (!dim) return;
    _openFullLawForCell(dim.cells[j], dim, j, sc);
  }

  function _boot(){
    _installVerbatimTable();
    _updateNavMode();
    // If user arrived via a #/verbatim... hash, re-render.
    if (typeof handleHash === 'function' && location.hash.indexOf('#/verbatim') === 0) {
      try { handleHash(); } catch(e){}
    }
  }
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', _boot);
  else _boot();
})();
</script>
"""


# --------------------------------------------------------------------------- #
# Main                                                                        #
# --------------------------------------------------------------------------- #

def main() -> None:
    import build_v18

    print("== v19 build ==")
    if not HTML_V18.exists():
        build_v18.main()
    html = HTML_V18.read_text(encoding="utf-8")

    html = _rename_concepts_nav_to_analysis(html)
    print("  nav:                 Concepts -> Analysis")

    html = _insert_verbatim_nav_button(html)
    print("  nav:                 + Verbatim button")

    html = _rewrite_home_cards(html)
    print("  home cards:          Browse analysis + Browse verbatim")

    parent_lookup, global_fallback = _build_parent_lookup()
    total = sum(len(v) for v in parent_lookup.values())
    print(f"  dim parent lookup:   {len(parent_lookup)} concepts, {total} sub->parent (per-concept)")
    print(f"  dim parent fallback: {len(global_fallback)} sub->parent (global + hardcoded)")
    html = _inject_dim_parent_blob(html, parent_lookup, global_fallback)

    html = html.replace("</body>", _v19_overrides_css_js() + "\n</body>", 1)
    print("  overrides:           CSS + JS appended")

    # Version tag.
    html = html.replace(
        "<!-- DAL v18 (final-edits pass on v17) -->",
        "<!-- DAL v19 (Analysis + Verbatim tabs) -->",
        1,
    )
    html = html.replace(
        "<title>Digital AI Lexicon v18 \u2014 CEPS</title>",
        "<title>Digital AI Lexicon v19 \u2014 CEPS</title>",
        1,
    )

    HTML_V19.write_text(html, encoding="utf-8")
    print(f"\nWrote {HTML_V19}  ({len(html):,} bytes)")

    # Smoke test.
    test_path = HERE / "test_lexicon_v19.py"
    if test_path.exists():
        print("\n[v19 smoke test]")
        rc = subprocess.run([sys.executable, str(test_path)], cwd=HERE).returncode
        if rc != 0:
            raise SystemExit(f"v19 smoke test failed (rc={rc})")


if __name__ == "__main__":
    main()
