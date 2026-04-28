"""test_lexicon_v24.py — acceptance tests for v24.

Extends v23. Adds:
  E0. The "Provider of limited-risk AI systems" sub-concept exists on the
      `provider-developer` concept with non-empty dimensions (the structural
      gap v23's E1 carve-out flagged for v24).
  E1 (raised). xlsx→HTML coverage rises to ≥ 95% now that the structural
      gap is closed and v24's continuation-aware resync ran.

Run:
    python3 -m pytest test_lexicon_v24.py -q
or standalone:
    python3 test_lexicon_v24.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
HTML = HERE / "digital_lexicon_v24.html"
LAWS = HERE / "laws"
# v24 builds from the Cross-checked xlsx — the test must use the same source.
XLSX = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis.xlsx"
)
XLSX_FALLBACK = HERE / "AI terminology and taxonomy-final.xlsx"


def _html() -> str:
    return HTML.read_text(encoding="utf-8")


def _get_const(html: str, name: str):
    for prefix in (rf"const\s+{name}\s*=\s*", rf"window\.{name}\s*=\s*"):
        m = re.search(prefix, html)
        if not m:
            continue
        s = m.end()
        depth = 0
        in_str = False
        esc = False
        quote = ""
        for i in range(s, len(html)):
            ch = html[i]
            if in_str:
                if esc:
                    esc = False
                elif ch == "\\":
                    esc = True
                elif ch == quote:
                    in_str = False
                continue
            if ch in '"\'':
                in_str = True
                quote = ch
            elif ch in "{[":
                depth += 1
            elif ch in "}]":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(html[s:i + 1])
                    except Exception:
                        return None
    return None


def _xlsx_path() -> Path:
    return XLSX if XLSX.exists() else XLSX_FALLBACK


def _walk_strings(o, out):
    if isinstance(o, str):
        out.append(o)
    elif isinstance(o, dict):
        for v in o.values():
            _walk_strings(v, out)
    elif isinstance(o, list):
        for v in o:
            _walk_strings(v, out)


def _norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip().lower()


# --------------------------------------------------------------------------- #
# E0.  Structural — limited-risk Provider sub-concept.                         #
# --------------------------------------------------------------------------- #

def test_E0_limited_risk_subconcept_present():
    """The limited-risk Provider slice (xlsx Provider_Developer_Analysis
    rows 1–20) is present as a sub-concept of `provider-developer`, has
    non-empty dimensions, and sits BEFORE the high-risk sub-concept."""
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS")
    assert CONCEPTS, "CONCEPTS missing"
    pd = next((c for c in CONCEPTS if c.get("id") == "provider-developer"),
              None)
    assert pd, "provider-developer concept missing"
    subs = pd.get("sub_concepts") or []
    assert subs, "provider-developer has no sub_concepts"

    # Find the limited-risk slot (id == 'provider', title contains 'limited').
    lim_idx = next((i for i, s in enumerate(subs)
                    if s.get("id") == "provider"), -1)
    assert lim_idx >= 0, "limited-risk Provider sub-concept (id=provider) missing"

    lim = subs[lim_idx]
    assert "limited" in (lim.get("title") or "").lower(), (
        f"unexpected title for limited-risk slot: {lim.get('title')!r}"
    )
    assert lim.get("dimensions"), "limited-risk slot has no dimensions"
    assert len(lim["dimensions"]) >= 10, (
        f"limited-risk slot has only {len(lim['dimensions'])} dims (expected ≥10)"
    )

    # Order: must come before provider-of-high-risk-ai-systems (high-risk).
    hr_idx = next((i for i, s in enumerate(subs)
                   if s.get("id") == "provider-of-high-risk-ai-systems"), -1)
    assert hr_idx >= 0, "provider-of-high-risk sub-concept missing"
    assert lim_idx < hr_idx, (
        f"limited-risk (idx={lim_idx}) must precede high-risk (idx={hr_idx})"
    )


def test_E0_limited_risk_dimensions_populated():
    """Every dim of the limited-risk slot has at least one populated cell
    (non-empty `analysis`), and the slot covers all three jurisdictions
    (eu, co, tx) per the xlsx."""
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    pd = next((c for c in CONCEPTS if c.get("id") == "provider-developer"),
              None)
    assert pd
    lim = next((s for s in (pd.get("sub_concepts") or [])
                if s.get("id") == "provider"), None)
    assert lim, "limited-risk Provider sub-concept missing"

    juris = lim.get("jurisdictions") or {}
    for j in ("eu", "co", "tx"):
        assert j in juris, f"limited-risk missing jurisdiction {j!r}"

    empty_dims = []
    for d in lim.get("dimensions") or []:
        cells = d.get("cells") or {}
        any_populated = any(
            ((c or {}).get("analysis") or "").strip() not in ("", "-", "–", "—")
            for c in cells.values()
        )
        if not any_populated:
            empty_dims.append(d.get("label"))
    # All-blank dim is allowed only for "Incident / risk reporting"
    # which is genuinely blank in the xlsx.
    blank_ok = {"incident / risk reporting"}
    real_empties = [d for d in empty_dims if (d or "").lower() not in blank_ok]
    assert not real_empties, (
        f"Limited-risk dims with no populated cells: {real_empties}"
    )


def test_E0_limited_risk_in_cluster_summary():
    """The matrix cluster_summary for `provider-developer` references the
    limited-risk Provider entry (term cells link back to sub_id='provider')."""
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    pd = next((c for c in CONCEPTS if c.get("id") == "provider-developer"),
              None)
    assert pd, "provider-developer missing"
    cs = pd.get("cluster_summary") or {}
    rows = cs.get("rows") or []
    has_lim_row = any(
        any(
            (var or {}).get("sub_id") == "provider"
            for cell in (r.get("cells") or {}).values()
            for var in (cell.get("variants") or [])
        )
        or r.get("sub_id") == "provider"
        for r in rows
    )
    assert has_lim_row, "cluster_summary has no row pointing at provider (limited-risk)"


# --------------------------------------------------------------------------- #
# E1.  xlsx→HTML coverage.                                                     #
# --------------------------------------------------------------------------- #

def test_E1_xlsx_to_html_coverage_high():
    """Now that the limited-risk slot is filled and continuations resync,
    coverage of dim-cell xlsx text in CONCEPTS must be ≥ 95%."""
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS")
    assert CONCEPTS, "CONCEPTS not present"
    corpus = []
    _walk_strings(CONCEPTS, corpus)
    big = _norm(" || ".join(corpus))

    wb = openpyxl.load_workbook(_xlsx_path(), data_only=True)
    misses = []
    checked = 0
    for sheet in [s for s in wb.sheetnames
                  if s.endswith("_ANALYSIS") or s.endswith("_ANALY") or
                  s.endswith("_Analys") or s.endswith("_Analysis")]:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not v:
                    continue
                if cell.column_letter in ("A", "E", "F", "G", "H"):
                    continue
                sv = str(v).strip()
                if len(sv) < 30 or len(sv) > 600:
                    continue
                if sv.lower().startswith(("interpretative notes",
                                          "this section")):
                    continue
                probe = _norm(sv)[:60]
                if not probe:
                    continue
                checked += 1
                if probe not in big:
                    misses.append((sheet, cell.coordinate, probe[:80]))
    rate = 1 - (len(misses) / checked) if checked else 1.0
    # v25 raised this threshold to ≥97% after closing the GL/truncation gaps.
    assert rate >= 0.97, (
        f"Only {rate:.1%} of dim cells found in HTML "
        f"({len(misses)}/{checked} missing); first misses: {misses[:5]}"
    )


# --------------------------------------------------------------------------- #
# E2.  Reg #12 still wired (carry-over from v23).                              #
# --------------------------------------------------------------------------- #

REG12 = "eu-guidelines-gpai-scope"


def test_E2_reg12_still_wired():
    html = _html()
    blob_re = re.compile(
        rf'<script[^>]*id="law-blob-{re.escape(REG12)}"', re.S
    )
    assert blob_re.search(html), "reg #12 blob lost in v24"
    laws_nav = _get_const(html, "LAWS")
    assert laws_nav
    found = False
    for region in laws_nav:
        for law in region.get("laws") or []:
            if law.get("law_id") == REG12:
                found = True
                break
    assert found, "reg #12 not on Law Sources page anymore"


# --------------------------------------------------------------------------- #
# E5.  References preserved on cells (carry-over from v23).                    #
# --------------------------------------------------------------------------- #

def test_E5_cells_carry_reference_field():
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    cells_with_ref = 0
    cells_total = 0
    for c in CONCEPTS:
        for sc in c.get("sub_concepts") or []:
            for d in sc.get("dimensions") or []:
                for jid, cell in (d.get("cells") or {}).items():
                    cells_total += 1
                    if (cell or {}).get("reference"):
                        cells_with_ref += 1
    assert cells_total > 0
    rate = cells_with_ref / cells_total
    assert rate >= 0.40, (
        f"Only {rate:.1%} of cells carry reference"
    )


# Standalone runner --------------------------------------------------------- #

if __name__ == "__main__":
    import inspect
    g = globals()
    tests = [(n, fn) for n, fn in g.items()
             if n.startswith("test_") and inspect.isfunction(fn)]
    failed = []
    for name, fn in tests:
        try:
            fn()
            print(f"  PASS  {name}")
        except AssertionError as e:
            print(f"  FAIL  {name}: {e}")
            failed.append(name)
        except Exception as e:
            print(f"  ERROR {name}: {type(e).__name__}: {e}")
            failed.append(name)
    print(f"\n{len(tests) - len(failed)}/{len(tests)} passed")
    sys.exit(1 if failed else 0)
