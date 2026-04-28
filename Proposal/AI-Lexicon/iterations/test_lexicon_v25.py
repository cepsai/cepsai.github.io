"""test_lexicon_v25.py — acceptance tests for v25.

v25 closes four gaps on top of v24:

  E0. `(GL, ...)` notation in references parses to regulation #12
      (`eu-guidelines-gpai-scope`) via build_v13.parse_reference.
  E1. At least 4 dim cells carry a reference that resolves to reg #12
      after parse_reference is applied.
  E2. The high-risk EU Definition cell starts with the full xlsx
      phrase "AI system placed on the market or put into service that
      1) acts as a safety component..." (no truncation).
  E3. Each of the 5 specific cells re-synced from the new xlsx appears
      verbatim in CONCEPTS.
  E4. Coverage rate is ≥ 97% (raises the v23/v24 E1 threshold).

Run:
    python3 -m pytest test_lexicon_v25.py -q
or standalone:
    python3 test_lexicon_v25.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
HTML = HERE / "digital_lexicon_v25.html"
LAWS = HERE / "laws"
XLSX = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis.xlsx"
)
XLSX_FALLBACK = HERE / "AI terminology and taxonomy-final.xlsx"

REG12 = "eu-guidelines-gpai-scope"


def _html() -> str:
    return HTML.read_text(encoding="utf-8")


def _get_const(html: str, name: str):
    for prefix in (rf"const\s+{name}\s*=\s*", rf"window\.{name}\s*=\s*"):
        m = re.search(prefix, html)
        if not m:
            continue
        s = m.end()
        depth = 0
        in_str = False
        esc = False
        quote = ""
        for i in range(s, len(html)):
            ch = html[i]
            if in_str:
                if esc:
                    esc = False
                elif ch == "\\":
                    esc = True
                elif ch == quote:
                    in_str = False
                continue
            if ch in '"\'':
                in_str = True
                quote = ch
            elif ch in "{[":
                depth += 1
            elif ch in "}]":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(html[s:i + 1])
                    except Exception:
                        return None
    return None


def _xlsx_path() -> Path:
    return XLSX if XLSX.exists() else XLSX_FALLBACK


def _walk_strings(o, out):
    if isinstance(o, str):
        out.append(o)
    elif isinstance(o, dict):
        for v in o.values():
            _walk_strings(v, out)
    elif isinstance(o, list):
        for v in o:
            _walk_strings(v, out)


def _norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip().lower()


# --------------------------------------------------------------------------- #
# E0.  parse_reference routes (GL, ...) to reg #12.                            #
# --------------------------------------------------------------------------- #

def test_E0_gl_notation_routes_to_reg12():
    sys.path.insert(0, str(HERE))
    import build_v13
    parse_reference = build_v13.parse_reference

    cases = [
        "(GL, (17))",
        "(GL, 3.2)",
        "(GL, 3.4)",
        "GL, 17",
        "(GL, (13))",
    ]
    for ref in cases:
        law, kind, anchor = parse_reference(ref)
        assert law == REG12, (
            f"parse_reference({ref!r}) returned law={law!r}, expected {REG12!r}"
        )
        assert kind == "section", (
            f"parse_reference({ref!r}) returned kind={kind!r}, expected 'section'"
        )
        assert anchor, f"anchor missing for {ref!r}"

    # Specific anchor mappings.
    assert parse_reference("(GL, (17))")[2] == "2", \
        "(GL, (17)) should resolve to section '2' (paragraph 17 lives in section 2)"
    assert parse_reference("(GL, 3.2)")[2] == "3", \
        "(GL, 3.2) should resolve to parent section '3'"


# --------------------------------------------------------------------------- #
# E1.  At least 4 dim cells carry a (GL, ...) reference resolving to reg #12. #
# --------------------------------------------------------------------------- #

def test_E1_gl_anchored_cells_link_to_reg12():
    sys.path.insert(0, str(HERE))
    import build_v13
    parse_reference = build_v13.parse_reference

    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []

    hits = 0
    examples = []
    for c in CONCEPTS:
        for sc in c.get("sub_concepts") or []:
            for d in sc.get("dimensions") or []:
                for jid, cell in (d.get("cells") or {}).items():
                    if not isinstance(cell, dict):
                        continue
                    ref = cell.get("reference") or ""
                    if not ref:
                        continue
                    pieces = [ref] + re.split(r"\s*;\s*", ref)
                    for piece in pieces:
                        law, _, _ = parse_reference(piece)
                        if law == REG12:
                            hits += 1
                            if len(examples) < 5:
                                examples.append(
                                    (c.get("id"), sc.get("id"),
                                     d.get("label"), jid, piece)
                                )
                            break
    assert hits >= 4, (
        f"Only {hits} dim cell(s) reference reg #12 via (GL, ...) — "
        f"expected ≥4. Examples: {examples}"
    )


# --------------------------------------------------------------------------- #
# E2.  No truncation — high-risk EU Definition has full xlsx phrase.          #
# --------------------------------------------------------------------------- #

def test_E2_no_truncation_high_risk_eu_definition():
    """The EU column of the high-risk Definition dim must start with the
    canonical xlsx phrase "AI system placed on the market or put into
    service that 1) acts as a safety component" — singular, with the
    "1) acts as a safety component" continuation, no truncation."""
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    c = next((x for x in CONCEPTS if x.get("id") == "model-system"), None)
    assert c, "model-system concept missing"
    sc = next((s for s in c.get("sub_concepts") or []
               if s.get("id") == "high-risk-ai-system"), None)
    assert sc, "high-risk-ai-system sub-concept missing"
    d = next((x for x in sc.get("dimensions") or []
              if (x.get("label") or "").lower() == "definition"), None)
    assert d, "Definition dim missing on high-risk"
    cell = (d.get("cells") or {}).get("eu") or {}
    text = cell.get("analysis") or ""
    expected_prefix = (
        "AI system placed on the market or put into service that 1) acts "
        "as a safety component"
    )
    assert text.startswith(expected_prefix), (
        f"high-risk EU Definition is truncated/abbreviated. "
        f"Got first 120 chars: {text[:120]!r}"
    )


def test_E2_no_truncation_gpai_sys_ca_regtrigger():
    """The CA column of GPAI-System Regulatory trigger must include the
    full "System and usership-based: targets 'covered providers'" text
    (probe miss GPAI-Sys C5 reported in v24 audit)."""
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    c = next((x for x in CONCEPTS if x.get("id") == "model-system"), None)
    assert c
    sc = next((s for s in c.get("sub_concepts") or []
               if s.get("id") == "general-purpose-ai-system"), None)
    assert sc
    d = next((x for x in sc.get("dimensions") or []
              if (x.get("label") or "").lower() == "regulatory trigger"), None)
    assert d, "Regulatory trigger dim missing on GPAI system"
    cell = (d.get("cells") or {}).get("ca") or {}
    text = cell.get("analysis") or ""
    assert "covered providers" in text.lower(), (
        f"GPAI-Sys CA Regulatory trigger truncated. Got: {text!r}"
    )


# --------------------------------------------------------------------------- #
# E3.  The five specific cells appear verbatim in CONCEPTS.                    #
# --------------------------------------------------------------------------- #

SPECIFIC_FRAGMENTS = [
    # Provider_Developer_Analysis B97 (Rebuttal — GPAISR EU)
    "Possibility to rebut GPAISR classification for not presenting "
    "systemic risks despite high-impact capabilities (Article 52)",
    # Modification_ANALYSIS B5 (Definition GPAI EU)
    "In the case of GPAI models, modification is legally relevant if it "
    "leads to a significant change in the model",
    # Modification_ANALYSIS B12 (Obligations triggered GPAI EU)
    "GPAI models: update technical documentation (Article 53).",
    # Modification_ANALYSIS B13 (Obligations triggered GPAISR EU)
    "In addition, providers of GPAISR models must notify Commission when "
    "compute threshold achieved (Article 52) and update Safety and "
    "Security Framework (Articles 53, 55)",
    # GPAI_Frontier_Foundation_Analys B8 (Compute threshold GPAI EU)
    "General-purpose AI model: 10^23 FLOPs plus capability to generate "
    "language, including video or text (GL, (17))",
]


def test_E3_specific_cells_present():
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    corpus = []
    _walk_strings(CONCEPTS, corpus)
    big = _norm(" || ".join(corpus))

    failures = []
    for fragment in SPECIFIC_FRAGMENTS:
        if _norm(fragment)[:80] not in big:
            failures.append(fragment[:80])
    assert not failures, (
        f"{len(failures)}/{len(SPECIFIC_FRAGMENTS)} expected fragments "
        f"absent from CONCEPTS:\n  " + "\n  ".join(failures)
    )


# --------------------------------------------------------------------------- #
# E4.  Coverage threshold raised to ≥97%.                                      #
# --------------------------------------------------------------------------- #

def test_E4_coverage_at_least_97_pct():
    html = _html()
    CONCEPTS = _get_const(html, "CONCEPTS") or []
    corpus = []
    _walk_strings(CONCEPTS, corpus)
    big = _norm(" || ".join(corpus))

    wb = openpyxl.load_workbook(_xlsx_path(), data_only=True)
    misses = []
    checked = 0
    for sheet in [s for s in wb.sheetnames
                  if s.endswith("_ANALYSIS") or s.endswith("_ANALY") or
                  s.endswith("_Analys") or s.endswith("_Analysis")]:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not v:
                    continue
                if cell.column_letter in ("A", "E", "F", "G", "H"):
                    continue
                sv = str(v).strip()
                if len(sv) < 30 or len(sv) > 600:
                    continue
                if sv.lower().startswith(("interpretative notes",
                                          "this section")):
                    continue
                probe = _norm(sv)[:60]
                if not probe:
                    continue
                checked += 1
                if probe not in big:
                    misses.append((sheet, cell.coordinate, probe[:80]))
    rate = 1 - (len(misses) / checked) if checked else 1.0
    assert rate >= 0.97, (
        f"Coverage only {rate:.1%} ({len(misses)}/{checked} missing); "
        f"first misses: {misses[:5]}"
    )


# --------------------------------------------------------------------------- #
# E5.  Carry-over: reg #12 still wired (sanity).                               #
# --------------------------------------------------------------------------- #

def test_E5_reg12_still_wired():
    html = _html()
    blob_re = re.compile(
        rf'<script[^>]*id="law-blob-{re.escape(REG12)}"', re.S
    )
    assert blob_re.search(html), "reg #12 blob lost in v25"
    laws_nav = _get_const(html, "LAWS")
    assert laws_nav
    found = False
    for region in laws_nav:
        for law in region.get("laws") or []:
            if law.get("law_id") == REG12:
                found = True
                break
    assert found, "reg #12 not on Law Sources page anymore"


# Standalone runner --------------------------------------------------------- #

if __name__ == "__main__":
    import inspect
    g = globals()
    tests = [(n, fn) for n, fn in g.items()
             if n.startswith("test_") and inspect.isfunction(fn)]
    failed = []
    for name, fn in tests:
        try:
            fn()
            print(f"  PASS  {name}")
        except AssertionError as e:
            print(f"  FAIL  {name}: {e}")
            failed.append(name)
        except Exception as e:
            print(f"  ERROR {name}: {type(e).__name__}: {e}")
            failed.append(name)
    print(f"\n{len(tests) - len(failed)}/{len(tests)} passed")
    sys.exit(1 if failed else 0)
