"""build_v18.py — Digital AI Lexicon v18

Final-edits pass on top of v17:

    1. Home-page text is sourced from the xlsx "About the Digital AI Lexicon" sheet.
    2. "Laws" and "Source laws" are renamed to "Regulations".
    3. The "Explore in full law" panel highlights the sub-paragraph that the
       cited cell points at (e.g. "Article 3 (3)" → paragraph (3) shown bold).
    4. Comparison table's Dimension column splits into two sub-columns
       (dim + sub-dim) for concepts whose legal sheet uses that hierarchy
       (e.g. Incident → Scope / High-risk AI systems · Scope / GPAI models).

v16 and v17 are untouched.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
HTML_V17 = HERE / "digital_lexicon_v17.html"
HTML_V18 = HERE / "digital_lexicon_v18.html"
XLSX = HERE / "AI terminology and taxonomy-final.xlsx"


# --------------------------------------------------------------------------- #
# 1. About text from the xlsx.                                                #
# --------------------------------------------------------------------------- #

def _read_about_blocks() -> list[str]:
    """Return the About sheet's A1 content split into paragraphs."""
    wb = load_workbook(XLSX, data_only=True)
    raw = (wb["About the Digital AI Lexicon"].cell(1, 1).value or "").strip()
    # Split on blank lines; keep each block as-is (newlines inside become <br>).
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", raw) if p.strip()]
    return paragraphs


def _format_home_prose(paragraphs: list[str]) -> tuple[str, str, str]:
    """Given the About paragraphs, return (hero_title, hero_tagline, hero_body_html).

    The first paragraph typically has "About the Digital AI Lexicon tool\\n\\n"
    followed by the lead sentence. We take the first-line as the tagline.
    """
    if not paragraphs:
        return "Digital AI Lexicon", "", ""
    # Take the first paragraph's non-title body as the tagline, rest as body.
    body_html_parts: list[str] = []
    tagline = ""
    for idx, p in enumerate(paragraphs):
        # Drop leading "About the Digital AI Lexicon tool" title.
        lines = [ln for ln in p.split("\n") if ln.strip()]
        if idx == 0 and lines and "about the digital ai lexicon tool" in lines[0].lower():
            lines = lines[1:]
        if not lines:
            continue
        para_text = " ".join(lines)
        # First real sentence = tagline; rest go into the body.
        if not tagline:
            tagline = para_text
            continue
        # Preserve dash-bullets within a paragraph.
        if "\n- " in p or p.startswith("- "):
            # Re-split the original paragraph so bullets keep their structure.
            bullet_lines: list[str] = []
            for ln in p.split("\n"):
                ln = ln.strip()
                if not ln:
                    continue
                if ln.startswith("- "):
                    bullet_lines.append(
                        f'<li>{_escape_html(ln[2:].strip())}</li>'
                    )
                else:
                    bullet_lines.append(
                        f'<p>{_escape_html(ln)}</p>'
                    )
            grouped = "\n".join(bullet_lines)
            # Wrap consecutive <li>s in a <ul>.
            grouped = re.sub(
                r"(<li>.*?</li>\s*)+",
                lambda m: f'<ul class="about-bullets">{m.group(0)}</ul>',
                grouped,
                flags=re.S,
            )
            body_html_parts.append(grouped)
        else:
            body_html_parts.append(f"<p>{_escape_html(para_text)}</p>")
    return "Digital AI Lexicon", tagline, "\n".join(body_html_parts)


def _escape_html(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# --------------------------------------------------------------------------- #
# 2. Main.                                                                    #
# --------------------------------------------------------------------------- #

def main() -> None:
    import build_v17

    print("== v18 build ==")
    if not HTML_V17.exists():
        build_v17.main()
    html = HTML_V17.read_text(encoding="utf-8")

    # ----- 1. Replace home landing copy with About-sheet prose --------------
    paragraphs = _read_about_blocks()
    title, tagline, body_html = _format_home_prose(paragraphs)
    new_landing = f"""    <h1 tabindex="-1">{_escape_html(title)}</h1>
    <p class="landing-tagline">{_escape_html(tagline)}</p>
    <div class="landing-sub about-body">{body_html}</div>"""
    old_landing_re = re.compile(
        r'\s*<h1 tabindex="-1">Digital AI Lexicon</h1>\s*'
        r'<p class="landing-tagline">[^<]*</p>\s*'
        r'<p class="landing-sub">[^<]*</p>',
        re.S,
    )
    m = old_landing_re.search(html)
    if not m:
        raise SystemExit("v18: could not locate v17 landing block to replace")
    html = html[: m.start()] + "\n    " + new_landing + html[m.end() :]
    print("  home prose:          replaced with About sheet content")

    # ----- 2. Rename Laws → Regulations -------------------------------------
    # Nav link text + Laws page heading + landing card title/desc + related.
    rename_pairs = [
        # nav link (visible text only — keep onclick="go('laws')" and page ids)
        ('onclick="go(\'laws\')">Laws</button>',
         'onclick="go(\'laws\')">Regulations</button>'),
        # home-card label
        ('<span class="v17-home-card-title">Source laws</span>',
         '<span class="v17-home-card-title">Regulations</span>'),
        ('<span class="v17-home-card-desc">11 regulatory frameworks</span>',
         '<span class="v17-home-card-desc">11 regulatory frameworks</span>'),
        # Laws page heading ("Primary Sources" → "Regulations")
        ("<h1 tabindex=\"-1\">Primary Sources</h1>",
         "<h1 tabindex=\"-1\">Regulations</h1>"),
    ]
    for old, new in rename_pairs:
        if old not in html:
            continue
        html = html.replace(old, new)
    print("  renamed:             Laws → Regulations")

    # ----- 3. Highlight the cited sub-paragraph in full-law --------------
    html = html.replace("</body>", _highlight_cited_js() + "\n</body>", 1)
    print("  added:               highlight cited sub-paragraph JS")

    # ----- 4. Split Dimension column into 2 sub-columns ------------------
    html = html.replace("</body>", _two_col_dim_js() + "\n</body>", 1)
    print("  added:               2-col dimension hierarchy override")

    # Tag version.
    html = html.replace(
        "<!-- DAL v17 (v16 data + v18.5 reference shell) -->",
        "<!-- DAL v18 (final-edits pass on v17) -->",
        1,
    )
    html = html.replace(
        "<title>Digital AI Lexicon v17 — CEPS</title>",
        "<title>Digital AI Lexicon v18 — CEPS</title>",
        1,
    )

    HTML_V18.write_text(html, encoding="utf-8")
    final = len(html)
    print(f"\nWrote {HTML_V18}  ({final:,} bytes)")

    import subprocess as _sp, sys as _sys
    test_path = HERE / "test_lexicon_v18.py"
    if test_path.exists():
        print("\n[v18 smoke test]")
        rc = _sp.run([_sys.executable, str(test_path)], cwd=HERE).returncode
        if rc != 0:
            raise SystemExit(f"v18 smoke test failed (rc={rc})")


# --------------------------------------------------------------------------- #
# JS helpers appended to the built HTML.                                      #
# --------------------------------------------------------------------------- #

def _highlight_cited_js() -> str:
    return r"""
<style>
.v18-cited{background:var(--accent-l);padding:3px 4px;border-radius:3px;font-weight:600;color:var(--accent-d)}
.v18-cited-para{background:linear-gradient(var(--accent-l), var(--accent-l));padding:2px 4px;border-radius:3px}
</style>
<script>
(function(){
  // Parse a reference string like "EU AI Act, Article 3 (3)" into
  // {article:'3', paragraph:'3'} or "California SB53, 22757.12. (c)" →
  // {section:'22757.12', paragraph:'c'}. The paragraph hint is what we
  // highlight inside the full article text.
  function _parseRef(s){
    if (!s) return null;
    s = String(s).trim();
    // Look for the parenthesised sub-part.
    var paraMatch = s.match(/\(([^)]+)\)\s*(?:[;.,]|$)/);
    var para = paraMatch ? paraMatch[1].trim().split(/[\s,]+/)[0] : '';
    return {paragraph: para};
  }

  // Highlight the matching sub-paragraph in a rendered full-article panel.
  function _highlightIn(el, paraHint){
    if (!el || !paraHint) return;
    // Strategy: wrap the first occurrence of "(<paraHint>)" or "<paraHint>."
    // at the start of a line/clause. Use a TreeWalker to avoid matching HTML.
    var walker = document.createTreeWalker(el, NodeFilter.SHOW_TEXT, null);
    var patterns = [
      // "(1)" or "(a)" — parenthesised paragraph markers
      new RegExp('\\((' + _escape(paraHint) + ')\\)', 'i'),
      // Section identifiers like "22757.12(c)" → "(c)"
      new RegExp('\\b(' + _escape(paraHint) + ')\\.', 'i'),
    ];
    var node;
    var done = false;
    while ((node = walker.nextNode()) && !done){
      var text = node.textContent;
      for (var i = 0; i < patterns.length && !done; i++){
        var m = text.match(patterns[i]);
        if (m && m.index >= 0){
          var before = text.slice(0, m.index);
          var hit = text.slice(m.index, m.index + m[0].length);
          var after = text.slice(m.index + m[0].length);
          // Extend the highlight to the end of the clause or paragraph.
          var endIdx = after.search(/(?:\n\s*\n|\.(?=\s)|;\s)/);
          var rest = endIdx >= 0 ? after.slice(0, endIdx + 1) : after;
          var tail = endIdx >= 0 ? after.slice(endIdx + 1) : '';
          var frag = document.createDocumentFragment();
          if (before) frag.appendChild(document.createTextNode(before));
          var mark = document.createElement('mark');
          mark.className = 'v18-cited-para';
          mark.textContent = hit + rest;
          frag.appendChild(mark);
          if (tail) frag.appendChild(document.createTextNode(tail));
          node.parentNode.replaceChild(frag, node);
          done = true;
        }
      }
    }
  }
  function _escape(s){return String(s).replace(/[.*+?^${}()|[\]\\]/g, '\\$&');}

  // Delegate on document: whenever the user clicks an Explore-in-full-law
  // button, wait for the .v17-full-article to appear, then highlight
  // the cited sub-paragraph inside it.
  function _handleExploreClick(){
    setTimeout(function(){
      var article = document.querySelector('.v17-explore-wrap .v17-full-article');
      if (!article) return;
      var refEl = document.getElementById('drawer-ref');
      var parsed = _parseRef(refEl ? refEl.textContent : '');
      if (parsed && parsed.paragraph) _highlightIn(article, parsed.paragraph);
    }, 60);
  }
  document.addEventListener('click', function(e){
    var btn = e.target && e.target.closest && e.target.closest('.v17-explore-btn');
    if (btn) _handleExploreClick();
  }, true);
})();
</script>
"""


def _two_col_dim_js() -> str:
    r"""Override renderAnalysisTable so dimensions whose cells' verbatim
    text carries a sub-category hint (e.g. "[High-risk AI systems] Providers
    shall report ...") render as two rows under the same Dimension header —
    one for each sub-category. Concepts whose dims don't split keep the
    single-column layout.

    The heuristic: for each dim row, look at each jurisdiction's verbatim.
    If it starts with a known prefix like "[High-risk AI systems]" or
    "High-risk AI systems:" or contains multiple semicolon-separated
    blocks prefixed by such, we split the dim into multiple rows. For now
    we apply the simpler rule — if a dim's rowLabel alone already carries
    a subLabel (Build v15 emits "rowLabel — subLabel" as dim.label when
    they differ), split on the " — " separator.
    """
    return r"""
<style>
.v18-dim-group{background:var(--bg2);font-weight:600;color:var(--ink);font-family:var(--mono);font-size:11px;text-transform:uppercase;letter-spacing:.05em}
.v18-subdim{font-family:var(--serif);font-weight:400;font-size:12px;color:var(--ink-s);text-transform:none;letter-spacing:0;padding-left:12px;font-style:italic}
.analysis-table th.analysis-dim-cell.v18-dim-col{min-width:160px}
.analysis-table th.analysis-subdim-cell{min-width:140px;background:var(--card);border-left:1px dashed var(--bd-s)}
.analysis-table td.analysis-subdim-cell{font-family:var(--serif);font-weight:500;font-size:12px;color:var(--ink-s);font-style:italic;border-left:1px dashed var(--bd-s);background:var(--card)}
</style>
<script>
(function(){
  function _hasSubDim(sc){
    if (!sc || !sc.dimensions) return false;
    return sc.dimensions.some(function(d){
      return (d.label || '').indexOf(' \u2014 ') !== -1
          || (d.sub_label || '').trim() !== '';
    });
  }
  function _split(d){
    // Returns {label, sub} honouring either an explicit d.sub_label field
    // or a label of form "X \u2014 Y".
    if (d.sub_label) return {label: d.label, sub: d.sub_label};
    var s = String(d.label || '');
    var idx = s.indexOf(' \u2014 ');
    if (idx === -1) return {label: s, sub: ''};
    return {label: s.slice(0, idx), sub: s.slice(idx + 3)};
  }
  function _escH(s){return String(s||'').replace(/[&<>"']/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];});}

  function _install(){
    if (typeof window.renderAnalysisTable !== 'function' || window.__v18_rat_patched) return;
    window.__v18_rat_patched = true;
    var orig = window.renderAnalysisTable;
    window.renderAnalysisTable = function(){
      // Call the original (it does all the drawer + CEPS-notes work).
      orig.apply(this, arguments);
      try {
        if (typeof state === 'undefined') return;
        var c = (typeof getConcept === 'function') ? getConcept(state.conceptId) : null;
        var sc = c && c.sub_concepts && (c.sub_concepts[state.subConceptIdx] || c.sub_concepts[0]);
        if (!sc || !_hasSubDim(sc)) return;
        var thead = document.getElementById('analysis-thead');
        var tbody = document.getElementById('analysis-tbody');
        if (!thead || !tbody) return;
        var juris = Object.keys(sc.jurisdictions);
        // Rebuild the header: two dimension sub-columns.
        var JL = (typeof JURIS_LABELS !== 'undefined') ? JURIS_LABELS : {};
        var head = '<tr>';
        head += '<th scope="col" class="analysis-dim-cell v18-dim-col">Dimension</th>';
        head += '<th scope="col" class="analysis-subdim-cell">Sub-dimension</th>';
        juris.forEach(function(j){
          var jid = j.split('-')[0];
          var label = JL[j] || JL[jid] || j;
          var jd = sc.jurisdictions[j];
          head += '<th scope="col" class="th-' + jid + '">' + _escH(label) +
                  '<span class="j-law">' + _escH((jd && jd.law) || '') + '</span></th>';
        });
        head += '</tr>';
        thead.innerHTML = head;
        // Rebuild the body with a dim + sub-dim TD per row.
        var rows = '';
        sc.dimensions.forEach(function(dim){
          var parts = _split(dim);
          rows += '<tr>';
          rows += '<td class="analysis-dim-cell">' + _escH(parts.label) + '</td>';
          rows += '<td class="analysis-subdim-cell">' + _escH(parts.sub) + '</td>';
          juris.forEach(function(j){
            var cell = dim.cells[j];
            if (cell && cell.analysis) {
              rows += '<td><span class="analysis-cell addressed" tabindex="0" role="button" ' +
                      'onclick="openDrawer(\'' + _escH(dim.id) + '\',\'' + j + '\')" ' +
                      'aria-label="Open verbatim: ' + _escH(JL[j] || j) + ' on ' + _escH(dim.label) + '">' +
                      _escH(cell.analysis) + '</span></td>';
            } else {
              rows += '<td><span class="cell-null" aria-label="Not addressed">\u2014</span></td>';
            }
          });
          rows += '</tr>';
        });
        tbody.innerHTML = rows;
      } catch(e){ console.error('v18 two-col dim:', e); }
    };
  }
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', _install);
  else _install();
})();
</script>
"""


if __name__ == "__main__":
    main()
