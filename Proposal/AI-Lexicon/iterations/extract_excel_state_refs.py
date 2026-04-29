"""extract_excel_state_refs.py — US-008 ground-truth extractor.

Walks each analysis sheet of the Excel reference workbook and emits, per
(sub-table, attribute, jurisdiction-column), the raw cell text and any
article/section references that appear inside it. The output is a JSON
file that downstream auditing can compare against `us008_state_cells.json`.

References are detected with a small regex set tuned to each
jurisdiction's citation style (Article N, §N, 22757.NN, §N-N-NN, etc.).
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

XLSX = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis.xlsx"
)
HERE = Path(__file__).parent
OUT = HERE / "outputs" / "us008_excel_state_refs.json"


# ---- Reference patterns --------------------------------------------------- #
# California: 22757.X / 22757.X.(letter) / §22757.X / 3110./ 3111. / 1107.1.
RE_CA = re.compile(
    r"(?:§\s*)?22757\.\d+(?:\.\s*\([a-z0-9]+\))?(?:\.|\b)"
    r"|(?:§\s*)?(?:3110|3111|1107\.1)(?:\.\s*\([a-z]\))?(?:\.|\b)"
)
# Colorado: §6-1-1701..§6-1-1707 with optional "(N)" "(N)(letter)"
RE_CO = re.compile(r"§\s*6-1-170[1-7](?:\s*\(\d+\)(?:\([a-z]\))?)?")
# New York: § 1420..§ 1430 (with optional space, optional "(N)" or "(N)(letter)")
RE_NY = re.compile(r"§\s*14[2-3]\d(?:\s*\((\d+)\)(?:\([a-z]\))?)?")
# Texas HB 149: 552.001..552.999 with optional ".(letter)"
RE_TX = re.compile(r"§?\s*552\.\d{3}(?:\.\s*\([a-z]\))?(?:\.|\b)")
# Utah SB 226: §13-75-101..§13-75-105 with optional "(N)"
RE_UT = re.compile(r"§\s*13-75-10[1-5](?:\s*\(\d+\))?")
# EU AIA: Article 1..Article 99, Annex I..XII
RE_EU = re.compile(
    r"Article\s+\d+(?:\s*\([0-9]+(?:[a-z])?\))?(?:\s*\([0-9]+\))?"
    r"|Annex\s+(?:I|II|III|IV|V|VI|VII|VIII|IX|X|XI|XII)"
    r"|Recital\s+\d+",
    re.IGNORECASE,
)


def find_refs(text: str) -> dict:
    if not text:
        return {}
    out = {}
    for jid, rx in (
        ("ca", RE_CA),
        ("co", RE_CO),
        ("ny", RE_NY),
        ("tx", RE_TX),
        ("ut", RE_UT),
        ("eu", RE_EU),
    ):
        hits = rx.findall(text)
        if hits:
            # findall returns tuples for groups; coerce back to strings
            normalised = []
            for h in hits:
                if isinstance(h, tuple):
                    h = next((g for g in h if g), "")
                normalised.append(h)
            # Get the actual matched substrings.
            spans = [m.group(0).strip() for m in rx.finditer(text)]
            out[jid] = sorted(set(spans))
    return out


def text_or_empty(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def walk_analysis_sheet(ws, header_rows: int = 2) -> list[dict]:
    """Walk a stacked-sub-table analysis sheet and return per-cell records.

    Each record is {row, col_letter, jid_guess, attribute, text, refs}.
    Sub-table boundaries: a row is a sub-table title if col A is non-empty
    AND no other column on that row holds a value (we treat single-cell
    rows as "title only"). Subsequent rows belong to that sub-table until
    the next title row is found.
    """
    records = []

    # Header: row 1 has jurisdiction names; row 2 has bill names. Combine.
    header_cells = []
    for col in range(1, ws.max_column + 1):
        h1 = text_or_empty(ws.cell(row=1, column=col))
        h2 = text_or_empty(ws.cell(row=2, column=col))
        header_cells.append((h1, h2))

    sub_title = ""
    attribute = ""
    for r in range(header_rows + 1, ws.max_row + 1):
        a_val = text_or_empty(ws.cell(row=r, column=1))
        # Detect sub-table title: column A non-empty and all other columns empty.
        other_vals = [
            text_or_empty(ws.cell(row=r, column=c))
            for c in range(2, ws.max_column + 1)
        ]
        is_title_row = bool(a_val) and not any(other_vals)
        if is_title_row:
            sub_title = a_val
            attribute = ""
            continue

        if a_val:
            attribute = a_val

        for c in range(2, ws.max_column + 1):
            txt = text_or_empty(ws.cell(row=r, column=c))
            if not txt:
                continue
            h1, h2 = header_cells[c - 1]
            records.append({
                "row": r,
                "col": c,
                "header_jur": h1,
                "header_bill": h2,
                "sub_table": sub_title,
                "attribute": attribute,
                "text": txt,
                "refs": find_refs(txt),
            })
    return records


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out: dict[str, list[dict]] = {}

    sheets = [
        # name in Excel (raw — caller must `.strip()` if leading-space)
        "Provider_Developer_Analysis",
        "Deployer_Supplier_Analysis",
        " High-risk AI system_ANALYSIS",  # leading space
        "GPAI system_Generative AI_ANALY",
        "GPAI_Frontier_Foundation_Analys",
        "Risk_ANALYSIS",
        "Modification_ANALYSIS",
        "Incident_ANALYSIS",
        "Prohibited_Practices",
    ]
    for sheet in sheets:
        # Find sheet by case-insensitive strip-equal match.
        target = None
        for name in wb.sheetnames:
            if name.strip().lower() == sheet.strip().lower():
                target = name
                break
        if target is None:
            print(f"  WARN: sheet not found: {sheet!r}")
            continue
        ws = wb[target]
        recs = walk_analysis_sheet(ws)
        out[target.strip()] = recs
        print(f"  {target!r:50s} {len(recs)} non-empty cells")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nwrote {OUT}")


if __name__ == "__main__":
    main()
