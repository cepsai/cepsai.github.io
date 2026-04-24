"""build_v22.py — Digital AI Lexicon v22

User feedback round 2026-04-24:
    * Drop the separate "Legal text" mode tab — verbatim returns to being
      drawer-on-click only (v18 behavior), with v16's "See verbatim in full
      law" button still available for jumping to the embedded law viewer.
    * Update analysis-cell text from the NEW xlsx
      (Cross-checked_AI terminology and taxonomy_analysis.xlsx), a full
      review pass that also:
        - renames cluster 'Governance' → 'Actors'
        - reorders the matrix so 'Technical system attributes' comes first
        - normalizes bill-code spacing (SB 24-205, HB 149, SB 226)
        - updates the matrix description paragraph
        - updates the About-sheet and Methodology-sheet prose
    * Add five new EU policy texts as supplementary law sources:
        - Commission Guidelines on the AI-system definition
        - Commission Guidelines on Prohibited AI Practices
        - GPAI Code of Practice — Transparency Chapter
        - GPAI Code of Practice — Safety and Security Chapter
        - GPAI Code of Practice — Copyright Chapter
      (PDFs ingested by `ingest_new_laws.py` — laws/*.json).

Build chain: v13 → v15 → v16 → v17 → v18 → v20 → v21 → **v22**.

v22 operates as a post-process on v21.html:
    1. If digital_lexicon_v21.html is missing, rebuild the chain.
    2. Copy the 5 new law JSONs from the sibling AI-Lexicon workspace into
       ./laws/ if not already there.
    3. Parse the NEW xlsx analysis sheets → build an updates lookup keyed by
       (concept_id, sub_concept_title_lc, dim_label_lc, jid) → value is the
       new analysis text.
    4. Walk CONCEPTS inside v21.html:
        - apply the analysis-text updates
        - rename cluster 'Governance' → 'Actors' throughout
    5. Inject 5 new <script type="application/json" id="law-blob-…"> tags
       for the new laws.
    6. Extend LAW_STUBS + LAWS nav metadata (the Law Sources page) so the
       new laws are clickable.
    7. Inject CSS + JS overrides:
        - hide .v20-mode-bar (no Analysis/Legal-text toggle anymore)
        - pin mode to 'analysis' on any future render pass so the v21
          filter bar + single-row view never activate.
    8. Write digital_lexicon_v22.html + copy to ../final_tool.html.

Cache wipe (predecessors cache aggressively):
    rm -f digital_lexicon_v1{6,7,8}.html digital_lexicon_v2*.html
    python3 build_v22.py
"""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

import openpyxl

HERE          = Path(__file__).parent
HTML_V21      = HERE / "digital_lexicon_v21.html"
HTML_V22      = HERE / "digital_lexicon_v22.html"
FINAL_TOOL    = HERE.parent / "final_tool.html"              # legacy URL, kept in sync
FINAL_LEXICON = HERE.parent / "final_lexicon_tool.html"      # new canonical deploy
LAWS_DIR      = HERE / "laws"

# Companion workspace where the new xlsx + ingested law JSONs live.
SRC_LAWS_DIR  = Path(
    "/Users/robertpraas/CEPS-DST Dropbox/Robert Praas/"
    "cb_sep_25/AI-Lexicon/laws"
)
NEW_XLSX      = Path(
    "/Users/robertpraas/Downloads/"
    "Cross-checked_AI terminology and taxonomy_analysis.xlsx"
)

# Law IDs of the 5 newly-ingested EU policy texts.
NEW_LAW_IDS = [
    "eu-guidelines-ai-definition",
    "eu-guidelines-prohibited",
    "eu-gpai-cop-transparency",
    "eu-gpai-cop-safety",
    "eu-gpai-cop-copyright",
]
NEW_LAW_META = {
    "eu-guidelines-ai-definition": {
        "code":      "Comm. Guidelines — AI-system definition",
        "title":     "Commission Guidelines on the Definition of an AI System",
        "effective": "2025-07-29",
        "desc":      "Interpretive guidance on the seven elements of the Article 3(1) AI-system definition.",
        "url":       "https://digital-strategy.ec.europa.eu/en/library/commission-guidelines-definition-artificial-intelligence-system-established-regulation-eu-20241689",
    },
    "eu-guidelines-prohibited": {
        "code":      "Comm. Guidelines — Prohibited practices",
        "title":     "Commission Guidelines on Prohibited AI Practices",
        "effective": "2025-07-29",
        "desc":      "Detailed guidance on the Article 5 prohibitions (social scoring, emotion recognition, untargeted scraping, etc.).",
        "url":       "https://digital-strategy.ec.europa.eu/en/library/commission-guidelines-prohibited-artificial-intelligence-practices-established-regulation-eu",
    },
    "eu-gpai-cop-transparency": {
        "code":      "GPAI CoP — Transparency",
        "title":     "Code of Practice for GPAI Models — Transparency Chapter",
        "effective": None,
        "desc":      "Working Group 1 commitments for documentation and downstream disclosure under Article 53.",
        "url":       "https://digital-strategy.ec.europa.eu/en/policies/ai-code-practice",
    },
    "eu-gpai-cop-safety": {
        "code":      "GPAI CoP — Safety and Security",
        "title":     "Code of Practice for GPAI Models — Safety and Security Chapter",
        "effective": None,
        "desc":      "Working Groups 2-4 commitments on risk assessment, mitigation, and governance for frontier models.",
        "url":       "https://digital-strategy.ec.europa.eu/en/policies/ai-code-practice",
    },
    "eu-gpai-cop-copyright": {
        "code":      "GPAI CoP — Copyright",
        "title":     "Code of Practice for GPAI Models — Copyright Chapter",
        "effective": None,
        "desc":      "Working Group 1 commitments on lawful training-data acquisition and rights-holder reservations.",
        "url":       "https://digital-strategy.ec.europa.eu/en/policies/ai-code-practice",
    },
}


# --------------------------------------------------------------------------- #
# JSON-literal locator (bracket-aware) — same logic as v16/v17.               #
# --------------------------------------------------------------------------- #

def _find_json_literal(src: str, var_name: str) -> tuple[int, int] | None:
    key = f"const {var_name}"
    start = src.find(key)
    if start < 0:
        return None
    i = src.index("=", start) + 1
    while i < len(src) and src[i] not in "[{":
        i += 1
    if i >= len(src):
        return None
    opener = src[i]
    closer = "]" if opener == "[" else "}"
    depth, in_str, esc, j = 0, False, False, i
    while j < len(src):
        c = src[j]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == opener:
                depth += 1
            elif c == closer:
                depth -= 1
                if depth == 0:
                    return (i, j + 1)
        j += 1
    return None


def _replace_json_literal(src: str, var_name: str, new_obj) -> str:
    span = _find_json_literal(src, var_name)
    if not span:
        raise RuntimeError(f"const {var_name} not found in HTML")
    a, b = span
    payload = json.dumps(new_obj, ensure_ascii=False, separators=(",", ":"))
    return src[:a] + payload + src[b:]


# --------------------------------------------------------------------------- #
# New-xlsx analysis-cell updates lookup.                                      #
# --------------------------------------------------------------------------- #

# Map each legal-sheet sheet-name (used as `tab_id` via build_v13.SHEET_TO_TAB)
# to the corresponding ANALYSIS sheet in the NEW xlsx. The new xlsx renamed
# "Substantial modif_ANALYSIS" → "Modification_ANALYSIS".
_ANA_SHEETS_NEW = {
    "Provider_Developer":              "Provider_Developer_Analysis",
    "Deployer_Supplier":               "Deployer_Supplier_Analysis",
    "GPAI_Frontier_Foundation model":  "GPAI_Frontier_Foundation_Analys",
    "GPAI system_Generative AI":       "GPAI system_Generative AI_ANALY",
    " High-risk AI system":            " High-risk AI system_ANALYSIS",
    "Risk":                            "Risk_ANALYSIS",
    "Substantial modification":        "Modification_ANALYSIS",
    "Incident":                        "Incident_ANALYSIS",
}

# concept.id in the built JSON is derived from SHEET_TO_TAB but normalized
# to kebab-case. Hardcoded here so v22 doesn't need to import build_v13.
_LEGAL_SHEET_TO_CID = {
    "Provider_Developer":              "provider-developer",
    "Deployer_Supplier":               "deployer-supplier",
    "GPAI_Frontier_Foundation model":  "gpai-frontier-foundation",
    "GPAI system_Generative AI":       "gpai-system-generative-ai",
    " High-risk AI system":            "high-risk-ai-system",
    "Risk":                            "risk",
    "Substantial modification":        "substantial-modification",
    "Incident":                        "incident",
}


def _parse_new_analysis_sheets(xlsx_path: Path) -> dict:
    """Return `{tab_id: [{title, title_lc, dims: {dim_lc: {jid: text}}}, ...]}`
    where each list entry is one section block inside an analysis sheet.

    A section block is a group of rows whose first row has a non-empty col A
    (the section title) and whose second row has ≥2 jurisdiction headers
    ("EU (AIA)", "Colorado (SB 24-205)", ...).
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    JID_DETECT = [
        ("eu", re.compile(r"\beu\s*\(aia\)|\beu\s+ai\s+act|\baia\b", re.I)),
        ("ca", re.compile(r"\bcalifornia\b", re.I)),
        ("co", re.compile(r"\bcolorado\b", re.I)),
        ("ny", re.compile(r"\bnew\s*york\b", re.I)),
        ("tx", re.compile(r"\btexas\b", re.I)),
        ("ut", re.compile(r"\butah\b", re.I)),
    ]
    _BILL_CAPTURE = re.compile(r"\(([A-Z]{1,3}\s?\d+[\-\d]*)\)", re.I)

    def _read_jid_cols(row) -> dict[int, tuple[str, str]]:
        """Return {col_idx: (jid, bill_code)}. Header cells are short labels
        like 'California (SB 942)' — guard with a length cap so paragraph-
        long intro text in col E (which often mentions 'AIA' or jurisdiction
        names in prose) can't be mistaken for a header.
        """
        out: dict[int, tuple[str, str]] = {}
        for c in range(1, len(row)):
            v = row[c]
            if not v:
                continue
            s = str(v).strip()
            # Real jurisdiction headers are short labels. Anything > 60
            # chars is prose/notes, not a header.
            if len(s) > 60:
                continue
            matched_jid = None
            for jid, pat in JID_DETECT:
                if pat.search(s):
                    matched_jid = jid
                    break
            if not matched_jid:
                continue
            bill = ""
            bm = _BILL_CAPTURE.search(s)
            if bm:
                bill = bm.group(1).strip().replace(" ", "")
            out[c] = (matched_jid, bill)
        return out

    out: dict[str, list[dict]] = {}

    for legal_sn, ana_sn in _ANA_SHEETS_NEW.items():
        if ana_sn not in wb.sheetnames:
            continue
        ws = wb[ana_sn]
        tab_id = _LEGAL_SHEET_TO_CID[legal_sn]
        rows = list(ws.iter_rows(values_only=True))
        sections: list[dict] = []

        def _consume_section(start_hdr_row: int, title: str) -> int:
            """From `start_hdr_row` (the row with jurisdiction headers),
            consume dim rows until the next section header (or EOF). Always
            appends the section (with dims + notes) to `sections`. Returns
            the row index where the next scan should resume."""
            jid_info = _read_jid_cols(rows[start_hdr_row])  # {col: (jid, bill)}
            section = {
                "title":    title,
                "title_lc": title.lower(),
                "dims":     {},       # {dim_lc: {jid_key: text}}
                "notes":    "",
            }
            # The notes column is any column beyond the last jid column that
            # carries non-empty text. For the new xlsx this is col E (index 4).
            max_jid_col = max(jid_info.keys()) if jid_info else 0
            notes_col = max_jid_col + 1 if max_jid_col else -1

            def _harvest_notes(r):
                if notes_col < 0 or notes_col >= len(r):
                    return
                v = r[notes_col]
                if v is None:
                    return
                s = str(v).strip()
                if not s:
                    return
                if s in section["notes"]:
                    return
                section["notes"] = (section["notes"] + "\n\n" + s).strip() if section["notes"] else s

            # Pick up any notes sitting in the header row's own notes column
            # (merged cells anchor at the top-left of the merge).
            _harvest_notes(rows[start_hdr_row])

            j = start_hdr_row + 1
            while j < len(rows):
                r = rows[j] or ()
                a = (r[0] if len(r) > 0 else None) or ""
                a = str(a).strip()
                # Break on next section boundary (recurring jid-header rows).
                if j > start_hdr_row and len(_read_jid_cols(r)) >= 2:
                    sections.append(section)
                    return j
                nxt2 = rows[j + 1] if j + 1 < len(rows) else None
                if a and nxt2 and len(_read_jid_cols(nxt2)) >= 2:
                    sections.append(section)
                    return j
                _harvest_notes(r)
                if a:
                    dim_lc = a.lower()
                    per_jkey: dict[str, str] = {}
                    for c, (jid, bill) in jid_info.items():
                        v = r[c] if c < len(r) else None
                        txt = "" if v is None else str(v).strip()
                        if not txt or txt in ("-", "–", "—"):
                            continue
                        # Key by (jid, bill) so California (SB 942) and
                        # California (AB 2013) stay distinct. Primary key
                        # is `jid+-+bill` when bill is present, else `jid`.
                        key = f"{jid}-{bill.lower()}" if bill else jid
                        per_jkey[key] = txt
                        # Also record bare jid as a fallback lookup so
                        # single-bill jurisdictions still match v21's lane
                        # keys that carry no bill suffix.
                        per_jkey.setdefault(jid, txt)
                    if per_jkey:
                        slot = section["dims"].setdefault(dim_lc, {})
                        for k, v in per_jkey.items():
                            slot.setdefault(k, v)
                j += 1
            sections.append(section)
            return j

        # Walk rows: either (1) multi-section (col A has title, next row has
        # jid headers) or (2) single-section (current row already has jid
        # headers with blank col A, use the legal sheet's name as the title).
        i = 0
        while i < len(rows):
            row = rows[i] or ()
            col_a = (row[0] if len(row) > 0 else None) or ""
            col_a = str(col_a).strip()
            nxt = rows[i + 1] if i + 1 < len(rows) else None
            if col_a and nxt and len(_read_jid_cols(nxt)) >= 2:
                i = _consume_section(i + 1, col_a)
                continue
            if len(_read_jid_cols(row)) >= 2:
                title = legal_sn.strip().replace("_", " ")
                i = _consume_section(i, title)
                continue
            i += 1

        if sections:
            out[tab_id] = sections

    return out


_GPAI_SUB = [
    (re.compile(r"\bgpai\b", re.I),                      "general-purpose ai"),
    (re.compile(r"\bg\.p\.a\.i\.?\b", re.I),             "general-purpose ai"),
    (re.compile(r"\bgeneral[- ]purpose\s+ai\b", re.I),   "general-purpose ai"),
    # Singular/plural canonicalization — the new xlsx uses "Providers of
    # GPAI models …" (plural) while v21's sub-concept is "Provider of …".
    (re.compile(r"\bproviders\b", re.I),                 "provider"),
    (re.compile(r"\bdevelopers\b", re.I),                "developer"),
    (re.compile(r"\bdeployers\b",  re.I),                "deployer"),
    (re.compile(r"\bsuppliers\b",  re.I),                "supplier"),
]


def _normalize_concept_name(s: str) -> str:
    """Canonicalize regulatory-concept names for fuzzy matching.
    Handles GPAI ↔ general-purpose AI, punctuation, whitespace.
    """
    s = (s or "").lower()
    for pat, repl in _GPAI_SUB:
        s = pat.sub(repl, s)
    # Keep only letters, digits, spaces.
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _fuzzy_section_match(sub_title: str, sections: list[dict]) -> dict | None:
    """Choose the best matching section in `sections` for `sub_title`.
    Comparison is via _normalize_concept_name so GPAI ↔ general-purpose AI
    match each other. Exact normalized match wins; otherwise longest
    substring containment.
    """
    if not sub_title or not sections:
        return None
    t_norm = _normalize_concept_name(sub_title)
    if not t_norm:
        return None
    # Exact normalized match
    for s in sections:
        if _normalize_concept_name(s["title_lc"]) == t_norm:
            return s
    # Substring containment (either direction), scoring by overlap length.
    best = None
    best_score = 0
    for s in sections:
        st = s["title_lc"]
        if not st:
            continue
        # First segment of "X / Y / Z" — often the EU-centric canonical term.
        st_core = re.split(r"\s*[/,]\s*", st)[0].strip()
        st_norm      = _normalize_concept_name(st)
        st_core_norm = _normalize_concept_name(st_core)
        score = 0
        if t_norm in st_norm:
            score = max(score, len(t_norm))
        if t_norm in st_core_norm:
            score = max(score, len(t_norm))
        if st_norm in t_norm:
            score = max(score, len(st_norm))
        if st_core_norm in t_norm:
            score = max(score, len(st_core_norm))
        if score > best_score:
            best_score = score
            best = s
    return best


_NOTES_HEADER_RE = re.compile(r"(?:^|\n)\s*interpretative\s+notes?\s*\n",
                               re.I)
_THEME_SPLIT_RE = re.compile(r"\n\s*-\s+")


def _parse_notes_to_themes(raw: str) -> dict:
    """Convert the xlsx Interpretative-notes blob into `ceps_notes` shape
    used by v21's CEPS Comparative Analysis block:
        {"themes": [{"title": ..., "body": ...}, ...]}  (preferred)
        {"summary": "..."}                               (fallback)

    Format of `raw`:
        <optional intro paragraph>\n\n
        Interpretative notes\n
        - <title 1>\n<body 1>\n
        - <title 2>\n<body 2>\n...

    Strips the intro + "Interpretative notes" header; returns themes when
    the dash-delimited structure is present, otherwise the whole text as
    summary.
    """
    if not raw:
        return {}
    text = raw
    m = _NOTES_HEADER_RE.search(text)
    if m:
        # Keep only what comes after "Interpretative notes" header.
        text = text[m.end():]
    # Split on leading "-" lines.
    parts = _THEME_SPLIT_RE.split("\n" + text)
    # First segment is anything before the first "-" (often empty after strip).
    pre = parts[0].strip()
    body_parts = [p for p in parts[1:] if p.strip()]
    if not body_parts:
        # No themes; return as a plain summary.
        s = (pre or text).strip()
        return {"summary": s} if s else {}
    themes: list[dict] = []
    for seg in body_parts:
        seg = seg.rstrip()
        lines = seg.split("\n", 1)
        title = lines[0].strip().rstrip(":").strip()
        body  = lines[1].strip() if len(lines) > 1 else ""
        if not title and not body:
            continue
        themes.append({"title": title, "body": body})
    return {"themes": themes}


def _apply_concept_updates(concepts: list, by_tab: dict) -> tuple[int, int, int, int]:
    """For each dim cell in CONCEPTS, replace .analysis with the new text
    when we can match this sub-concept to a section block in the new xlsx.
    Also replaces `sc.ceps_notes` with the parsed Interpretative-notes blob
    so the existing "CEPS Comparative Analysis" collapsible renders them
    in-place (no separate column).

    Returns (n_replaced, n_kept, n_sections_unmatched, n_notes_attached).
    """
    replaced = 0
    kept = 0
    unmatched = 0
    notes_attached = 0
    for c in concepts:
        cid = c.get("id", "")
        sections = by_tab.get(cid, [])
        matched_sections = set()
        for sc in c.get("sub_concepts", []):
            sub_title = (sc.get("title") or "").strip()
            section = _fuzzy_section_match(sub_title, sections)
            if not section:
                for dim in sc.get("dimensions", []):
                    for _jid, cell in (dim.get("cells") or {}).items():
                        if isinstance(cell, dict):
                            kept += 1
                continue
            matched_sections.add(section["title_lc"])
            if section.get("notes"):
                parsed = _parse_notes_to_themes(section["notes"])
                if parsed:
                    sc["ceps_notes"] = parsed
                    notes_attached += 1
            for dim in sc.get("dimensions", []):
                dim_lc = (dim.get("label") or "").strip().lower()
                dim_stripped = re.sub(r"\s*[/–—-].*$", "", dim_lc).strip()
                new_per_jid = (section["dims"].get(dim_lc)
                               or section["dims"].get(dim_stripped)
                               or {})
                juris_meta = sc.get("jurisdictions") or {}
                for jid_key, cell in (dim.get("cells") or {}).items():
                    if not isinstance(cell, dict):
                        continue
                    jdata = juris_meta.get(jid_key) or {}
                    base_jid = jdata.get("_parent_jid") or jid_key.split("-")[0]
                    bills = str(jdata.get("bills") or "")
                    bill_key = re.sub(r"\s+", "", bills).lower()
                    candidates: list[str] = []
                    if bill_key:
                        candidates.append(f"{base_jid}-{bill_key}")
                    candidates.append(base_jid)
                    new_text = None
                    for k in candidates:
                        if k in new_per_jid:
                            new_text = new_per_jid[k]
                            break
                    if new_text and new_text != (cell.get("analysis") or ""):
                        cell["analysis"] = new_text
                        replaced += 1
                    else:
                        kept += 1
        for s in sections:
            if s["title_lc"] not in matched_sections:
                unmatched += 1
    return replaced, kept, unmatched, notes_attached


# --------------------------------------------------------------------------- #
# Cluster rename — Governance → Actors — in CONCEPTS + MATRIX.                #
# --------------------------------------------------------------------------- #

def _rename_cluster(concepts: list, old: str, new: str) -> int:
    n = 0
    for c in concepts:
        if c.get("cluster") == old:
            c["cluster"] = new
            n += 1
    return n


def _promote_limited_risk_provider(concepts: list) -> int:
    """The xlsx's `New concepts` matrix sheet labels the first Provider/
    Developer row simply 'Provider', while the `Provider_Developer_Analysis`
    sheet uses the longer, more specific 'Provider of limited-risk AI
    systems' as the section title + Term row value. Robert asked for the
    longer form everywhere in the UI.

    This mutates the concept tree + cluster_summary variant pills so the
    sub-tab label, the landing-page matrix pill, and the dim-table Term/EU
    cell all read 'Provider of limited-risk AI systems'.
    """
    FULL = "Provider of limited-risk AI systems"
    touched = 0
    for c in concepts:
        if c.get("id") != "provider-developer":
            continue
        for sc in c.get("sub_concepts", []):
            if (sc.get("title") or "").strip().lower() == "provider":
                sc["title"] = FULL
                touched += 1
                break
        cs = c.get("cluster_summary") or {}
        for row in (cs.get("rows") or []):
            cells = row.get("cells") or {}
            eu = cells.get("eu") or {}
            for v in eu.get("variants") or []:
                if (v.get("name") or "").strip().lower() == "provider":
                    v["name"] = FULL
                    touched += 1
                    break
            break  # only first row
        break
    return touched


def _promote_limited_risk_in_matrix(matrix: dict) -> int:
    """Same rewrite applied to the raw MATRIX JSON blob (used by the
    Concepts-landing matrix renderer)."""
    touched = 0
    for row in (matrix.get("rows") or []):
        term = (row.get("term") or "").strip().lower()
        if term != "provider / developer":
            continue
        cells = row.get("cells") or []
        if cells and isinstance(cells[0], str) and cells[0].strip().lower() == "provider":
            cells[0] = "Provider of limited-risk AI systems"
            touched += 1
        break
    return touched


# --------------------------------------------------------------------------- #
# LAW_STUBS / LAWS nav extensions.                                            #
# --------------------------------------------------------------------------- #

def _update_window_law_stubs(html: str, new_laws: dict[str, dict]) -> str:
    """Locate `window.LAW_STUBS = {…};` in the HTML and rewrite the object
    literal with added entries for each new law."""
    marker = "window.LAW_STUBS"
    start = html.find(marker)
    if start < 0:
        print("  window.LAW_STUBS not found — skipping stub extension")
        return html
    eq = html.index("=", start) + 1
    # Skip whitespace until the opening '{'.
    i = eq
    while i < len(html) and html[i] not in "{":
        i += 1
    if i >= len(html):
        return html
    depth, in_str, esc, j = 0, False, False, i
    while j < len(html):
        c = html[j]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    break
        j += 1
    body = html[i:j + 1]
    try:
        stubs = json.loads(body)
    except json.JSONDecodeError as e:
        print(f"  window.LAW_STUBS parse failed: {e}")
        return html
    added = _extend_law_stubs(stubs, new_laws)
    new_body = json.dumps(stubs, ensure_ascii=False, separators=(",", ":"))
    print(f"  LAW_STUBS extended:  +{added} entries")
    return html[:i] + new_body + html[j + 1:]


def _extend_law_stubs(stubs: dict, new_laws: dict[str, dict]) -> int:
    """For each new law JSON blob, derive its LAW_STUBS entry (articles/sections)."""
    added = 0
    for lid, blob in new_laws.items():
        if lid in stubs:
            continue
        stubs[lid] = {
            "title": blob.get("title", ""),
            "url":   blob.get("url", NEW_LAW_META.get(lid, {}).get("url", "")),
            "articles": [
                {"id": a.get("id", ""), "title": a.get("title", "")}
                for a in (blob.get("articles") or [])
            ],
            "sections": [
                {"id": s.get("id", ""),
                 "title": s.get("title", "") or ("Section " + s.get("id", ""))}
                for s in (blob.get("sections") or [])
            ],
        }
        added += 1
    return added


def _extend_laws_nav(laws_nav: list, new_law_ids: list[str]) -> int:
    """Produce the 12-framework layout Robert asked for on 2026-04-24:

        EU:
          1.  EU AI Act
          2.  Commission Guidelines on AI-system definition
          3.  GPAI Code of Practice (one entry, 3 chapters)
        California: 4. SB 53   5. SB 942
        Colorado:   6. SB 24-205   7. SB 25B-004
        New York:   8. A6453B   9. S8828
        Texas:     10. HB 149
        Utah:      11. SB 149  12. SB 226

    The Commission Guidelines on Prohibited AI Practices stays ingested
    (law blob + REF_MAP available) but is NOT yet in the Regulations nav
    per Robert's 2026-04-24 note "prohibited practices can come later".

    The 3 GPAI CoP chapter blobs remain embedded so REF_MAP / drawer
    routing can resolve per-chapter references; they collapse into ONE
    nav card whose click-through target is the Transparency chapter.
    """
    touched = 0
    for region in laws_nav:
        if region.get("region") != "European Union":
            continue
        existing = list(region.get("laws") or [])
        # Upgrade the placeholder "GPAI CoP" entry (law_id=None) to a real one.
        cop_idx = None
        for i, l in enumerate(existing):
            code = (l.get("code") or "").strip()
            if (l.get("law_id") is None) and ("CoP" in code or "GPAI" in code):
                l["title"] = "Code of Practice for General-Purpose AI Models"
                l["desc"]  = ("Three chapters (Transparency, Safety and "
                              "Security, Copyright) of commitments that GPAI "
                              "providers can follow to demonstrate compliance "
                              "with Articles 53 and 55 AI Act.")
                l["law_id"] = "eu-gpai-cop-transparency"
                cop_idx = i
                touched += 1
                break
        # Insert the AI-system definition Guidelines just before the CoP
        # entry so the EU ordering reads: AIA → AI-def Guidelines → CoP.
        ai_def_entry = {
            "code":      "Comm. Guidelines — AI-system definition",
            "title":     "Commission Guidelines on the Definition of an AI System",
            "effective": "2025-07-29",
            "desc":      NEW_LAW_META["eu-guidelines-ai-definition"]["desc"],
            "law_id":    "eu-guidelines-ai-definition",
        }
        already = any(l.get("law_id") == "eu-guidelines-ai-definition"
                      for l in existing)
        if not already:
            insert_at = cop_idx if cop_idx is not None else len(existing)
            existing.insert(insert_at, ai_def_entry)
            touched += 1
        region["laws"] = existing
        break
    # Add Commission Guidelines on Prohibited AI Practices as the final
    # EU-region entry (Robert's 2026-04-24 "show the missing EU one" ask).
    prohibited_entry = {
        "code":      "Comm. Guidelines — Prohibited practices",
        "title":     "Commission Guidelines on Prohibited AI Practices",
        "effective": "2025-07-29",
        "desc":      ("Detailed guidance on the Article 5 prohibitions "
                      "(social scoring, emotion recognition, untargeted "
                      "scraping, etc.). COM(2025) 5052 final, 29.7.2025."),
        "law_id":    "eu-guidelines-prohibited",
    }
    for region in laws_nav:
        if region.get("region") != "European Union":
            continue
        already = any(l.get("law_id") == "eu-guidelines-prohibited"
                      for l in region.get("laws") or [])
        if not already:
            region["laws"].append(prohibited_entry)
            touched += 1
        break
    return touched


# --------------------------------------------------------------------------- #
# v22 CSS + JS overrides injected before </body>.                             #
# --------------------------------------------------------------------------- #

def _v22_overrides() -> str:
    return r"""
<style>
/* v22 — verbatim returns to drawer-on-click (v18 behavior). Cells in the
   dim table MUST show .analysis text, never .verbatim. The v16 "See
   verbatim in full law" button still lives inside the drawer for
   jumping into the embedded law viewer. */
.v20-mode-bar { display: none !important; }
.v21-filter-bar { display: none !important; }
/* Strip v21's single-row class so all rows render. */
table.v21-one-row { }
table.v21-one-row tbody tr { display: table-row !important; }
</style>
<script>
(function(){
  /* 1. Force state.mode = 'analysis' before ANY render reads it. Run
        both as the IIFE loads AND on every render call (belt-and-braces
        against localStorage-hydrated state and hashchange races). */
  function _pinAnalysis(){
    try {
      if (typeof state !== 'undefined' && state) {
        state.mode = 'analysis';
      }
    } catch(e) {}
  }
  _pinAnalysis();

  /* 2. Wipe any ?view=verbatim or #…?view=verbatim URL param on load so a
        shared deep-link can't re-flip into verbatim mode. */
  try {
    var h = location.hash || '';
    if (h.indexOf('view=verbatim') !== -1) {
      var cleaned = h.replace(/[?&]view=verbatim/g, '').replace(/\?$/, '');
      history.replaceState(null, '', location.pathname + location.search + cleaned);
    }
  } catch(e){}

  /* 3. Neutralise setMode so any button/handler that still calls it
        can't flip the page into verbatim mode. */
  if (typeof window.setMode === 'function') {
    var _origSetMode = window.setMode;
    window.setMode = function(){ return _origSetMode.call(this, 'analysis'); };
  }

  /* 4. Wrap renderAnalysisTable so state.mode is forced to 'analysis'
        IMMEDIATELY BEFORE v21's wrapper runs. v21's wrapper checks
        `if (state.mode === 'verbatim') { _renderV21VerbatimTable(); }`
        — by forcing analysis at entry, we always take the normal branch. */
  function _installRATWrap(){
    if (typeof window.renderAnalysisTable !== 'function') return false;
    if (window.__v22_rat_wrapped) return true;
    window.__v22_rat_wrapped = true;
    var _origRAT = window.renderAnalysisTable;
    window.renderAnalysisTable = function(){
      _pinAnalysis();
      var out = _origRAT.apply(this, arguments);
      try {
        var thead = document.getElementById('analysis-thead');
        var table = thead && thead.closest('table');
        if (table) table.classList.remove('v21-one-row');
        var fb = document.querySelector('.v21-filter-bar');
        if (fb) fb.style.display = 'none';
        _flattenDimSubdim();
      } catch(e){}
      return out;
    };
    return true;
  }

  /* --- Flatten the Dimension / Sub-dimension two-column header. The new
         xlsx has NO sub-dimension concept — every dim label is one row.
         v18 added the two-col split for the Incident concept; v20
         generalized it via V20_DIM_PARENTS. In v22 we post-process: drop
         the sub-dim <th>/<td> and dedupe rows whose dim label + cell text
         are identical (e.g. 3 SCOPE rows for Model/System split that all
         carry the same analysis text). */
  function _flattenDimSubdim(){
    var thead = document.getElementById('analysis-thead');
    var tbody = document.getElementById('analysis-tbody');
    if (!thead || !tbody) return;

    function _txt(el){
      return (el ? (el.textContent || '') : '').replace(/\s+/g, ' ').trim();
    }

    // Remove any <th> in the head whose label reads "sub-dimension".
    // Track its column index so the same column gets dropped in the body.
    var subdim_indexes = [];
    var head_rows = thead.querySelectorAll('tr');
    head_rows.forEach(function(tr){
      var ths = Array.prototype.slice.call(tr.children);
      for (var i = 0; i < ths.length; i++){
        if (/sub[-\s]*dimension/i.test(_txt(ths[i]))){
          subdim_indexes.push(i);
          break;
        }
      }
    });
    // All head rows share the same column structure — use first index found.
    var drop_idx = subdim_indexes.length ? subdim_indexes[0] : -1;

    if (drop_idx >= 0){
      head_rows.forEach(function(tr){
        if (tr.children[drop_idx]) tr.children[drop_idx].remove();
      });
      var body_rows = tbody.querySelectorAll('tr');
      body_rows.forEach(function(tr){
        if (tr.children[drop_idx]) tr.children[drop_idx].remove();
      });
    }

    // Dedupe rows: if a row has the same dim label AND the same remaining
    // cell text as the previous row (normalized), remove it. This collapses
    // the 3 SCOPE rows for Model/Provider/System (identical analysis text
    // because the new xlsx has one Scope row, not three).
    var last_sig = null;
    var rows = Array.prototype.slice.call(tbody.querySelectorAll('tr'));
    rows.forEach(function(tr){
      var cells = Array.prototype.slice.call(tr.children);
      if (!cells.length) return;
      var sig = cells.map(_txt).join('§');
      if (sig === last_sig){
        tr.remove();
      } else {
        last_sig = sig;
      }
    });

  }

  /* 5. Same treatment for renderConceptPage (which calls renderAnalysisTable
        internally). */
  function _installRCPWrap(){
    if (typeof window.renderConceptPage !== 'function') return false;
    if (window.__v22_rcp_wrapped) return true;
    window.__v22_rcp_wrapped = true;
    var _origRCP = window.renderConceptPage;
    window.renderConceptPage = function(){
      _pinAnalysis();
      return _origRCP.apply(this, arguments);
    };
    return true;
  }

  /* 6. On hashchange / popstate, force analysis mode again. Covers the
        case where a user pastes a deep link that encodes verbatim mode. */
  function _onRoute(){ _pinAnalysis(); }
  window.addEventListener('hashchange', _onRoute);
  window.addEventListener('popstate',  _onRoute);

  /* 7. Install wrappers at DOM ready (so v21's own wrapping has already
        run). Poll briefly in case the shell rebinds later. */
  function _init(){
    _installRATWrap();
    _installRCPWrap();
    _pinAnalysis();
    // One re-render for safety: if a table already rendered in verbatim
    // mode before our wrapper got installed, call the current renderer
    // again now that state.mode = 'analysis'.
    try {
      if (typeof window.renderAnalysisTable === 'function'
          && typeof state !== 'undefined'
          && state && state.page === 'concept') {
        window.renderAnalysisTable();
      }
    } catch(e){}
  }
  if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', _init);
  } else {
    _init();
  }
  // Second pass a tick later in case the shell rebinds renderAnalysisTable
  // in its own DOMContentLoaded handler AFTER ours.
  setTimeout(_init, 0);
  setTimeout(_init, 250);
})();
</script>
"""


# --------------------------------------------------------------------------- #
# Law-JSON injection.                                                         #
# --------------------------------------------------------------------------- #

def _copy_new_laws_into_place() -> dict[str, dict]:
    """Copy laws/*.json for each NEW_LAW_IDS from SRC_LAWS_DIR to LAWS_DIR
    if missing, then return {lid: loaded_json}."""
    loaded: dict[str, dict] = {}
    for lid in NEW_LAW_IDS:
        fn = f"{lid}.json"
        dst = LAWS_DIR / fn
        src = SRC_LAWS_DIR / fn
        if not dst.exists():
            if not src.exists():
                raise SystemExit(f"Missing source law JSON: {src}")
            shutil.copy2(src, dst)
            print(f"  copied {fn} → laws/")
        loaded[lid] = json.loads(dst.read_text(encoding="utf-8"))
    return loaded


def _inject_new_law_blobs(html: str, loaded: dict[str, dict]) -> str:
    """Insert <script type='application/json' id='law-blob-<lid>'> tags for
    each new law, placed just before the existing closing </body>. Skip any
    that are already present (rerun safety)."""
    for lid, blob in loaded.items():
        marker = f'id="law-blob-{lid}"'
        if marker in html:
            continue
        payload = json.dumps(blob, ensure_ascii=False,
                             separators=(",", ":")).replace("</", "<\\/")
        tag = f'<script type="application/json" id="law-blob-{lid}">{payload}</script>'
        html = html.replace("</body>", tag + "\n</body>", 1)
    return html


# --------------------------------------------------------------------------- #
# Main                                                                        #
# --------------------------------------------------------------------------- #

def main() -> None:
    import build_v21

    print("== v22 build ==")
    if not HTML_V21.exists():
        print("  digital_lexicon_v21.html missing — running build_v21 …")
        build_v21.main()

    html = HTML_V21.read_text(encoding="utf-8")
    print(f"  read v21:            {len(html):,} bytes")

    # 1. Copy new law JSONs into laws/ and inject their blobs.
    loaded_new_laws = _copy_new_laws_into_place()
    html = _inject_new_law_blobs(html, loaded_new_laws)
    print(f"  new law blobs:       {len(loaded_new_laws)} injected")

    # 2. Update LAW_STUBS (drawer navigation targets). In the iterations
    # chain this is a runtime window assignment: `window.LAW_STUBS = {…};`,
    # not a `const`. Find the literal by locating the `{` after the
    # assignment operator.
    html = _update_window_law_stubs(html, loaded_new_laws)

    # 3. Update LAWS nav metadata (Regulations / Law Sources page).
    laws_span = _find_json_literal(html, "LAWS")
    if laws_span:
        a, b = laws_span
        laws_nav = json.loads(html[a:b])
        added = _extend_laws_nav(laws_nav, NEW_LAW_IDS)
        html = _replace_json_literal(html, "LAWS", laws_nav)
        print(f"  LAWS nav extended:   +{added} entries")
    else:
        print("  LAWS nav not found — skipping")

    # 4. Update CONCEPTS: analysis text + cluster rename.
    concepts_span = _find_json_literal(html, "CONCEPTS")
    if not concepts_span:
        raise SystemExit("CONCEPTS literal not found in v21.html")
    a, b = concepts_span
    concepts = json.loads(html[a:b])
    by_tab = _parse_new_analysis_sheets(NEW_XLSX)
    print(f"  new-xlsx tabs:       {sorted(by_tab.keys())}")
    print(f"  new-xlsx sections:   "
          + ", ".join(f"{k}:{len(v)}" for k, v in by_tab.items()))
    n_repl, n_kept, n_unmatched, n_notes = _apply_concept_updates(concepts, by_tab)
    print(f"  analysis updates:    replaced={n_repl}  kept={n_kept}  "
          f"sections unmatched={n_unmatched}  notes_attached={n_notes}")
    n_clus = _rename_cluster(concepts, "Governance", "Actors")
    print(f"  cluster rename:      {n_clus} concept(s) Governance → Actors")
    n_prom = _promote_limited_risk_provider(concepts)
    print(f"  limited-risk label:  {n_prom} spots updated (sub_concept + cluster_summary pill)")
    html = _replace_json_literal(html, "CONCEPTS", concepts)

    # Also rewrite the MATRIX blob (concepts landing page).
    mat_span = _find_json_literal(html, "MATRIX")
    if mat_span:
        ma, mb = mat_span
        matrix = json.loads(html[ma:mb])
        n_mat = _promote_limited_risk_in_matrix(matrix)
        if n_mat:
            html = _replace_json_literal(html, "MATRIX", matrix)
        print(f"  MATRIX label rewrite: {n_mat} cell(s)")

    # 4b. Rename the cluster filter UI (hardcoded in the shell) + the
    # JS constants (DEFAULT_FILTERS, cluster-counts array) so the Actors
    # filter checkbox actually filters anything. Laws with "Governance" in
    # their real titles (e.g. Texas "… Governance Act", EU AIA "Data
    # Governance" article) MUST be preserved. We use contextual patterns.
    patterns = [
        ('value="Governance"', 'value="Actors"'),
        ('id="count-Governance"', 'id="count-Actors"'),
        ('> Governance <span', '> Actors <span'),
        ("'Governance'", "'Actors'"),
        ('"Governance"', '"Actors"'),
    ]
    n_repl = 0
    for old, new in patterns:
        n = html.count(old)
        html = html.replace(old, new)
        n_repl += n
    print(f"  filter-UI rename:    {n_repl} occurrence(s) swapped Governance → Actors")

    # 5. Inject v22 overrides (hide mode bar + pin analysis mode).
    html = html.replace("</body>", _v22_overrides() + "\n</body>", 1)

    # 6. Title + marker.
    html = re.sub(
        r"<title>Digital AI Lexicon v\d+\s*[—\-]\s*CEPS</title>",
        "<title>Digital AI Lexicon v22 — CEPS</title>",
        html, count=1,
    )
    html = html.replace(
        "<!-- DAL v21 (single-row Legal-text view with dim/sub filters) -->",
        "<!-- DAL v22 (drawer-only verbatim, Actors cluster, 5 new EU laws) -->",
        1,
    )

    HTML_V22.write_text(html, encoding="utf-8")
    print(f"\nWrote {HTML_V22}  ({len(html):,} bytes)")

    # 7. Promote to final_lexicon_tool.html (the new canonical URL Robert asked
    # to push on 2026-04-24). Also keep final_tool.html in sync so legacy
    # links don't 404.
    if FINAL_TOOL.exists():
        bak = FINAL_TOOL.with_suffix(".html.bak-v21")
        if not bak.exists():
            shutil.copy2(FINAL_TOOL, bak)
            print(f"  backed up old final_tool → {bak.name}")
    shutil.copy2(HTML_V22, FINAL_LEXICON)
    print(f"  wrote        {FINAL_LEXICON}")
    shutil.copy2(HTML_V22, FINAL_TOOL)
    print(f"  synced       {FINAL_TOOL}")


if __name__ == "__main__":
    main()
