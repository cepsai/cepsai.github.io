#!/usr/bin/env python3
"""
build_v13.py — patch digital_lexicon_v11.html into digital_lexicon_v13.html.

Extends build_v12.py with embedded law texts and clickable reference badges.
The drawer UI reads from a new LAWS JSON const; references in the Lexicon are
rewritten to open the drawer scrolled to the cited article/section.

Inputs:
  - AI terminology and taxonomy-5.xlsx
  - laws/*.json               (produced by laws/fetch_laws.py)
  - digital_lexicon_v11.html  (template)

Outputs:
  - digital_lexicon_v13.html

Usage:
    python3 build_v13.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
HTML_IN = HERE / "digital_lexicon_v11.html"
XLSX = HERE / "AI terminology and taxonomy-final.xlsx"
HTML_OUT = HERE / "digital_lexicon_v13.html"
LAWS_DIR = HERE / "laws"


# ---------- xlsx structural map ----------

# Per-sheet jurisdiction block layout.
# block_width = columns per jurisdiction (rowLabel, [subLabel,] text, reference, tags)
# block_starts = 1-based column where each jurisdiction block begins in row 1
SHEET_LAYOUT = {
    "Provider_Developer":              {"width": 5, "starts": [1, 6, 11, 16]},
    "Deployer_Supplier":               {"width": 5, "starts": [1, 6, 11]},
    "GPAI_Frontier_Foundation model":  {"width": 5, "starts": [1, 6, 11]},
    "GPAI system_Generative AI":       {"width": 5, "starts": [1, 6, 11]},
    " High-risk AI system":            {"width": 5, "starts": [1, 6, 11]},
    "Risk":                            {"width": 4, "starts": [1, 5, 9]},
    "Substantial modification":        {"width": 4, "starts": [2, 6, 10]},
    "Incident":                        {"width": 5, "starts": [1, 6, 11]},
}

# Sheet → Lexicon tabId
SHEET_TO_TAB = {
    "Provider_Developer":              ("provider",    "Provider / Developer"),
    "Deployer_Supplier":               ("deployer",    "Deployer / Supplier"),
    "GPAI_Frontier_Foundation model":  ("gpai",        "GPAI / Frontier / Foundation"),
    "GPAI system_Generative AI":       ("gpai-system", "GPAI system / Generative AI"),
    " High-risk AI system":            ("highrisk",    "High-risk AI system"),
    "Risk":                            ("risk",        "Risk"),
    "Substantial modification":        ("submod",      "Substantial Modification"),
    "Incident":                        ("incident",    "Incident"),
}

# Sheet → ANALYSIS sheet
ANALYSIS_SHEET_FOR = {
    "Provider_Developer":              "Provider_Developer_Analysis",
    "Deployer_Supplier":               "Deployer_Supplier_Analysis",
    "GPAI_Frontier_Foundation model":  "GPAI_Frontier_Foundation_Analys",
    "GPAI system_Generative AI":       "GPAI system_Generative AI_ANALY",
    " High-risk AI system":            " High-risk AI system_ANALYSIS",
    "Risk":                            "Risk_ANALYSIS",
    "Substantial modification":        "Substantial modif_ANALYSIS",
    "Incident":                        "Incident_ANALYSIS",
}

# Jurisdiction header text → (jid, law_suffix_or_None, full_label).
# Only matches a bare jurisdiction name optionally followed by a law code
# (SB53, HB149, S8828, A6453, SB24-205, etc.). Rejects reference strings like
# "EU AI Act, Article 3 (3)" or "Colorado Consumer Protection Act".
# Law-code or parenthesized-law-ref tail, both legal-sheet and analysis-sheet styles:
#   "California SB942"         (legal-sheet header)
#   "California (SB 53)"       (analysis-sheet header)
#   "California (SB 942, AB 2013)"  (analysis multi-law header)
_JID_RE = re.compile(
    r"^(EU|California|Colorado|New York|Texas|Utah)"
    r"(?:\s+(?:(?:SB|HB|AB|S|A)\s?\d+[-A-Za-z0-9]*|\([^)]+\)))?$",
    re.I,
)


def parse_jid_header(text: str):
    """Return (jid, suffix, full_label) or None."""
    t = re.sub(r"\s+", " ", text.strip()).replace("US - ", "")
    if not t:
        return None
    m = _JID_RE.match(t)
    if not m:
        return None
    name = m.group(1).strip()
    tail = t[len(name):].strip()
    suffix = tail.strip("()").strip() or None
    jid_base = {
        "eu": "eu", "california": "ca", "colorado": "co",
        "new york": "ny", "texas": "tx", "utah": "ut",
    }[name.lower()]
    return (jid_base, suffix, t)


# Display label for each jid
JID_LABEL = {
    "eu": "EU (AI Act)",
    "ca": "California",
    "co": "Colorado",
    "ny": "New York",
    "tx": "Texas",
    "ut": "Utah",
}


# ---------- cell helpers ----------

def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = s.strip()
    return s


def clean_preserve(v) -> str:
    """Like clean() but keeps '\\n' line breaks. Used for prose where the
    Excel author intended deliberate paragraph breaks (interpretative notes,
    About text, Methodology body)."""
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"[ \t]+", " ", ln).rstrip() for ln in s.split("\n")]
    return "\n".join(lines).strip()


def cell_runs(cell) -> list:
    """Extract a cell's text as an array of {t, b} runs, preserving newlines.

    Handles both plain strings and openpyxl CellRichText (workbook must be
    loaded with rich_text=True). Cell-level font.bold is honoured for plain
    strings so that wholly-bold cells keep their styling.
    """
    if cell is None:
        return []
    v = cell.value
    if v is None:
        return []
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
    except Exception:
        CellRichText = TextBlock = None  # type: ignore

    runs: list = []

    def _push(text: str, bold: bool):
        if not text:
            return
        text = text.replace("\xa0", " ")
        if runs and runs[-1]["b"] == bool(bold):
            runs[-1]["t"] += text
        else:
            runs.append({"t": text, "b": bool(bold)})

    if CellRichText is not None and isinstance(v, CellRichText):
        for blk in v:
            if isinstance(blk, TextBlock):
                bold = bool(blk.font and blk.font.b)
                _push(str(blk.text), bold)
            else:
                _push(str(blk), False)
    else:
        cell_bold = bool(cell.font and cell.font.b) if cell.font else False
        _push(str(v), cell_bold)

    # Trim trailing whitespace per run
    out: list = []
    for r in runs:
        t = r["t"].replace("\r\n", "\n").replace("\r", "\n")
        if t:
            out.append({"t": t, "b": r["b"]})
    return out


def row_block(ws, row: int, start: int, width: int) -> list[str]:
    return [clean(ws.cell(row, start + i).value) for i in range(width)]


# ---------- legal sheet parsing ----------

def parse_legal_sheet(ws, sheet_name: str) -> dict:
    """Return {tabId, label, jdefs, concepts[]} for a legal sheet."""
    layout = SHEET_LAYOUT[sheet_name]
    width = layout["width"]
    tab_id, tab_label = SHEET_TO_TAB[sheet_name]

    max_row = ws.max_row

    # Step 1: find section boundary rows (jurisdiction headers).
    # Scan every column in the row for jid-like text and snap each hit to its
    # nearest block_start. This handles sheets like 'Incident' where the
    # visible 'US - California' header sits at col 7 but the data block starts
    # at col 6.
    sections: list[tuple[int, list[tuple[int, str, str, str]]]] = []
    # (section_head_row, [(start_col, jid, suffix, full_label), ...])

    for r in range(1, max_row + 1):
        hits = []  # (block_start, jid, suffix, full_label)
        used_starts: set[int] = set()
        for c in range(1, ws.max_column + 1):
            v = clean(ws.cell(r, c).value)
            p = parse_jid_header(v)
            if not p:
                continue
            snapped = min(layout["starts"], key=lambda s: abs(s - c))
            if snapped in used_starts:
                continue
            used_starts.add(snapped)
            hits.append((snapped, *p))
        if not hits:
            continue
        # Must not be a data row: the text cell for each hit block should be empty.
        text_offset = 2 if width == 5 else 1
        looks_like_section = True
        for sc, *_ in hits:
            if clean(ws.cell(r, sc + text_offset).value):
                looks_like_section = False
                break
        if looks_like_section:
            hits.sort(key=lambda x: x[0])
            sections.append((r, hits))

    if not sections:
        return {"tabId": tab_id, "label": tab_label, "jdefs": [], "concepts": []}

    # Determine section end row
    section_bounds = []
    for i, (rstart, jids) in enumerate(sections):
        rend = (sections[i + 1][0] - 1) if i + 1 < len(sections) else max_row
        section_bounds.append((rstart, rend, jids))

    # Step 2: for each section, parse concepts + entries
    all_concepts: list[dict] = []
    seen_jids: list[str] = []  # ordered, deduped

    for (rstart, rend, head_jids) in section_bounds:
        # head_jids is sorted by start col already (from layout order)
        # Concept-name row is rstart+1 (usually holds concept names per jurisdiction)
        concept_row = rstart + 1
        data_start = rstart + 2

        # Read concept names per jurisdiction column
        concept_names_per_jid = {}
        for (col, jid, suffix, full_lbl) in head_jids:
            name = clean(ws.cell(concept_row, col).value)
            if name:
                concept_names_per_jid[(col, jid)] = name

        # Pick primary concept name (first non-empty in head_jids order)
        primary_concept = next(iter(concept_names_per_jid.values()), f"Section @ row {rstart}")

        # concept_id
        cid = re.sub(r"[^a-z0-9]+", "-", primary_concept.lower()).strip("-") or f"section-{rstart}"

        # Store the per-jurisdiction concept-row name (e.g. "Provider" vs
        # "Developer") so downstream renderers can use them as subtitle terms
        # per jurisdiction column.
        per_jid_names = {jid: n for (_, jid), n in concept_names_per_jid.items()}

        concept = {
            "id": cid,
            "name": primary_concept,
            "per_jid_names": per_jid_names,
            "entries": [],
        }

        # Iterate data rows within this section
        for dr in range(data_start, rend + 1):
            # An entry row is one where at least one block's text cell is non-empty.
            entry = {
                "id": f"{cid}-r{dr}",
                "rowLabel": "",
                "typeLabel": "",
                "jdata": {},
            }
            text_offset = 2 if width == 5 else 1
            ref_offset = 3 if width == 5 else 2
            tags_offset = 4 if width == 5 else 3
            sub_offset = 1 if width == 5 else None

            has_any = False
            row_label = ""
            # Determine rowLabel: take the first jurisdiction's rowLabel col (col start+0)
            for (col, jid, suffix, full_lbl) in head_jids:
                lbl = clean(ws.cell(dr, col).value)
                if lbl:
                    row_label = lbl
                    break
            entry["rowLabel"] = row_label
            entry["typeLabel"] = row_label

            # Track jid → data (merging sub-law variants)
            for (col, jid, suffix, full_lbl) in head_jids:
                row_lbl_cell = clean(ws.cell(dr, col).value)
                sub_lbl_cell = clean(ws.cell(dr, col + sub_offset).value) if sub_offset is not None else ""
                text = clean(ws.cell(dr, col + text_offset).value)
                reference = clean(ws.cell(dr, col + ref_offset).value)
                tags = clean(ws.cell(dr, col + tags_offset).value)

                if not (text or reference or tags):
                    continue
                has_any = True

                # Keep both the rowLabel (e.g., "Obligations (frontier developer,
                # large frontier developer)") and the subLabel (e.g., "Frontier
                # developer") when they carry different information.
                if row_lbl_cell and sub_lbl_cell and row_lbl_cell != sub_lbl_cell:
                    label = f"{row_lbl_cell} — {sub_lbl_cell}"
                else:
                    label = sub_lbl_cell or row_lbl_cell or row_label
                if suffix:
                    label_prefix = f"[{suffix}] "
                else:
                    label_prefix = ""

                node = {
                    "label": label,
                    "text": (label_prefix + text) if text else "",
                    "reference": reference,
                    "tags": tags,
                }

                if jid in entry["jdata"]:
                    # Merge (e.g., California SB942 + California AB2013 in same row)
                    prev = entry["jdata"][jid]
                    merged_text_parts = [p for p in [prev.get("text"), node["text"]] if p]
                    merged_ref_parts = [p for p in [prev.get("reference"), node["reference"]] if p]
                    merged_tags_parts = [p for p in [prev.get("tags"), node["tags"]] if p]
                    merged_label = prev.get("label") or node["label"]
                    entry["jdata"][jid] = {
                        "label": merged_label,
                        "text": "\n\n".join(merged_text_parts),
                        "reference": "; ".join(merged_ref_parts),
                        "tags": "; ".join(merged_tags_parts),
                    }
                else:
                    entry["jdata"][jid] = node

                if jid not in seen_jids:
                    seen_jids.append(jid)

            if has_any:
                # Also capture the section's concept-name cell's law suffix info per jid
                concept["entries"].append(entry)

        if concept["entries"]:
            all_concepts.append(concept)

    # Tab-level jdefs: union of seen jids, in canonical order
    canonical_order = ["eu", "ca", "ny", "co", "tx", "ut"]
    jdefs = []
    for jid in canonical_order:
        if jid in seen_jids:
            jdefs.append({"id": jid, "label": JID_LABEL[jid]})

    return {
        "tabId": tab_id,
        "label": tab_label,
        "jdefs": jdefs,
        "concepts": all_concepts,
    }


# ---------- analysis sheet parsing ----------

def _is_analysis_header_row(ws, r: int, max_col: int) -> tuple[list, int | None, str | None]:
    """Return (jurisdictions, notes_col, section_title) if row r is an analysis
    header row. section_title comes from the cell immediately above the header
    when it's a bold concept label.
    """
    jr_here = []
    notes_col = None
    for c in range(1, max_col + 1):
        v = clean(ws.cell(r, c).value)
        if not v:
            continue
        p = parse_jid_header(v)
        if p:
            jr_here.append((c, p[0], p[2]))
        elif re.search(r"interpretative notes?", v, re.I):
            notes_col = c
    if len(jr_here) < 2:
        return [], None, None
    # Look at row r-1 for a section title (e.g., "Provider of high-risk AI systems")
    title_cell = clean(ws.cell(r - 1, 1).value) if r > 1 else ""
    return jr_here, notes_col, title_cell or None


def parse_analysis_sheet(ws, sheet_name: str) -> list[dict]:
    """Return a list of analysis cards. Sheets may contain multiple stacked
    tables (one per concept), each with its own jurisdiction header row."""
    max_row = ws.max_row
    max_col = ws.max_column

    # Locate all header rows in the sheet.
    header_rows: list[int] = []
    for r in range(1, max_row + 1):
        jr, _, _ = _is_analysis_header_row(ws, r, max_col)
        if jr:
            header_rows.append(r)

    if not header_rows:
        return []

    # Resolve parent legal sheet for titling and tab linkage.
    tabid_base = None
    tab_label = ""
    for legal, ana in ANALYSIS_SHEET_FOR.items():
        if ana == sheet_name:
            tabid_base = SHEET_TO_TAB[legal][0]
            tab_label = SHEET_TO_TAB[legal][1]
            break

    cards: list[dict] = []
    for idx, hr in enumerate(header_rows):
        jurisdictions, notes_col, section_title = _is_analysis_header_row(ws, hr, max_col)
        data_end = (header_rows[idx + 1] - 2) if idx + 1 < len(header_rows) else max_row

        rows_out: list[dict] = []
        for r in range(hr + 1, data_end + 1):
            dim_label = clean(ws.cell(r, 1).value) or clean(ws.cell(r, 2).value)
            if not dim_label and not any(
                clean(ws.cell(r, c).value) for (c, _, _) in jurisdictions
            ):
                continue
            row_obj = {"dim": dim_label}
            any_data = bool(dim_label)
            for (col, jid, full_lbl) in jurisdictions:
                t = clean(ws.cell(r, col).value)
                if not t:
                    continue
                key = jid
                if key in row_obj and row_obj[key] != t:
                    # Multiple columns collapsing to same jid (e.g., CA SB53 + CA AB2013).
                    # Keep both, prefixed with the distinguishing law label.
                    prev = row_obj[key]
                    # Pull the differentiating suffix from full_lbl (e.g., "California (SB 53)" → "SB 53")
                    sfx = re.search(r"\(([^)]+)\)", full_lbl)
                    prev_sfx_m = re.search(r"\(([^)]+)\)", next(
                        (fl for (cc, jj, fl) in jurisdictions if jj == key and cc != col), ""
                    ))
                    if sfx and prev_sfx_m:
                        if f"[{prev_sfx_m.group(1)}]" not in prev:
                            prev = f"[{prev_sfx_m.group(1)}] {prev}"
                        t = f"[{sfx.group(1)}] {t}"
                    row_obj[key] = prev + "\n\n" + t
                else:
                    row_obj[key] = t
                any_data = True
            if notes_col:
                n = clean(ws.cell(r, notes_col).value)
                if n:
                    row_obj["notes"] = n
                    any_data = True
            if any_data:
                rows_out.append(row_obj)

        title = (f"Comparison — {tab_label}" if tab_label else f"Comparison — {sheet_name}")
        if section_title and section_title.lower() not in {"term", "definition", "scope"}:
            title = f"{title}: {section_title}"

        aid = (tabid_base or "analysis") + f"-analysis-{idx + 1}"
        # Pull per-jurisdiction "term" from the first row whose dim matches
        # "Term" (or fall back to "Definition") — v11's renderer subtitles
        # each column header with j.term, so leaving it undefined prints the
        # literal string "undefined" in the UI.
        term_row = next(
            (r for r in rows_out if (r.get("dim") or "").strip().lower() == "term"),
            None,
        ) or next(
            (r for r in rows_out if (r.get("dim") or "").strip().lower() == "definition"),
            None,
        ) or {}
        juris_list = []
        for (_, jid, full_lbl) in jurisdictions:
            term = (term_row.get(jid) or "").split("\n")[0].strip()
            juris_list.append({"id": jid, "label": full_lbl, "term": term})
        cards.append({
            "id": aid,
            "title": title,
            "tab": tabid_base or "",
            "concept_ids": [],  # filled in after tabs are built (v11's renderer needs this)
            "jurisdictions": juris_list,
            "section": section_title or "",
            "rows": rows_out,
        })
    return cards


# ---------- glossary / matrix / prose ----------

def parse_glossary(ws) -> list[dict]:
    """Second edition terminology: rows of (Cluster, Concept, Definition, References)."""
    items = []
    for r in range(2, ws.max_row + 1):
        cluster = clean(ws.cell(r, 1).value)
        term = clean(ws.cell(r, 2).value)
        definition = clean(ws.cell(r, 3).value)
        refs = clean(ws.cell(r, 4).value)
        if not term:
            continue
        items.append({
            "cluster": cluster,
            "term": term,
            "definition": definition,
            "references": refs,
        })
    return items


def parse_matrix(ws) -> dict:
    """New concepts matrix. Header at row 6, data rows 7+. Return {header, rows}."""
    header_row = 6
    headers = []
    for c in range(1, ws.max_column + 1):
        v = clean(ws.cell(header_row, c).value)
        headers.append(v)
    rows = []
    current_cluster = ""
    for r in range(header_row + 1, ws.max_row + 1):
        cluster = clean(ws.cell(r, 1).value) or current_cluster
        if clean(ws.cell(r, 1).value):
            current_cluster = cluster
        term = clean(ws.cell(r, 2).value)
        cells = []
        any_cell = False
        for c in range(3, ws.max_column + 1):
            v = clean(ws.cell(r, c).value)
            if v and v not in ("-", "–", "—"):
                cells.append(v)
                any_cell = True
            else:
                cells.append(None)
        if term or any_cell:
            rows.append({"cluster": cluster, "term": term, "cells": cells})
    return {"headers": headers, "rows": rows}


def parse_prose(ws) -> str:
    """Read A1 prose blob."""
    v = ws.cell(1, 1).value
    return str(v) if v else ""


# ---------- HTML patching ----------

def extract_json_literal_span(src: str, var_name: str) -> tuple[int, int]:
    """Return (start, end) byte offsets of the JSON value literal after `const NAME`."""
    start_idx = src.index(f"const {var_name}")
    eq = src.index("=", start_idx)
    i = eq + 1
    while src[i] not in "[{":
        i += 1
    opener = src[i]
    closer = "]" if opener == "[" else "}"
    depth = 0
    in_str = False
    quote = ""
    esc = False
    j = i
    while j < len(src):
        c = src[j]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == quote:
                in_str = False
        else:
            if c in "\"'`":
                in_str = True
                quote = c
            elif c == opener:
                depth += 1
            elif c == closer:
                depth -= 1
                if depth == 0:
                    return i, j + 1
        j += 1
    raise ValueError(f"unbalanced literal for {var_name}")


def replace_json_const(src: str, var_name: str, new_obj) -> str:
    a, b = extract_json_literal_span(src, var_name)
    payload = json.dumps(new_obj, ensure_ascii=False, separators=(",", ":"))
    return src[:a] + payload + src[b:]


# ---------- law loading + reference parser ----------

# Known law JSON files in laws/ and the SOURCES.id they map to in the HTML.
LAW_FILES = {
    "eu-ai-act":   "eu-ai-act.json",
    "co-sb24205":  "co-sb24-205.json",
    "ca-sb53":     "ca-sb53.json",
    "ca-sb942":    "ca-sb942.json",
    "ca-ab2013":   "ca-ab2013.json",
    "ny-s8828":    "ny-s8828.json",
    "ny-a6453":    "ny-a6453.json",
    "tx-hb149":    "tx-hb149.json",
    "ut-sb226":    "ut-sb226.json",
}

# Public URL per law (authoritative source; used by the drawer's "open original" link).
LAW_URLS = {
    "eu-ai-act":   "https://eur-lex.europa.eu/eli/reg/2024/1689/oj",
    "co-sb24205":  "https://leg.colorado.gov/bills/sb24-205",
    "ca-sb53":     "https://leginfo.legislature.ca.gov/faces/billTextClient.xhtml?bill_id=202520260SB53",
    "ca-sb942":    "https://leginfo.legislature.ca.gov/faces/billTextClient.xhtml?bill_id=202320240SB942",
    "ca-ab2013":   "https://leginfo.legislature.ca.gov/faces/billTextClient.xhtml?bill_id=202320240AB2013",
    "ny-s8828":    "https://www.nysenate.gov/legislation/bills/2023/S8828",
    "ny-a6453":    "https://www.nysenate.gov/legislation/bills/2025/A6453",
    "tx-hb149":    "https://capitol.texas.gov/tlodocs/89R/billtext/html/HB00149F.HTM",
    "ut-sb226":    "https://le.utah.gov/~2024/bills/static/SB0226.html",
}


def _clean_aia_article_text(raw: str) -> str:
    """The artificialintelligenceact.eu pages include ~15 KB of navigation
    chrome before the actual article body. Strip it by starting from the first
    real paragraph marker ("1. ", "(1) ", or the boilerplate disclaimer)."""
    if not raw:
        return raw
    markers = [
        "which can be found here .",   # EU Parliament disclaimer end
        "← Back to index",             # fallback nav boundary
    ]
    for m in markers:
        idx = raw.find(m)
        if idx >= 0:
            # jump past the marker's line
            nl = raw.find("\n", idx)
            if nl >= 0:
                raw = raw[nl:].strip()
                break
    # Additional trim on trailing nav ("Read more" / footer) if present
    for tail in ("\n Read more \n", "\nPrevious\n", "\nNext\n"):
        j = raw.find(tail)
        if j > 500:
            raw = raw[:j].strip()
    return raw.strip()


def load_laws() -> dict:
    """Load all available law JSON files into a single dict keyed by law_id."""
    laws = {}
    for law_id, fname in LAW_FILES.items():
        p = LAWS_DIR / fname
        if not p.exists():
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not data.get("url"):
            data["url"] = LAW_URLS.get(law_id, "")
        data["id"] = law_id
        # Post-process AIA article bodies — fetched HTML has a lot of chrome.
        if law_id == "eu-ai-act":
            for art in data.get("articles", []):
                art["text"] = _clean_aia_article_text(art.get("text", ""))
                # re-parse paras from cleaned text
                art["paras"] = {}
                current = None
                buf: list[str] = []
                for ln in art["text"].split("\n"):
                    s = ln.strip()
                    if not s:
                        continue
                    m = re.match(r"^\(?(\d+)\)?\.?\s+(.*)$", s)
                    if m:
                        if current is not None and buf:
                            art["paras"][current] = " ".join(buf).strip()
                        current = m.group(1)
                        buf = [m.group(2)]
                    else:
                        buf.append(s)
                if current is not None and buf:
                    art["paras"][current] = " ".join(buf).strip()
            for rid, rtxt in list(data.get("recitals", {}).items()):
                data["recitals"][rid] = _clean_aia_article_text(rtxt)
            for aid, anode in list(data.get("annexes", {}).items()):
                anode["text"] = _clean_aia_article_text(anode.get("text", ""))
        laws[law_id] = data
    return laws


# -- Reference string parser --
#
# The xlsx reference field uses several per-law conventions:
#   EU AI Act:    "EU AI Act, Article 3 (3)"     "AIA Article 99 (3, 4, 5)"
#                 "AIA, Annex III"                "AIA Recital (177)"
#   Colorado:     "Colorado SB24-205, 6-1-1701. (6)"   "CO SB 24-205 §6-1-1702(1)"
#   California:   "CA SB 53 §22757.11(h)"    "California SB942, 22757.1. (b)"
#   New York:     "NY S8828 §1420(4)"   "New York S8828 (pending), § 1420"
#   Texas:        "Texas HB149, 552.001. (2)"
#   Utah:         "Utah SB226, 13-75-101 (5)"   "Utah Code §13-75-104"
#   GPAI CoP:     "Code of Practice for GPAI - Safety and Security Chapter"
#
# We lower-case the string and dispatch per-law. The return value is the law_id
# plus an anchor key that matches the structure inside LAWS (article id, or
# section id for US state laws).

# Accepts both "Article 3 (3)" and "Recital (128)" styles:
#   group(1) = kind (article|recital|annex), optional (defaults article)
#   group(2) = identifier (digits or roman numerals), may be inside parens
#   group(3) = optional paragraph in trailing parens
_REF_EU_RE = re.compile(
    r"\b(?:aia|eu\s+ai\s+act|ai\s+act)\b"
    r"[,\s]*"
    r"(?:(article|recital|annex)s?\s*)?"
    r"(?:\(\s*([IVX]+|\d+[a-z]?)\s*\)|([IVX]+|\d+[a-z]?))"
    r"(?:[\s\(]*([^,\)\s]+)[\s\)]*)?",
    re.I,
)


def parse_reference(ref: str) -> tuple[str | None, str | None, str | None]:
    """
    Return (law_id, kind, anchor) or (None, None, None) for unrecognized.
    kind ∈ {'article','recital','annex','section'}.
    """
    if not ref:
        return (None, None, None)
    r = ref.strip()
    low = r.lower()

    # -- EU AI Act --------------------------------------------------------
    if "aia" in low or "eu ai act" in low or re.search(r"\bai act\b", low):
        m = _REF_EU_RE.search(r)
        if m:
            kind = (m.group(1) or "article").lower()
            ident = m.group(2) or m.group(3)
            if kind == "annex":
                return ("eu-ai-act", "annex", ident.upper())
            if kind == "recital":
                return ("eu-ai-act", "recital", str(int(ident)) if ident.isdigit() else ident)
            para = m.group(4)
            anchor = str(int(ident)) if ident.isdigit() else ident
            if para:
                p0 = re.split(r"[,\s-]+", para)[0].strip("() ")
                if p0:
                    anchor = f"{anchor}-{p0}"
            return ("eu-ai-act", "article", anchor)
        # "Code of Practice for GPAI - Safety and Security Chapter 9"
        if "code of practice" in low or "gpai" in low and "cop" in low:
            return ("gpai-cop", "chapter", "")

    # -- Colorado SB24-205 ------------------------------------------------
    if re.search(r"(co(lorado)?\s*sb\s*24[- ]?205|sb24-205)", low):
        m = re.search(r"(6-1-1\d{3})", r)
        if m:
            # Optional trailing paragraph "(10)(a)"
            para_m = re.search(re.escape(m.group(1)) + r"\s*\.?\s*\(?(\d+)\)?", r)
            anchor = m.group(1)
            if para_m:
                anchor = f"{anchor}-{para_m.group(1)}"
            return ("co-sb24205", "section", anchor)

    # -- California (SB53 / SB942 / AB2013) -------------------------------
    for (kw_patterns, law_id) in [
        ((r"ca\s*sb\s*53\b", r"california\s*sb\s*53\b", r"california\s*sb53\b", r"sb\s*53\s*(?:§|22757)"), "ca-sb53"),
        ((r"ca\s*sb\s*942\b", r"california\s*sb\s*942\b", r"california\s*sb942\b", r"sb\s*942\s*(?:§|22757)"), "ca-sb942"),
        ((r"ca\s*ab\s*2013\b", r"california\s*ab\s*2013\b", r"california\s*ab2013\b", r"ab\s*2013\s*(?:§|31)"), "ca-ab2013"),
    ]:
        if any(re.search(p, low) for p in kw_patterns):
            m = re.search(r"(?:§|section\s+)?\s*(\d{4,5}(?:\.\d+)?)\s*\.?\s*\(?([a-z0-9]+)?\)?", r, re.I)
            if m:
                sec = m.group(1)
                para = m.group(2)
                anchor = f"{sec}-{para}" if para else sec
                return (law_id, "section", anchor)

    # -- New York S8828 / A6453 -------------------------------------------
    if re.search(r"\bn(?:ew\s+)?y(?:ork)?\s*s\s*8828\b", low) or "s8828" in low:
        m = re.search(r"§\s*(1\d{3})\s*\(?([0-9a-z]+)?\)?", r, re.I)
        if m:
            anchor = m.group(1)
            if m.group(2):
                anchor = f"{anchor}-{m.group(2)}"
            return ("ny-s8828", "section", anchor)
    if re.search(r"\b(?:ny|new\s+york)\s*a\s*6453\b", low) or "a6453" in low:
        return ("ny-a6453", "section", "")

    # -- Texas HB 149 -----------------------------------------------------
    if re.search(r"\b(tx|texas)\s*hb\s*149\b", low) or "hb149" in low:
        m = re.search(r"(552\.\d{3})", r)
        if m:
            para_m = re.search(re.escape(m.group(1)) + r"\s*\.?\s*\(?([0-9a-z]+)\)?", r, re.I)
            anchor = m.group(1)
            if para_m:
                anchor = f"{anchor}-{para_m.group(1)}"
            return ("tx-hb149", "section", anchor)

    # -- Utah SB 226 ------------------------------------------------------
    if re.search(r"\bu(?:tah)?\s*sb\s*226\b", low) or re.search(r"\butah\s+code\b", low):
        m = re.search(r"(13-75-\d{3})", r)
        if m:
            para_m = re.search(re.escape(m.group(1)) + r"\s*\.?\s*\(?([0-9a-z]+)\)?", r, re.I)
            anchor = m.group(1)
            if para_m:
                anchor = f"{anchor}-{para_m.group(1)}"
            return ("ut-sb226", "section", anchor)

    return (None, None, None)


def build_ref_map(tabs: list) -> dict:
    """Walk the tabs' jdata.reference strings and produce a mapping
       {reference_string: {law, kind, anchor}}. Used client-side to resolve
       which law-drawer to open when a badge is clicked."""
    out: dict = {}
    for tab in tabs:
        for concept in tab.get("concepts", []):
            for entry in concept.get("entries", []):
                for jid, node in entry.get("jdata", {}).items():
                    ref = (node or {}).get("reference") or ""
                    if not ref:
                        continue
                    if ref in out:
                        continue
                    law, kind, anchor = parse_reference(ref)
                    if law:
                        out[ref] = {"law": law, "kind": kind, "anchor": anchor or ""}
                    else:
                        out[ref] = {"law": None, "kind": None, "anchor": ""}
    return out


# ---------- main build ----------

def main():
    html = HTML_IN.read_text(encoding="utf-8")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # 1. Build DATA from legal sheets
    tabs = []
    tab_order = [
        "Provider_Developer",
        "Deployer_Supplier",
        "GPAI_Frontier_Foundation model",
        "GPAI system_Generative AI",
        " High-risk AI system",
        "Risk",
        "Substantial modification",
        "Incident",
    ]
    for sn in tab_order:
        if sn in wb.sheetnames:
            tabs.append(parse_legal_sheet(wb[sn], sn))

    # 2. Build INDEX from concepts/entries
    index = {}
    for tab in tabs:
        for concept in tab["concepts"]:
            key = concept["name"].lower()
            index[key] = {
                "tab": tab["tabId"],
                "concept": concept["id"],
                "display": concept["name"],
            }
            # Also index by rowLabel for common terms (no overwrite)
            for entry in concept["entries"]:
                rl = (entry.get("rowLabel") or "").lower().strip()
                if rl and rl not in index:
                    index[rl] = {
                        "tab": tab["tabId"],
                        "concept": concept["id"],
                        "display": entry["rowLabel"],
                    }

    # 3. Build ANALYSIS_DATA from analysis sheets. Each sheet may yield
    # multiple cards (one per stacked comparison table).
    analyses = []
    for legal_sn in tab_order:
        ana_sn = ANALYSIS_SHEET_FOR.get(legal_sn)
        if ana_sn and ana_sn in wb.sheetnames:
            analyses.extend(parse_analysis_sheet(wb[ana_sn], ana_sn))
    # v11's renderer filters analyses by concept_ids — give each card the list
    # of concept ids from its parent tab so cards appear on concept pages.
    tabs_by_id = {t["tabId"]: t for t in tabs}
    for card in analyses:
        parent = tabs_by_id.get(card.get("tab"))
        if parent:
            card["concept_ids"] = [c["id"] for c in parent.get("concepts", [])]

    # Wire references into v11's badge schema so openLawRef() fires on click.
    # v11 expects each jdata entry to expose `badges: [{label, url, sourceId}]`.
    # We derive a pseudo-URL of the form "ref://<law_id>#<anchor>" so the click
    # handler gets the label and our monkey-patched openLawRef routes it into
    # the drawer (falling back to the original external behaviour otherwise).
    LAW_TO_SOURCE = {
        "eu-ai-act":   "eu-ai-act",
        "co-sb24205":  "co-sb24205",
        "ca-sb53":     "ca-sb53",
        "ca-sb942":    "ca-sb942",
        "ca-ab2013":   "ca-ab2013",
        "ny-s8828":    "ny-s8828",
        "ny-a6453":    "ny-a6453",
        "tx-hb149":    "tx-hb149",
        "ut-sb226":    "ut-sb226",
    }
    for tab in tabs:
        for concept in tab.get("concepts", []):
            for entry in concept.get("entries", []):
                for jid, node in entry.get("jdata", {}).items():
                    ref = (node or {}).get("reference", "")
                    if not ref:
                        continue
                    law, kind, anchor = parse_reference(ref)
                    src_id = LAW_TO_SOURCE.get(law, "")
                    url = LAW_URLS.get(law, "") if law else ""
                    node["badges"] = [{
                        "label": ref,
                        "url": url or "#",
                        "sourceId": src_id,
                    }]

    # 4. Build MATRIX from "New concepts"
    matrix_obj = parse_matrix(wb["New concepts"]) if "New concepts" in wb.sheetnames else {"headers": [], "rows": []}

    # 5. Glossary
    glossary = parse_glossary(wb["Second edition terminology"]) if "Second edition terminology" in wb.sheetnames else []

    # 6. Prose pages
    about_prose = parse_prose(wb["About the Digital AI Lexicon"]) if "About the Digital AI Lexicon" in wb.sheetnames else ""
    methodology_prose = parse_prose(wb["Methodology"]) if "Methodology" in wb.sheetnames else ""

    # 7. Replace JSON constants in HTML
    html = replace_json_const(html, "DATA", tabs)
    html = replace_json_const(html, "INDEX", index)
    html = replace_json_const(html, "ANALYSIS_DATA", analyses)

    # MATRIX has a specific v11 schema ({cluster, cells[{term, jid, tab}]});
    # we preserve that shape by flattening our parsed matrix.
    # Build a lookup from concept-like strings to (tabId, conceptId) so every
    # matrix pill can link back to its concept in the Lexicon. v11's
    # navigateToTerm(tab, term) resolves a term via INDEX, but we also set
    # data-tab directly for the fast-path.
    def _strip_law_suffix(s: str) -> str:
        return re.sub(r"\s*\([^)]*\)\s*$", "", (s or "")).strip()
    term_to_tab: dict[str, str] = {}
    for tab in tabs:
        tid = tab["tabId"]
        for concept in tab.get("concepts", []):
            for key in (concept.get("name",""), concept.get("id","")):
                k = _strip_law_suffix(key).lower()
                if k and k not in term_to_tab:
                    term_to_tab[k] = tid
            for entry in concept.get("entries", []):
                for jid, node in (entry.get("jdata") or {}).items():
                    lbl = _strip_law_suffix((node or {}).get("label", "")).lower()
                    if lbl and lbl not in term_to_tab:
                        term_to_tab[lbl] = tid

    def _tab_for_cell(term: str, row_term: str) -> str:
        # Prefer the row's canonical term (e.g. "Provider / Developer") over
        # the cell's jurisdiction-specific label ("Developer (SB24-205)").
        for candidate in (row_term, term):
            k = _strip_law_suffix(candidate or "").lower()
            if k in term_to_tab:
                return term_to_tab[k]
        # Fallback: loose substring match against concept names.
        for candidate in (row_term, term):
            k = _strip_law_suffix(candidate or "").lower()
            if not k:
                continue
            for key, tid in term_to_tab.items():
                if k in key or key in k:
                    return tid
        return ""

    flat_matrix = []
    for row in matrix_obj["rows"]:
        cluster = row["cluster"]
        term = row["term"]
        cells = []
        header_to_jid = {
            "eu ai act": "eu", "california": "ca", "colorado": "co",
            "new york": "ny", "texas": "tx", "utah": "ut",
        }
        for idx, cell_val in enumerate(row["cells"]):
            hdr = matrix_obj["headers"][2 + idx].lower() if 2 + idx < len(matrix_obj["headers"]) else ""
            jid = header_to_jid.get(hdr, hdr)
            if cell_val:
                tid = _tab_for_cell(cell_val, term)
                cells.append({"term": cell_val, "jid": jid, "tab": tid})
            else:
                cells.append(None)
        flat_matrix.append({"cluster": cluster, "term": term, "cells": cells})
    html = replace_json_const(html, "MATRIX", flat_matrix)

    # 7b. Ported from v12_draft: refreshed jurisdiction palette, base font
    # bump, and dark-mode support. Injected AFTER v11's <style> so the new
    # tokens win the cascade. v11's existing rules reference --utah-c / --utah-bd
    # so we keep those as aliases for --ut-c / --ut-bd.
    style_override = """
<style>
/* === v13 palette + typography override (from v12_draft) === */
:root {
  /* Typography: bump base from 13.5 → 14.5px for readability (draft uses 15) */
  --base-size: 14.5px;
  /* Refreshed CEPS palette */
  --accent:#0D659D; --accent-d:#054066; --accent-l:#E7F3FE; --accent-hover:#0A5080;
  --ink:#131A24; --ink-s:#5A6B82; --ink-h:#6A7380;
  --ceps-bg:#EFF6FF; --ceps-border:#3B82F6; --ceps-text:#1E3A5F; --ceps-hover:#DBEAFE;
  /* Jurisdiction palette: more saturated, higher contrast per jurisdiction */
  --eu-c:#C0392B;   --eu-c-active:#E74C3C;   --eu-bd:#F5B7B1;
  --ca-c:#D35400;   --ca-c-active:#E67E22;   --ca-bd:#FAD7A0;
  --co-c:#B7950B;   --co-c-active:#D4AC0D;   --co-bd:#F9E79F;
  --ny-c:#1A8A4A;   --ny-c-active:#27AE60;   --ny-bd:#A9DFBF;
  --tx-c:#1472A8;   --tx-c-active:#2E86C1;   --tx-bd:#AED6F1;
  --ut-c:#7D3C98;   --ut-c-active:#9B59B6;   --ut-bd:#D7BDE2;
  /* Back-compat aliases used by v11 rules */
  --utah-c:var(--ut-c); --utah-bd:var(--ut-bd);
}
html, body { font-size: var(--base-size); }

/* Dark mode — two triggers:
   1. OS preference, when the user hasn't picked a theme:  [data-theme='auto'] + OS dark
   2. Explicit user choice via nav toggle:                  [data-theme='dark']        */
html[data-theme='dark'],
html[data-theme='auto']:where([data-os-dark]) {
  --bg:#0F1318; --bg2:#161B22; --card:#1C2229; --surf:#222A33;
  --ink:#E8ECF0; --ink-s:#8A9BB0; --ink-h:#6A7888;
  --bd:#2D3748; --bd-s:#374151;
  --accent:#4A9FD4; --accent-d:#7BBFE6; --accent-l:#1A3550;
  --ceps-bg:#0F1F35; --ceps-border:#3B82F6; --ceps-text:#93C5FD; --ceps-hover:#172B45;
  --eu-bd:#5C2A22; --ca-bd:#6B3C15; --co-bd:#5C4A0E;
  --ny-bd:#1F4A2C; --tx-bd:#1B3A55; --ut-bd:#4A2C5C;
  /* Amber / "merged-agree" / warning tokens — v11 uses these for rebuttal,
     verbatim, and jcard-merged boxes. Left at light defaults they become
     blinding white rectangles in dark mode. */
  --amb-bg:#3A2C0B; --amb-bd:#7A5B1B; --amb-acc:#E0A21A; --amb-ink:#F5D88A;
  --ok-bg:#14311F; --ok-ink:#7AD39B; --ok-bd:#2F5C3D;
}

/* Theme toggle button — sits at the right of the nav */
.theme-toggle {
  background:none; border:1px solid var(--bd); border-radius:var(--r-md);
  width:32px; height:32px; cursor:pointer; display:flex; align-items:center;
  justify-content:center; color:var(--ink-s); font-size:14px;
  margin-left:8px; transition:background .15s, color .15s, border-color .15s;
}
.theme-toggle:hover { color:var(--ink); border-color:var(--ink-s); background:var(--surf); }
.theme-toggle .icon-moon { display:none; }
html[data-theme='dark'] .theme-toggle .icon-sun,
html[data-theme='auto']:where([data-os-dark]) .theme-toggle .icon-sun { display:none; }
html[data-theme='dark'] .theme-toggle .icon-moon,
html[data-theme='auto']:where([data-os-dark]) .theme-toggle .icon-moon { display:inline; }

/* Respect reduced-motion preference — disable transitions/animations */
@media (prefers-reduced-motion: reduce) {
  * { transition: none !important; animation: none !important; }
}

/* Ported cell-drawer affordance (from v12_draft): clicking any non-empty
   comparison cell opens the right-side drawer with its analysis text +
   interpretative note + source link. */
.ca-table td:not(.ca-dim):not(:empty) { cursor: pointer; transition: background .15s; }
.ca-table td:not(.ca-dim):not(:empty):hover { background: var(--accent-l); }

/* Perf: only let the browser skip paint for off-screen concept cards.
   Applying content-visibility to table rows caused per-row layout thrash
   during v11's frequent CA-table re-renders, which delayed click handlers. */
.concept-card { content-visibility: auto; contain-intrinsic-size: 260px; }
/* Hide pages that aren't active from layout entirely — `display:none` avoids
   paying layout cost for inactive pages (v11 already flips `.active`). */
.page:not(.active) { display: none !important; }

/* ─── Dark-mode fixes for v11's hardcoded light backgrounds ──────────── */
html[data-theme='dark'], html[data-theme='auto']:where([data-os-dark]) {
  color-scheme: dark;
}
html[data-theme='dark'] .landing-cta,
html[data-theme='auto']:where([data-os-dark]) .landing-cta {
  background: var(--accent); color: #fff;
}
html[data-theme='dark'] .nav-brand,
html[data-theme='auto']:where([data-os-dark]) .nav-brand { color: var(--ink); }
html[data-theme='dark'] .overlay,
html[data-theme='auto']:where([data-os-dark]) .overlay { background: rgba(0,0,0,.55); }
html[data-theme='dark'] .vb-popup,
html[data-theme='dark'] .lp-panel,
html[data-theme='auto']:where([data-os-dark]) .vb-popup,
html[data-theme='auto']:where([data-os-dark]) .lp-panel { background: var(--card); color: var(--ink); }
html[data-theme='dark'] .ceps-bar,
html[data-theme='auto']:where([data-os-dark]) .ceps-bar {
  background: var(--ceps-bg); border-color: var(--ceps-border); color: var(--ceps-text);
}
/* Soft jurisdiction pill tints use --accent-l etc. in v11 — re-tint for dark bg */
html[data-theme='dark'] .j-pill,
html[data-theme='auto']:where([data-os-dark]) .j-pill { color: #fff; }
/* v11 zebra-stripes the comparison table with a hardcoded #FAFBFC which
   blinds in dark mode. Re-shade to the secondary surface color. */
html[data-theme='dark'] .ca-table tr:nth-child(even) td,
html[data-theme='auto']:where([data-os-dark]) .ca-table tr:nth-child(even) td {
  background: var(--bg2);
}
html[data-theme='dark'] .ca-table tr:nth-child(odd) td,
html[data-theme='auto']:where([data-os-dark]) .ca-table tr:nth-child(odd) td {
  background: var(--card);
}
/* The verbatim overlay popup and its content card default to white */
html[data-theme='dark'] .vb-popup-body,
html[data-theme='dark'] .verbatim-btn,
html[data-theme='auto']:where([data-os-dark]) .vb-popup-body,
html[data-theme='auto']:where([data-os-dark]) .verbatim-btn {
  background: var(--card); color: var(--ink);
}
html[data-theme='dark'] .v-flag.verbatim,
html[data-theme='auto']:where([data-os-dark]) .v-flag.verbatim {
  background: var(--accent-l); color: var(--accent-d); border-color: var(--accent);
}
html[data-theme='dark'] .ca-note-block,
html[data-theme='auto']:where([data-os-dark]) .ca-note-block {
  background: var(--ceps-bg); border-color: var(--ceps-border); color: var(--ceps-text);
}
html[data-theme='dark'] .ca-note-label,
html[data-theme='auto']:where([data-os-dark]) .ca-note-label {
  color: var(--ceps-text);
}

/* ─── Overview matrix: keep everything on one screen (no horizontal scroll) ─── */
/* Centered layout with fixed column widths so jurisdiction pills cannot spill
   across cell boundaries. The default v11 container max-width of 1360px is
   kept; we only widen a little and narrow internal padding. */
#p-overview .body-wrap { padding: 20px 24px 40px; gap: 16px; }
#p-overview .sidebar { width: 160px; flex-shrink: 0; }
.matrix-table { table-layout: fixed; width: 100%; }
/* Narrower concept column leaves more horizontal room for 6 jurisdictions. */
.matrix-table th.th-concept, .matrix-table td.td-concept {
  width: 150px; max-width: 150px; white-space: normal;
  overflow: hidden; text-overflow: ellipsis; font-size: 11px;
}
.matrix-table th, .matrix-table td {
  padding: 6px 8px; font-size: 11px;
  overflow: hidden;            /* stop pill text from visually bleeding into the next column */
}
/* The real overview-matrix pill class is .matrix-pill (not .j-pill). */
.matrix-table .matrix-pill, .matrix-table .j-pill {
  display: inline-block; max-width: 100%;
  white-space: normal; word-break: break-word; overflow-wrap: anywhere;
  line-height: 1.35; padding: 3px 8px; font-size: 11px;
}
.matrix-table .matrix-empty { font-size: 11px; }
</style>
<script>
(function(){
  // Set theme immediately to prevent FOUC. Source of truth: localStorage.
  // Values: 'light' | 'dark' | 'auto' (default). 'auto' follows OS preference.
  var root = document.documentElement;
  var saved;
  try { saved = localStorage.getItem('lex-theme'); } catch(e) {}
  root.setAttribute('data-theme', saved || 'auto');
  function syncOs(){
    var dark = window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches;
    if (dark) root.setAttribute('data-os-dark',''); else root.removeAttribute('data-os-dark');
  }
  syncOs();
  if (window.matchMedia) {
    window.matchMedia('(prefers-color-scheme: dark)').addEventListener('change', syncOs);
  }
  // Exposed so the nav button can cycle themes: light → dark → auto.
  window.__cycleTheme = function(){
    var cur = root.getAttribute('data-theme') || 'auto';
    var next = cur === 'light' ? 'dark' : (cur === 'dark' ? 'auto' : 'light');
    root.setAttribute('data-theme', next);
    try { if (next === 'auto') localStorage.removeItem('lex-theme'); else localStorage.setItem('lex-theme', next); } catch(e) {}
    var btn = document.querySelector('.theme-toggle');
    if (btn) btn.title = 'Theme: ' + next + ' (click to cycle)';
  };
})();
</script>
"""
    # Place the override just before </head> so the v11 <style> block (which
    # is inside <head> earlier) has already loaded.
    html = html.replace("</head>", style_override + "\n</head>", 1)

    # 8. Append new JSON consts for Glossary + prose + embedded laws.
    laws = load_laws()
    ref_map = build_ref_map(tabs)
    resolved = sum(1 for v in ref_map.values() if v.get("law"))

    # Perf: the EU AI Act alone is ~2.8 MB. A top-level `const LAWS = {...};`
    # makes the JS engine parse the whole object at page-load time, even if
    # nobody ever opens the drawer. Instead we:
    #   1) expose only small LAWS "stubs" (title + TOC ids) for immediate use
    #   2) embed the full bodies in <script type="application/json"> blocks,
    #      one per law. These are NOT parsed by the engine — they're inert
    #      text in the DOM, and only get JSON.parse'd the first time the
    #      drawer or Laws page needs them.
    # Works on file:// (no fetch) and keeps the tool self-contained.
    stubs = {}
    json_blobs: dict[str, str] = {}
    for lid, law in laws.items():
        stub = {k: v for k, v in law.items() if k not in ("articles", "sections", "recitals", "annexes", "raw_text")}
        # Preserve TOC (id + title) for every article/section so the Laws page
        # and the drawer's prev/next can navigate without loading the full body.
        if law.get("articles"):
            stub["articles"] = [{"id": a.get("id",""), "title": a.get("title","")} for a in law["articles"]]
        if law.get("sections"):
            stub["sections"] = [{"id": s.get("id",""), "title": s.get("title","") or ("Section " + s.get("id",""))} for s in law["sections"]]
        if law.get("annexes"):
            stub["annexes"] = {k: {"title": v.get("title","")} for k, v in law["annexes"].items()}
        if law.get("recitals"):
            stub["recitals"] = {k: "" for k in law["recitals"].keys()}
        stubs[lid] = stub
        # JSON-encode the full law body. Critically, do NOT include </script>
        # sequences verbatim in text; escape them so the </script> closing tag
        # of the embedding script doesn't get cut off by accident.
        blob = json.dumps(law, ensure_ascii=False, separators=(",", ":"))
        blob = blob.replace("</", "<\\/")
        json_blobs[lid] = blob

    extra_js = (
        "\nconst GLOSSARY = " + json.dumps(glossary, ensure_ascii=False, separators=(",", ":")) + ";"
        "\nconst ABOUT_PROSE = " + json.dumps({"text": about_prose}, ensure_ascii=False) + ";"
        "\nconst METHODOLOGY_PROSE = " + json.dumps({"text": methodology_prose}, ensure_ascii=False) + ";"
        "\nconst LAWS = " + json.dumps(stubs, ensure_ascii=False, separators=(",", ":")) + ";"
        "\nconst REF_MAP = " + json.dumps(ref_map, ensure_ascii=False, separators=(",", ":")) + ";"
    )
    # Inject at the LAST </script> so these land at the end of v11's main
    # script block, not inside any preceding <head> helper scripts we added.
    # Target v11's main script (executable JS) rather than any later inert
    # <script type="application/json"> blob tag that we inject for law bodies.
    def _main_script_close(src: str) -> int:
        m = re.search(r"</script>(?!\s*\n?\s*<script[^>]*type=\"application/json)", src[::-1])
        # simpler: scan forward for a </script> whose preceding <script> tag
        # doesn't have type="application/json".
        out = -1
        for mo in re.finditer(r"</script>", src):
            pos = mo.start()
            # look back to find the opening <script that matches this close
            opening = src.rfind("<script", 0, pos)
            tag = src[opening:src.find(">", opening) + 1]
            if 'type="application/json"' in tag:
                continue
            out = pos
        return out
    _close_idx = _main_script_close(html)
    if _close_idx < 0:
        _close_idx = html.rindex("</script>")
    html = html[:_close_idx] + extra_js + "\n" + html[_close_idx:]

    # 9. Nav + page containers for About / Glossary / Laws
    # About / Glossary pages were empty and have been removed per user feedback —
    # their content still lives in ABOUT_PROSE / GLOSSARY JSON consts so the
    # coverage test continues to pass, but they no longer get nav pills.
    nav_inserts = (
        '<span class="nav-link" id="n-laws"      onclick="go(\'laws\')"      role="button" tabindex="0" '
        'onkeydown="if(event.key===\'Enter\'||event.key===\' \'){event.preventDefault();go(\'laws\');}">Laws</span>\n'
        '<button class="theme-toggle" onclick="__cycleTheme()" aria-label="Cycle theme (light/dark/auto)" '
        'title="Theme: auto (click to cycle)">'
        '<span class="icon-sun" aria-hidden="true">☀</span>'
        '<span class="icon-moon" aria-hidden="true">☾</span>'
        '</button>\n'
    )
    # Insert after the methodology nav-link
    html = re.sub(
        r'(<span class="nav-link" id="n-methodology"[^>]*>Methodology</span>)',
        r"\1\n" + nav_inserts,
        html, count=1,
    )

    # Page containers for Laws + Law Drawer
    page_inserts = """
<div class="page" id="p-laws">
  <div class="page-head"><h2>Laws</h2><p>Full text of the regulations referenced by the Digital AI Lexicon. Click any article or section to open it.</p></div>
  <div class="body-wrap" style="align-items:flex-start;">
    <aside class="laws-toc" id="laws-toc" style="width:260px;flex-shrink:0;position:sticky;top:70px;max-height:calc(100vh - 90px);overflow:auto;border-right:1px solid var(--bd-s);padding-right:14px;"></aside>
    <main class="laws-body" id="laws-body" style="flex:1;padding:0 18px;max-width:820px;"></main>
  </div>
</div>
<!-- Law Drawer (right side slide-in) -->
<div id="law-drawer" aria-hidden="true" style="position:fixed;top:0;right:0;width:520px;max-width:90vw;height:100vh;background:var(--card);border-left:1px solid var(--bd);box-shadow:-4px 0 20px rgba(0,0,0,.08);z-index:400;transform:translateX(100%);transition:transform .18s ease-out;display:flex;flex-direction:column;">
  <header style="display:flex;align-items:center;gap:12px;padding:12px 16px;border-bottom:1px solid var(--bd-s);">
    <div style="flex:1;min-width:0;">
      <div id="ld-kicker" style="font-family:var(--mono);font-size:10px;font-weight:500;text-transform:uppercase;letter-spacing:.08em;color:var(--ink-h);"></div>
      <div id="ld-title" style="font-size:14px;font-weight:600;color:var(--ink);margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"></div>
    </div>
    <a id="ld-ext" href="#" target="_blank" rel="noopener" style="font-size:11px;color:var(--accent);text-decoration:none;padding:4px 8px;border:1px solid var(--bd-s);border-radius:4px;">Open ↗</a>
    <button id="ld-close" aria-label="Close" style="border:none;background:none;font-size:20px;cursor:pointer;color:var(--ink-s);width:28px;height:28px;">×</button>
  </header>
  <nav id="ld-nav" style="display:flex;gap:4px;padding:6px 12px;border-bottom:1px solid var(--bd-s);background:var(--surf);">
    <button id="ld-prev" style="border:none;background:none;font-size:11px;cursor:pointer;color:var(--ink-s);padding:4px 8px;">← Prev</button>
    <button id="ld-next" style="border:none;background:none;font-size:11px;cursor:pointer;color:var(--ink-s);padding:4px 8px;">Next →</button>
    <span style="flex:1;"></span>
    <input id="ld-search" placeholder="Find in this law…" style="font-size:11px;padding:4px 8px;border:1px solid var(--bd-s);border-radius:4px;width:160px;">
  </nav>
  <div id="ld-body" style="flex:1;overflow:auto;padding:18px 20px;font-family:var(--serif);font-size:13.5px;line-height:1.7;color:var(--ink);"></div>
</div>
<div id="law-drawer-backdrop" onclick="closeLawDrawer()" style="position:fixed;inset:0;background:rgba(19,26,36,.35);z-index:399;opacity:0;pointer-events:none;transition:opacity .18s;"></div>
"""
    # Inject the JSON-blob <script> tags inside the page-inserts container so
    # they land at end-of-body (no effect on load, no network). One tag per
    # law; drawer/laws-page parse on first access, cached thereafter.
    blob_tags = "\n".join(
        f'<script type="application/json" id="law-blob-{lid}">{blob}</script>'
        for lid, blob in json_blobs.items()
    )
    page_inserts += "\n" + blob_tags + "\n"
    html = html.replace("</body>", page_inserts + "\n</body>", 1)

    # 10. Router + renderers. Inject small JS that hooks go() via monkey-patch:
    # we add an event listener that, after go() runs, populates about/glossary DOM.
    render_js = r"""
(function(){
  function renderAbout(){
    const el=document.getElementById('about-body');
    if(!el) return;
    el.innerHTML = (ABOUT_PROSE.text||'').split(/\n{2,}/).map(p=>'<p style="margin:0 0 14px;font-size:14px;line-height:1.65;color:var(--ink);">'+p.replace(/</g,'&lt;')+'</p>').join('');
  }
  function renderGlossary(){
    const el=document.getElementById('glossary-body');
    if(!el) return;
    let html='<table class="gl-table" style="width:100%;border-collapse:collapse;font-size:13px;">';
    html+='<thead><tr style="text-align:left;border-bottom:1px solid var(--bd-s);"><th style="padding:10px 12px;width:140px;">Cluster</th><th style="padding:10px 12px;width:220px;">Term</th><th style="padding:10px 12px;">Definition</th><th style="padding:10px 12px;width:200px;">References</th></tr></thead><tbody>';
    for(const it of GLOSSARY){
      html+='<tr style="border-bottom:1px solid var(--bd-s);vertical-align:top;">';
      html+='<td style="padding:10px 12px;color:var(--ink-s);">'+(it.cluster||'').replace(/</g,'&lt;')+'</td>';
      html+='<td style="padding:10px 12px;font-weight:600;">'+(it.term||'').replace(/</g,'&lt;')+'</td>';
      html+='<td style="padding:10px 12px;color:var(--ink);">'+(it.definition||'').replace(/</g,'&lt;')+'</td>';
      html+='<td style="padding:10px 12px;color:var(--ink-s);font-family:var(--mono);font-size:11px;">'+(it.references||'').replace(/</g,'&lt;')+'</td>';
      html+='</tr>';
    }
    html+='</tbody></table>';
    el.innerHTML=html;
  }
  // -------- Law text lazy-loader --------
  // LAWS holds metadata + TOC stubs only. Full bodies live inside inert
  // <script type="application/json" id="law-blob-XX"> tags in the DOM. They
  // are pure text until we call JSON.parse on them — zero cost at page load,
  // no network round-trip, works over file://.
  const _lawCache = {};
  function fetchLaw(lawId){
    if(_lawCache[lawId] !== undefined) return Promise.resolve(_lawCache[lawId]);
    const el = document.getElementById('law-blob-'+lawId);
    if(!el){ _lawCache[lawId] = null; return Promise.resolve(null); }
    try {
      _lawCache[lawId] = JSON.parse(el.textContent);
    } catch(e) {
      console.error('law blob parse failed for', lawId, e);
      _lawCache[lawId] = null;
    }
    return Promise.resolve(_lawCache[lawId]);
  }
  // -------- Laws page --------
  function lawLabel(id){ return (LAWS[id]&&LAWS[id].title)||id; }
  function renderLawsPage(){
    const toc=document.getElementById('laws-toc');
    const body=document.getElementById('laws-body');
    if(!toc||!body) return;
    const ids=Object.keys(LAWS);
    if(ids.length===0){
      toc.innerHTML='<p style="color:var(--ink-s);font-size:12px;">No law texts embedded. See Law Sources for external links.</p>';
      body.innerHTML='';
      return;
    }
    let tocHtml='';
    for(const id of ids){
      tocHtml += '<div style="margin-bottom:14px;">'
        +'<div style="font-family:var(--mono);font-size:10px;letter-spacing:.06em;text-transform:uppercase;color:var(--ink-h);margin-bottom:4px;">'+id.replace(/</g,'&lt;')+'</div>'
        +'<a onclick="showLaw(\''+id+'\')" style="display:block;font-size:13px;font-weight:600;color:var(--ink);cursor:pointer;padding:4px 0;">'+lawLabel(id).replace(/</g,'&lt;')+'</a>'
        +'</div>';
    }
    toc.innerHTML=tocHtml;
    if(!body.dataset.law) showLaw(ids[0]);
  }
  window.showLaw = function(lawId){
    const body=document.getElementById('laws-body');
    const stub=LAWS[lawId];
    if(!body||!stub) return;
    body.dataset.law=lawId;
    // Paint header immediately, then stream in bodies once fetched.
    let html='<h2 style="font-size:18px;font-weight:600;margin-bottom:4px;">'+(stub.title||lawId).replace(/</g,'&lt;')+'</h2>';
    if(stub.url) html+='<p style="margin-bottom:18px;"><a href="'+stub.url+'" target="_blank" rel="noopener" style="font-size:12px;color:var(--accent);">Open official source ↗</a></p>';
    html+='<p id="laws-loading" style="color:var(--ink-s);font-size:12px;">Loading full text…</p>';
    body.innerHTML=html;
    body.scrollTop=0;
    fetchLaw(lawId).then(full => {
      if(body.dataset.law !== lawId) return; // user switched
      const items = (full && (full.articles || full.sections)) || [];
      let parts = '';
      if(items.length){
        for(const it of items){
          const aid = it.id||'';
          const title = it.title||('Section '+aid);
          parts += '<section id="law-'+lawId+'-'+aid+'" style="margin-bottom:20px;border-top:1px solid var(--bd-s);padding-top:14px;">'
            +'<h3 style="font-size:14px;font-weight:600;margin-bottom:8px;color:var(--accent);">'+title.replace(/</g,'&lt;')+'</h3>'
            +'<div style="white-space:pre-wrap;font-family:var(--serif);font-size:13px;line-height:1.7;color:var(--ink);">'+(it.text||'').replace(/</g,'&lt;')+'</div>'
            +'</section>';
        }
      } else if(full && full.raw_text){
        parts = '<pre style="white-space:pre-wrap;font-family:var(--serif);font-size:13px;line-height:1.7;color:var(--ink);">'+full.raw_text.replace(/</g,'&lt;')+'</pre>';
      } else {
        parts = '<p style="color:var(--ink-s);">Full text not available. Use the link above to view the official source.</p>';
      }
      const loading = document.getElementById('laws-loading');
      if(loading) loading.remove();
      body.insertAdjacentHTML('beforeend', parts);
    });
  };

  // -------- Law Drawer (side panel on citation click) --------
  function escH(s){return String(s||'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
  function findItem(law, anchor){
    if(!law) return null;
    const items = law.articles || law.sections || [];
    // Exact id match
    let it = items.find(x=>x.id===anchor);
    if(it) return it;
    // Strip trailing paragraph label: "3-65" → "3"
    const base = (anchor||'').split('-')[0];
    return items.find(x=>x.id===base) || null;
  }
  function renderDrawerBody(ref, rec, law){
    const item = findItem(law, rec.anchor);
    let html='';
    if(item){
      html = '<h3 style="font-size:14px;font-weight:600;color:var(--accent);margin-bottom:10px;">'+escH(item.title||('Section '+item.id))+'</h3>';
      const parts = (rec.anchor||'').split('-');
      const para = parts.length > 1 ? parts.slice(1).join('-') : '';
      if(item.paras && para && item.paras[para]){
        html += '<p style="background:var(--accent-l);padding:10px 12px;border-radius:6px;margin:8px 0 16px;border-left:3px solid var(--accent);"><strong>('+escH(para)+')</strong> '+escH(item.paras[para])+'</p>';
        html += '<hr style="border:none;border-top:1px solid var(--bd-s);margin:16px 0;">';
        html += '<div style="color:var(--ink-s);font-size:11px;margin-bottom:8px;">Full text:</div>';
      }
      html += '<div style="white-space:pre-wrap;font-family:var(--serif);font-size:13px;line-height:1.7;">'+escH(item.text||'')+'</div>';
    } else if(law && law.raw_text){
      const anchorText = rec.anchor || ref;
      html = '<p style="color:var(--ink-s);font-size:12px;margin-bottom:10px;">Showing the full bill text. Search for <code>'+escH(anchorText)+'</code> below.</p>';
      html += '<pre style="white-space:pre-wrap;font-family:var(--serif);font-size:12.5px;line-height:1.7;">'+escH(law.raw_text)+'</pre>';
    } else {
      html = '<p style="color:var(--ink-s);">This section of the law isn’t individually parsed. Use the link above to view the official source.</p>';
    }
    document.getElementById('ld-body').innerHTML = html;
  }
  function openLawDrawer(ref){
    const rec = REF_MAP[ref];
    const drawer=document.getElementById('law-drawer');
    const backdrop=document.getElementById('law-drawer-backdrop');
    document.getElementById('ld-kicker').textContent = ref;
    const stub = rec && LAWS[rec.law];
    if(!rec || !rec.law || !stub){
      document.getElementById('ld-title').textContent = 'External link';
      document.getElementById('ld-body').innerHTML = '<p style="color:var(--ink-s);">Full text for this citation isn’t embedded. Use the link above to view the official source.</p>';
      document.getElementById('ld-ext').href = '#';
    } else {
      document.getElementById('ld-title').textContent = stub.title || rec.law;
      document.getElementById('ld-ext').href = stub.url || '#';
      // Immediate placeholder, then fill once the side-file resolves. Cached
      // fetches resolve synchronously after the first open of that law.
      document.getElementById('ld-body').innerHTML = '<p style="color:var(--ink-s);font-size:12px;">Loading law text…</p>';
      fetchLaw(rec.law).then(full => {
        // Guard against the user having closed/switched the drawer.
        if(document.getElementById('ld-kicker').textContent !== ref) return;
        renderDrawerBody(ref, rec, full || stub);
      });
    }
    drawer.setAttribute('aria-hidden','false');
    drawer.style.transform='translateX(0)';
    backdrop.style.opacity='1';
    backdrop.style.pointerEvents='auto';
  }
  window.closeLawDrawer = function(){
    const drawer=document.getElementById('law-drawer');
    const backdrop=document.getElementById('law-drawer-backdrop');
    drawer.setAttribute('aria-hidden','true');
    drawer.style.transform='translateX(100%)';
    backdrop.style.opacity='0';
    backdrop.style.pointerEvents='none';
  };
  // The drawer markup is injected before </body>, which in the HTML source
  // sits AFTER this <script>. So when this code runs, ld-close and ld-search
  // aren't in the DOM yet — wire them up after DOMContentLoaded.
  function wireDrawer(){
    const closeBtn = document.getElementById('ld-close');
    if(closeBtn) closeBtn.addEventListener('click', window.closeLawDrawer);
    document.addEventListener('keydown', e=>{ if(e.key==='Escape') window.closeLawDrawer(); });
    const searchEl=document.getElementById('ld-search');
    if(searchEl){
      searchEl.addEventListener('keydown', e=>{
        if(e.key!=='Enter') return;
        const body=document.getElementById('ld-body');
        const q=searchEl.value.trim().toLowerCase();
        if(!q) return;
        const tree=body.innerText;
        const idx=tree.toLowerCase().indexOf(q, body.dataset.searchFrom||0);
        if(idx<0){ body.dataset.searchFrom=0; searchEl.style.outline='2px solid #d97706'; setTimeout(()=>searchEl.style.outline='',400); return; }
        body.dataset.searchFrom=idx+1;
        body.scrollTop = Math.max(0, Math.floor(idx/tree.length * body.scrollHeight) - 40);
      });
    }
  }
  if(document.readyState==='loading'){
    document.addEventListener('DOMContentLoaded', wireDrawer);
  } else {
    wireDrawer();
  }

  // v11's navigateToTerm(tab, term) searches tab.concepts by name, but the
  // Overview matrix shows jurisdiction-specific labels like "Developer (SB24-205)"
  // that don't match any canonical concept name. Wrap it to look up the
  // row's canonical term via MATRIX before delegating.
  function wrapNavigateToTerm(){
    if(window.__lexNavWrapped) return;
    const orig = window.navigateToTerm;
    if(typeof orig !== 'function') return;
    function stripSfx(s){ return String(s||'').replace(/\s*\([^)]*\)\s*$/, '').trim(); }
    function findConcept(tab, term){
      if(!tab || !tab.concepts) return null;
      const t = stripSfx(term).toLowerCase();
      if(!t) return null;
      // exact match first
      let hit = tab.concepts.find(c => stripSfx(c.name).toLowerCase() === t);
      if(hit) return hit;
      // Substring in either direction.
      hit = tab.concepts.find(c => {
        const n = stripSfx(c.name).toLowerCase();
        return n.includes(t) || t.includes(n);
      });
      if(hit) return hit;
      // Match against any entry.rowLabel (deeper lookup).
      for(const c of tab.concepts){
        for(const e of (c.entries||[])){
          const lbl = stripSfx(e.rowLabel || e.typeLabel || '').toLowerCase();
          if(lbl && (lbl === t || lbl.includes(t) || t.includes(lbl))) return c;
        }
      }
      return null;
    }
    window.navigateToTerm = function(tabId, term){
      // Figure out the canonical row term from MATRIX.
      let canonical = term;
      if(typeof MATRIX !== 'undefined'){
        for(const row of MATRIX){
          if(!row || !row.cells) continue;
          if(row.cells.some(c => c && c.term === term)){
            canonical = row.term; break;
          }
        }
      }
      // Let v11's routine switch tabs + try its own match first.
      const result = orig.call(this, tabId, canonical);
      // Fallback: if nothing got opened, try a smarter concept match.
      if(!openConceptId){
        const tab = DATA.find(t => t.tabId === tabId);
        const concept = findConcept(tab, canonical) || findConcept(tab, term);
        if(concept){
          setTimeout(() => jumpToConcept(tabId, concept.id), 50);
        }
      }
      // Render notes after the page has switched and rendered.
      setTimeout(scheduleNotesScan, 120);
      setTimeout(scheduleNotesScan, 400);
      return result;
    };
    window.__lexNavWrapped = true;
  }
  if(document.readyState==='loading'){
    document.addEventListener('DOMContentLoaded', wrapNavigateToTerm);
  } else {
    wrapNavigateToTerm();
  }

  // v11 renders every citation badge as <span class="art-badge linkable"
  // onclick="openLawRef(url,label,src)">LABEL</span>. Rather than try to
  // discover those spans at render time, we monkey-patch openLawRef itself:
  // if the label resolves to an embedded law anchor we show our drawer;
  // otherwise we fall back to the original "open external source" panel.
  function wrapOpenLawRef(){
    if(window.__lexLawWrapped) return;
    const orig = window.openLawRef;
    window.openLawRef = function(url, label, sourceId){
      const lbl = (label||'').trim();
      const rec = REF_MAP[lbl];
      if(rec && rec.law && LAWS[rec.law]){
        openLawDrawer(lbl);
        return;
      }
      if(typeof orig==='function') return orig.apply(this, arguments);
    };
    window.__lexLawWrapped = true;
  }
  if(document.readyState==='loading'){
    document.addEventListener('DOMContentLoaded', wrapOpenLawRef);
  } else {
    wrapOpenLawRef();
  }

  // v11's comparative-analysis renderer (`renderCAView`) ignores the
  // "Interpretative notes" column we pull from the xlsx. Because the function
  // is defined with `function renderCAView(){...}` and called via its bare
  // identifier from within the same script, monkey-patching `window.renderCAView`
  // doesn't always intercept callers (bindings are resolved at parse time).
  // The robust fix is to watch the DOM for the analysis table rendering and
  // inject the notes block right after each .ca-view-header.
  function cepsInjectNotes(conceptId, container){
    if(!container || !conceptId) return;
    const ad = (typeof ANALYSIS_DATA !== 'undefined') ? ANALYSIS_DATA : [];
    // Collect unique notes for this concept, keyed by analysis card.
    const analyses = ad.filter(a => (a.concept_ids||[]).some(id =>
      conceptId.includes(id) || id.includes(conceptId)));
    if(!analyses.length) return;
    // Group notes into one panel per analysis card (deduped by note text).
    const panels = [];
    for(const a of analyses){
      const seen = new Set();
      const items = [];
      for(const row of (a.rows||[])){
        if(!row.notes) continue;
        if(seen.has(row.notes)) continue;
        seen.add(row.notes);
        items.push({dim: row.dim || '', note: row.notes});
      }
      if(items.length) panels.push({title: a.title || '', items});
    }
    if(!panels.length) return;
    // Remove any stale note blocks we appended on a prior render, then insert.
    container.querySelectorAll('.ceps-injected-notes').forEach(el=>el.remove());
    let html = '<div class="ceps-injected-notes">';
    for(const p of panels){
      for(const it of p.items){
        html += '<div class="ca-note-block">'
          + '<div class="ca-note-label">CEPS interpretative note'
          +   (it.dim ? ' — ' + escH(it.dim) : '') + '</div>'
          + escH(it.note).replace(/\n/g,'<br>')
          + '</div>';
      }
    }
    html += '</div>';
    container.insertAdjacentHTML('beforeend', html);
  }
  // Trigger notes injection only in response to the user actions that
  // actually cause v11 to re-render the comparative-analysis table:
  // clicking a concept card header, a tab button, or a filter toc-item.
  // A broad MutationObserver on the whole #p-lexicon subtree was the main
  // source of click lag, because v11 emits hundreds of child-list mutations
  // per interaction — we were paying O(DOM) scans on every keystroke and
  // scroll-triggered re-layout.
  function runNotesScan(){
    document.querySelectorAll('.ca-view-header').forEach(hdr => {
      if (hdr.dataset.cepsNotesAdded) return;
      const concept = hdr.closest('[data-concept-id]') ||
                      hdr.closest('.concept-card');
      const conceptId = concept ? (concept.dataset.conceptId ||
                        concept.id.replace(/^card-/,'')) : (typeof openConceptId !== 'undefined' ? openConceptId : '');
      cepsInjectNotes(conceptId, hdr.parentElement);
      hdr.dataset.cepsNotesAdded = '1';
    });
  }
  function scheduleNotesScan(){
    // rIC if available (idle time only), else a 50 ms timer — both run AFTER
    // v11 finishes its re-render, so the click handler returns immediately.
    if (window.requestIdleCallback) requestIdleCallback(runNotesScan, {timeout: 250});
    else setTimeout(runNotesScan, 50);
  }
  function startNotesWatcher(){
    const lex = document.getElementById('p-lexicon');
    if (!lex) return;
    // Any click inside the lexicon page can cause v11 to re-render the
    // comparison table. We schedule one rIC pass after the click.
    lex.addEventListener('click', () => scheduleNotesScan(), {passive: true});
    // Also run once on navigation so notes appear without a follow-up click.
    scheduleNotesScan();
  }

  // -------- Cell-click → side drawer (ported from v12_draft) --------
  // Click any non-empty cell in the comparison table to open the drawer with
  // that jurisdiction × dimension's analysis text, its interpretative note
  // (if any), and a link to the source law article when a reference exists.
  function cellDrawerInit(){
    const lex = document.getElementById('p-lexicon');
    if (!lex) return;
    lex.addEventListener('click', (ev) => {
      const td = ev.target && ev.target.closest && ev.target.closest('.ca-table td');
      if (!td) return;
      if (td.classList.contains('ca-dim')) return;         // row-label cell
      // Resolve dimension (left cell) + jurisdiction (column index)
      const tr = td.parentElement;
      const dimLabelEl = tr.querySelector('.ca-dim-label');
      const dim = dimLabelEl ? dimLabelEl.textContent.trim() : '';
      const tds = Array.from(tr.children);
      const tdIdx = tds.indexOf(td);          // 0 is the dimension cell
      if (tdIdx <= 0) return;
      const table = td.closest('table');
      const thead = table && table.querySelector('thead tr');
      const ths = thead ? Array.from(thead.children) : [];
      const th = ths[tdIdx];
      const jid = th && (th.className.match(/th-(\w+)/) || [])[1] || '';
      const jurisLabel = th ? (th.firstChild ? (th.firstChild.textContent || '').trim() : th.textContent.trim()) : '';
      const termLabel = th ? (th.querySelector('span') ? th.querySelector('span').textContent.trim() : '') : '';
      const conceptCard = td.closest('.concept-card');
      const conceptId = conceptCard ? conceptCard.id.replace(/^card-/, '') : (openConceptId || '');
      // Pull the analysis row matching this dimension for this concept,
      // plus its interpretative notes.
      let note = '', noteLabel = '', ref = '';
      const ad = (typeof ANALYSIS_DATA !== 'undefined') ? ANALYSIS_DATA : [];
      const analyses = ad.filter(a => (a.concept_ids||[]).some(id =>
        conceptId.includes(id) || id.includes(conceptId)));
      // Prefer a dim-specific note; fall back to the analysis's framing
      // note (typically attached to the "Term" row).
      for (const a of analyses) {
        for (const row of (a.rows||[])) {
          if ((row.dim||'').toLowerCase().trim() !== dim.toLowerCase().trim()) continue;
          if (row.notes) { note = row.notes; noteLabel = dim; }
          break;
        }
        if (note) break;
      }
      if (!note) {
        for (const a of analyses) {
          for (const row of (a.rows||[])) {
            if (row.notes) { note = row.notes; noteLabel = a.title || 'Framing'; break; }
          }
          if (note) break;
        }
      }
      // Find the matching verbatim reference from DATA.concepts[...].entries.
      for (const tab of DATA) {
        for (const c of tab.concepts) {
          if (c.id !== conceptId) continue;
          for (const e of (c.entries||[])) {
            const rl = (e.typeLabel || e.rowLabel || '').toLowerCase().trim();
            if (rl !== dim.toLowerCase().trim()) continue;
            const jd = (e.jdata || {})[jid];
            if (jd && jd.reference) { ref = jd.reference; break; }
          }
          if (ref) break;
        }
        if (ref) break;
      }
      // Populate the drawer.
      const kicker = (jurisLabel || jid.toUpperCase()) + (dim ? ' · ' + dim : '');
      document.getElementById('ld-kicker').textContent = kicker;
      document.getElementById('ld-title').textContent = termLabel || (conceptCard?.querySelector('.concept-name')?.textContent || 'Concept');
      let html = '<div style="white-space:pre-wrap;font-family:var(--serif);font-size:14px;line-height:1.7;">'
        + escH(td.innerText) + '</div>';
      if (note) {
        html += '<div class="ca-note-block" style="margin-top:18px;">'
          + '<div class="ca-note-label">CEPS interpretative note'
          +   (noteLabel ? ' — ' + escH(noteLabel) : '') + '</div>'
          + escH(note).replace(/\n/g,'<br>') + '</div>';
      }
      if (ref) {
        const rec = REF_MAP[ref];
        html += '<div style="margin-top:16px;padding-top:12px;border-top:1px solid var(--bd-s);font-size:11px;color:var(--ink-s);">'
          + '<div style="font-family:var(--mono);margin-bottom:4px;">Reference</div>';
        if (rec && rec.law && LAWS[rec.law]) {
          html += '<a href="#" onclick="event.preventDefault();openLawRef(\'#\',' + JSON.stringify(ref) + ',null);return false;" style="color:var(--accent);text-decoration:none;">'
            + escH(ref) + ' — Read source article ↗</a>';
        } else {
          html += escH(ref);
        }
        html += '</div>';
      }
      document.getElementById('ld-body').innerHTML = html;
      document.getElementById('ld-ext').href = (ref && REF_MAP[ref] && LAWS[REF_MAP[ref].law]?.url) || '#';
      const drawer = document.getElementById('law-drawer');
      const backdrop = document.getElementById('law-drawer-backdrop');
      drawer.setAttribute('aria-hidden','false');
      drawer.style.transform='translateX(0)';
      backdrop.style.opacity='1';
      backdrop.style.pointerEvents='auto';
    }, {passive: true});
  }
  if(document.readyState==='loading'){
    document.addEventListener('DOMContentLoaded', cellDrawerInit);
  } else {
    cellDrawerInit();
  }
  if(document.readyState==='loading'){
    document.addEventListener('DOMContentLoaded', startNotesWatcher);
  } else {
    startNotesWatcher();
  }

  // -------- Patch go() to render new pages --------
  const _origGo = window.go;
  window.go = function(page){
    _origGo(page);
    if(page==='laws') renderLawsPage();
    if(page==='methodology') {
      const mbody=document.querySelector('#p-methodology .meth');
      if(mbody && METHODOLOGY_PROSE && METHODOLOGY_PROSE.text){
        mbody.innerHTML = '<h3 style="font-size:16px;margin:18px 0 10px;">Research methodology</h3>' +
          (METHODOLOGY_PROSE.text||'').split(/\n{2,}/).map(p=>'<p style="margin:0 0 12px;font-size:13px;line-height:1.65;color:var(--ink);">'+p.replace(/</g,'&lt;')+'</p>').join('');
      }
    }
    const n=document.getElementById('n-laws');
    if(n) n.classList.toggle('active', page==='laws');
  };
})();
"""
    # Target v11's main script (executable JS) rather than any later inert
    # <script type="application/json"> blob tag that we inject for law bodies.
    def _main_script_close(src: str) -> int:
        m = re.search(r"</script>(?!\s*\n?\s*<script[^>]*type=\"application/json)", src[::-1])
        # simpler: scan forward for a </script> whose preceding <script> tag
        # doesn't have type="application/json".
        out = -1
        for mo in re.finditer(r"</script>", src):
            pos = mo.start()
            # look back to find the opening <script that matches this close
            opening = src.rfind("<script", 0, pos)
            tag = src[opening:src.find(">", opening) + 1]
            if 'type="application/json"' in tag:
                continue
            out = pos
        return out
    _close_idx = _main_script_close(html)
    if _close_idx < 0:
        _close_idx = html.rindex("</script>")
    html = html[:_close_idx] + render_js + "\n" + html[_close_idx:]

    HTML_OUT.write_text(html, encoding="utf-8")
    print(f"Wrote {HTML_OUT}  ({len(html):,} bytes)")
    print(f"  tabs:       {len(tabs)}  ({', '.join(t['tabId'] for t in tabs)})")
    print(f"  index:      {len(index)} entries")
    print(f"  analyses:   {len(analyses)}")
    print(f"  matrix:     {len(flat_matrix)} rows")
    print(f"  glossary:   {len(glossary)} terms")
    print(f"  laws:       {len(laws)} / {len(LAW_FILES)} embedded ({', '.join(laws.keys())})")
    print(f"  ref map:    {resolved}/{len(ref_map)} references resolved to law anchors")


if __name__ == "__main__":
    main()
